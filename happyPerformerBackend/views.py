<<<<<<< HEAD
import json
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseNotAllowed , JsonResponse, HttpResponseBadRequest
from .models import *
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.db import transaction
from django.core import serializers
from django.db import connection, IntegrityError
from django.utils import timezone
from datetime import datetime
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.contrib.auth.decorators import login_required
from .decorators import role_required
from django.db.models import Q
from django.db.models import Sum
from django.core.files.uploadedfile import UploadedFile
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.core.exceptions import ObjectDoesNotExist

import openpyxl
from django.core.files.storage import default_storage
import pandas as pd
from django.db.models import Max
from datetime import timedelta
import traceback



import logging
#added
from datetime import timedelta
from functools import wraps
from django.utils.decorators import method_decorator
import random, string
from django.contrib.auth.hashers import make_password
from django.utils.dateparse import parse_datetime
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import os

# decorators
# def role_required(allowed_roles):
#     def decorator(view_func):
#         @wraps(view_func)
#         def _wrapped_view(request, *args, **kwargs):
#             if not request.user.is_authenticated:
#                 return JsonResponse({'error': 'Authentication required'}, status=401)
            
#             user_role = getattr(request.user, 'emp_role', None)
#             if user_role not in allowed_roles:
#                 return JsonResponse({'error': 'Permission denied'}, status=403)
            
#             return view_func(request, *args, **kwargs)
#         return _wrapped_view
#     return decorator

from django.core.files.storage import default_storage
from datetime import timedelta
from functools import wraps
from django.utils.decorators import method_decorator
import random, string
from django.contrib.auth.hashers import make_password
import urllib.parse
import logging
from django.core.serializers import serialize
from rest_framework.response import Response
from rest_framework.decorators import api_view





# Set up logging
logger = logging.getLogger(__name__)



@csrf_exempt
def Home(request):
    return JsonResponse({"data":"Welcome to happy backsideeee hehe"})


@csrf_exempt
def MeetTheTeam(request):
    return JsonResponse({"data":"Meet the team here!"})

@csrf_exempt
def Contact(request):
    return JsonResponse({"data":"Contact Us Page reached without any issues"})

@csrf_exempt
def About(request):
    return JsonResponse({"data":"About Us Page reached without any issues"})

@csrf_exempt
def TermsAndConditions(request):
    return JsonResponse({"data":"Terms and Conditions Page reached without any issues"})

# Working login without checking department
# @csrf_exempt
# def Login(request):
#     if request.method == 'POST':
#         if request.session.get('user_id'):
#             return JsonResponse({'error': 'User already logged in'}, status=403)  # Changed to 403

#         data = json.loads(request.body)
#         emp_emailid = data.get('email')
#         emp_pwd = data.get('password')

#         # Validate input
#         if not emp_emailid or not emp_pwd:
#             return JsonResponse({'error': 'Email and password are required'}, status=400)

#         try:
#             # Check if the employee exists with the provided email and password
#             user = Employee.objects.get(emp_emailid=emp_emailid, emp_pwd=emp_pwd)

#             # Set session data
#             department = user.d_id
#             company = department.c_id

#             request.session['user_id'] = user.emp_emailid
#             request.session['emp_name'] = user.emp_name
#             request.session['emp_emailid'] = user.emp_emailid
#             request.session['emp_role'] = user.emp_role
#             request.session['d_id'] = department.d_id
#             request.session['c_id'] = company.c_id
#             request.session['c_name'] = company.c_name

#             request.session.save()

#             # Prepare response data
#             profile_url = settings.MEDIA_URL + str(user.emp_profile) if user.emp_profile else None

#             response_data = {
#                 'message': 'Login successful',
#                 'user_id': user.emp_emailid,
#                 'emp_name': user.emp_name,
#                 'emp_emailid': user.emp_emailid,
#                 'emp_role': user.emp_role,
#                 'd_id': department.d_id,
#                 'c_id': company.c_id,
#                 'c_name': company.c_name,
#                 'profile_url': profile_url,
#             }

#             return JsonResponse(response_data, status=200)

#         except Employee.DoesNotExist:
#             return JsonResponse({'error': 'Invalid email or password'}, status=401)
#     else:
#         return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
@csrf_exempt
def Login(request):
    if request.method == 'POST':
        if request.session.get('user_id'):
            return JsonResponse({'message': 'User already logged in'}, status=200)

        data = json.loads(request.body)
        emp_emailid = data.get('email')
        emp_pwd = data.get('password')
        emp_role = data.get('department')  # Get department (emp_role) from the request

        # Validate input
        if not emp_emailid or not emp_pwd or not emp_role:
            return JsonResponse({'error': 'Email, password, and department are required'}, status=400)

        try:
            # Check if the employee exists with the provided email, password, and emp_role
            user = Employee.objects.get(emp_emailid=emp_emailid, emp_pwd=emp_pwd, emp_role=emp_role)

            # Set session data
            department = user.d_id
            company = department.c_id

            request.session['user_id'] = user.emp_emailid
            request.session['emp_name'] = user.emp_name
            request.session['emp_emailid'] = user.emp_emailid
            request.session['emp_role'] = user.emp_role
            request.session['d_id'] = department.d_id
            request.session['c_id'] = company.c_id
            request.session['c_name'] = company.c_name

            request.session.save()

            # Prepare response data
            profile_url = settings.MEDIA_URL + str(user.emp_profile) if user.emp_profile else None

            response_data = {
                'message': 'Login successful',
                'user_id': user.emp_emailid,
                'emp_name': user.emp_name,
                'emp_emailid': user.emp_emailid,
                'emp_role': user.emp_role,
                'd_id': department.d_id,
                'c_id': company.c_id,
                'c_name': company.c_name,
                'profile_url': profile_url,
            }

            return JsonResponse(response_data, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Invalid email, password, or department'}, status=401)
    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)

from django.http import JsonResponse
from .models import Employee

@csrf_exempt
def GetDepartments(request):
    if request.method == 'GET':
        # Fetch unique emp_role values from the Employee table
        departments = Employee.objects.values_list('emp_role', flat=True).distinct()
        departments_data = [{"d_id": idx, "d_name": dept} for idx, dept in enumerate(departments)]
        return JsonResponse({"departments": departments_data}, status=200)
    else:
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)


def Users(request):
    company_id = request.session.get('c_id')
    user_id = request.session.get('user_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    employees = Employee.objects.filter(d_id__c_id=company_id, emp_emailid=user_id).values(
        'emp_name', 'emp_emailid', 'emp_phone', 'emp_role', 'd_id', 'emp_profile'
    )
    data = {
        'employees': list(employees)
    }

    return JsonResponse(data)


@csrf_exempt
def Logout(request):
    if request.method == 'POST':
        if 'user_id' in request.session:
            request.session.flush()
            return JsonResponse({'message': 'Logout successful'}, status=200)
        else:
            return JsonResponse({'error': 'No user is logged in'}, status=400)
    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


# NOTE
# ***Department name should contain Super Manager at last ***

    

    
# @csrf_exempt
# def Register(request):
#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)

#             # Extract company details
#             name = data.get('companyName')
#             addr = data.get('companyAddress')
#             phone = data.get('companyPhone')
#             dept_names_str = data.get('deptName')
#             dept_names = [name.strip() for name in dept_names_str.split(',')]

#             # Extract employee details
#             emp_name = data.get('empName')
#             emp_email = data.get('empMail')
#             emp_phone = data.get('empNum')
#             emp_pwd = data.get('emp_pwd', 'changeme')  # Default to 'changeme' if not provided
#             emp_role = data.get('empRole', 'Super Manager')  # Default to 'Super Manager' if not provided
#             emp_skills = data.get('empSkills')

#             # Check if the company with the exact details exists
#             try:
#                 company = Company.objects.get(c_name=name, c_addr=addr, c_phone=phone)
#                 created = False
#             except Company.DoesNotExist:
#                 company = Company.objects.create(c_name=name, c_addr=addr, c_phone=phone)
#                 created = True

#             # Only create departments if the company is newly created
#             if created:
#                 # Create departments and get the first department's ID
#                 first_dept_id = None
#                 for dept_name in dept_names:
#                     # Create department with the Company instance
#                     new_dept = Department.objects.create(d_name=dept_name, c_id=company)
#                     if first_dept_id is None:
#                         first_dept_id = new_dept
#             else:
#                 # If the company already exists, use the existing first department ID
#                 first_dept_id = Department.objects.filter(c_id=company).first()
                
#                 if not first_dept_id:
#                     return JsonResponse({'error': 'No departments found for existing company'}, status=404)

#             # Create the employee
#             Employee.objects.create(
#                 emp_name=emp_name,
#                 emp_emailid=emp_email,
#                 emp_skills=emp_skills,
#                 emp_role=emp_role,
#                 emp_phone=emp_phone,
#                 emp_pwd=emp_pwd,
#                 d_id=first_dept_id
#             )

#             return JsonResponse({'message': 'Company registration successful'}, status=201)

#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

#     else:
#         return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
    
@csrf_exempt
def Register(request):
    if request.method == 'POST':
        
        try:
            data = json.loads(request.body)

            # Extract company details
            name = data.get('companyName')
            addr = data.get('companyAddress')
            phone = data.get('companyPhone')
            dept_names_str = data.get('deptName')
            dept_names = [name.strip() for name in dept_names_str.split(',')]

            # Extract employee details
            emp_name = data.get('empName')
            emp_email = data.get('empMail')
            emp_phone = data.get('empNum')
            emp_pwd = data.get('empPwd')
            emp_role = data.get('deptName')
            emp_skills = data.get('empSkills')

            # Capture the client's IP address
            #office_ip = request.META.get('REMOTE_ADDR')

            # Check if the company with the exact details exists
            try:
                company = Company.objects.get(c_name=name, c_addr=addr, c_phone=phone)
                created = False
            except Company.DoesNotExist:
                # Create a new company with the fetched IP address
                company = Company.objects.create(c_name=name, c_addr=addr, c_phone=phone)
                created = True
            # Only create departments if the company is newly created
            if created:
                first_dept_id = None
                for dept_name in dept_names:
                    new_dept = Department.objects.create(d_name=dept_name, c_id=company)
                    if first_dept_id is None:
                        first_dept_id = new_dept
            else:
                first_dept_id = Department.objects.filter(c_id=company).first()
                
                if not first_dept_id:
                    return JsonResponse({'error': 'No departments found for existing company'}, status=404)

            # Create the employee
            Employee.objects.create(
                emp_name=emp_name,
                emp_emailid=emp_email,
                emp_skills=emp_skills,
                emp_role=emp_role,
                emp_phone=emp_phone,
                emp_pwd=emp_pwd,
                d_id=first_dept_id
            )
            return JsonResponse({'message': 'Company registration successful'}, status=201)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

            return JsonResponse({'message': 'Company registration successful'}, status=201)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)






# incomplete , need to complete and add session
@csrf_exempt
def Profile(request, id):
    try:
        personal_details =  personal_details.objects.filter(mail=id).values()
        if not personal_details.exists():
            return JsonResponse({'error': 'Employee Details not found'}, status=404)
        return JsonResponse(list(personal_details), safe=False)
    except Exception as e:
        print("Error fetching employee data:", e)
        return JsonResponse({'error': 'Internal Server Error'}, status=500)


@csrf_exempt
def SopAndPolicies(request):
    # sop and policies display page. to be visible to everyone
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if user_id and company_id:
        tasks_data = Tasks.objects.filter(emp_emailid=user_id).values('ratings', 'remark', 'sop_id', 'selfratings')

        sop_ids = [task['sop_id'] for task in tasks_data]

        sop_data = Sop.objects.filter(sop_id__in=sop_ids).values('sop_id', 'type', 's_name', 'sdate')
        sop_data_dict = {sop['sop_id']: sop for sop in sop_data}

        files_data = Files.objects.filter(sop_id__in=sop_ids).values('file_name', 'sop_id')
        sop_files_dict = {}
        for file in files_data:
            sop_id = file['sop_id']
            if sop_id not in sop_files_dict:
                sop_files_dict[sop_id] = []
            sop_files_dict[sop_id].append(file['file_name'])

        combined_data = []
        for task in tasks_data:
            sop_id = task['sop_id']
            sop_info = sop_data_dict.get(sop_id, {})
            files_info = sop_files_dict.get(sop_id, [])

            combined_task = {
                'ratings': task['ratings'],
                'remark': task['remark'],
                'sop_id': sop_id,
                'selfratings': task['selfratings'],
                'sop_info': sop_info,
                'files': files_info
            }
            combined_data.append(combined_task)

        return JsonResponse({'data': 
            combined_data}, safe=False)
    else:
        return JsonResponse({'error': 'User not logged in'}, status=401)


@csrf_exempt
@require_http_methods(["PUT"])
def UpdateSelfratings(request, sop_id):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    try:
        task = Tasks.objects.get(sop_id=sop_id, emp_emailid=user_id)

        data = json.loads(request.body)
        self_rating = int(data.get('selfratings'))
        task.selfratings = self_rating
        task.save()

        return JsonResponse({'message': 'Rating updated successfully'}, status=200)
    except Tasks.DoesNotExist:
        return JsonResponse({'error': 'Task does not exist'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def KraList(request):
    if request.method == 'GET':
        kras = Kra.objects.all().values('kra_no', 'kra', 'weightage', 'kpi', 'target', 'ratings_scale', 'ratings', 'selfratings', 'remarks', 'status', 'email_id', 'kra_id')
        kra_list = list(kras)
        return JsonResponse(kra_list, safe=False)
    else:
        return JsonResponse({'error': 'Invalid HTTP method'}, status=400)


@csrf_exempt
def Forms(request):
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)
    user_name = request.session.get('emp_name')

    if not c_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    if request.method == 'GET':
        try:
            employee = get_object_or_404(Employee, emp_name=user_name, d_id__c_id=company_id)

            allocated_forms = []

            forms = Custom_forms.objects.filter(c_id=company_id)

            for form in forms:
                allocated_users = [user.strip() for user in form.alloc.split(',') if user.strip()]
                if allocated_users and user_name in allocated_users:
                    formatted_form_name = form.form_name.replace(' ', '_').lower()
                    response_table_name = f"resp_{company_id}{formatted_form_name}"

                    with connection.cursor() as cursor:
                        cursor.execute(f"SELECT COUNT(*) FROM {response_table_name} WHERE emp_name = %s", [user_name])
                        responded = cursor.fetchone()[0] > 0

                    allocated_forms.append({
                        'form_name': form.form_name,
                        'responded': responded
                    })

            return JsonResponse({'forms': allocated_forms}, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def FormReviewRespose(request):
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)
    user_name = request.session.get('emp_name')

    if not c_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    if request.method == 'GET':
        try:
            form_name = request.GET.get('form_name')

            if not form_name:
                return JsonResponse({'error': 'Form name is required'}, status=400)

            form = Custom_forms.objects.get(c_id=company_id, form_name = form_name)

            formatted_form_name = form.form_name.replace(' ', '_').lower()
            response_table_name = f"resp_{company_id}{formatted_form_name}"
            response_data = []

            with connection.cursor() as cursor:
                cursor.execute(f"SELECT * FROM {response_table_name} WHERE emp_name = %s", [user_name])
                response_row = cursor.fetchone()
                if not response_row:
                    return JsonResponse({'error': 'Response not found'}, status=404)

                columns = [col[0] for col in cursor.description]
                for col, val in zip(columns, response_row):
                    if col != 'emp_name':
                        cursor.execute(f"SELECT label FROM Custom_forms_questions WHERE ID = %s AND form_name = %s AND c_id = %s", [col, form_name, c_id])
                        question = cursor.fetchone()
                        if question:
                            response_data.append({'question': question[0], 'answer': val})

            return JsonResponse({'response': response_data}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def FormsSubmitResponse(request):
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)
    user_name = request.session.get('emp_name')

    if not c_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            form_name = data.get('form_name')

            if not form_name:
                return JsonResponse({'error': 'Form name is required'}, status=400)

            form = Custom_forms.objects.get(c_id=company_id, form_name = form_name)

            formatted_form_name = form.form_name.replace(' ', '_').lower()
            response_table_name = f"resp_{company_id}{formatted_form_name}"

            columns = ', '.join(data.keys())
            values = ', '.join(['%s'] * len(data))
            query = f"INSERT INTO {response_table_name} ({columns}) VALUES ({values})"
            params = list(data.values())

            with connection.cursor() as cursor:
                cursor.execute(query, params)

            return JsonResponse({'message': 'Response submitted successfully'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def Letters(request):
    if 'emp_emailid' in request.session and 'c_id' in request.session:
        emp_emailid = request.session['emp_emailid']
        c_id = request.session['c_id']

        try:
            user_name = Employee.objects.get(emp_emailid=emp_emailid).emp_name
            company = get_object_or_404(Company, pk=c_id)

            allocated_letters = []

            custom_letters = Custom_letters.objects.filter(c_id=company)

            for letter in custom_letters:
                allocated_emails = [email.strip() for email in letter.alloc.split(',') if email.strip()]
                if allocated_emails and emp_emailid in allocated_emails:
                    allocated_letters.append({
                        'letter_name': letter.letter_name.replace('_', ' ').title()
                    })

            return JsonResponse({'allocated_letters': allocated_letters})

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Required session data not found'}, status=401)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin', 'Employee'])
def FAQsView(request):
    c_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')

    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            faqs = Faqs.objects.filter(c_id=c_id).values('faq_id', 'question', 'answer', 'emp_emailid', 'imp')
            data = list(faqs)
            return JsonResponse({'faqs': data})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            question = data.get('question')
            emp_emailid = get_object_or_404(Employee, emp_emailid=emp_emailid)
            c_id = get_object_or_404(Company, c_id=c_id)

            with transaction.atomic():
                new_faq = Faqs.objects.create(
                    c_id=c_id,
                    emp_emailid=emp_emailid,
                    question=question,
                    answer='',
                    imp=False
                )
                return JsonResponse({'success': 'FAQ added successfully', 'faq_id': new_faq.faq_id}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

@csrf_exempt
def ApplyLeave(request):
    if request.method == 'POST':
        try:
            # Parse the request body to get the data
            data = json.loads(request.body)
            leavetype = data.get('leaveType')
            fromdate = data.get('fromDate')
            todate = data.get('toDate')
            description = data.get('leaveDescription')

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        # Check if user is authenticated through session
        emp_email = request.session.get('emp_emailid')
        if not emp_email:
            return JsonResponse({'error': 'User not authenticated'}, status=401)

        try:
            # Retrieve employee record based on email
            employee = Employee.objects.get(emp_emailid=emp_email)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

        try:
            # Parse and validate the date formats
            date1 = datetime.strptime(fromdate, "%Y-%m-%d")
            date2 = datetime.strptime(todate, "%Y-%m-%d")

            # Validate that fromdate is not later than todate
            if date1 > date2:
                return JsonResponse({'error': 'From date cannot be later than To date'}, status=400)

            days = (date2 - date1).days + 1  # Include both start and end dates
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)

        try:
            # Access the company through the employee's department
            leave_type = Leavetype.objects.get(LeaveType__iexact=leavetype, company_id=employee.d_id.c_id)


        except Leavetype.DoesNotExist:
            return JsonResponse({'error': 'Leave type does not exist'}, status=400)

        # Check if the Leavecounttemp record exists
        if Leavecounttemp.objects.filter(emp_emailid=emp_email).exists():
            leave_count = Leavecounttemp.objects.get(emp_emailid=emp_email)
        else:
            # Create a new Leavecounttemp record with default values
            leave_count = Leavecounttemp.objects.create(
                emp_emailid=emp_email,
                casualleave=15,
                medicalleave=15,
                lopleave=365,
                earnedleave=20  # Set default values as needed
            )

        # Mapping the leave type to the correct field
        leave_type_mapping = {
            'Casual Leave': 'casualleave',
            'Medical Leave': 'medicalleave',
            'LOP Leave': 'lopleave',
            'Earned Leave': 'earnedleave'
        }

        leave_limit_field = leave_type_mapping.get(leavetype, '').lower()

        # Validate if the field exists in the Leavecounttemp model
        if not leave_limit_field:
            return JsonResponse({'error': 'Invalid leave type provided'}, status=400)

        leave_limit = getattr(leave_count, leave_limit_field, 0)

        final = leave_limit - days
        if final < 0:
            return JsonResponse({'error': 'Exceeding leave limits'}, status=400)

        # Create the leave record
        leave = Tblleaves.objects.create(
            LeaveType=leave_type,
            FromDate=fromdate,
            ToDate=todate,
            Days=days,
            Description=description,
            Status=0,
            IsRead=0,
            emp_emailid=employee
        )

        return JsonResponse({'message': 'Leave submitted successfully'}, status=200)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)





@csrf_exempt
def LeaveHistory(request):
    emp_emailid = request.session.get('emp_emailid')

    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    leaves = Tblleaves.objects.filter(emp_emailid=emp_emailid).values('id', 'LeaveType__LeaveType', 'FromDate', 'ToDate', 'PostingDate', 'AdminRemark', 'Status')
    leaves_data = list(leaves)
    print(leaves_data)

    return JsonResponse({'Leaves': leaves_data}, status=200)


@csrf_exempt
def AddLoan(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'Employee email not found in session'}, status=400)

    if request.method == 'GET':
        loans = Loan.objects.filter(emp_emailid=emp_emailid).values('name', 'department', 'lamt', 'mamt', 'startdate', 'reason', 'status')

        data = []
        for index, loan in enumerate(loans, start=1):
            status = "Pending" if loan['status'] == 0 else "Approved"
            data.append({
                'name': loan['name'],
                'department': loan['department'],
                'loan_amount': loan['lamt'],
                'monthly_amount': loan['mamt'],
                'start_date': loan['startdate'],
                'reason': loan['reason'],
                'status': status
            })

        return JsonResponse({'loans': data})

    elif request.method == 'POST':
        try:
            loan_data = json.loads(request.body)
            name = loan_data.get('name')
            department = loan_data.get('department')
            lamt = loan_data.get('lamt')
            mamt = loan_data.get('mamt')
            startdate = loan_data.get('startdate')
            reason = loan_data.get('reason')
            status = loan_data.get('status')

            Loan.objects.create(name=name, emp_emailid=emp_emailid, department=department, lamt=lamt, mamt=mamt, startdate=startdate, reason=reason, status=status)

            return JsonResponse({'message': 'Loan applied successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddExpenses(request):
    emp_emailid = request.session.get('emp_emailid')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email not found in session'}, status=400)

    if request.method == 'POST':
        expense_data = json.loads(request.body)

        expense = expense_data.get('expense')
        expensedate = expense_data.get('expensedate')
        expenseitm = expense_data.get('expenseitm')

        Expenses.objects.create(emp_emailid=emp_emailid, expense=expense, expensedate=expensedate, expenseitm=expenseitm)

        return JsonResponse({'message': 'Expense added successfully'})

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin', 'Employee'])
def ManageExpenses(request):
    emp_emailid = request.session.get('emp_emailid')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email not found in session'}, status=400)

    if request.method == 'GET':
        expenses = Expenses.objects.filter(emp_emailid=emp_emailid).values('expensedate', 'expenseitm', 'expense_id').annotate(total_expense=Sum('expense')).order_by('expensedate')

        data = []
        for expense in expenses:
            data.append({
                'date': expense['expensedate'].strftime('%Y-%m-%d'),
                'amount': expense['total_expense'],
                'expense_item': expense['expenseitm'],
                'expense_id': expense['expense_id']
            })

        return JsonResponse({'expenses': data})

    elif request.method == 'POST':
        expense_data = json.loads(request.body)
        expense_id = request.GET.get('expense_id')

        if expense_id is not None:
            expense = get_object_or_404(Expenses, expense_id=expense_id, emp_emailid=emp_emailid)
            expense.expensedate = expense_data.get('date', expense.expensedate)
            expense.expense = expense_data.get('amount', expense.expense)
            expense.expenseitm = expense_data.get('expense_item', expense.expenseitm)
            expense.save()
            return JsonResponse({'message': 'Expense updated successfully'})
        else:
            return JsonResponse({'error': 'Expense ID not provided'}, status=400)


    elif request.method == 'DELETE':
        expense_id = request.GET.get('expense_id')
        expense = get_object_or_404(Expenses, expense_id=expense_id, emp_emailid=emp_emailid)
        expense.delete()
        return JsonResponse({'message': 'Expense deleted successfully'})

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ExpenseReport(request):
    emp_emailid = request.session.get('emp_emailid')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email not found in session'}, status=400)

    try:
        expenses = Expenses.objects.filter(emp_emailid=emp_emailid).values('expensedate', 'expenseitm', 'expense_id').annotate(total_expense=Sum('expense')).order_by('expensedate')

        data = []
        for expense in expenses:
            data.append({
                'date': expense['expensedate'].strftime('%Y-%m-%d'),
                'amount': expense['total_expense'],
                'expense_item': expense['expenseitm'],
                'expense_id': expense['expense_id']
            })

        return JsonResponse({'expenses': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def Resign(request):
    c_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')

    if not c_id or not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    email = get_object_or_404(Resignation, emp_emailid=emp_emailid)

    if request.method == 'GET':
        resignations = Resignation.objects.all()
        print(resignations)
        resignation_data = []
        for resignation in resignations:
            resignation_data.append({
                'emp_emailid': email,
                'submit_date': resignation.submit_date,
                'exp_leave': resignation.exp_leave,
                'leave_reason': resignation.leave_reason,
                'leave_reason_2': resignation.leave_reason_2,
                'leave_reason_3': resignation.leave_reason_3,
                'leave_date': resignation.leave_date,
                'notice_serve': resignation.notice_serve,
                'settle_date': resignation.settle_date,
                'shortfall_date': resignation.shortfall_date,
                'exit_interview': resignation.exit_interview,
                'last_working': resignation.last_working,
                'status': resignation.status,
                'approved_by': resignation.approved_by,
                'notice_per': resignation.notice_per,
            })
        return JsonResponse(resignation_data, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            emp_emailid = data.get('emp_emailid')
            exp_leave = data.get('exp_leave')
            leave_reason = data.get('leave_reason')
            leave_reason_2 = data.get('leave_reason_2')
            leave_reason_3 = data.get('leave_reason_3')
            leave_date = data.get('leave_date')
            settle_date = data.get('settle_date')
            exit_interview = data.get('exit_interview')
            last_working = data.get('last_working')
            status = data.get('status')
            approved_by = data.get('approved_by')

            if not all([emp_emailid, exp_leave, leave_reason, exit_interview, last_working]):
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            submit_date = timezone.now().date()
            date1 = submit_date
            date2 = timezone.now().date()
            diff = (date2 - date1).days
            notice_serve = abs(diff)
            notice_per = 30
            shortfall_date = max(0, 30 - notice_serve)

            resignation, created = Resignation.objects.update_or_create(
                emp_emailid=emp_emailid,
                defaults={
                    'submit_date': submit_date,
                    'exp_leave': exp_leave,
                    'leave_reason': leave_reason,
                    'leave_reason_2': leave_reason_2,
                    'leave_reason_3': leave_reason_3,
                    'leave_date': leave_date,
                    'notice_serve': notice_serve,
                    'settle_date': settle_date,
                    'shortfall_date': shortfall_date,
                    'exit_interview': exit_interview,
                    'last_working': last_working,
                    'status': status,
                    'approved_by': approved_by,
                    'notice_per': notice_per,
                }
            )

            response = {'message': 'Resignation created successfully'} if created else {'message': 'Resignation updated successfully'}
            return JsonResponse({'status': 1, 'message': response})

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['Super Admin', 'Manager', 'HR','Employee'])
def AddCourses(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    # Check if user is logged in and has a valid company ID
    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        try:
            course_title = request.POST.get('course_title')
            description = request.POST.get('description')
            thumbnail = request.FILES.get('thumbnail')  # Optional field

            # Check for required fields
            if not all([course_title, description]):
                return HttpResponseBadRequest("Missing required fields")

            # Fetch the company based on company_id from session
            company = get_object_or_404(Company, pk=company_id)

            # Ensure that the course is being added to the correct company
            if company.c_id != company_id:
                return JsonResponse({'error': 'Invalid company ID'}, status=403)

            # Prepare course data
            course_data = {
                'course_title': course_title,
                'description': description,
                'c_id': company,  # Ensure that the company is the same as user's company
                'c_name': company.c_name
            }

            # Include thumbnail only if provided
            if thumbnail:
                course_data['thumbnail'] = thumbnail

            # Create the course within the company
            course = Courses.objects.create(**course_data)

            response_data = {
                'message': 'Course added successfully',
                'course_id': course.course_id
            }

            return JsonResponse(response_data, status=201)

        except Exception as e:
            # Handle exceptions gracefully and return an error message
            return JsonResponse({'error': str(e)}, status=400)

    else:
        # Only allow POST requests
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
    


@csrf_exempt
@role_required(['Super Admin', 'Manager', 'HR','Employee'])
def GetCourses(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    # Check if user is logged in and has a valid company ID
    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in or no company ID found'}, status=401)

    if request.method == 'GET':
        try:
            # Fetch the company based on company_id from session
            company = get_object_or_404(Company, pk=company_id)

            # Fetch all courses for the logged-in user's company
            courses = Courses.objects.filter(c_id=company)

            # Prepare the response data
            course_list = []
            for course in courses:
                course_list.append({
                    'id': course.course_id,         # Ensure 'id' is used as expected in frontend
                    'title': course.course_title,   # Ensure 'title' is used as expected in frontend
                })

            return JsonResponse({'courses': course_list}, status=200)

        except Exception as e:
            # Handle exceptions gracefully and return an error message
            return JsonResponse({'error': str(e)}, status=400)

    else:
        # Only allow GET requests
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)

@csrf_exempt
@role_required(['Super Admin', 'Manager', 'HR','Employee'])
def GetCourses(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    # Check if user is logged in and has a valid company ID
    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in or no company ID found'}, status=401)

    if request.method == 'GET':
        try:
            # Fetch the company based on company_id from session
            company = get_object_or_404(Company, pk=company_id)

            # Fetch all courses for the logged-in user's company
            courses = Courses.objects.filter(c_id=company)

            # If no courses found, return a message
            if not courses.exists():
                return JsonResponse({'message': 'No courses found for this company'}, status=200)

            # Prepare the response data
            course_list = []
            for course in courses:
                course_list.append({
                    'course_id': course.course_id,
                    'course_title': course.course_title,
                    'description': course.description,
                    'thumbnail': course.thumbnail.url if course.thumbnail else None,
                    'company_name': course.c_name,
                })

            return JsonResponse({'courses': course_list}, status=200)

        except Exception as e:
            # Handle exceptions gracefully and return an error message
            return JsonResponse({'error': str(e)}, status=400)

    else:
        # Only allow GET requests
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UploadMedia(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            # video_name = data.get('video_name')
            location = data.get('location')
            descr = data.get('descr')
            course_id = data.get('course_id')

            try:
                course = Courses.objects.get(course_id=course_id)
            except Courses.DoesNotExist:
                return JsonResponse({'error': 'Course not found'}, status=404)

            existing_videos_count = Video.objects.filter(course_id=course_id).count()
            if existing_videos_count > 1:
                return JsonResponse({'message': 'Video with the same details already exists.'}, status=400)

            video = Video.objects.create(
                location=location,
                descr=descr,
                course_id=course,
                # video_name=video_name
            )

            return JsonResponse({'message': 'Video added successfully'}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UploadPdf(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'POST':
        try:
            pdf_file = request.FILES.get('pdf_file')
            pdf_description = request.POST.get('pdf_description')
            course_id = request.POST.get('course_id')

            if not pdf_file or not pdf_description or not course_id:
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            try:
                # Fetch the course using the provided course_id
                course = Courses.objects.get(course_id=course_id)
            except Courses.DoesNotExist:
                return JsonResponse({'error': 'Course not found'}, status=404)

            # Ensure the 'pdfs/' directory exists
            pdf_dir = os.path.join(settings.MEDIA_ROOT, 'pdfs')
            if not os.path.exists(pdf_dir):
                os.makedirs(pdf_dir)

            # Save the PDF file
            pdf_path = os.path.join('pdfs', pdf_file.name)
            with open(os.path.join(settings.MEDIA_ROOT, pdf_path), 'wb+') as destination:
                for chunk in pdf_file.chunks():
                    destination.write(chunk)

            # Save PDF details in the database
            pdf = Pdf.objects.create(
                pdf_name=pdf_file.name,
                location=pdf_path,
                descr=pdf_description,
                course=course
            )

            return JsonResponse({'message': 'PDF uploaded successfully'}, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
    



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddMediaContent(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        c_id = request.session.get('c_id')

        if not c_id:
            return JsonResponse({'error': 'Company ID not found in session'}, status=400)

        courses_data = Courses.objects.filter(c_id=c_id).values('course_title', 'course_id')
        data = list(courses_data)

        return JsonResponse(data, safe=False)
    else:
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UpdateDeleteMedia(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        courses_data = Courses.objects.filter(c_id=company_id).values('course_title', 'description', 'course_id')
        data = list(courses_data)
        return JsonResponse(data, safe=False)
    else:
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UpdateMedia(request, course_id):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    try:
        course = Courses.objects.get(course_id=course_id, c_id=company_id)
    except Courses.DoesNotExist:
        return JsonResponse({'error': 'Course not found or you do not have permission to access it'}, status=404)

    if request.method == 'GET':
        try:
            videos = Video.objects.filter(course_id=course_id).values('video_id', 'location', 'descr')

            course_data = {
                'course_id': course.course_id,
                'course_title': course.course_title,
                'videos': list(videos)
            }
            return JsonResponse(course_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            course = Courses.objects.get(course_id=course_id)

            if 'course_title' in data:
                course.course_title = data['course_title']
                course.save()

            if 'videos' in data:
                for video_data in data['videos']:
                    video_id = video_data.get('video_id')
                    if video_id:
                        video = Video.objects.get(video_id=video_id, course_id=course_id)
                        video.location = video_data.get('location', video.location)
                        video.descr = video_data.get('descr', video.descr)
                        video.save()

            return JsonResponse({'message': 'Course updated successfully'}, status=200)
        except (Courses.DoesNotExist, Video.DoesNotExist):
            return JsonResponse({'error': 'Course or video not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'DELETE':
        try:
            media_content = Courses.objects.get(course_id=course_id)
            media_content.delete()
            return JsonResponse({'message': 'Media content deleted successfully'})
        except Courses.DoesNotExist:
            return JsonResponse({'error': 'Media content not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ReportingStructureForm(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        employees = Employee.objects.filter(d_id__c_id=company_id, Status='Active').values('emp_emailid', 'd_id__c_id')
        employees_list = list(employees)
        return JsonResponse(employees_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            selected_employees = data.get('eemail', [])
            reporting_employee = data.get('remail')
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)

        if not selected_employees or not reporting_employee:
            return JsonResponse({'error': 'Select employees and a reporting employee first!'}, status=400)

        with transaction.atomic():
            try:
                for emp_email in selected_employees:
                    Reporting.objects.update_or_create(
                        c_id=company_id,
                        Reporting_from=emp_email,
                        defaults={'Reporting_to': reporting_employee}
                    )
                return JsonResponse({'success': 'Reporting structure updated successfully'}, status=200)
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CeoHrAnnouncements(request):
    company_id = request.session.get('c_id')
    print(f"Company ID from session: {company_id}")  # Debugging

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'POST':
        sender = request.POST.get('email')  # Correct syntax for calling 'get'
        cc = request.POST.get('cc', '')
        departments = request.POST.getlist('dept')
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        files = request.FILES.getlist('files')
        send_to_all = 'check' in request.POST and request.POST['check'] == 'yes'

        # Debugging statements to verify POST data
        print(f"Sender Email: {sender}")
        print(f"CC Email: {cc}")
        print(f"Departments: {departments}")
        print(f"Subject: {subject}")
        print(f"Message: {message}")
        print(f"Files: {files}")
        print(f"Send to All: {send_to_all}")


        emails_sent = 0
        separator = getattr(settings, 'EMAIL_SEPARATOR', ',')  # Use default if not set
        print(f"Email Separator: {separator}")  # Debugging
        try:
            if send_to_all:
                employees = Employee.objects.filter(d_id__c_id=company_id, Status='Active')
                print(f"Employees (send_to_all=True): {employees.count()} found")  # Debugging
            else:
                employees = Employee.objects.filter(d_id__in=departments, Status='Active')
                print(f"Employees (send_to_all=False): {employees.count()} found")  # Debugging
            for employee in employees:
                print(f"Employee Email: {employee.emp_emailid}")  # Debugging

            for employee in employees:
                email = EmailMessage(
                    subject=subject,
                    body=message,
                    from_email=sender,
                    to=[employee.emp_emailid],
                    cc=[cc],  # Use a list directly
                )

                for file in files:
                    email.attach(file.name, file.read(), file.content_type)

                email.send()
                print(f"Email sent to: {employee.emp_emailid}")  # Debugging
                emails_sent += 1

            return JsonResponse({'success': f'Total {emails_sent} emails sent successfully'}, status=200)

        except Exception as e:
            # Improved exception handling and logging
            print(f"Error sending emails: {str(e)}")  # Use logging in production
            return JsonResponse({'error': 'An error occurred while sending emails. Check server logs for details.'}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)



# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def CeoHrAnnouncements(request):
#     company_id = request.session.get('c_id')
#     if not company_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'GET':
#         # Fetch departments for the logged-in company
#         departments = Department.objects.filter(c_id=company_id).values('d_id', 'd_name')
#         return JsonResponse(list(departments), safe=False)

#     if request.method == 'POST':
#         sender = request.POST.get('email')
#         cc = request.POST.get('cc', '')
#         departments = request.POST.getlist('dept')  # Expecting d_id values
#         subject = request.POST.get('subject')
#         message = request.POST.get('message')
#         files = request.FILES.getlist('files')
#         send_to_all = request.POST.get('check') == 'yes'

#         try:
#             if send_to_all:
#                 employees = Employee.objects.filter(d_id__c_id=company_id, Status='Active')
#             else:
#                 employees = Employee.objects.filter(d_id__in=departments, Status='Active')

#             emails_sent = 0
#             for employee in employees:
#                 email = EmailMessage(
#                     subject=subject,
#                     body=message,
#                     from_email=sender,
#                     to=[employee.emp_emailid],
#                     cc=[cc] if cc else None,
#                 )
#                 for file in files:
#                     email.attach(file.name, file.read(), file.content_type)
#                 email.send()
#                 emails_sent += 1

#             return JsonResponse({'success': f'Total {emails_sent} emails sent successfully'}, status=200)
        
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

#     return JsonResponse({'error': 'Invalid request method'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddNewEmployee(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        emp_count = Employee.objects.filter(d_id__c_id=company_id).count()
        company_data = Company.objects.filter(c_id=company_id).values('payment_type', 'emp_limit').first()
        departments = Department.objects.filter(c_id=company_id).values('d_id', 'd_name')


        response_data = {
            'emp_count': emp_count,
            'company': company_data,
            'departments': list(departments)
        }

        return JsonResponse(response_data, status=200)

    elif request.method == 'POST':
        data = json.loads(request.body)
        name = data.get('fullName')
        email = data.get('email')
        phone = data.get('phone')
        department_id = data.get('department')
        skills = data.get('skills')
        emp_role = data.get('emp_role')

        valid_roles = ['HR', 'Super Admin', 'Manager', 'Employee']
        if emp_role not in valid_roles:
            return JsonResponse({'error': 'Invalid role selected'}, status=400)


        company = Company.objects.filter(c_id=company_id).first()
        department = Department.objects.filter(d_id=department_id).first()

        # if company.emp_limit <= Employee.objects.filter(d_id__c_id=company_id).count():
        #     return JsonResponse({'error': 'Employee limit reached'}, status=403)

        Employee.objects.create(
            emp_name=name,
            emp_emailid=email,
            emp_phone=phone,
            emp_skills=skills,
            emp_role=emp_role,  # Assign the validated role
            d_id=department,
        )

        return JsonResponse({'success': 'Employee created successfully'}, status=201)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UpdateDeleteEmployee(request):
    company_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            employees = Employee.objects.filter(d_id__c_id=company_id)
            email_ids = [employee.emp_emailid for employee in employees]
            return JsonResponse({'email_ids': email_ids}, status=200)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employees not found or you do not have permission to access them'}, status=404)

    elif request.method == 'DELETE':
        try:
            email_id = request.GET.get('emailId')
            employee = Employee.objects.get(emp_emailid=email_id, d_id__c_id=company_id)
            if Quiz.objects.filter(eid=employee).exists():
                return JsonResponse({'error': 'Cannot delete employee with existing quizzes'}, status=400)

            employee.delete()
            return JsonResponse({'success': 'Employee deleted successfully'}, status=200)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found or you do not have permission to delete it'}, status=404)


# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def UpdateEmployeeDetails(request):
#     company_id = request.session.get('c_id')
#     emp_emailid = request.session.get('emp_emailid')

#     if not company_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'GET':
#         try:
#             employee = Employee.objects.get(emp_emailid=emp_emailid, d_id__c_id=company_id)
#             employee_data = {
#                 'emp_name': employee.emp_name,
#                 'emp_emailid': employee.emp_emailid,
#                 'emp_phone': employee.emp_phone,
#                 'd_id': employee.d_id.d_id,
#                 'emp_skills': employee.emp_skills,
#             }

#             departments = Department.objects.filter(c_id=company_id).values('d_id', 'd_name')

#             response_data = {
#                 'employee_data': employee_data,
#                 'departments': list(departments)
#             }

#             return JsonResponse(response_data, status=200)
#         except Employee.DoesNotExist:
#             return JsonResponse({'error': 'Employee not found or you do not have permission to access it'}, status=404)

#     elif request.method == 'POST':
#         data = json.loads(request.body)
#         emp_name = data.get('emp_name')
#         emp_emailid = data.get('emp_emailid')
#         emp_phone = data.get('emp_phone')
#         d_id = data.get('d_id')
#         emp_skills = data.get('emp_skills')

#         try:
#             employee = Employee.objects.get(emp_emailid=emp_emailid, d_id__c_id=company_id)
#             employee.emp_name = emp_name
#             employee.emp_phone = emp_phone
#             employee.d_id = Department.objects.get(d_id=d_id, c_id=company_id)
#             employee.emp_skills = emp_skills
#             employee.save()
#             return JsonResponse({'success': 'Employee details updated successfully'}, status=200)
#         except Employee.DoesNotExist:
#             return JsonResponse({'error': 'Employee not found or you do not have permission to update it'}, status=404)
@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UpdateEmployeeDetails(request):
    company_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')
    to_be_updated_employee_email = request.GET.get('emp_emailid')


    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            employee = Employee.objects.get(emp_emailid=to_be_updated_employee_email or emp_emailid, d_id__c_id=company_id)
            
            employee_data = {
                'emp_name': employee.emp_name,
                'emp_emailid': employee.emp_emailid,
                'emp_phone': employee.emp_phone,
                'd_id': employee.d_id.d_id,
                'emp_skills': employee.emp_skills,
                'emp_role': employee.emp_role,
            }

            departments = Department.objects.filter(c_id=company_id).values('d_id', 'd_name')

            response_data = {
                'employee_data': employee_data,
                'departments': list(departments)
            }

            return JsonResponse(response_data, status=200)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found or you do not have permission to access it'}, status=404)

    elif request.method == 'POST':
        data = json.loads(request.body)
        emp_name = data.get('emp_name')
        emp_emailid = data.get('emp_emailid')
        emp_phone = data.get('emp_phone')
        emp_role = data.get('emp_role')
        d_id = data.get('d_id')
        emp_skills = data.get('emp_skills')

        try:
            employee = Employee.objects.get(emp_emailid=emp_emailid, d_id__c_id=company_id)
            employee.emp_name = emp_name
            employee.emp_phone = emp_phone
            employee.emp_role = emp_role
            employee.d_id = Department.objects.get(d_id=d_id, c_id=company_id)
            employee.emp_skills = emp_skills
            employee.save()
            return JsonResponse({'success': 'Employee details updated successfully'}, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found or you do not have permission to update it'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def JDForms(request):
    if request.method == 'POST':
        res = request.POST.getlist('res[]')
        role = request.POST.get('role')
        sdate = request.POST.get('date')
        emp_emails = request.POST.getlist('eemail[]')

        if not (res and role and sdate and emp_emails):
            return JsonResponse({'error': 'All fields are required!'}, status=400)

        try:
            sdate = datetime.datetime.strptime(sdate, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Invalid date format!'}, status=400)

        for emp_email in emp_emails:
            if not Employee.objects.filter(emp_emailid=emp_email).exists():
                return JsonResponse({'error': f'Employee with email {emp_email} does not exist!'}, status=400)

        try:
            jd_instance = Job_desc.objects.create(jd_name=role, sdate=sdate)

            for emp_email in emp_emails:
                for responsibility in res:
                    Tasks.objects.create(job_desc_id=jd_instance, responsiblities=responsibility, emp_emailid=emp_email, sdate=sdate)

            return JsonResponse({'message': 'Form submitted successfully!'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)



@csrf_exempt
def LeaveDashboard(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        # Fetch department, employee, and leave type counts
        departments = Department.objects.filter(c_id=company_id)
        dptcount = departments.count()
        employees = Employee.objects.filter(d_id__in=departments).values('emp_emailid')
        empcount = employees.count()
        leavtypcount = Leavetype.objects.count()

        emp_emails = [e['emp_emailid'] for e in employees]
        leaves_fetched = Tblleaves.objects.filter(emp_emailid__in=emp_emails).order_by('-id')

        # Define status mapping from numbers to meaningful strings
        status_mapping = {
            0: 'Pending',
            1: 'Approved',
            2: 'Rejected',
            # Add more status mappings as needed
        }

        leaves_data = []
        for leave in leaves_fetched:
            leave_dict = {
                'lid': leave.id,
                'emp_name': leave.emp_emailid.emp_name,
                'emp_emailid': leave.emp_emailid.emp_emailid,
                'LeaveType': leave.LeaveType.LeaveType,
                'PostingDate': leave.PostingDate.strftime('%Y-%m-%d %H:%M:%S'),
                'Status': status_mapping.get(leave.Status, 'Unknown'),  # Map status number to string
            }
            leaves_data.append(leave_dict)

        response_data = {
            'empcount': empcount,
            'dptcount': dptcount,
            'leavtypcount': leavtypcount,
            'leaves_fetched': leaves_data,
        }
        return JsonResponse(response_data, safe=False)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


#leaveaction

@csrf_exempt
def update_leave_status(request, leave_id):
    # Check for session authentication
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == "POST":
        try:
            # Parse JSON body instead of using request.POST
            data = json.loads(request.body)
            
            status = data.get('status')  # 'approved' or 'not-approved'
            admin_remark = data.get('description')

            # Fetch the leave record based on the ID passed from the frontend
            leave = Tblleaves.objects.get(id=leave_id)

            # Map the frontend status to backend integer values
            if status == 'approved':
                leave.Status = 1  # Approved
            elif status == 'not-approved':
                leave.Status = 2  # Rejected
            else:
                return JsonResponse({"error": "Invalid status"}, status=400)

            # Update the remark and the date of action
            leave.AdminRemark = admin_remark
            leave.AdminRemarkDate = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

            # Save changes to the database
            leave.save()

            return JsonResponse({"message": "Leave status updated successfully"}, status=200)

        except Tblleaves.DoesNotExist:
            return JsonResponse({"error": "Leave record not found"}, status=404)

    return JsonResponse({"error": "Invalid request method"}, status=405)

# Configure logging
logging.basicConfig(level=logging.DEBUG)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def LeaveDetails(request, id):  # Accept the 'id' parameter from the URL
    logging.debug(f"Received request method: {request.method}")

    if request.method == 'GET':
        try:
            logging.debug(f"Leave ID: {id}")
            if not id:
                return JsonResponse({'error': 'Leave ID is required'}, status=400)

            company_id = request.session.get('c_id')
            logging.debug(f"Company ID from session: {company_id}")

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            leave_dict = Tblleaves.objects.filter(id=id).values(
                'id', 'LeaveType__LeaveType', 'ToDate', 'FromDate', 'Description', 'PostingDate', 'Status',
                'AdminRemark', 'AdminRemarkDate', 'emp_emailid__emp_name', 'emp_emailid__emp_phone',
                'emp_emailid__d_id__c_id', 'emp_emailid__emp_role'
            ).first()

            logging.debug(f"Leave Details: {leave_dict}")

            if not leave_dict:
                return JsonResponse({'error': 'Leave not found'}, status=404)

            user_company_id = request.session.get('c_id')
            if leave_dict['emp_emailid__d_id__c_id'] != user_company_id:
                return JsonResponse({'error': 'You are not authorized to view this leave details'}, status=403)

            # Mapping statuses: 0 = Pending, 1 = Approved, 2 = Rejected
            status_mapping = {0: "Pending", 1: "Approved", 2: "Rejected"}
            leave_dict['Status'] = status_mapping.get(leave_dict['Status'], "Unknown")

            leave_dict['emp_emailid__emp_name'] = leave_dict['emp_emailid__emp_name'].strip()
            leave_dict['PostingDate'] = leave_dict['PostingDate'].isoformat()
            leave_dict['AdminRemarkDate'] = leave_dict['AdminRemarkDate'].isoformat() if leave_dict['AdminRemarkDate'] else None

            return JsonResponse(leave_dict, safe=False)
        except Exception as e:
            logging.error(f"Exception occurred: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    else:
        logging.error(f"Invalid request method: {request.method}")
        return JsonResponse({'error': 'Invalid request method'}, status=405)





@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ManageLeaveType(request):
    if request.method == 'GET':
        try:
            company_id = request.session.get('c_id')

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            leavetypes = Leavetype.objects.filter(company_id=company_id)
            leavetypes_list = []

            for leavetype in leavetypes:
                leavetype_dict = {
                    'id': leavetype.id,
                    'LeaveType': leavetype.LeaveType,
                    'Description': leavetype.Description,
                    'Limit': leavetype.Limit,
                    'CreationDate': leavetype.CreationDate.strftime('%Y-%m-%d %H:%M:%S')
                }
                leavetypes_list.append(leavetype_dict)

            return JsonResponse(leavetypes_list, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            leave_id = data.get('id')
            
            if not leave_id:
                return JsonResponse({'error': 'Leave ID is required for updating'}, status=400)
            
            leavetype = Leavetype.objects.get(id=leave_id, company_id=request.session.get('c_id'))
            leavetype.LeaveType = data.get('LeaveType', leavetype.LeaveType)
            leavetype.Description = data.get('Description', leavetype.Description)
            leavetype.Limit = data.get('Limit', leavetype.Limit)
            leavetype.save()

            return JsonResponse({'message': 'Leave Type updated successfully'}, status=200)
        except Leavetype.DoesNotExist:
            return JsonResponse({'error': 'Leave Type not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'DELETE':
        try:
            leave_id = request.GET.get('id')
            if leave_id:
                leavetype = Leavetype.objects.get(id=leave_id, company_id=request.session.get('c_id'))
                leavetype.delete()
                return JsonResponse({'message': 'Leave Type deleted successfully'})
            else:
                return JsonResponse({'error': 'Leave ID is required for deletion'}, status=400)
        except Leavetype.DoesNotExist:
            return JsonResponse({'error': 'Leave Type not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EditLeaveType(request, id):
    if request.method == 'GET':
        try:
            leavetype = Leavetype.objects.get(id=id)

            company_id = request.session.get('c_id')
            if not company_id or leavetype.company_id != company_id:
                return JsonResponse({'error': 'Leave Type not found'}, status=404)

            leavetype_data = {
                'LeaveType': leavetype.LeaveType,
                'Description': leavetype.Description,
                'Limit': leavetype.Limit,
            }

            return JsonResponse(leavetype_data, safe=False)
        except Leavetype.DoesNotExist:
            return JsonResponse({'error': 'Leave Type not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            leavetype = Leavetype.objects.get(id=id)

            company_id = request.session.get('c_id')
            if not company_id or leavetype.company_id != company_id:
                return JsonResponse({'error': 'Leave Type not found'}, status=404)

            if 'LeaveType' in data:
                leavetype.LeaveType = data['LeaveType']
            if 'Description' in data:
                leavetype.Description = data['Description']
            if 'Limit' in data:
                leavetype.Limit = data['Limit']

            leavetype.save()

            return JsonResponse({'message': 'Leave Type updated successfully'}, status=200)
        except Leavetype.DoesNotExist:
            return JsonResponse({'error': 'Leave Type not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def Leaves(request):
    if request.method == 'GET':
        try:
            company_id = request.session.get('c_id')

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            departments = Department.objects.filter(c_id=company_id)
            dptcount = departments.count()
            employees = Employee.objects.filter(d_id__in=departments).values('emp_emailid')
            empcount = employees.count()

            leavtypcount = Leavetype.objects.count()

            emp_emails = [e['emp_emailid'] for e in employees]
            leaves_fetched = Tblleaves.objects.filter(emp_emailid__in=emp_emails).order_by('-id')

            leaves_data = []
            for leave in leaves_fetched:
                leave_dict = {
                    'lid': leave.id,
                    'emp_name': leave.emp_emailid.emp_name,
                    'emp_emailid': leave.emp_emailid.emp_emailid,
                    'LeaveType': leave.LeaveType_id,
                    'PostingDate': leave.PostingDate.strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': leave.Status,
                }
                leaves_data.append(leave_dict)

            response_data = {
                'empcount': empcount,
                'dptcount': dptcount,
                'leavtypcount': leavtypcount,
                'leaves_fetched': leaves_data,
            }
            return JsonResponse(response_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def PendingLeaves(request):
    if request.method == 'GET':
        try:
            company_id = request.session.get('c_id')

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            departments = Department.objects.filter(c_id=company_id)
            employees = Employee.objects.filter(d_id__in=departments).values('emp_emailid')

            emp_emails = [e['emp_emailid'] for e in employees]
            leaves_fetched = Tblleaves.objects.filter(emp_emailid__in=emp_emails, Status=0).order_by('-id')

            leaves_data = []
            for leave in leaves_fetched:
                leave_dict = {
                    'lid': leave.id,
                    'emp_name': leave.emp_emailid.emp_name,
                    'emp_emailid': leave.emp_emailid.emp_emailid,
                    'LeaveType': leave.LeaveType_id,
                    'PostingDate': leave.PostingDate.strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': leave.Status,
                }
                leaves_data.append(leave_dict)

            response_data = {
                'leaves_fetched': leaves_data
            }
            return JsonResponse(response_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ApprovedLeaves(request):
    if request.method == 'GET':
        try:
            company_id = request.session.get('c_id')

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            departments = Department.objects.filter(c_id=company_id)
            employees = Employee.objects.filter(d_id__in=departments).values('emp_emailid')

            emp_emails = [e['emp_emailid'] for e in employees]
            leaves_fetched = Tblleaves.objects.filter(emp_emailid__in=emp_emails, Status=1).order_by('-id')

            leaves_data = []
            for leave in leaves_fetched:
                leave_dict = {
                    'lid': leave.id,
                    'emp_name': leave.emp_emailid.emp_name,
                    'emp_emailid': leave.emp_emailid.emp_emailid,
                    'LeaveType': leave.LeaveType_id,
                    'PostingDate': leave.PostingDate.strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': leave.Status,
                }
                leaves_data.append(leave_dict)

            response_data = {
                'leaves_fetched': leaves_data
            }
            return JsonResponse(response_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def RejectedLeaves(request):
    if request.method == 'GET':
        try:
            company_id = request.session.get('c_id')

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            departments = Department.objects.filter(c_id=company_id)
            employees = Employee.objects.filter(d_id__in=departments).values('emp_emailid')

            emp_emails = [e['emp_emailid'] for e in employees]
            leaves_fetched = Tblleaves.objects.filter(emp_emailid__in=emp_emails, Status=2).order_by('-id')

            leaves_data = []
            for leave in leaves_fetched:
                leave_dict = {
                    'lid': leave.id,
                    'emp_name': leave.emp_emailid.emp_name,
                    'emp_emailid': leave.emp_emailid.emp_emailid,
                    'LeaveType': leave.LeaveType_id,
                    'PostingDate': leave.PostingDate.strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': leave.Status,
                }
                leaves_data.append(leave_dict)

            response_data = {
                'leaves_fetched': leaves_data
            }
            return JsonResponse(response_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


# Resignation Management
@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AllResignation(request):
    emp_emailid = request.session.get('emp_emailid')
    c_id = request.session.get('c_id')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            resignations = Resignation.objects.filter(emp_emailid__d_id__c_id=c_id).select_related('emp_emailid', 'emp_emailid__d_id', 'job_info')

            resignations_data = []

            for resignation in resignations:
                employee = resignation.emp_emailid
                job_info = resignation.emp_emailid.job_info
                if job_info:
                    interval = (job_info.start_date - resignation.submit_date.date()).days
                else:
                    interval = None

                resignation_data = {
                    'emp_profile': employee.emp_profile.url,
                    'emp_name': employee.emp_name,
                    'emp_role': employee.emp_role,
                    'emp_emailid': employee.emp_emailid,
                    'department_name': employee.d_id.d_name,
                    'job_title': job_info.job_title if job_info else '',
                    'start_date': job_info.start_date if job_info else None,
                    'leave_date': resignation.leave_date,
                    'exp_leave': resignation.exp_leave,
                    'shortfall_date': resignation.shortfall_date,
                    'days_since_resignation': interval,
                }

                resignations_data.append(resignation_data)

            return JsonResponse({'resignations': resignations_data}, status=200)

        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AllExitClearance(request):
    emp_emailid = request.session.get('emp_emailid')
    c_id = request.session.get('c_id')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email ID not found in session'}, status=401)

    try:
        employees_in_company = Employee.objects.filter(d_id__c_id=c_id).values_list('emp_emailid', flat=True)

        resignations = Resignation.objects.filter(
            emp_emailid__in=employees_in_company,
            status='Approved',
            emp_emailid__d_id__c_id=c_id
        ).select_related('emp_emailid', 'emp_emailid__d_id', 'clearance')

        clearances_data = []

        for resignation in resignations:
            employee = resignation.emp_emailid
            clearance = resignation.clearance
            try:
                job_info = Job_info.objects.get(emp_emailid=employee.emp_emailid)
            except Job_info.DoesNotExist:
                job_info = None

            today = timezone.now().date()
            interval = (today - resignation.submit_date.date()).days

            clearance_status = clearance.status if clearance else 'Pending'

            clearance_data = {
                'emp_profile': employee.emp_profile.url,
                'emp_name': employee.emp_name,
                'emp_role': employee.emp_role,
                'emp_emailid': employee.emp_emailid,
                'department_name': employee.d_id.d_name,
                'job_title': job_info.job_title if job_info else '',
                'days_since_resignation': interval,
                'clearance_status': clearance_status,
            }

            clearances_data.append(clearance_data)

        return JsonResponse({'clearances': clearances_data}, status=200)

    except Exception as e:
        return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AllFinalSettlement(request):
    emp_emailid = request.session.get('emp_emailid')
    c_id = request.session.get('c_id')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email ID not found in session'}, status=401)

    try:
        employees_in_company = Employee.objects.filter(d_id__c_id=c_id).values_list('emp_emailid', flat=True)

        resignations = Resignation.objects.filter(
            emp_emailid__in=employees_in_company,
            clearance__status='Approved',
            emp_emailid__d_id__c_id=c_id
        ).select_related('emp_emailid', 'emp_emailid__d_id', 'clearance')

        settlements_data = []

        for resignation in resignations:
            employee = resignation.emp_emailid
            clearance = resignation.clearance
            try:
                job_info = Job_info.objects.get(emp_emailid=employee.emp_emailid)
            except Job_info.DoesNotExist:
                job_info = None

            today = timezone.now().date()
            interval = (today - resignation.submit_date.date()).days

            clearance_status = clearance.status if clearance else 'Pending'

            settlement_data = {
                'emp_profile': employee.emp_profile.url,
                'emp_name': employee.emp_name,
                'emp_role': employee.emp_role,
                'emp_emailid': employee.emp_emailid,
                'department_name': employee.d_id.d_name,
                'job_title': job_info.job_title if job_info else '',
                'days_since_resignation': interval,
                'clearance_status': clearance_status,
            }

            settlements_data.append(settlement_data)

        return JsonResponse({'settlements': settlements_data}, status=200)

    except Exception as e:
        return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EditResignation(request):
    c_id = request.session.get('c_id')

    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        departments = Department.objects.filter(c_id=c_id)
        employees = Employee.objects.filter(d_id__in=departments)

        resignations = Resignation.objects.select_related('emp_emailid').filter(emp_emailid__d_id__in=departments)

        resignation_data = []
        for resignation in resignations:
            job_info = Job_info.objects.get(emp_emailid=resignation.emp_emailid.emp_emailid)
            interval = (timezone.now().date() - job_info.start_date).days
            length_of_service = '{} years {} months'.format(interval // 365, (interval % 365) // 30)
            resignation_data.append({
                'emp_name': resignation.emp_emailid.emp_name,
                'emp_profile': resignation.emp_emailid.emp_profile,
                'emp_emailid': resignation.emp_emailid.emp_emailid,
                'emp_role': resignation.emp_emailid.emp_role,
                'status': resignation.status,
                'start_date': job_info.start_date,
                'leave_date': resignation.leave_date,
                'exp_leave': resignation.exp_leave,
                'length_of_service': length_of_service,
                'shortfall_date': resignation.shortfall_date,
            })

        return JsonResponse(resignation_data, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            emp_emailid = data.get('emp_emailid')
            leave_reason = data.get('leave_reason')
            exp_leave = data.get('exp_leave')
            notice_per = data.get('notice_per')
            status = data.get('status')

            if not all([emp_emailid, leave_reason, exp_leave, notice_per, status]):
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            resignation, created = Resignation.objects.update_or_create(
                emp_emailid_id=emp_emailid,
                defaults={
                    'leave_reason': leave_reason,
                    'exp_leave': exp_leave,
                    'notice_per': notice_per,
                    'status': status
                }
            )

            response = {'message': 'Resignation created successfully'} if created else {'message': 'Resignation updated successfully'}
            return JsonResponse(response, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EditExitClearnace(request):
    c_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')
    given_to_email = request.GET.get('given_to_email')

    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            clearance = Clearance.objects.get(given_to=given_to_email)

            clearance_data = {
                'Accounts': clearance.Accounts,
                'Hr': clearance.Hr,
                'Hr_Plant': clearance.Hr_Plant,
                'IT': clearance.IT,
                'Project': clearance.Project,
                'status': clearance.status,
                'given_by': clearance.given_by.emp_emailid if clearance.given_by else None,
                'given_to': clearance.emp_emailid.emp_emailid
            }

            return JsonResponse(clearance_data, status=200)

        except Clearance.DoesNotExist:
            return JsonResponse({'error': 'Clearance not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            clearance, created = Clearance.objects.get_or_create(emp_emailid=given_to_email)

            if 'Accounts' in data:
                clearance.Accounts = data['Accounts']
            if 'Hr' in data:
                clearance.Hr = data['Hr']
            if 'Hr_Plant' in data:
                clearance.Hr_Plant = data['Hr_Plant']
            if 'IT' in data:
                clearance.IT = data['IT']
            if 'Project' in data:
                clearance.Project = data['Project']

            clearance.given_to = given_to_email
            clearance.given_by = emp_emailid

            clearance.save()

            return JsonResponse({'success': 'Clearance updated successfully'}, status=200)

        except Clearance.DoesNotExist:
            return JsonResponse({'error': 'Clearance not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def GenerateFnf(request):
    c_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
    personal_details = get_object_or_404(Personal_details, mail=emp_emailid)
    job_info = get_object_or_404(Job_info, emp_emailid=emp_emailid)
    resignation = get_object_or_404(Resignation, emp_emailid=emp_emailid)
    bank_details = get_object_or_404(Bank_details, emp_emailid=emp_emailid)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="full_and_final_settlement_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    data = [
        ["Name:", employee.emp_name, "Contact No.:", employee.emp_phone],
        ["Email Id:", employee.emp_emailid, "Role:", employee.emp_role],
        ["Gender:", personal_details.gender, "Birth Date:", personal_details.birth_date],
        ["Department:", job_info.department, "Joining Date:", job_info.start_date],
        ["Resigned On:", resignation.submit_date, "Leaving Date:", resignation.leave_date],
        ["Bank Name:", bank_details.bank_name, "Account No.:", bank_details.acc_no],
        ["Account Type:", bank_details.acc_type, "IFSC:", bank_details.ifsc],
        ["Pan no:", bank_details.Pan_no]
    ]

    table = Table(data)
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                               ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('FONTNAME', (0, 0), (-1, 0), 'Courier-Bold'),
                               ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                               ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                               ('GRID', (0, 0), (-1, -1), 1, colors.black)]))

    elements.append(table)
    doc.build(elements)

    return response


import logging
logger = logging.getLogger(__name__)

@csrf_exempt
def DisplayTraining(request):
    c_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')

    logger.info(f'c_id from session: {c_id}, emp_emailid from session: {emp_emailid}')

    if not c_id or not emp_emailid:
        return JsonResponse({'error': 'Company ID or Employee Email not found in session'}, status=401)

    if request.method == 'GET':
        courses = Courses.objects.filter(
            c_id=c_id,
            course_employee__emp_emailid=emp_emailid
        ).distinct().values('course_id', 'course_title', 'thumbnail', 'description')

        return JsonResponse(list(courses), safe=False, status=200)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


# Pages section
@csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
def CreateCase(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')
    print(f"User ID: {user_id}, Company ID: {company_id}")
    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            create_for = data.get('createFor')
            case_type = data.get('caseType')
            case_title = data.get('caseTitle')
            case_desc = data.get('detailedDescription')

            if not all([create_for, case_type, case_title, case_desc]):
                return HttpResponseBadRequest("Missing required fields")

            employee = get_object_or_404(Employee, emp_emailid=user_id)

            new_case = Case(
                create_for=create_for,
                case_type=case_type,
                case_title=case_title,
                case_desc=case_desc,
                created_by=employee,
                case_status='New'
            )

            new_case.save()

            return JsonResponse({"message": "Case created successfully"}, status=201)
        except json.JSONDecodeError:
            return HttpResponseBadRequest("Invalid JSON")
        except Employee.DoesNotExist:
            return HttpResponseBadRequest("Employee not found")
    else:
        return HttpResponseBadRequest("Only POST requests are allowed")


@csrf_exempt
def MyCases(request):
    def TimeAgo(ptime):
        current_time = timezone.now()
        estimate_time = current_time - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(Q(created_by__emp_emailid=user_id) | Q(assigned_to__emp_emailid=user_id))

    case_data = []
    for case in cases:
        timeago = TimeAgo(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
        })

    return JsonResponse({'cases': case_data})


# Need to add signals ( triggers) to the employee table
@csrf_exempt
def UpdatePersonalDetails(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    # print(f"Logged in user email: {emp_emailid}")  # Debugging

    # Fetch the Employee instance using the logged-in user's email
    try:
        employee = Employee.objects.get(emp_emailid=emp_emailid)
    except Employee.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)

    # print(f"Logged in user email: {emp_emailid}")  # Debugging

    # Fetch the Employee instance using the logged-in user's email
    try:
        employee = Employee.objects.get(emp_emailid=emp_emailid)
    except Employee.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)

    if request.method == 'GET':
        try:
            # personal_detail = Personal_details.objects.get(mail=employee)
            personal_detail = Personal_details.objects.get(mail=employee)
            data = {
                'first_name': personal_detail.first_name,
                'last_name': personal_detail.last_name,
                'address': personal_detail.address,
                'state': personal_detail.state,
                'city': personal_detail.city,
                'district': personal_detail.district,
                'post_code': personal_detail.post_code,
                'Contact': personal_detail.Contact,
                'birth_date': personal_detail.birth_date,
                'gender': personal_detail.gender,
                'emergency_name': personal_detail.emergency_name,
                'emergency_contact': personal_detail.emergency_contact,
                'emp_emailid': emp_emailid,
            }
            return JsonResponse({'status': 'success', 'data': data})
        except Personal_details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)

            # Extract data from request body
            # Extract data from request body
            first_name = data.get('first_name')
            last_name = data.get('last_name')
            address = data.get('address')
            state = data.get('state')
            city = data.get('city')
            district = data.get('district')
            post_code = data.get('post_code')
            Contact = data.get('Contact')
            birth_date = data.get('birth_date')
            gender = data.get('gender')
            emergency_name = data.get('emergency_name')
            emergency_contact = data.get('emergency_contact')

            # Use get_or_create to find or create a new Personal_details entry
            personal_detail, created = Personal_details.objects.get_or_create(mail=employee)

            # Update or set details
            personal_detail.first_name = first_name
            personal_detail.last_name = last_name
            personal_detail.address = address
            personal_detail.state = state
            personal_detail.city = city
            personal_detail.district = district
            personal_detail.post_code = post_code
            personal_detail.Contact = Contact
            personal_detail.birth_date = birth_date
            personal_detail.gender = gender
            personal_detail.emergency_name = emergency_name
            personal_detail.emergency_contact = emergency_contact
            personal_detail.save()

            # Send appropriate response based on whether the record was created or updated
            if created:
                return JsonResponse({'status': 'success', 'message': 'Personal details created successfully.'}, status=201)
            else:
                return JsonResponse({'status': 'success', 'message': 'Personal details updated successfully.'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)




@csrf_exempt
def UpdateJobDetails(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':
        try:
            # Try to retrieve the job_info for the employee's email
            job_info = Job_info.objects.get(emp_emailid=emp_emailid)

            data = {
                'job_title': job_info.job_title,
                'department': job_info.department,
                'working_type': job_info.working_type,
                'start_date': job_info.start_date,
            }

            return JsonResponse({'status': 'success', 'data': data})

        except Job_info.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Job information not found. You can add your own details.'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    elif request.method == 'PUT':
        try:
            # Parse JSON data from the request body
            data = json.loads(request.body.decode('utf-8'))

            # Collect data from the parsed JSON
            job_title = data.get('job_title')
            department = data.get('department')
            working_type = data.get('working_type')
            start_date = data.get('start_date')

            # Validate the data
            if not all([job_title, department, working_type, start_date]):
                return JsonResponse({'status': 'error', 'message': 'All fields are required'}, status=400)

            # Retrieve the employee instance using the email
            employee = Employee.objects.get(emp_emailid=emp_emailid)

            # Get or create a new Job_info object for the employee
            job_info, created = Job_info.objects.get_or_create(emp_emailid=employee)
            
            # Update the job_info fields with the new data
            job_info.job_title = job_title
            job_info.department = department
            job_info.working_type = working_type
            job_info.start_date = start_date

            # Save the updated or newly created job_info object
            job_info.save()

            # Respond with appropriate message based on whether a new job_info was created or updated
            if created:
                return JsonResponse({'status': 'success', 'message': 'Job information created successfully'})
            else:
                return JsonResponse({'status': 'success', 'message': 'Job information updated successfully'})

        except Employee.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)



@csrf_exempt
def UpdateBankDetails(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':
        try:
            bank_details = Bank_details.objects.get(emp_emailid=emp_emailid)
            data = {
                'holder_name': bank_details.holder_name,
                'bank_name': bank_details.bank_name,
                'acc_no': bank_details.acc_no,
                'branch': bank_details.branch,
                'acc_type': bank_details.acc_type,
                'ifsc': bank_details.ifsc,
                'Pan_no': bank_details.Pan_no,
            }
            return JsonResponse({'status': 'success', 'data': data})

        
        except Bank_details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Bank details not found'}, status=404)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            holder_name = data.get('holder_name')
            bank_name = data.get('bank_name')
            acc_no = data.get('acc_no')
            branch = data.get('branch')
            acc_type = data.get('acc_type')
            ifsc = data.get('ifsc')
            Pan_no = data.get('Pan_no')

            if not all([holder_name, bank_name, acc_no, branch, acc_type, ifsc, Pan_no, emp_emailid]):
                return JsonResponse({'status': 'error', 'message': 'Missing fields'}, status=400)
            # Fetch the Employee instance based on the email ID
            try:
                employee = Employee.objects.get(emp_emailid=emp_emailid)

                # Check if bank details already exist and update them
                try:
                    bank_details = Bank_details.objects.get(emp_emailid=employee)
                    bank_details.holder_name = holder_name
                    bank_details.bank_name = bank_name
                    bank_details.acc_no = acc_no
                    bank_details.branch = branch
                    bank_details.acc_type = acc_type
                    bank_details.ifsc = ifsc
                    bank_details.Pan_no = Pan_no
                    bank_details.save()

                    return JsonResponse({'status': 'success', 'message': 'Bank details updated successfully'})

                except Bank_details.DoesNotExist:
                    # Create new bank details if not found
                    Bank_details.objects.create(
                        emp_emailid=employee,
                        holder_name=holder_name,
                        bank_name=bank_name,
                        acc_no=acc_no,
                        branch=branch,
                        acc_type=acc_type,
                        ifsc=ifsc,
                        Pan_no=Pan_no
                    )
                    return JsonResponse({'status': 'success', 'message': 'Bank details created successfully'})

            except Employee.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)



@csrf_exempt
def UpdateWorkExperience(request):
    # print("View is being accessed")
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    # Fetch the Employee instance using the emp_emailid from the session
    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

    if request.method == 'GET':
        # Fetch only the logged-in user's work experience
        work_experiences = Work_exp.objects.filter(emp_emailid=employee)

        work_exp_list = []
        for work_exp in work_experiences:
            work_exp_list.append({
                'W_Id': work_exp.W_Id,
                'start_date': work_exp.start_date,
                'end_date': work_exp.end_date,
                'comp_name': work_exp.comp_name,
                'comp_location': work_exp.comp_location,
                'designation': work_exp.designation,
                'gross_salary': work_exp.gross_salary,
                'leave_reason': work_exp.leave_reason
            })

        return JsonResponse(work_exp_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Fetch the Employee instance using the emp_emailid from the session
            emp_emailid = request.session.get('emp_emailid')
            if not emp_emailid:
                return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

            employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

            # Check if work experience with the same details already exists for the logged-in employee
            existing_work_exp = Work_exp.objects.filter(
                emp_emailid=employee,
                start_date=data['start_date'],
                end_date=data['end_date'],
                comp_name=data['comp_name'],
                comp_location=data['comp_location'],
                designation=data['designation'],
                gross_salary=data['gross_salary'],
                leave_reason=data['leave_reason']
            ).exists()

            if existing_work_exp:
                return JsonResponse({'status': 'error', 'message': 'Work experience with these details already exists'}, status=400)

            # If W_Id is provided, try to update the existing work experience
            if 'W_Id' in data:
                work_exp = Work_exp.objects.filter(W_Id=data['W_Id'], emp_emailid=employee).first()
                if not work_exp:
                    return JsonResponse({'status': 'error', 'message': 'Work experience not found for the provided W_Id'}, status=404)

                # Update the existing work experience
                work_exp.start_date = data['start_date']
                work_exp.end_date = data['end_date']
                work_exp.comp_name = data['comp_name']
                work_exp.comp_location = data['comp_location']
                work_exp.designation = data['designation']
                work_exp.gross_salary = data['gross_salary']
                work_exp.leave_reason = data['leave_reason']
                work_exp.save()

                return JsonResponse({'message': 'Work experience updated successfully', 'W_Id': work_exp.W_Id})

            else:
                # If W_Id is not provided, create a new work experience entry
                new_work_exp = Work_exp(
                    emp_emailid=employee,
                    start_date=data['start_date'],
                    end_date=data['end_date'],
                    comp_name=data['comp_name'],
                    comp_location=data['comp_location'],
                    designation=data['designation'],
                    gross_salary=data['gross_salary'],
                    leave_reason=data['leave_reason']
                )
                new_work_exp.save()

                return JsonResponse({'message': 'Work experience created successfully', 'W_Id': new_work_exp.W_Id})

        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    
    elif request.method == 'PUT':
        try:
            # Parse the incoming JSON data
            data = json.loads(request.body)

            # Retrieve the email and work experience ID
            email_to_delete = data.get('emp_emailid')  # The email ID of the employee whose record is to be deleted
            work_exp_id = data.get('W_Id')  # The ID of the specific work experience to delete

            # Check if the user's role has permission to delete
            if employee.emp_role not in ['Manager', 'Super Admin', 'HR','Employee']:
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

            # Ensure emp_emailid is provided
            if not email_to_delete:
                return JsonResponse({'status': 'error', 'message': 'emp_emailid is required'}, status=400)

            # Ensure W_Id is provided
            if not work_exp_id:
                return JsonResponse({'status': 'error', 'message': 'W_Id is required'}, status=400)

            # Fetch the specific work experience for the given email and W_Id
            work_experience = get_object_or_404(Work_exp, W_Id=work_exp_id, emp_emailid=email_to_delete)

            # Delete the specific work experience
            work_experience.delete()

            return JsonResponse({'message': 'Work experience deleted successfully'})

        except KeyError as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")
        except Work_exp.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Work experience not found'}, status=404)

    else:
        return HttpResponseNotAllowed(['GET', 'POST', 'PUT'])



@csrf_exempt
def UpdateDependent(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

    if request.method == 'GET':
        dependents = Dependent.objects.filter(emp_emailid=emp_emailid)

        dependent_list = []
        for dep in dependents:
            dependent_list.append({
                'D_name': dep.D_name,
                'D_gender': dep.D_gender,
                'D_dob': dep.D_dob,
                'D_relation': dep.D_relation,
                'D_desc': dep.D_desc,
                'D_Id': dep.D_Id,
            })

        return JsonResponse(dependent_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
        
            # Fetch the Employee instance
            employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
        
        # Create and save the new dependent
            dependent = Dependent(
                emp_emailid=employee,  # Use the Employee instance, not the email string
                D_name=data['D_name'],
                D_gender=data['D_gender'],
                D_dob=data['D_dob'],
                D_relation=data['D_relation'],
                D_desc=data['D_desc']
            )
            dependent.save()

            return JsonResponse({'message': 'Dependent added successfully', 'D_Id': dependent.D_Id})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")


    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            D_Id = data.get('D_Id')  # Get D_Id from the request body instead of query params
            if not D_Id:
                return JsonResponse({'status': 'error', 'message': 'D_Id is required'}, status=400)
            

            dependent = get_object_or_404(Dependent, D_Id=D_Id, emp_emailid=emp_emailid)
            employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
            
            # Update the dependent's information
            dependent.D_name = data['D_name']
            dependent.D_gender = data['D_gender']
            dependent.D_dob = data['D_dob']
            dependent.D_relation = data['D_relation']
            dependent.D_desc = data['D_desc']
            dependent.save()

            return JsonResponse({'message': 'Dependent updated successfully'})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    elif request.method == 'DELETE':
        try:
            D_Id = request.GET.get('D_Id')
            dependent = get_object_or_404(Dependent, D_Id = D_Id, emp_emailid=emp_emailid)

            if employee.emp_role not in ['Manager', 'Super Admin','Employee', 'HR']:
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
            
            dependent.delete()

            return JsonResponse({'message': 'Dependent deleted successfully'})
        except KeyError as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    else:
        return HttpResponseNotAllowed(['GET', 'POST', 'PUT', 'DELETE'])


@csrf_exempt
def UpdateAdhaar(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':

        adhaar = Adhaar.objects.filter(emp_emailid=emp_emailid).first()
        if adhaar:
            data = {
                'adhaar_no': adhaar.adhaar_no,
                'adhaar_name': adhaar.adhaar_name,
                'enroll_no': adhaar.enroll_no,
                'adhaar_pic': adhaar.adhaar_pic.url if adhaar.adhaar_pic else ''
            }
            return JsonResponse(data)
        else:
            return JsonResponse({'error': 'Adhaar details not found!'}, status=404)

    elif request.method == 'POST':
        data = request.POST
        file = request.FILES.get('adhaar_pic')

        employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
        adhaar, created = Adhaar.objects.get_or_create(
            emp_emailid=employee,
            defaults={
                'adhaar_no': data['adhaar_no'],
                'adhaar_name': data['adhaar_name'],
                'enroll_no': data['enroll_no'],
                'adhaar_pic': file
            }
        )
        if created:
            return JsonResponse({'message': 'Adhaar details created successfully!'}, status=201)
        else:
            adhaar.adhaar_no = data['adhaar_no']
            adhaar.adhaar_name = data['adhaar_name']
            adhaar.enroll_no = data['enroll_no']
            if file:
                adhaar.adhaar_pic = file
            adhaar.save()
            return JsonResponse({'message': 'Adhaar details updated successfully!'})

    elif request.method == 'DELETE':
        adhaar = Adhaar.objects.filter(emp_emailid=emp_emailid).first()
        if adhaar:
            adhaar.delete()
            return JsonResponse({'message': 'Adhaar details deleted successfully!'})
        else:
            return JsonResponse({'error': 'Adhaar details not found!'}, status=404)

    else:
        return JsonResponse({'error': 'Method not allowed!'}, status=405)


@csrf_exempt
def UpdateLicence(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

    if request.method == 'GET':
        licence = Licence.objects.filter(emp_emailid=employee).first()
        if licence:
            data = {
                'licence_no': licence.licence_no,
                'licence_name': licence.licence_name,
                'expiry_date': licence.expiry_date,
                'licence_pic': licence.licence_pic.url if licence.licence_pic else None,
            }
            return JsonResponse(data)
        else:
            return JsonResponse({'message': 'No licence found'}, status=404)

    elif request.method == 'POST':
        data = request.POST
        file = request.FILES.get('licence_pic')
        licence_record, created = Licence.objects.update_or_create(
            emp_emailid=employee,
            defaults={
                'licence_no': data['licence_no'],
                'licence_name': data['licence_name'],
                'expiry_date': data['expiry_date'],
                'licence_pic': file
            }
        )
        if created:
            return JsonResponse({'message': 'Licence details created successfully!'}, status=201)
        else:
            # Update the licence details if already exists
            licence_record.licence_no = data['licence_no']
            licence_record.licence_name = data['licence_name']
            licence_record.expiry_date = data['expiry_date']
            if file:
                licence_record.licence_pic = file
            licence_record.save()
            return JsonResponse({'message': 'Licence details updated successfully!'})

    elif request.method == 'DELETE':
        licence = get_object_or_404(Licence, emp_emailid=employee)
        if licence.licence_pic:
            licence.licence_pic.delete(save=False)
        licence.delete()
        return JsonResponse({'message': 'Licence details deleted successfully!'})

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def UpdatePassport(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':
        passport = Passport.objects.filter(emp_emailid=emp_emailid).first()
        if passport:
            data = {
                'passport_no': passport.passport_no,
                'passport_name': passport.passport_name,
                'passport_validity': passport.passport_validity,
                'passport_pic': passport.passport_pic.url if passport.passport_pic else ''
            }
            return JsonResponse(data)
        else:
            return JsonResponse({'error': 'Passport details not found!'}, status=404)

    elif request.method == 'POST':
        data = request.POST
        file = request.FILES.get('passport_pic')

        employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
        passport, created = Passport.objects.get_or_create(
            emp_emailid=employee,
            defaults={
                'passport_no': data['passport_no'],
                'passport_name': data['passport_name'],
                'passport_validity': data['passport_validity'],
                'passport_pic': file
            }
        )
        if created:
            return JsonResponse({'message': 'Passport details created successfully!'}, status=201)
        else:
            passport.passport_no = data['passport_no']
            passport.passport_name = data['passport_name']
            passport.passport_validity = data['passport_validity']
            if file:
                passport.passport_pic = file
            passport.save()
            return JsonResponse({'message': 'Passport details updated successfully!'})

    elif request.method == 'DELETE':
        passport = Passport.objects.filter(emp_emailid=emp_emailid).first()
        if passport:
            passport.delete()
            return JsonResponse({'message': 'Passport details deleted successfully!'})
        else:
            return JsonResponse({'error': 'Passport details not found!'}, status=404)

    else:
        return JsonResponse({'error': 'Method not allowed!'}, status=405)


@csrf_exempt
def UpdatePan(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':
        pan = Pan.objects.filter(emp_emailid=emp_emailid).first()
        if pan:
            data = {
                'pan_no': pan.pan_no,
                'pan_name': pan.pan_name,
                'pan_pic': pan.pan_pic.url if pan.pan_pic else ''
            }
            return JsonResponse(data)
        else:
            return JsonResponse({'error': 'PAN details not found!'}, status=404)

    elif request.method == 'POST':
        data = request.POST
        file = request.FILES.get('pan_pic')

        employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
        pan, created = Pan.objects.get_or_create(
            emp_emailid=employee,
            defaults={
                'pan_no': data['pan_no'],
                'pan_name': data['pan_name'],
                'pan_pic': file
            }
        )
        if created:
            return JsonResponse({'message': 'PAN details created successfully!'}, status=201)
        else:
            pan.pan_no = data['pan_no']
            pan.pan_name = data['pan_name']
            if file:
                pan.pan_pic = file
            pan.save()
            return JsonResponse({'message': 'PAN details updated successfully!'})

    elif request.method == 'DELETE':
        pan = Pan.objects.filter(emp_emailid=emp_emailid).first()
        if pan:
            pan.delete()
            return JsonResponse({'message': 'PAN details deleted successfully!'})
        else:
            return JsonResponse({'error': 'PAN details not found!'}, status=404)

    else:
        return JsonResponse({'error': 'Method not allowed!'}, status=405)


@csrf_exempt
def UpdateQualification(request):
    emp_emailid = request.session.get('emp_emailid')
    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':
        qualifications = Qualification.objects.filter(emp_emailid=emp_emailid)

        qualification_list = []
        for qual in qualifications:
            qualification_list.append({
                'Q_Id': qual.Q_Id,
                'q_type': qual.q_type,
                'q_degree': qual.q_degree,
                'q_clg': qual.q_clg,
                'q_uni': qual.q_uni,
                'q_duration': qual.q_duration,
                'q_yop': qual.q_yop,
                'q_comment': qual.q_comment,
            })

        return JsonResponse(qualification_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

            qualification = Qualification(
                emp_emailid = employee,
                q_type=data['q_type'],
                q_degree=data['q_degree'],
                q_clg=data['q_clg'],
                q_uni=data['q_uni'],
                q_duration=data['q_duration'],
                q_yop=data['q_yop'],
                q_comment=data['q_comment']
            )
            qualification.save()

            return JsonResponse({'message': 'Qualification added successfully', 'Q_Id': qualification.Q_Id})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            Q_Id = request.GET.get('Q_Id')
            qualification = Qualification.objects.get(Q_Id=Q_Id, emp_emailid=emp_emailid)
            emp_emailid=employee,
            qualification.q_type = data['q_type']
            qualification.q_degree = data['q_degree']
            qualification.q_clg = data['q_clg']
            qualification.q_uni = data['q_uni']
            qualification.q_duration = data['q_duration']
            qualification.q_yop = data['q_yop']
            qualification.q_comment = data['q_comment']
            qualification.save()

            return JsonResponse({'message': 'Qualification updated successfully'})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")
        except Qualification.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Qualification not found'}, status=404)

    elif request.method == 'DELETE':
        try:
            Q_Id = request.GET.get('Q_Id')
            qualification = Qualification.objects.get(Q_Id=Q_Id, emp_emailid=emp_emailid)

            employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

            if employee.emp_role not in ['Manager', 'Super Admin','Employee', 'HR']:
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
            
            qualification.delete()

            return JsonResponse({'message': 'Qualification deleted successfully'})
        except (KeyError, Qualification.DoesNotExist) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    else:
        return HttpResponseNotAllowed(['GET', 'POST', 'PUT', 'DELETE'])


@csrf_exempt
def UpdateFamilyDetails(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)
    

     # Fetch the Employee instance using emp_emailid
    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

    if request.method == 'GET':
        family_details = Family_details.objects.filter(emp_emailid=emp_emailid)

        family_list = []
        for fam in family_details:
            family_list.append({
                'F_Id': fam.F_Id,
                'F_name': fam.F_name,
                'F_gender': fam.F_gender,
                'F_dob': fam.F_dob,
                'F_contact': fam.F_contact,
                'F_mail': fam.F_mail,
                'F_relation': fam.F_relation,
                'F_comment': fam.F_comment,
            })

        return JsonResponse(family_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            family_detail = Family_details(
                emp_emailid=employee,
                F_name=data['F_name'],
                F_gender=data['F_gender'],
                F_dob=data['F_dob'],
                F_contact=data['F_contact'],
                F_mail=data['F_mail'],
                F_relation=data['F_relation'],
                F_comment=data['F_comment']
            )
            family_detail.save()

            return JsonResponse({'message': 'Family detail added successfully', 'F_Id': family_detail.F_Id})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            F_Id = request.GET.get('F_Id')
            family_detail = Family_details.objects.get(F_Id=F_Id, emp_emailid=emp_emailid)

            family_detail.F_name = data['F_name']
            family_detail.F_gender = data['F_gender']
            family_detail.F_dob = data['F_dob']
            family_detail.F_contact = data['F_contact']
            family_detail.F_mail = data['F_mail']
            family_detail.F_relation = data['F_relation']
            family_detail.F_comment = data['F_comment']
            family_detail.save()

            return JsonResponse({'message': 'Family detail updated successfully'})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")
        except Family_details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Family detail not found'}, status=404)

    elif request.method == 'DELETE':
        try:
            F_Id = request.GET.get('F_Id')
            family_detail = Family_details.objects.get(F_Id=F_Id, emp_emailid=emp_emailid)

            if employee.emp_role not in ['Manager', 'Super Admin','Employee', 'HR']:
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)


            family_detail.delete()

            return JsonResponse({'message': 'Family detail deleted successfully'})
        except (KeyError, Family_details.DoesNotExist) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    else:
        return HttpResponseNotAllowed(['GET', 'POST', 'PUT', 'DELETE'])


# Case Management section
@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AllCases(request):
    def TimeAgo(ptime):
        current_time = timezone.now()
        estimate_time = current_time - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(created_by__d_id__c_id=company_id)

    case_data = []
    for case in cases:
        timeago = TimeAgo(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by_profile': created_by.emp_profile.url
        })
    return JsonResponse({'cases': case_data})


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CaseInfo(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        case_id = request.GET.get('case_id')
        if not case_id:
            return JsonResponse({"error": "Case ID not provided"}, status=400)

        case = get_object_or_404(Case, case_id=case_id)
        # if case.created_by.d_id.c_id != company_id:
        #     print(case.created_by.d_id.c_id)
        #     print(company_id)
        #     return JsonResponse({"error": "Unauthorized access"}, status=403)


        case_data = {
            'create_for': case.create_for,
            'case_type': case.case_type,
            'case_title': case.case_title,
            'case_desc': case.case_desc,
            'case_date': case.case_date.strftime('%Y-%m-%d %H:%M:%S'),
            'case_status': case.case_status,
            'case_id': case.case_id,
            'created_by': {
                'emp_name': case.created_by.emp_name,
                'emp_emailid': case.created_by.emp_emailid,
                'emp_phone': case.created_by.emp_phone,
                'emp_profile': case.created_by.emp_profile.url,
            }
        }

        return JsonResponse({'case': case_data})

    elif request.method == 'POST':
        user_email = request.session.get('emp_emailid')
        case_id = request.GET.get('case_id')

        if not case_id or not user_email:
            return JsonResponse({"error": "Case ID or User Email not provided"}, status=400)

        case = get_object_or_404(Case, case_id=case_id)
        employee = get_object_or_404(Employee, emp_emailid=user_email)

        case.assigned_to = employee
        case.save()
        return JsonResponse({"message": "Case assigned successfully"}, status=200)

    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def BenefitsCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='Benefits'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url
        })

    context = {
        'cases': case_data
    }

    return JsonResponse({'cases': case_data})


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def TravelExpenseCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='Travel and Expense'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url
        })

    context = {
        'cases': case_data
    }

    return JsonResponse({'cases': case_data})


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CompensationPayrollCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='Compensation and payroll'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url
        })

    context = {
        'cases': case_data
    }

    return JsonResponse({'cases': case_data})


# incomplete pending uncomment , need to complete this (Employee Jobs)
@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ManagerRating(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        # Tasks
        tasks_data = Tasks.objects.filter(emp_emailid__d_id__c_id=c_id).values('emp_emailid', 'selfratings', 'status', 'tid__description' )
        print(tasks_data)
        tasks = list(tasks_data)
        return JsonResponse(tasks_data, safe=False)


@csrf_exempt
def FAQManagement(request):
    if request.method == 'GET':
        c_id = request.session.get('c_id')  # Retrieve company ID from session
        if not c_id:
            return JsonResponse({'error': 'User not authenticated or company ID not found'}, status=401)

        faqs = Faqs.objects.filter(c_id=c_id)  # Filter FAQs by the logged-in user's company ID

        if not faqs.exists():
            return JsonResponse({'message': 'No FAQs pending'})

        faqs_list = []
        for faq in faqs:
            faqs_list.append({
                'faq_id': faq.faq_id,
                'question': faq.question,
                'answer': faq.answer if faq.answer else None,
                'emp_emailid': faq.emp_emailid.emp_emailid,
                'imp': faq.imp,
                'c_id': faq.c_id.c_id,
            })

        return JsonResponse({'faqs': faqs_list})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            faq_id = data.get('faq_id')  # Get the FAQ ID from the POST data
            answer = data.get('answer')    # Get the answer from the POST data
            
            if not faq_id or answer is None:
                return JsonResponse({'error': 'FAQ ID and answer are required'}, status=400)

            # Update or create the FAQ with the provided FAQ ID and answer
            faq, created = Faqs.objects.update_or_create(
                faq_id=faq_id,
                defaults={'answer': answer}
            )
            return JsonResponse({'message': 'FAQ updated successfully' if not created else 'FAQ created successfully'})

        except Exception as e:
            return HttpResponseBadRequest({'error': str(e)})

    elif request.method == 'DELETE':
        faq_id = request.GET.get('faq_id')  # Get FAQ ID from the query parameters
        if not faq_id:
            return JsonResponse({'error': 'FAQ ID is required'}, status=400)

        try:
            faq = Faqs.objects.filter(faq_id=faq_id, c_id=request.session.get('c_id'))  # Ensure the user can only delete FAQs from their company
            if faq.exists():
                faq.delete()  # Delete the FAQ
                return JsonResponse({'message': 'FAQ deleted successfully.'}, status=204)  # No Content response
            return JsonResponse({'error': 'FAQ not found.'}, status=404)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)




@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EnrollEmployee(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        courses = Courses.objects.filter(c_id=c_id).values('course_id', 'course_title').distinct()

        employees = Employee.objects.filter(d_id__c_id=c_id).values('emp_emailid')

        return JsonResponse({'courses': list(courses), 'employees': list(employees)})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            course_id = data.get('course_id')
            course_title = data.get('course_title')
            emp_emailid = data.get('emp_emailid')

            if not course_title:
                return JsonResponse({'error': 'Course title cannot be empty'}, status=400)

            if not Employee.objects.filter(emp_emailid=emp_emailid, d_id__c_id=c_id).exists():
                return JsonResponse({'error': 'Employee does not belong to the same company'}, status=400)

            course_company_id = Courses.objects.get(course_id=course_id).c_id_id
            if course_company_id != c_id:
                return JsonResponse({'error': 'This course does not belong to your company'}, status=400)

            if Course_employee.objects.filter(course_id=course_id, emp_emailid=emp_emailid).exists():
                return JsonResponse({'error': 'Employee is already enrolled in this course'}, status=400)

            Course_employee.objects.create(
                course_id_id=course_id,
                emp_emailid_id=emp_emailid,
                status=0,
                course_title=course_title
            )

            return JsonResponse({'message': 'Employee enrolled successfully'}, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ViewAllEnrollments(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        enrollments = Course_employee.objects.filter(course_id__c_id=c_id).values('id', 'course_title', 'emp_emailid')

        return JsonResponse({'enrollments': list(enrollments)})

    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            id = data.get('id')

            enrollment = Course_employee.objects.get(id=id)
            enrollment.delete()
            return JsonResponse({'message': 'Enrollment deleted successfully'}, status=200)
        except Course_employee.DoesNotExist:
            return JsonResponse({'error': 'Enrollment not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AdhocPayments(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        allowances = Adhoc.objects.all().values('name', 'dept', 'type', 'amt', 'mon', 'year')
        allowances_list = list(allowances)
        return JsonResponse(allowances_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            adhoc_payment = Adhoc(
                name=data.get('name'),
                dept=data.get('dept'),
                type=data.get('type'),
                amt=data.get('amt'),
                mon=data.get('mon'),
                year=data.get('year')
            )
            adhoc_payment.save()

            return JsonResponse({'message': "Allowance or Deduction added successfully"}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except IntegrityError as e:
            return JsonResponse({"error": "Database integrity error: " + str(e)}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def LoanPayments(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        all_loans = list(Loan.objects.all().values('id', 'name', 'department', 'lamt', 'mamt', 'startdate', 'reason', 'status'))
        pending_loans = list(Loan.objects.filter(status=0).values('id', 'name', 'department', 'lamt', 'mamt', 'startdate', 'reason', 'status'))
        approved_loans = list(Loan.objects.filter(status=1).values('id', 'name', 'department', 'lamt', 'mamt', 'startdate', 'reason', 'status'))

        response_data = {
            'all_loans': all_loans,
            'pending_loans': pending_loans,
            'approved_loans': approved_loans
        }

        return JsonResponse(response_data, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            loan_id = data.get('id')
            if not loan_id:
                return JsonResponse({"error": "Loan ID not provided"}, status=400)

            loan = Loan.objects.get(id=loan_id)
            loan.status = 1
            loan.save()

            return JsonResponse({'message': "Loan approved successfully"}, status=200)
        except Loan.DoesNotExist:
            return JsonResponse({"error": "Loan not found"}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def LeaveEncashment(request):
#     user_id = request.session.get('user_id')
#     company_id = request.session.get('c_id')

#     if not user_id or not company_id:
#         return JsonResponse({'error': 'User not logged in'}, status=401)

#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             employee = Employee.objects.get(emp_emailid=user_id)

#             leave_encashment = Leave_Encashment(
#                 txndt = data.get('txndt'),
#                 refn = data.get('refn'),
#                 effdt = data.get('effdt'),
#                 emp = data.get('emp'),
#                 type = data.get('type'),
#                 blnc = data.get('blnc'),
#                 pdays = data.get('pdays'),
#                 sal = data.get('sal'),
#                 emp_emailid = employee
#             )
#             leave_encashment.save()

#             return JsonResponse({"message": "Leave inserted successfully"}, status=201)
#         except Employee.DoesNotExist:
#             return JsonResponse({"error": "Employee not found"}, status=404)
#         except Exception as e:
#             return JsonResponse({"error": str(e)}, status=400)
#     else:
#         return JsonResponse({"error": "Invalid request method"}, status=405)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def LeaveEncashment(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            employee = Employee.objects.get(emp_emailid=user_id)

            leave_encashment = Leave_Encashment(
                txndt=data.get('txndate'),  # Changed to match the frontend
                refn=data.get('referencenum'),
                effdt=data.get('effectivedate'),
                emp=data.get('empname'),
                type=data.get('leaveType'),
                blnc=data.get('leavesallowed'),
                pdays=data.get('totalleaves'),
                sal=data.get('salary'),
                emp_emailid=employee
            )
            leave_encashment.save()

            return JsonResponse({"message": "Leave inserted successfully"}, status=201)
        except Employee.DoesNotExist:
            return JsonResponse({"error": "Employee not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    elif request.method == 'GET':
        # Fetch the leave encashment data for the logged-in user
        try:
            employee = Employee.objects.get(emp_emailid=user_id)
            leave_encashments = Leave_Encashment.objects.filter(emp_emailid=employee)
            data = list(leave_encashments.values())
            return JsonResponse(data, safe=False)
        except Employee.DoesNotExist:
            return JsonResponse({"error": "Employee not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    
    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@role_required(['HR', 'Manager', 'Super Admin','Employee'])
@csrf_exempt
def ViewLeaveEncashment(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            leave_encashments = Leave_Encashment.objects.filter(
                emp_emailid__d_id__c_id=company_id
            ).values(
                'emp',
                'emp_emailid__emp_name',
                'blnc',
                'pdays',
                'enclve'
            ).distinct('emp')
            
            employee_list = [
                {
                    'emp': item['emp'],
                    'name': item['emp'],
                    'totalLeaves': item['blnc'],
                    'leavesTaken': item['pdays'],
                    'encashment': item['enclve']
                } for item in leave_encashments
            ]
            
            return JsonResponse({'employees': employee_list}, safe=False)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def OffCyclePayments(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        off_cycle_payments = Off_cycle.objects.all().values()
        off_cycle_list = list(off_cycle_payments)
        return JsonResponse(off_cycle_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            off_cycle_payment = Off_cycle(
                name=data.get('name'),
                tname=data.get('tname'),
                amt=data.get('amt'),
                effdt=data.get('effdt'),
                note=data.get('note')
            )
            off_cycle_payment.save()

            return JsonResponse({"message": "Off-cycle payment inserted successfully"}, status=201)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    elif request.method == 'DELETE':
        try:
            # Extract the ID from the URL
            payment_id = request.GET.get('id')

            if not payment_id:
                return JsonResponse({"error": "Payment ID is required"}, status=400)

            # Fetch the off-cycle payment by ID
            off_cycle_payment = Off_cycle.objects.get(id=payment_id)
            off_cycle_payment.delete()

            return JsonResponse({"message": "Off-cycle payment deleted successfully"}, status=200)
        except Off_cycle.DoesNotExist:
            return JsonResponse({"error": "Off-cycle payment not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)  
    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def BankTransferPayout(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        payouts = Salary.objects.filter(paymentmethod='bank').order_by('-payout_month').values('sal_id', 'payout_month').distinct()
        return JsonResponse({'payouts': list(payouts)})
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def BankTransfer(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        month = request.GET.get('month')
        if not month:
            return JsonResponse({'error': 'Month parameter is required'}, status=400)

    try:
        salaries = Salary.objects.filter(paymentmethod='bank', payout_month=month)
        total_employees = salaries.count()
        total_amount = salaries.filter(holdsalary=0, paid=0).aggregate(Sum('Net_Salary'))['Net_Salary__sum'] or 0
        total_employees_with_errors = salaries.filter(holdsalary=1).count()

        salary_details = salaries.select_related('emp_emailid').values(
            'sal_id', 'emp_emailid__holder_name', 'emp_emailid__bank_name', 'emp_emailid__branch',
            'emp_emailid__ifsc', 'emp_emailid__acc_no', 'emp_emailid__acc_type', 'Net_Salary',
            'holdsalary', 'paid'
        )

        response_data = {
            'month': month,
            'total_employees': total_employees,
            'total_amount': total_amount,
            'total_employees_with_errors': total_employees_with_errors,
            'salary_details': list(salary_details)
        }

        return JsonResponse(response_data, status=200)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def BankTransferUpdate(request):
    c_id = request.session.get('c_id')
    if not c_id:
        print('Company ID not found in session')
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    month = request.GET.get('month')
    if not month:
        print('Month parameter is required')
        return JsonResponse({'error': 'Month parameter is required'}, status=400)

    if request.method == 'POST':
        try:
            narration = request.POST.get('narration')
            debit_account = request.POST.get('debitac')
            file_name_start = request.POST.get('filename')
            value_date = request.POST.get('valuedate')

            print(f"Received data: narration={narration}, debit_account={debit_account}, file_name_start={file_name_start}, value_date={value_date}, month={month}")

            batch_id = Banktransferstatement.objects.all().count() + 1

            salaries = Salary.objects.filter(holdsalary=0, paymentmethod='bank', paid=0, payout_month=month)
            for salary in salaries:
                print(f"Creating Banktransferstatement for {salary.emp_emailid.holder_name}")
                Banktransferstatement.objects.create(
                    batchid=batch_id,
                    name=salary.emp_emailid.holder_name,
                    bank=salary.emp_emailid.bank_name,
                    branch=salary.emp_emailid.branch,
                    ifsc=salary.emp_emailid.ifsc,
                    accountno=salary.emp_emailid.acc_no,
                    amount=salary.Net_Salary
                )

            Banktransferstatement.objects.filter(batchid=batch_id).update(
                debitno=debit_account,
                date=datetime.strptime(value_date, '%Y-%m-%d').date(),
                narration=narration,
                filename=file_name_start
            )

            salaries.update(paid=1)

            print("Bank Transfer Statement generated successfully.")
            return JsonResponse({'message': 'Bank Transfer Statement Generated successfully.'}, status=200)

        except Exception as e:
            print(f"Error during Bank Transfer Update: {str(e)}")
            return JsonResponse({'error': str(e)}, status=400)

    else:
        print('Method not allowed')
        return JsonResponse({'error': 'Method not allowed'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CashChequeTransferPayout(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            payouts = Salary.objects.filter(paymentmethod='cash').order_by('-payout_month').values('sal_id', 'payout_month').distinct()

            payout_list = [{'sal_id': payout['sal_id'], 'payout_month': payout['payout_month']} for payout in payouts]

            return JsonResponse({'payroll_months': payout_list}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def HoldSalaryPayout(request):
#     c_id = request.session.get('c_id')
#     if not c_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'GET':
#         try:
#             payouts = Salary1.objects.filter().values('sal_id', 'holdsalary', 'emp_emailid__holder_name', 'emp_emailid__bank_name', 'emp_emailid__branch', 'emp_emailid__ifsc', 'emp_emailid__acc_no').distinct()

#             payout_list = []
#             for payout in payouts:
#                 payout_dict = {
#                     'sal_id': payout['sal_id'],
#                     'holder_name': payout['emp_emailid__holder_name'],
#                     'bank_name': payout['emp_emailid__bank_name'],
#                     'branch': payout['emp_emailid__branch'],
#                     'ifsc': payout['emp_emailid__ifsc'],
#                     'acc_no': payout['emp_emailid__acc_no'],
#                     'hold_salary': payout['holdsalary']
#                 }
#                 payout_list.append(payout_dict)

#             return JsonResponse({'payouts': payout_list}, status=200)

#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)

#     else:
#         return JsonResponse({'error': 'Method not allowed'}, status=405)







@csrf_exempt
def HoldSalaryPayout(request):
    c_id = request.session.get('c_id')
    
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        emailid = request.GET.get('emailid')

        # Decode URL-encoded email ID for debugging purposes
        decoded_emailid = urllib.parse.unquote(emailid)
        print(f"Decoded email ID: {decoded_emailid}")
        
        if not emailid:
            return JsonResponse({'error': 'Email ID not provided'}, status=400)

        try:
            payouts = Salary1.objects.filter(
                emp_emailid=emailid
            ).values(
                'emp_emailid__holder_name',
                'emp_emailid__bank_name',
                'emp_emailid__branch',
                'emp_emailid__ifsc',
                'emp_emailid__acc_no',
                'sal_id',
                'holdsalary'
            ).distinct()

            if not payouts:
                return JsonResponse({'error': 'No salary details found for this employee'}, status=404)

            payout_list = []
            for payout in payouts:
                payout_dict = {
                    'sal_id': payout['sal_id'],
                    'holder_name': payout['emp_emailid__holder_name'],
                    'bank_name': payout['emp_emailid__bank_name'],
                    'branch': payout['emp_emailid__branch'],
                    'ifsc': payout['emp_emailid__ifsc'],
                    'acc_no': payout['emp_emailid__acc_no'],
                    'holdsalary': payout['holdsalary']
                }
                payout_list.append(payout_dict)
                print("Payout List:")
                for item in payout_list:
                    print(item)


            return JsonResponse({'payouts': payout_list}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Method not allowed'}, status=405)



# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def HoldSalary(request):
#     c_id = request.session.get('c_id')
#     emailid = request.GET.get('emailid')
#     if not c_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'GET':
#         try:
#             emp_name = Salary1.objects.filter(emp_emailid=emailid).values('emp_emailid__holder_name').first()

#             if not emp_name:
#                 return JsonResponse({'error': 'Employee not found'}, status=404)
#             return JsonResponse(emp_name, status=200)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)

#     elif request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             hold_reason = data.get('hold_reason')
#             remarks = data.get('remarks')

#             if not hold_reason or not remarks:
#                 return JsonResponse({'error': 'Hold reason and remarks are required'}, status=400)

#             salary = Salary.objects.filter(emp_emailid=emailid, emp_emailid__emp_emailid__d_id__c_id=c_id).first()
#             if not salary:
#                 return JsonResponse({'error': 'Salary or Bank details not found for the given email ID'}, status=404)

#             salary.holdsalary = 1
#             salary.notes = hold_reason
#             salary.remarks = remarks
#             salary.save()

#             return JsonResponse({'message': f'{salary.emp_emailid.holder_name} has been put on hold for salary.'}, status=200)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)
#     else:
#         return JsonResponse({'error': 'Method not allowed'}, status=405)
    






@csrf_exempt
def HoldSalary(request):
    c_id = request.session.get('c_id')

    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            employees = Employee.objects.filter(d_id__c_id=c_id).values('emp_name', 'emp_emailid')
            if not employees:
                return JsonResponse({'error': 'No employees found for this company'}, status=404)
            return JsonResponse({'employees': list(employees)}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            emp_emailid = data.get('employeeId')
            hold_reason = data.get('holdReason')
            remarks = data.get('remarks')

            if not emp_emailid or not hold_reason or not remarks:
                return JsonResponse({'error': 'Employee ID, hold reason, and remarks are required'}, status=400)

            # Retrieve the Bank_details instance for the given email
            bank_details = Bank_details.objects.filter(emp_emailid=emp_emailid).first()
            if not bank_details:
                return JsonResponse({'error': 'Bank details not found for the given employee ID'}, status=404)

            # Find the salary record for the given employee and c_id
            salary, created = Salary1.objects.update_or_create(
                emp_emailid=bank_details,  # Use the Bank_details instance
                defaults={
                    'holdsalary': 1,
                    'notes': hold_reason,
                    'remarks': remarks
                }
            )

            message = 'Salary held successfully' if not created else 'Salary record created and held successfully'
            return JsonResponse({'message': message}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)




# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def UnholdSalary(request):
#     c_id = request.session.get('c_id')
#     emailid = request.GET.get('emailid')
#     if not c_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'POST':
#         try:
#             salary = Salary1.objects.filter(emp_emailid=emailid, emp_emailid_emp_emailidd_id_c_id=c_id).first()
#             if not salary:
#                 return JsonResponse({'error': 'Salary or Bank details not found for the given email ID'}, status=404)

#             salary.holdsalary = 0
#             salary.save()

#             return JsonResponse({'message': f'{salary.emp_emailid.holder_name} is no more on hold for salary.'}, status=200)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)
#     else:
#         return JsonResponse({'error': 'Method not allowed'}, status=405)
    


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])  # Assuming this decorator checks roles
def UnholdSalary(request):
    c_id = request.session.get('c_id')  # Get the company ID from session
    emailid = request.GET.get('emailid')  # Get employee email from GET parameters
    
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'POST':
        try:
            # Query to check if employee exists within the given company through their department
            employee = Employee.objects.filter(emp_emailid=emailid, d_id__c_id=c_id).first()
            
            if not employee:
                return JsonResponse({'error': 'Employee not found in the specified company'}, status=404)

            # Now that we have the employee, we fetch their salary record
            salary = Salary1.objects.filter(emp_emailid=employee.emp_emailid).first()

            if not salary:
                return JsonResponse({'error': 'Salary details not found for the given email ID'}, status=404)

            # Unhold the salary by setting the holdsalary field to 0
            salary.holdsalary = 0
            salary.save()

            return JsonResponse({'message': f'{employee.emp_name} is no more on hold for salary.'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Method not allowed'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def PayslipPayout(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        payouts = Salary.objects.order_by('-payout_month').values('sal_id', 'payout_month').distinct()
        return JsonResponse({'payouts': list(payouts)})
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def GeneratePayslip(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        month = request.GET.get('month')
        if not month:
            return JsonResponse({'error': 'Month parameter is required'}, status=400)

        payslip = Salary.objects.filter(emp_emailid__emp_emailid__d_id__c_id=c_id, payout_month=month).select_related('emp_emailid').values(
            'sal_id',
            'emp_emailid__holder_name',
            'emp_emailid__bank_name',
            'emp_emailid__branch',
            'emp_emailid__acc_type',
            'emp_emailid__acc_no'
        )

        return JsonResponse({'payouts': list(payslip)})
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def HomeSalary(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        employees = Employee.objects.filter(d_id__in=Department.objects.filter(c_id=c_id)).values('emp_name', 'emp_emailid', 'emp_role')
        return JsonResponse({'employees': list(employees)})
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddSalary(request):
    print("View reached")
    if request.method == 'POST':
        print("View reached")
        emp_emailid = request.GET.get('emp_emailid')
        print("View reached")

        try:
            employee = Employee.objects.get(emp_emailid=emp_emailid)
            personal_details = Personal_details.objects.get(emp_emailid=emp_emailid)
            poifiles_new = Poifiles_new.objects.get(emp_emailid=emp_emailid)

            age = datetime.now().year - personal_details.birth_date.year
            actual80c = poifiles_new.actualAmount_80C
            actual80d = poifiles_new.actualAmount_80D
            actualoie = poifiles_new.OIE_actualAmount
            actualosi = poifiles_new.OSI_actualAmount

            eligible_deductions = actual80c + actual80d + actualoie + actualosi

            annual_ctc = 0
            remarks = ""
            notes = ""

            # Retrieve salary details
            salary_instance = Salary.objects.filter(emp_emailid=emp_emailid).order_by('-sal_id').first()
            if salary_instance:
                annual_ctc = salary_instance.annual_ctc
                remarks = salary_instance.remarks
                notes = salary_instance.notes

            payout_month = request.POST.get('payoutMonth', '')
            effective_from = request.POST.get('effectiveFrom', '')
            revise = int(request.POST.get('revision', 0))

            if not annual_ctc:
                annual_ctc = int(request.POST.get('annualCTC', 0))

            annual_ctc += annual_ctc * (revise / 100)
            monthly_ctc = annual_ctc / 12
            basic = monthly_ctc * 0.4
            hra = basic * 0.4
            conveyance = 1600
            da = 0
            special_allowance = monthly_ctc - (hra + conveyance + da + basic)

            annual_taxable = annual_ctc - eligible_deductions
            tax_liability = tax_calculation_to_add_salary(age, annual_taxable)

            monthly_tds = tax_liability / 12
            monthly_epf = basic * 0.12
            monthly_prof_tax = 200
            net_salary = monthly_ctc - (monthly_tds + monthly_epf + monthly_prof_tax)

            # Save salary details to database
            salary_instance = Salary(
                emp_emailid=emp_emailid,
                basic=basic,
                hra=hra,
                conveyance=conveyance,
                da=da,
                special_allowance=special_allowance,
                monthly_ctc=monthly_ctc,
                annual_ctc=annual_ctc,
                Eligible_Deductions=eligible_deductions,
                Yearly_Taxable_Salary=annual_taxable,
                otal_Tax_Liability=tax_liability,
                Monthly_TDS=monthly_tds,
                Monthly_EPF=monthly_epf,
                Monthly_Professional_Tax=monthly_prof_tax,
                Net_Salary=net_salary,
                payout_month=payout_month,
                effective_from=effective_from,
                notes=notes,
                remarks=remarks,
                revision=revise
            )
            salary_instance.save()
            return JsonResponse({'message': 'Salary details saved successfully!'})

        except ObjectDoesNotExist:
            return JsonResponse({'error': 'Employee with the provided email ID does not exist'}, status=404)

    else:
        emp_emailid = request.GET.get('emp_emailid')

        # Retrieve employee details from models
        employee = Employee.objects.get(emp_emailid=emp_emailid)
        personal_details = Personal_details.objects.get(emp_emailid=emp_emailid)
        poifiles_new = Poifiles_new.objects.get(emp_emailid=emp_emailid)

        age = datetime.now().year - personal_details.birth_date.year
        actual80c = poifiles_new.actualAmount_80C
        actual80d = poifiles_new.actualAmount_80D
        actualoie = poifiles_new.OIE_actualAmount
        actualosi = poifiles_new.OSI_actualAmount

        eligible_deductions = actual80c + actual80d + actualoie + actualosi

        annual_ctc = 0
        remarks = ""
        notes = ""

        # Retrieve salary details
        salary_instance = Salary.objects.filter(emp_emailid=emp_emailid).order_by('-sal_id').first()
        if salary_instance:
            annual_ctc = salary_instance.annual_ctc
            remarks = salary_instance.remarks
            notes = salary_instance.notes

        context = {
            'employee': employee,
            'personal_details': personal_details,
            'poifiles_new': poifiles_new,
            'age': age,
            'eligible_deductions': eligible_deductions,
            'annual_ctc': annual_ctc,
            'remarks': remarks,
            'notes': notes,
        }

        return JsonResponse(context)





# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def AddSalary1(request):
#     if request.method == 'GET':
#         # Get company id from session
#         c_id = request.session.get('c_id')
#         if not c_id:
#             return JsonResponse({'error': 'Company ID not found in session'}, status=400)

#         # Retrieve all employees that belong to departments under this company
#         employees = Employee.objects.filter(d_id__c_id=c_id)

#         # Prepare employee data to return to the frontend
#         employee_data = [{'emp_emailid': emp.emp_emailid, 'name': emp.emp_name} for emp in employees]

#         return JsonResponse({'employees': employee_data})

#     elif request.method == 'POST':
#         data = json.loads(request.body)
#         emp_emailid = data.get('emp_emailid')
#         payout_month = data.get('payoutMonth')
#         revision = int(data.get('revision', 0))
#         annual_ctc = int(data.get('annualCTC', 0))
#         payment_method = data.get('paymentMethod')
#         employee_remarks = data.get('employeeRemarks')
#         notes = data.get('notes')
#         effective_from = data.get('effectiveFrom')

#         c_id = request.session.get('c_id')

#         if not emp_emailid or not c_id:
#             return JsonResponse({'error': 'Missing employee or company details'}, status=400)

#         try:
#             # Retrieve the employee's bank details from Bank_details using emp_emailid
#             bank_details = Bank_details.objects.get(emp_emailid__emp_emailid=emp_emailid)

#             # Retrieve any additional employee information if necessary
#             personal_details = Personal_details.objects.get(mail__emp_emailid=emp_emailid)
#             poifiles_new = Poifiles_new1.objects.get(Emp_id__emp_emailid=emp_emailid)

#             # # Calculate salary details
#             age = datetime.now().year - personal_details.birth_date.year
#             actual80c = poifiles_new.actualAmount_80C
#             actual80d = poifiles_new.actualAmount_80D
#             actualoie = poifiles_new.OIE_actualAmount
#             actualosi = poifiles_new.OSI_actualAmount

#             eligible_deductions = actual80c + actual80d + actualoie + actualosi

#             # If revision is provided, apply it to the annual CTC
#             annual_ctc += annual_ctc * (revision / 100)
#             monthly_ctc = annual_ctc / 12
#             basic = monthly_ctc * 0.4
#             hra = basic * 0.4
#             conveyance = 1600
#             da = 0
#             special_allowance = monthly_ctc - (hra + conveyance + da + basic)

#             # Calculate tax liability
#             annual_taxable = annual_ctc - eligible_deductions
#             tax_liability = tax_calculation_to_add_salary(age, annual_taxable)

#             monthly_tds = tax_liability / 12
#             monthly_epf = basic * 0.12
#             monthly_prof_tax = 200
           
#             net_salary = monthly_ctc - (monthly_tds + monthly_epf + monthly_prof_tax)

#             # Save salary details to the database
#             salary_instance = Salary1.objects.create(
#                 emp_emailid=bank_details,  # Assign Bank_details instance to Salary
#                 basic=basic,
#                 hra=hra,
#                 conveyance=conveyance,
#                 da=da,
#                 special_allowance=special_allowance,
#                 monthly_ctc=monthly_ctc,
#                 annual_ctc=annual_ctc,
#                 Eligible_Deductions=eligible_deductions,
#                 Yearly_Taxable_Salary=annual_taxable,
#                 Total_Tax_Liability=tax_liability,
#                 Monthly_TDS=monthly_tds,
#                 Monthly_EPF=monthly_epf,
#                 Monthly_Professional_Tax=monthly_prof_tax,
#                 Net_Salary=net_salary,
#                 payout_month=payout_month,
#                 effective_from=effective_from,
#                 notes=notes,
#                 remarks=employee_remarks,
#                 revision=revision,
#                 paymentmethod=payment_method
#             )
#             salary_instance.save()

#             return JsonResponse({'message': 'Salary details saved successfully.'})

#         except Bank_details.DoesNotExist:
#             return JsonResponse({'error': 'Employee bank details not found.'}, status=404)

#         except Employee.DoesNotExist:
#             return JsonResponse({'error': 'Employee not found or does not belong to this company.'}, status=404)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddSalary1(request):
    if request.method == 'GET':
        # Get company id from session
        c_id = request.session.get('c_id')
        if not c_id:
            return JsonResponse({'error': 'Company ID not found in session'}, status=400)

        # Retrieve all employees that belong to departments under this company
        employees = Employee.objects.filter(d_id__c_id=c_id)

        # Prepare employee data to return to the frontend
        employee_data = [{'emp_emailid': emp.emp_emailid, 'name': emp.emp_name} for emp in employees]

        return JsonResponse({'employees': employee_data})

    elif request.method == 'POST':
        data = json.loads(request.body)
        emp_emailid = data.get('emp_emailid')
        payout_month = data.get('payoutMonth')
        revision = int(data.get('revision', 0))
        annual_ctc = int(data.get('annualCTC', 0))
        payment_method = data.get('paymentMethod')
        employee_remarks = data.get('employeeRemarks')
        notes = data.get('notes')
        effective_from = data.get('effectiveFrom')

        c_id = request.session.get('c_id')

        if not emp_emailid or not c_id:
            return JsonResponse({'error': 'Missing employee or company details'}, status=400)

        try:
            # Retrieve the employee's bank details from Bank_details using emp_emailid
            bank_details = Bank_details.objects.get(emp_emailid__emp_emailid=emp_emailid)

            # Retrieve any additional employee information if necessary
            personal_details = Personal_details.objects.get(mail__emp_emailid=emp_emailid)
            poifiles_new = Poifiles_new1.objects.get(Emp_id__emp_emailid=emp_emailid)

            # Calculate salary details
            age = datetime.now().year - personal_details.birth_date.year
            actual80c = poifiles_new.actualAmount_80C
            actual80d = poifiles_new.actualAmount_80D
            actualoie = poifiles_new.OIE_actualAmount
            actualosi = poifiles_new.OSI_actualAmount

            eligible_deductions = actual80c + actual80d + actualoie + actualosi

            # If revision is provided, apply it to the annual CTC
            annual_ctc += annual_ctc * (revision / 100)
            monthly_ctc = annual_ctc / 12
            basic = monthly_ctc * 0.4
            hra = basic * 0.4
            conveyance = 1600
            da = 0
            special_allowance = monthly_ctc - (hra + conveyance + da + basic)

            # Calculate tax liability
            annual_taxable = annual_ctc - eligible_deductions
            tax_liability = tax_calculation_to_add_salary(age, annual_taxable)

            monthly_tds = tax_liability / 12
            monthly_epf = basic * 0.12
            monthly_prof_tax = 200

            net_salary = monthly_ctc - (monthly_tds + monthly_epf + monthly_prof_tax)

            # Check if salary record exists for the given employee and payout month
            salary_instance, created = Salary1.objects.update_or_create(
                emp_emailid=bank_details,  # Use Bank_details instance
                # payout_month=payout_month,  # Use the payout month for identifying the record
                defaults={
                    'basic': basic,
                    'hra': hra,
                    'conveyance': conveyance,
                    'da': da,
                    'special_allowance': special_allowance,
                    'monthly_ctc': monthly_ctc,
                    'annual_ctc': annual_ctc,
                    'Eligible_Deductions': eligible_deductions,
                    'Yearly_Taxable_Salary': annual_taxable,
                    'Total_Tax_Liability': tax_liability,
                    'Monthly_TDS': monthly_tds,
                    'Monthly_EPF': monthly_epf,
                    'Monthly_Professional_Tax': monthly_prof_tax,
                    'Net_Salary': net_salary,
                    'effective_from': effective_from,
                    'notes': notes,
                    'remarks': employee_remarks,
                    'revision': revision,
                    'paymentmethod': payment_method
                }
            )

            if created:
                message = 'Salary details created successfully.'
            else:
                message = 'Salary details updated successfully.'

            return JsonResponse({'message': message})

        except Bank_details.DoesNotExist:
            return JsonResponse({'error': 'Employee bank details not found.'}, status=404)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found or does not belong to this company.'}, status=404)



# Function to calculate tax while adding salary. Check Function above Name: AddSalary
def tax_calculation_to_add_salary(age, annual_taxable):
    tax = 0
    tcredit = 0
    surch = 0
    mtr = 0

    if age < 60:
        if annual_taxable > 1000000:
            tax = 112500 + (annual_taxable - 1000000) * 30 / 100
        elif annual_taxable > 500000 and annual_taxable <= 1000000:
            tax = 12500 + (annual_taxable - 500000) * 20 / 100
        elif annual_taxable > 200000 and annual_taxable <= 500000:
            tax = (annual_taxable - 250000) * 5 / 100
        elif annual_taxable <= 250000:
            tax = 0
    elif age >= 80:
        if annual_taxable > 1000000:
            tax = 100000 + (annual_taxable - 1000000) * 30 / 100
        elif annual_taxable > 500000 and annual_taxable <= 1000000:
            tax = (annual_taxable - 500000) * 20 / 100
        elif annual_taxable <= 500000:
            tax = 0
    elif age >= 60:
        if annual_taxable > 1000000:
            tax = 110000 + (annual_taxable - 1000000) * 30 / 100
        elif annual_taxable > 500000 and annual_taxable <= 1000000:
            tax = 10000 + (annual_taxable - 500000) * 20 / 100
        elif annual_taxable > 250000 and annual_taxable <= 500000:
            tax = (annual_taxable - 300000) * 5 / 100
        elif annual_taxable < 300000:
            tax = 0

    if age < 60 and annual_taxable <= 500000 and tax > 0:
        tcredit = tax

    if age >= 60 and age < 80 and annual_taxable <= 500000 and tax > 0:
        tcredit = tax

    tax = tax - tcredit

    if annual_taxable > 5000000 and annual_taxable <= 10000000:
        surch = tax * 10 / 100
        if surch > 0 and age < 60 and (annual_taxable * 1 - tax * 1.10) <= 3687500:
            mtr = surch * 1 - (annual_taxable - 3687500 - tax)
        if surch > 0 and age >= 60 and (annual_taxable - tax * 1.10) <= 3690000:
            mtr = surch * 1 - (annual_taxable - 3690000 * 1 - tax * 1)
        if surch > 0 and age >= 80 and (annual_taxable - tax * 1.10) <= 3700000:
            mtr = surch - (annual_taxable - 3700000 * 1 - tax * 1)
    if annual_taxable > 10000000:
        surch = tax * 15 / 100
        if surch > 0 and (age < 60) and (annual_taxable - tax * 1.15) <= 6906250:
            mtr = surch - (annual_taxable - 6906250 * 1 - tax)
        if surch > 0 and age >= 60 and (annual_taxable - tax * 1.15) <= 6909000:
            mtr = surch - (annual_taxable - 6909000 * 1 - tax)
        if surch > 0 and age >= 80 and (annual_taxable - tax * 1.15) <= 6920000:
            mtr = surch - (annual_taxable - 6920000 * 1 - tax)
    if annual_taxable <= 5000000:
        surch = 0
    surch = surch - mtr
    ttax = tax + surch
    ecess = ttax * .04
    liab = ttax + ecess

    return liab


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def SalaryRevisionHistory(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        emailid=request.GET.get('emailid')

        decoded_emailid = urllib.parse.unquote(emailid)
        print(f"Decoded email ID: {decoded_emailid}")

        if not emailid:
            return JsonResponse({'error':'Email ID not provided'}, status=400)

        try:
            # Retrieve salaries based on employee email ID
            salaries = Salary1.objects.filter(
               emp_emailid=emailid
            ).values(
                    'sal_id',         
                    'emp_emailid__emp_emailid',                     
                    'revision',                     
                    'effective_from',              
                    'basic',           
                    'hra', 
                    'conveyance',             
                    'da',                     
                    'special_allowance',                      
                    'annual_ctc',         
                    'paymentmethod',
                    'notes',
                    'remarks'                                                                                       
            ).distinct()

            if not salaries:
                return JsonResponse({'error': 'No salary details found for this employee'}, status=404)

            salary_list = []
            for salary in salaries:
                salary_dict = {
                    'sal_id': salary['sal_id'],  # Correctly access email ID from Bank_details
                    'emp_emailid': salary['emp_emailid__emp_emailid'],  # Correctly access email ID from Bank_details
                    'revision': salary['revision'],
                    'effective_from': salary['effective_from'],
                    'basic': salary['basic'],
                    'hra': salary['hra'],
                    'conveyance': salary['conveyance'],
                    'da': salary['da'],
                    'special_allowance': salary['special_allowance'],
                    'annual_ctc': salary['annual_ctc'],
                    'paymentmethod': salary['paymentmethod'],
                    'notes': salary['notes'],
                    'remarks': salary['remarks'],
                }
                salary_list.append(salary_dict)
                print("salary list: ")
                for item in salary_list:
                    print(item)

            return JsonResponse({'salaries': salary_list}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)






@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def DisplaySalaryDetails(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        emailid=request.GET.get('emailid')

        decoded_emailid = urllib.parse.unquote(emailid)
        print(f"Decoded email ID: {decoded_emailid}")

        if not emailid:
            return JsonResponse({'error':'Email ID not provided'}, status=400)
        try:
           # Retrieve salaries based on employee email ID
            salaries = Salary1.objects.filter(
               emp_emailid=emailid
            ).values(
                    'emp_emailid__emp_emailid',         # Employee email ID from Bank_details model
                    'payout_month',                     # Payout month
                    'monthly_ctc',                      # Monthly CTC
                    'Eligible_Deductions',              # Eligible Deductions
                    'Yearly_Taxable_Salary',            # Yearly Taxable Salary
                    'Total_Tax_Liability',              # Total Tax Liability
                    'Monthly_TDS',                      # Monthly TDS
                    'Monthly_EPF',                      # Monthly EPF
                    'Monthly_Professional_Tax',         # Monthly Professional Tax
                    'Net_Salary'                        # Net Salary
            ).distinct()

            if not salaries:
                return JsonResponse({'error': 'No salary details found for this employee'}, status=404)
            
            salary_list = []
            for salary in salaries:
                salary_dict = {
                    'emp_emailid': salary['emp_emailid__emp_emailid'],  # Correctly access email ID from Bank_details
                    'payout_month': salary['payout_month'],
                    'monthly_ctc': salary['monthly_ctc'],
                    'Eligible_Deductions': salary['Eligible_Deductions'],
                    'Yearly_Taxable_Salary': salary['Yearly_Taxable_Salary'],
                    'Total_Tax_Liability': salary['Total_Tax_Liability'],
                    'Monthly_TDS': salary['Monthly_TDS'],
                    'Monthly_EPF': salary['Monthly_EPF'],
                    'Monthly_Professional_Tax': salary['Monthly_Professional_Tax'],
                    'Net_Salary': salary['Net_Salary'],
                }
                salary_list.append(salary_dict)
                print("salary list: ")
                for item in salary_list:
                    print(item)

            return JsonResponse({'salaries': salary_list}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)




# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def DisplaySalaryDetails(request):
#     c_id = request.session.get('c_id')
#     if not c_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'GET':
#         # sid = request.GET.get('id')
#         # if not sid:
#         #     return JsonResponse({'error': 'Employee ID not provided'}, status=400)

#         try:
#             salaries = Salary1.objects.filter(emp_emailid__emp_emailid=sid)

#             salary_list = []
#             for salary in salaries:
#                 salary_dict = {
#                     'emp_emailid': salary.emp_emailid.emp_emailid,  # Accessing the email id through Bank_details
#                     'payout_month': salary.payout_month,
#                     'monthly_ctc': salary.monthly_ctc,
#                     'Eligible_Deductions': salary.Eligible_Deductions,
#                     'Yearly_Taxable_Salary': salary.Yearly_Taxable_Salary,
#                     'Total_Tax_Liability': salary.Total_Tax_Liability,
#                     'Monthly_TDS': salary.Monthly_TDS,
#                     'Monthly_EPF': salary.Monthly_EPF,
#                     'Monthly_Professional_Tax': salary.Monthly_Professional_Tax,
#                     'Net_Salary': salary.Net_Salary,
#                 }
#                 salary_list.append(salary_dict)

#             return JsonResponse({'salaries': salary_list}, status=200)

#         except Salary1.DoesNotExist:
#             return JsonResponse({'error': 'Salary records not found for the provided employee ID'}, status=404)

#     else:
#         return JsonResponse({'error': 'Method not allowed'}, status=405)







@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CustomForms(request):
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)
    user_name = request.session.get('emp_name')

    if not c_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    if request.method == 'GET':
        try:
            forms = Custom_forms.objects.filter(c_id=company_id).order_by('-seq').values('form_name')
            forms_list = [{'form_name': form['form_name']} for form in forms]
            return JsonResponse({'forms': forms_list})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    # elif request.method == 'POST':
    #     data = json.loads(request.body)
    #     form_name = data.get('newFormName')
    #     if not form_name:
    #         return JsonResponse({'error': 'Form name is required'}, status=400)

    #     try:
    #         formatted_form_name = form_name.replace(' ', '_').lower()
    #         response_table_name = f"resp_{c_id}{formatted_form_name}"

    #         with connection.cursor() as cursor:
    #             cursor.execute(f"""
    #                 CREATE TABLE IF NOT EXISTS "{response_table_name}" (
    #                     "emp_name" VARCHAR(255) PRIMARY KEY
    #                 );
    #             """)

    #         if Custom_forms.objects.filter(c_id=company_id, form_name=formatted_form_name).exists():
    #             return JsonResponse({'error': 'Form name already exists'}, status=400)

    #         new_form = Custom_forms(c_id=c_id, form_name=formatted_form_name)
    #         new_form.save()
    #         return JsonResponse({'message': 'Form created successfully'})
    #     except Exception as e:
    #         return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            form_name = data.get('newFormName')
            alloc = data.get('alloc', None)

            if not form_name:
                return JsonResponse({'error': 'Form name is required'}, status=400)

            # Create the Custom_forms instance using the company_id from the session
            custom_form = Custom_forms(
                form_name=form_name,
                alloc=alloc,
                c_id=c_id  # Use the company instance obtained from the session
            )
            custom_form.save()

            return JsonResponse({'message': 'Custom form created successfully', 'form_id': custom_form.seq}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

        # return JsonResponse({'error': 'Invalid request method'}, status=405)

    elif request.method == 'DELETE':
        form_name = request.GET.get('form_name')
        question_id = request.GET.get('idOfQuestion')

        if not form_name:
            return JsonResponse({'error': 'Form name is required'}, status=400)

        try:
            formatted_form_name = form_name.replace(' ', '_').lower()
            response_table_name = f"resp_{company_id}{formatted_form_name}"

            if question_id:
                # Delete specific question of the form
                Custom_forms_questions.objects.filter(ID=question_id, form_name=formatted_form_name, c_id=company_id).delete()

                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        ALTER TABLE "{response_table_name}" DROP COLUMN "{question_id}";
                    """)
                return JsonResponse({'message': 'Question deleted successfully'})
            else:
                # Delete the entire form
                Custom_forms.objects.filter(form_name=formatted_form_name, c_id=company_id).delete()
                Custom_forms_questions.objects.filter(form_name=formatted_form_name, c_id=company_id).delete()

                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        DROP TABLE IF EXISTS "{response_table_name}";
                    """)
                return JsonResponse({'message': 'Form deleted successfully'})
        except Custom_forms.DoesNotExist:
            return JsonResponse({'error': 'Form not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EditFormView(request):
    user_name = request.session.get('emp_name')
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)

    if not user_name or not c_id:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    form_name = request.GET.get('form_name')
    if not form_name:
        return JsonResponse({'error': 'Form name is required'}, status=400)

    formatted_form_name = form_name.replace(' ', '_').lower()
    try:
        questions = Custom_forms_questions.objects.filter(form_name=formatted_form_name, c_id=company_id).values()
        employees = Employee.objects.filter(d_id__c_id=c_id).values_list('emp_name', flat=True)
        return JsonResponse({'questions': list(questions), 'employees': list(employees)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AllocateFormView(request):
    user_name = request.session.get('emp_name')
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)

    if not user_name or not c_id:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    # Get form name from POST body
    data = json.loads(request.body)
    form_name = data.get('form_name')
    print("Form name from POST:", form_name)

    if not form_name:
        return JsonResponse({'error': 'Form name is required'}, status=400)

    if request.method == 'POST':
        try:
            allocated_employees = data.get('allocated_employee')
            if not allocated_employees:
                return JsonResponse({'error': 'Allocated employees are missing'}, status=400)

            allocated_employee_str = ", ".join(allocated_employees)
            print("Allocated employees as string:", allocated_employee_str)

            custom_form = Custom_forms.objects.get(form_name=form_name, c_id=company_id)
            print("Custom form found:", custom_form)

            custom_form.alloc = allocated_employee_str
            custom_form.save()
            print("Form allocation updated successfully.")

            return JsonResponse({'message': 'Form allocated successfully'}, status=200)  # Added explicit status code

        except Custom_forms.DoesNotExist:
            print("Form not found with name:", form_name, "and company_id:", company_id)
            return JsonResponse({'error': 'Form not found'}, status=404)
        except Exception as e:
            print("Unexpected error:", str(e))
            return JsonResponse({'error': str(e)}, status=500)

    else:
        print("Received non-POST request. Method:", request.method)
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddTextQuestionFormView(request):
    user_name = request.session.get('emp_name')
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)

    if not user_name or not c_id:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    form_name = request.GET.get('form_name')
    if not form_name:
        return JsonResponse({'error': 'Form name is required'}, status=400)

    if request.method == 'POST':
        data = json.loads(request.body)
        label = data.get('question')
        ID = data.get('ID')
        type = data.get('Type')
        name = ID

        if not label or not ID or not type:
            return JsonResponse({'error': 'All fields are required'}, status=400)

        if Custom_forms_questions.objects.filter(ID=ID, name=name, form_name=form_name, c_id=company_id).exists():
                return JsonResponse({'error': 'Question with the same ID already exists'}, status=400)

        formatted_form_name = form_name.replace(' ', '_').lower()
        response_table_name = f"resp_{company_id}{formatted_form_name}"

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name='{response_table_name}' AND column_name='{name}'")
                if not cursor.fetchone():
                    cursor.execute(f"ALTER TABLE {response_table_name} ADD {name} VARCHAR(255);")
                else:
                    return JsonResponse({'error': f"Column '{name}' already exists in the table"}, status=400)

            custom_form_question = Custom_forms_questions.objects.create(
                label=label,
                type=type,
                ID=ID,
                name=name,
                form_name=formatted_form_name,
                c_id=c_id
            )
            return JsonResponse({'success': 'Text question added successfully'}, status=201)
        except IntegrityError as e:
            return JsonResponse({'error': 'Database integrity error: ' + str(e)}, status=500)
        except Exception as e:
            return JsonResponse({'error': 'An unexpected error occurred: ' + str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddRadioQuestionView(request):
    user_name = request.session.get('emp_name')
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)

    if not user_name or not c_id:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    form_name = request.GET.get('form_name')
    if not form_name:
        return JsonResponse({'error': 'Form name is required'}, status=400)

    if request.method == 'POST':
        data = json.loads(request.body)
        label = data.get('radioLabel')
        ID = data.get('radioID')
        options_arr = data.get('radioOptionName')
        options_str = ", ".join(options_arr) if options_arr else ''
        name = ID

        if Custom_forms_questions.objects.filter(ID=ID, form_name=form_name, c_id=company_id).exists():
            return JsonResponse({'error': 'Question with the same ID already exists'}, status=400)

        formatted_form_name = form_name.replace(' ', '_').lower()
        response_table_name = f"resp_{company_id}{formatted_form_name}"

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name='{response_table_name}' AND column_name='{name}'")
                if not cursor.fetchone():
                    cursor.execute(f"ALTER TABLE {response_table_name} ADD {name} VARCHAR(255);")
                else:
                    return JsonResponse({'error': f"Column '{name}' already exists in the table"}, status=400)


            custom_form_question = Custom_forms_questions.objects.create(
                label=label,
                type='radio',
                ID=ID,
                name=name,
                optionName=options_str,
                form_name=formatted_form_name,
                c_id=c_id
            )
            return JsonResponse({'success': 'Radio question added successfully'}, status=201)
        except IntegrityError as e:
            return JsonResponse({'error': 'Database integrity error: ' + str(e)}, status=500)
        except Exception as e:
            return JsonResponse({'error': 'An unexpected error occurred: ' + str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CustomLetter(request):
    company_id = request.session.get('c_id')
    user_name = request.session.get('emp_name')

    if not company_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    c_id = get_object_or_404(Company, pk=company_id)
    employee = get_object_or_404(Employee, emp_name=user_name)

    if request.method == 'GET':
        try:
            letters = Custom_letters.objects.filter(c_id=c_id).order_by('-seq').values('letter_name')
            if not letters.exists():
                return JsonResponse({'message': 'No letters available. Please create a letter to see it here.'}, status=200)

            letters_list = [{'letter_name': letter['letter_name'], 'display_name': letter['letter_name'].replace('_', ' ').title()} for letter in letters]
            return JsonResponse({'letters': letters_list}, status=200)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_letter_name = data.get('newLetterName')

            if not new_letter_name:
                return JsonResponse({'error': 'Letter name is required'}, status=400)

            new_letter_name = new_letter_name.replace(' ', '_').lower()

            if Custom_letters.objects.filter(letter_name=new_letter_name, c_id=c_id).exists():
                return JsonResponse({'error': 'Letter name already exists!'}, status=400)

            Custom_letters.objects.create(letter_name=new_letter_name, c_id=c_id)
            return JsonResponse({ 'message': 'Letter deleted successfully'}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)
    


    elif request.method == 'DELETE':
        this_letter = request.GET.get('letter_name')

        if not this_letter:
            return JsonResponse({'error': 'Letter name is required'}, status=400)

        formatted_letter_name = this_letter.replace(' ', '_').lower()

        try:
            letter_to_delete = get_object_or_404(Custom_letters, letter_name=formatted_letter_name, c_id=c_id)
            letter_to_delete.delete()
            return JsonResponse({'message': 'Letter deleted successfully'}, status=200)
        except Custom_letters.DoesNotExist:
            return JsonResponse({'error': 'Letter not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EditLetterView(request):
    company_id = request.session.get('c_id')
    user_name = request.session.get('emp_name')
    emp_emailid = request.session.get('emp_emailid')
    department_id = request.session.get('d_id')

    if not company_id or not user_name or not department_id:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    company = get_object_or_404(Company, pk=company_id)
    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

    if request.method == 'GET':
        this_letter = request.GET.get('letter_name')
        if not this_letter:
            return JsonResponse({'error': 'Letter name is required'}, status=400)

        this_letter = this_letter.replace(' ', '_').lower()
        custom_letter = get_object_or_404(Custom_letters, letter_name=this_letter, c_id=company)
        departments = Department.objects.filter(c_id=company).values('d_id')

        employees = []
        for dept in departments:
            emp_in_dept = Employee.objects.filter(Q(d_id=dept['d_id']) & ~Q(emp_role='HR')).values('emp_name', 'emp_emailid')
            employees.extend(emp_in_dept)

        return JsonResponse({
            'letter_content': custom_letter.letter_content,
            'letter_name': custom_letter.letter_name.replace('_', ' ').title(),
            'employees': [{'name': emp['emp_name'], 'email': emp['emp_emailid']} for emp in employees]
        }, status=200)

    elif request.method == 'POST':
        data = json.loads(request.body)
        this_letter = data.get('thisLetter')
        letter_content = data.get('letterContent')

        if not this_letter or not letter_content:
            return JsonResponse({'error': 'Letter name and content are required'}, status=400)

        this_letter = this_letter.replace(' ', '_').lower()

        try:
            custom_letter = get_object_or_404(Custom_letters, letter_name=this_letter, c_id=company)
            custom_letter.letter_content = letter_content
            custom_letter.save()

            return JsonResponse({'success': 'Letter updated successfully'}, status=201)

        except Custom_letters.DoesNotExist:
            return JsonResponse({'error': f'Letter "{this_letter}" not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EmployeeDetails(request):
    company_id = request.session.get('c_id')
    logged_in_email = request.session.get('emp_emailid')  # Ensure user is logged in

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if not logged_in_email:  # If the user is not logged in
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        try:
            # Fetch departments that belong to the logged-in user's company
            departments = Department.objects.filter(c_id=company_id)

            # Fetch employees who are in the departments of the logged-in user's company
            employees = Employee.objects.filter(d_id__in=departments)

            # Apply search filter if search term is provided
            search_term = request.GET.get('search', '')
            if search_term:
                employees = employees.filter(emp_name__icontains=search_term)

            # Sorting logic
            sort_column = request.GET.get('sort_column', 'emp_name')  # Default to 'emp_name' if not provided
            sort_direction = request.GET.get('sort_direction', 'asc')
            if sort_direction == 'desc':
                sort_column = '-' + sort_column
            employees = employees.order_by(sort_column)

            # Pagination logic
            page = request.GET.get('page', 1)
            entries_per_page = request.GET.get('entries_per_page', 10)
            paginator = Paginator(employees, entries_per_page)

            try:
                employees_page = paginator.page(page)
            except PageNotAnInteger:
                employees_page = paginator.page(1)
            except EmptyPage:
                employees_page = paginator.page(paginator.num_pages)

            # Serialize employee data
            employee_data = employees_page.object_list.values(
                'emp_name', 'emp_emailid', 'emp_phone', 'emp_role'
            )

            data = {
                'employees': list(employee_data),
                'total_pages': paginator.num_pages,
                'current_page': employees_page.number,
                'total_entries': paginator.count,
            }
            return JsonResponse(data)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'DELETE':
        emp_emailid = request.GET.get('emp_emailid')

        if not emp_emailid:
            return JsonResponse({'error': 'Employee email ID is required'}, status=400)

        try:
            # Ensure the employee to be deleted belongs to the same company as the logged-in user
            employee = Employee.objects.get(emp_emailid=emp_emailid, d_id__c_id=company_id)
            # Ensure the employee to be deleted belongs to the same company as the logged-in user
            employee = Employee.objects.get(emp_emailid=emp_emailid, d_id__c_id=company_id)
            employee.delete()
            return JsonResponse({'message': 'Employee deleted successfully'})
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found or does not belong to your company'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)



@csrf_exempt
def AttendanceDetails(request):
    user_email = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_email or not company_id:
        return JsonResponse({'error': 'User not authenticated or company ID not found'}, status=403)

    if request.method == 'GET':
        try:
            user = Employee.objects.get(emp_emailid=user_email)

            # If user is HR, fetch attendance for all employees in their company
            if user.emp_role == 'HR':
                # Fetch all departments under the user's company
                departments = Department.objects.filter(c_id=company_id).values_list('d_id', flat=True)
                
                # Fetch all employees under those departments
                employees = Employee.objects.filter(d_id__in=departments)
                
                # Fetch attendance records for those employees
                attendance_data = Attendance.objects.filter(emp_emailid__in=employees.values_list('emp_emailid', flat=True)).values('id', 'datetime_log', 'log_type', 'emp_emailid')
                
                # Create a mapping of employee emails to names
                employee_name_mapping = {employee.emp_emailid: employee.emp_name for employee in employees}
                
            else:
                # If not HR, fetch attendance only for the logged-in user
                attendance_data = Attendance.objects.filter(emp_emailid=user_email).values('id', 'datetime_log', 'log_type', 'emp_emailid')
                
                # Only need the name of the logged-in user
                employee_name_mapping = {user.emp_emailid: user.emp_name}

            data = []
            for attendance in attendance_data:
                datetime_log = attendance['datetime_log']

                # Parse the datetime_log if it's a string
                if isinstance(datetime_log, str):
                    datetime_log = parse_datetime(datetime_log)
                    if datetime_log is None:
                        raise ValueError(f"Invalid datetime format: {attendance['datetime_log']}")

                emp_email = attendance['emp_emailid']
                emp_name = employee_name_mapping.get(emp_email, "Unknown")  # Fetch the name from the mapping

                data.append({
                    'id': attendance['id'],
                    'emp_email': emp_email,
                    'emp_name': emp_name,  # Include name
                    'date': datetime_log.date(),
                    'log_type': attendance['log_type'],
                    'time': datetime_log.time()
                })

            # Return 'results' and 'total' to match the frontend expectation
            return JsonResponse({
                'results': data,
                'total': len(data)
            })

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User does not exist'}, status=404)
        except ValueError as ve:
            return JsonResponse({'error': str(ve)}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'DELETE':
        attendance_id = request.GET.get('id')
        if not attendance_id:
            return JsonResponse({'error': 'Attendance ID is required'}, status=400)

        try:
            # Fetch the attendance record and ensure it belongs to the correct company
            attendance = get_object_or_404(Attendance, id=attendance_id)
            
            # Debug: Print attendance details
            print(f"Attendance Record: {attendance.id}, Employee Email: {attendance.emp_emailid.emp_emailid}")

            # Fetch the employee associated with the attendance record
            employee = get_object_or_404(Employee, emp_emailid=attendance.emp_emailid.emp_emailid)
            
            # Debug: Print employee details
            print(f"Employee Found: {employee.emp_emailid}")

            # Check if the employee's department belongs to the user's company
            if employee.d_id.c_id.c_id != company_id:
                return JsonResponse({'error': 'Unauthorized action'}, status=403)

            # Delete the attendance record
            attendance.delete()
            return JsonResponse({'message': 'Attendance record deleted successfully'})

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee associated with attendance record not found'}, status=404)
        except Attendance.DoesNotExist:
            return JsonResponse({'error': 'Attendance record not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)




@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EmployeeMaster(request):
    company_id = request.session.get('c_id')
    print(company_id)
    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    employees = Employee.objects.filter(d_id__c_id=company_id).values(
        'emp_name', 'emp_emailid', 'emp_phone', 'emp_role', 'd_id', 'emp_profile'
    )

    data = {
        'employees': list(employees)
    }

    return JsonResponse(data)


from psycopg2.extras import DateTimeTZRange
from datetime import datetime
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Todotasks, Employee, Tasks

@csrf_exempt
def task_list(request):
    emp_emailid = request.session.get('emp_emailid')

    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            print(f"📩 [RECEIVED] Data from frontend: {data}")

            title = data.get('title')
            description = data.get('description')
            priority = data.get('priority', False)
            completed = data.get('completed', False)
            start_date = data.get('start_date')
            end_date = data.get('end_date')

            if not title or not description or not start_date or not end_date:
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            # Convert date strings to datetime objects
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d")

            # Store in proper PostgreSQL range format
            date_range = DateTimeTZRange(start_date_obj, end_date_obj, '[)')  # Inclusive start, exclusive end

            new_task = Todotasks(
                title=title,
                description=description,
                date=date_range,  # Corrected format
                priority=priority,
                completed=completed
            )
            new_task.save()
            print(f"✅ [DB SUCCESS] Task Created: {new_task.__dict__}")

            # Associate with Employee
            employee = Employee.objects.get(emp_emailid=emp_emailid)
            task_entry = Tasks.objects.create(tid=new_task, emp_emailid=employee)
            print(f"✅ [DB SUCCESS] Task linked in Tasks table: {task_entry.__dict__}")

            return JsonResponse({'message': 'Task added successfully!'}, status=201)

        except Exception as e:
            print(f"❌ [ERROR] Task creation failed: {str(e)}")
            return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def task_detail(request, task_id):
    """
    Handles requests for a specific to-do task.

    GET: Retrieves details of a task.
    PUT: Updates a task.
    DELETE: Deletes a task.
    """
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    try:
        # Get the task, ensuring it belongs to the user
        task = Todotasks.objects.get(tid=task_id, tasks_tid__emp_emailid=emp_emailid)
    except Todotasks.DoesNotExist:
        return HttpResponse(status=404)

    if request.method == 'GET':
        return JsonResponse({'id': task.tid, 'description': task.description, 'date': task.date})

    elif request.method == 'PUT':
        data = json.loads(request.body)
        task.description = data.get('description', task.description)
        task.save()
        return JsonResponse({'id': task.tid, 'description': task.description, 'date': task.date})

    elif request.method == 'DELETE':
        task.delete()
        return HttpResponse(status=204)

@csrf_exempt
def task_search(request):
    """
    Handles search requests for to-do tasks.

    GET: Retrieves tasks matching the search term for the authenticated user.
    """
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        search_term = request.GET.get('search', '')
        tasks = Todotasks.objects.filter(
            description__icontains=search_term,
            tasks_tid__emp_emailid=emp_emailid
        ).values()
        tasks_list = list(tasks)
        return JsonResponse(tasks_list, safe=False)
    
    
@csrf_exempt
@require_http_methods(["GET"])
def get_performance_data(request, employee_email, year):
    # Authenticate user
    user_email = request.session.get('user_id')
    
    # Fetch the employee record using the email from the session
    try:
        logged_in_employee = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'User not authenticated'}, status=401)
    
    # Fetch the employee record for the requested performance data
    try:
        employee = Employee.objects.get(emp_emailid=employee_email, d_id=logged_in_employee.d_id)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)
    
    # Filter tasks based on the employee and year
    tasks = Tasks.objects.filter(emp_emailid=employee, sop__sdate__year=year)
    
    # Calculate task statistics
    total_sop_assigned = tasks.count()
    total_sop_accomplished = tasks.filter(status=True).count()
    total_kra_assigned = tasks.filter(kra_id__isnull=False).count()
    total_kra_accomplished = tasks.filter(kra_id__isnull=False, status=True).count()
    total_jd_assigned = tasks.filter(job_desc_id__isnull=False).count()
    total_jd_accomplished = tasks.filter(job_desc_id__isnull=False, status=True).count()
    total_training_assigned = tasks.filter(tid__isnull=False).count()
    total_training_accomplished = tasks.filter(tid__isnull=False, status=True).count()
    
    # Prepare the response data
    data = {
        'totalSopAssigned': total_sop_assigned,
        'totalSopAccomplished': total_sop_accomplished,
        'totalKraAssigned': total_kra_assigned,
        'totalKraAccomplished': total_kra_accomplished,
        'totalJdAssigned': total_jd_assigned,
        'totalJdAccomplished': total_jd_accomplished,
        'totalTrainingAssigned': total_training_assigned,
        'totalTrainingAccomplished': total_training_accomplished,
    }

    return JsonResponse(data, status=200)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def jd_form(request):
    """
    Handles requests related to Job Description forms.

    GET: Not supported for this endpoint.
    POST: Creates a new JD form and associated tasks.
    PUT: Not supported for this endpoint.
    DELETE: Not supported for this endpoint.
    """
    company_id = request.session.get('c_id')
    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            jd_name = data.get('jdName')
            date = data.get('date')
            responsibilities = data.get('responsibilities')

            if not all([jd_name, date, responsibilities]):
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            # Assuming you are sending employee emails in the request
            employee_emails = data.get('employeeEmails', [])  # Get employee emails from data

            with transaction.atomic():
                jd_instance = Job_desc.objects.create(
                    jd_name=jd_name,
                    responsiblities=', '.join(responsibilities),  # Store responsibilities as comma-separated string
                    sdate=date
                )

                for emp_email in employee_emails:
                    employee = Employee.objects.get(emp_emailid=emp_email)
                    if employee.d_id.c_id_id != company_id:
                        raise ValueError(f"Employee {emp_email} does not belong to this company.")

                    Tasks.objects.create(
                        job_desc_id=jd_instance,
                        responsiblities=', '.join(responsibilities),  # Store responsibilities as comma-separated string
                        emp_emailid=employee,
                        sdate=date,
                        d_id=employee.d_id
                    )

            return JsonResponse({'message': 'JD form and tasks created successfully'}, status=201)

        except (Employee.DoesNotExist, ValueError) as e:
            return JsonResponse({'error': str(e)}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AttritionReport(request):
    """
    Generates an attrition report for the company.

    GET: Retrieves the attrition report data.
    """
    company_id = request.session.get('c_id')
    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            # Calculate attrition rate for the past 12 months
            today = datetime.now()
            start_date = today - timedelta(days=365)  # Past 12 months

            # Get all employees who joined within the past year
            joined_employees = Job_info.objects.filter(
                emp_emailid__d_id__c_id=company_id, 
                start_date__gte=start_date
            ).annotate(
                join_month=TruncMonth('start_date')
            )

            # Get employees who resigned within the past year
            resigned_employees = Resignation.objects.filter(
                emp_emailid__d_id__c_id=company_id,
                submit_date__gte=start_date
            ).annotate(
                resignation_month=TruncMonth('submit_date')
            )

            # Aggregate data by month
            attrition_data = joined_employees.values('join_month').annotate(
                joined_count=Count('emp_emailid'),
                resigned_count=Count('emp_emailid__R_Id', filter=Q(emp_emailid__R_Id__in=resigned_employees)),
            ).order_by('join_month')

            # Calculate attrition rate and format the data for response
            report_data = []
            for data in attrition_data:
                month_name = calendar.month_name[data['join_month'].month]
                attrition_rate = (data['resigned_count'] / data['joined_count']) * 100 if data['joined_count'] > 0 else 0
                report_data.append({
                    'month': month_name,
                    'joined_count': data['joined_count'],
                    'resigned_count': data['resigned_count'],
                    'attrition_rate': round(attrition_rate, 2)
                })

            return JsonResponse({'attrition_report': report_data}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
@csrf_exempt
@require_http_methods(["GET"])
def get_performance_data(request, year):
    # Fetch the user ID from the session
    user_email = request.session.get('user_id')

    if not user_email:
        return JsonResponse({'error': 'User not authenticated'}, status=401)

    # Fetch the employee record using the email from the session
    try:
        logged_in_employee = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'User not authenticated'}, status=401)
    
    # Filter tasks based on the employee and year, assuming sop_id has a field like `sdate`
    tasks = Tasks.objects.filter(emp_emailid=logged_in_employee, sop_id__sdate__year=year)
    
    # Calculate task statistics
    total_sop_assigned = tasks.count()
    total_sop_accomplished = tasks.filter(status=True).count()
    total_kra_assigned = tasks.filter(kra_id__isnull=False).count()
    total_kra_accomplished = tasks.filter(kra_id__isnull=False, status=True).count()
    total_jd_assigned = tasks.filter(job_desc_id__isnull=False).count()
    total_jd_accomplished = tasks.filter(job_desc_id__isnull=False, status=True).count()
    total_training_assigned = tasks.filter(tid__isnull=False).count()
    total_training_accomplished = tasks.filter(tid__isnull=False, status=True).count()
    
    # Prepare the response data
    data = {
        'totalSopAssigned': total_sop_assigned,
        'totalSopAccomplished': total_sop_accomplished,
        'totalKraAssigned': total_kra_assigned,
        'totalKraAccomplished': total_kra_accomplished,
        'totalJdAssigned': total_jd_assigned,
        'totalJdAccomplished': total_jd_accomplished,
        'totalTrainingAssigned': total_training_assigned,
        'totalTrainingAccomplished': total_training_accomplished,
    }

    return JsonResponse(data, status=200)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def BenefitsCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='Benefits'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url
        })

    context = {
        'cases': case_data
    }

    return JsonResponse({'cases': case_data})


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def TravelExpenseCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='TravelCases'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url
        })

    context = {
        'cases': case_data
    }

    return JsonResponse({'cases': case_data})


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CompensationPayrollCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='CompensationPayroll'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url if created_by.emp_profile else None
        })

    return JsonResponse({'cases': case_data})

from django.contrib.auth import get_user_model
from django.db import transaction
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import Company, Department, Employee

@csrf_exempt
def RegisterHR(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Extract HR details
            username = data.get('username')
            email = data.get('email')
            password = data.get('password')
            first_name = data.get('first_name')
            last_name = data.get('last_name')
            phone = data.get('phone')
            skills = data.get('skills', '')

            # Extract company details
            company_name = data.get('company_name')
            company_address = data.get('company_address')
            company_phone = data.get('company_phone')

            # Validate required fields
            if not all([username, email, password, first_name, last_name, phone, company_name, company_address, company_phone]):
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            with transaction.atomic():
                # Create or get the company
                company, created = Company.objects.get_or_create(
                    c_name=company_name,
                    defaults={
                        'c_addr': company_address,
                        'c_phone': company_phone
                    }
                )

                # Create or get the HR department
                hr_department, _ = Department.objects.get_or_create(
                    d_name='Human Resources',
                    c_id=company
                )

                # Create the HR employee
                hr_employee = Employee.objects.create(
                    emp_name=f"{first_name} {last_name}",
                    emp_emailid=email,
                    emp_phone=phone,
                    emp_pwd=password,  # Note: You should hash this password
                    emp_role='HR',
                    emp_skills=skills,
                    d_id=hr_department
                )

            return JsonResponse({
                'message': 'HR registration successful',
                'hr_id': hr_employee.emp_emailid,
                'company_id': company.c_id
            }, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except IntegrityError as e:
            return JsonResponse({'error': f'Database integrity error: {str(e)}'}, status=409)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
    
    
    
@csrf_exempt
@require_http_methods(["GET", "POST"])
def Resign(request):
    if request.method == "GET":
        try:
            resignations = Resignation.objects.all().values()
            return JsonResponse(list(resignations), safe=False)
        except Exception as e:
            print("Error fetching resignations:", str(e))  # Debugging statement
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            print("Received data:", data)  # Debugging statement

            # Validate required fields
            required_fields = ["emp_emailid", "exp_leave", "leave_reason", "leave_date"]
            for field in required_fields:
                if field not in data:
                    print(f"Missing required field: {field}")  # Debugging statement
                    return JsonResponse({"error": f"Missing required field: {field}"}, status=400)

            # Convert date fields to datetime objects
            try:
                exp_leave = datetime.strptime(data.get("exp_leave"), "%Y-%m-%d").date()
                leave_date = datetime.strptime(data.get("leave_date"), "%Y-%m-%d").date()
                settle_date = datetime.strptime(data.get("settle_date"), "%Y-%m-%d").date() if data.get("settle_date") else None
                exit_interview = datetime.strptime(data.get("exit_interview"), "%Y-%m-%d").date() if data.get("exit_interview") else None
                last_working = datetime.strptime(data.get("last_working"), "%Y-%m-%d").date() if data.get("last_working") else None
            except ValueError as e:
                print("Date conversion error:", str(e))  # Debugging statement
                return JsonResponse({"error": "Invalid date format"}, status=400)

            resignation, created = Resignation.objects.update_or_create(
                emp_emailid=data.get("emp_emailid"),
                defaults={
                    'exp_leave': exp_leave,
                    'leave_reason': data.get("leave_reason"),
                    'leave_reason_2': data.get("leave_reason_2"),
                    'leave_reason_3': data.get("leave_reason_3"),
                    'leave_date': leave_date,
                    'settle_date': settle_date,
                    'exit_interview': exit_interview,
                    'last_working': last_working,
                    'status': data.get("status"),
                    'approved_by': data.get("approved_by"),
                    'notice_per': data.get("notice_per"),
                }
            )

            response = {'message': 'Resignation created successfully'} if created else {'message': 'Resignation updated successfully'}
            return JsonResponse({'status': 1, 'message': response})

        except json.JSONDecodeError:
            print("Invalid JSON data")  # Debugging statement
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            print("Error:", str(e))  # Debugging statement
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    

@csrf_exempt
def social_submit_feedback_get(request):
    if request.method == 'GET':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            user = Employee.objects.select_related('d_id__c_id').get(emp_emailid=user_email)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)

        feedbacks = Feedback.objects.filter(emp_emailid=user).values(
            'fid', 'reason', 'from_email', 'date'
        )

        feedback_list = []
        for fb in feedbacks:
            reactions = Reaction.objects.filter(feedback_id=fb['fid']).values('user_email', 'emoji', 'date')
            feedback_list.append({
                'feedback_id': fb['fid'],
                'content': fb['reason'],
                'sender': fb['from_email'],
                'created_at': fb['date'],
                'reactions': list(reactions)  # Include reactions for each feedback
            })

        return JsonResponse(feedback_list, safe=False)

    return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)



@csrf_exempt
def social_add_reaction(request, feedback_id):  # Add feedback_id here
    if request.method == 'POST':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            data = json.loads(request.body)
            emoji = data.get('emoji')

            if not emoji:
                return JsonResponse({'error': 'Emoji is required'}, status=400)

            # Check if feedback exists
            try:
                feedback = Feedback.objects.get(fid=feedback_id)
            except Feedback.DoesNotExist:
                return JsonResponse({'error': 'Feedback not found'}, status=404)

            # Add the reaction
            Reaction.objects.create(
                feedback=feedback,
                user_email=user_email,
                emoji=emoji
            )

            return JsonResponse({'message': 'Reaction added successfully'}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)





@csrf_exempt
def social_submit_feedback_post(request):
    if request.method == 'POST':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        # Get the authenticated user's company
        try:
            user = Employee.objects.select_related('d_id__c_id').get(emp_emailid=user_email)
            user_company = user.d_id.c_id
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)

        try:
            data = json.loads(request.body)
            emp_emailid = data.get('emp_emailid')
            skill = data.get('skill')
            from_email = user_email
            reason = data.get('reason')

            if not emp_emailid or not skill or not from_email or not reason:
                return JsonResponse({'error': 'All fields are required'}, status=400)

            # Ensure the employee exists and belongs to the same company
            try:
                employee = Employee.objects.select_related('d_id__c_id').get(emp_emailid=emp_emailid, d_id__c_id=user_company)
            except Employee.DoesNotExist:
                return JsonResponse({'error': 'Employee does not exist or does not belong to your company'}, status=404)

            # Save feedback, using the correct field for the foreign key reference
            feedback = Feedback.objects.create(
                emp_emailid_id=employee.emp_emailid,  # Use emp_emailid_id instead of emp_emailid
                skill=skill,
                from_email=from_email,
                reason=reason
            )

            # Retrieve feedbacks only for employees in the same company
            employee_emails = Employee.objects.filter(d_id__c_id=user_company).values_list('emp_emailid', flat=True)
            feedbacks = Feedback.objects.filter(emp_emailid__in=employee_emails).values(
                'emp_emailid__emp_emailid', 'skill', 'from_email', 'reason', 'date'
            )

            feedback_list = list(feedbacks)

            return JsonResponse({'message': 'Feedback submitted successfully', 'feedbacks': feedback_list}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except TypeError:
            return JsonResponse({'error': 'TypeError occurred'}, status=500)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)
    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)




@csrf_exempt
def employee_view(request):
    if request.method == 'GET':
        # Check if user is authenticated
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        # Get the authenticated user's company
        try:
            user = Employee.objects.select_related('d_id__c_id').get(emp_emailid=user_email)
            user_company = user.d_id.c_id
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)

        # Retrieve all employee details for employees in the same company
        employees = Employee.objects.filter(d_id__c_id=user_company).values(
            'emp_emailid', 'd_id'
        )
        employee_list = [{'id': emp['d_id'], 'name': emp['emp_emailid']} for emp in employees]
        
        return JsonResponse(employee_list, safe=False)

    else:
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)





@csrf_exempt
def allquiz(request):
    if request.method == 'GET':
        # Check if user is authenticated
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            # Retrieve the user from the Employee model
            user = Employee.objects.get(emp_emailid=user_email)

            # Get the user's company through the department
            company = user.d_id.c_id

            # Get the list of courses that belong to the user's company
            company_courses = Courses.objects.filter(c_id=company).values_list('course_title', flat=True)

            # Fetch only quizzes that are related to courses in the user's company
            quizzes = Quiz.objects.filter(course_title__in=company_courses)
            quiz_list = quizzes.values('id', 'title', 'course_title', 'total', 'total_marks', 'passing', 'time', 'status')

            return JsonResponse(list(quiz_list), safe=False)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User does not exist'}, status=404)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    
# Configure logging
logging.basicConfig(level=logging.DEBUG)

@csrf_exempt
def markattendance(request):
    if request.method == 'POST':
        try:
            # Parse request body
            data = json.loads(request.body)
            emp_email = data.get('email')
            latitude = data.get('latitude')
            longitude = data.get('longitude')
            log_type = data.get('log_type')
            employee_ip = request.META.get('REMOTE_ADDR')

            logging.debug(f"Employee IP: {employee_ip}")

            # Check for missing data
            if not emp_email or not latitude or not longitude:
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            # Fetch employee details
            try:
                employee = Employee.objects.get(emp_emailid=emp_email)
                logging.debug(f"Employee found: {employee}")
            except Employee.DoesNotExist:
                logging.error(f"Employee with email {emp_email} not found")
                return JsonResponse({'error': 'Employee not found'}, status=404)

            # Proceed with marking attendance
            Attendance.objects.create(
                emp_emailid=employee,  # Use the correct field name
                user_ip=employee_ip,  # Use the correct field name
                log_type=log_type,
                latitude=latitude,
                longitude=longitude,
                datetime_log=timezone.now().isoformat(),  # Use current datetime in ISO format
                date_updated=timezone.now()
            )
            logging.debug("Attendance marked successfully")
            return JsonResponse({'message': 'Attendance marked successfully'}, status=201)

        except Exception as e:
            logging.error(f"Error marking attendance: {str(e)}")
            logging.error(traceback.format_exc())  # Log the full traceback for debugging
            return JsonResponse({'error': 'Internal Server Error'}, status=500)

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)





#Settings views
@csrf_exempt
@require_http_methods(["POST"])
def update_settings_account(request):
    try:
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'Authentication required'}, status=401)

        new_name = request.POST.get('name')
        new_phone = request.POST.get('phone')
        new_skills = request.POST.get('skills')
        new_photo = request.FILES.get('photo')  # Get the uploaded file

        if not any([new_name, new_phone, new_skills, new_photo]):
            return JsonResponse({'error': 'No data provided for update'}, status=400)

        try:
            employee = Employee.objects.get(emp_emailid=user_email)

            should_save = False
            if new_name:
                employee.emp_name = new_name
                should_save = True
            if new_phone:
                employee.emp_phone = new_phone
                should_save = True
            if new_skills:
                employee.emp_skills = new_skills
                should_save = True
            if new_photo:
                employee.emp_photo = new_photo  # Assuming you have a field named 'emp_photo'
                should_save = True

            if should_save:
                employee.save()
                return JsonResponse({'message': 'Account details updated successfully'}, status=200)
            else:
                return JsonResponse({'message': 'No changes made'}, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        return JsonResponse({'error': 'An internal error occurred'}, status=500)



@csrf_exempt
@require_http_methods(["POST"])
def update_settings_password(request):
    try:
        # Check if the user is logged in
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'Authentication required'}, status=401)

        # Extract form data
        old_password = request.POST.get('oldPassword')
        new_password = request.POST.get('newPassword')
        new_confirm_password = request.POST.get('confirmPassword')

        # Validate password fields
        if not old_password:
            return JsonResponse({'error': 'Old password is required to change password'}, status=400)
        if not new_password or not new_confirm_password:
            return JsonResponse({'error': 'New password and confirm password are required'}, status=400)

        try:
            # Fetch the employee object using the email from the session
            employee = Employee.objects.get(emp_emailid=user_email)

            if employee.emp_pwd == old_password:  # Check if the old password matches
                if new_password == new_confirm_password:
                    employee.emp_pwd = new_password  # Update to the new password (plain text storage)
                    employee.save()
                    return JsonResponse({'message': 'Password updated successfully'}, status=200)
                else:
                    return JsonResponse({'error': 'Passwords do not match'}, status=400)
            else:
                return JsonResponse({'error': 'Old password is incorrect'}, status=400)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        return JsonResponse({'error': 'An internal error occurred'}, status=500)
    

    

def generate_otp():
    return ''.join(random.choices(string.digits, k=6))

@csrf_exempt
def send_otp(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')

        try:
            # Verify if the email exists in the Employee model
            user = Employee.objects.get(emp_emailid=email)
            
            # Generate OTP
            otp = generate_otp()

            # Check if an OTP record already exists for the given email
            otp_record, created = OTPVerification.objects.update_or_create(
                emp_emailid=email,
                defaults={
                    'otp': otp,
                    'created_at': timezone.now()
                    }
                
            )
            
            # Send the OTP via email
            send_mail(
                'Your OTP for Password Reset',
                f'Your OTP is: {otp}',
                settings.EMAIL_HOST_USER,
                [email],
                fail_silently=False,
            )
            
            return JsonResponse({'message': 'OTP sent successfully!'}, status=200)
        
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Email not found'}, status=404)
    
    return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)

@csrf_exempt
def verify_otp(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')
        otp = data.get('otp')

        try:
            # Fetch the OTPVerification object
            otp_record = OTPVerification.objects.get(emp_emailid=email, otp=otp)
            
            # Get the current time
            now = timezone.now()
            created_at = otp_record.created_at
            
            # Set the OTP expiration time (e.g., 10 minutes from creation)
            expiration_time = created_at + timedelta(minutes=10)
            
            # Debugging information
            print(f"Current Time: {now}")
            print(f"Created At: {created_at}")
            print(f"Expiration Time: {expiration_time}")

            # Check if the OTP is still valid
            if now > expiration_time:
                return JsonResponse({'error': 'OTP expired'}, status=400)
            
            # OTP is valid; allow the user to reset their password
            return JsonResponse({'message': 'OTP verified successfully'}, status=200)
        
        except OTPVerification.DoesNotExist:
            return JsonResponse({'error': 'Invalid OTP'}, status=400)
    
    return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


@csrf_exempt
def reset_password(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')
        password = data.get('password')  # Update this line

        try:
            user = Employee.objects.get(emp_emailid=email)

            if user.emp_pwd == password:
                return JsonResponse({'error': 'New password cannot be the same as the current password'}, status=400)

            user.emp_pwd = password
            user.save()
            return JsonResponse({'message': 'Password reset successful'}, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)

    return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)

def employee_details(request, emp_emailid):
    if request.method == 'GET':
        # Retrieve the employee based on their email ID
        employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
        
        # Serialize the data to send as JSON
        employee_data = {
            "emp_name": employee.emp_name,
            "emp_emailid": employee.emp_emailid,
            "emp_skills": employee.emp_skills,
            "emp_role": employee.emp_role,
            "emp_phone": employee.emp_phone,
            "emp_profile": employee.emp_profile.url if employee.emp_profile else None,
            "add_date": employee.add_date,
            "pay_sts": employee.pay_sts,
            "status": employee.Status,
            "department_id": employee.d_id.d_id,  # ForeignKey field
            "department_name": employee.d_id.d_name,  # Fetching department name
        }

        return JsonResponse({"employee": employee_data}, status=200)

# @csrf_exempt
# with hashing
# def reset_password(request):
#     if request.method == 'POST':
#         data = json.loads(request.body)
#         email = data.get('email')
#         new_password = data.get('new_password')

#         try:
#             user = Employee.objects.get(emp_emailid=email)
#             user.emp_pwd = make_password(new_password)
#             user.save()
#             return JsonResponse({'message': 'Password reset successful'}, status=200)
        
#         except Employee.DoesNotExist:
#             return JsonResponse({'error': 'User not found'}, status=404)
    
#     return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


# Calender
@csrf_exempt
def event_handler(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        # Add new event
        data = json.loads(request.body)
        event = Events.objects.create(
            evt_type=data.get('evt_type'),
            evt_start=data.get('evt_start'),
            evt_end=data.get('evt_end'),
            evt_text=data.get('evt_text'),
            evt_color=data.get('evt_color'),
            status=data.get('status', False),
            emp_emailid=emp_emailid
        )
        return JsonResponse({
            'evt_id': event.evt_id,
            'evt_type': event.evt_type,
            'evt_start': event.evt_start,
            'evt_end': event.evt_end,
            'evt_text': event.evt_text,
            'evt_color': event.evt_color,
            'status': event.status
        }, status=201)


    elif request.method == 'PUT':
        # Update existing event
        data = json.loads(request.body)
        evt_id = data.get('evt_id')
        event = get_object_or_404(Events, pk=evt_id, emp_emailid=emp_emailid)

        event.evt_type = data.get('evt_type', event.evt_type)
        event.evt_start = data.get('evt_start', event.evt_start)
        event.evt_end = data.get('evt_end', event.evt_end)
        event.evt_text = data.get('evt_text', event.evt_text)
        event.evt_color = data.get('evt_color', event.evt_color)
        event.status = data.get('status', event.status)
        event.save()

        return JsonResponse({
            'evt_id': event.evt_id,
            'evt_type': event.evt_type,
            'evt_start': event.evt_start,
            'evt_end': event.evt_end,
            'evt_text': event.evt_text,
            'evt_color': event.evt_color,
            'status': event.status
        }, status=200)

    elif request.method == 'DELETE':
        # Delete an event
        data = json.loads(request.body)
        evt_id = data.get('evt_id')
        event = get_object_or_404(Events, pk=evt_id, emp_emailid=emp_emailid)
        event.delete()
        return HttpResponse(status=204)

    elif request.method == 'GET':
        # Retrieve all events for the logged-in user
        events = Events.objects.filter(emp_emailid=emp_emailid)
        events_list = list(events.values(
            'evt_id', 'evt_type', 'evt_start', 'evt_end', 'evt_text', 'evt_color', 'status'
        ))
        return JsonResponse(events_list, safe=False, status=200)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


# IT Declaration
# Utility function to get the employee object
def get_employee(email):
    try:
        return Employee.objects.get(emp_emailid=email)
    except Employee.DoesNotExist:
        return None

@csrf_exempt
def investment_80C(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)
        
        if employee:
            ITDeclaration80c.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Life Insurance Premium'),
                Investment1_Amount=data.get('amount', 0),
            )
            return JsonResponse({"message": "Investment 80C data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)

@csrf_exempt
def investment_80D(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)
        
        if employee:
            Itdeclaration80d_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Medi claim Policy for Self,Spouse,Children-80D'),
                Investment1_Amount=data.get('amount', 0),
                emp_emailid=employee,
            )
            return JsonResponse({"message": "Investment 80D data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)

@csrf_exempt
def other_investments(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)
        
        if employee:
            Itdeclaration_oie_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Additional Exemptions on voluntary NPS'),
                Investment1_Amount=data.get('amount', 0),
                emp_emailid=employee,
            )
            return JsonResponse({"message": "Other investments data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)

@csrf_exempt
def other_income(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)
        
        if employee:
            Itdeclaration_osi_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('income', 'Income from other sources'),
                Investment1_Amount=data.get('amount', 0),
                Investment2_Amount=data.get('savingsInterest', 0),
                Investment3_Amount=data.get('fixedDepositInterest', 0),
                Investment4_Amount=data.get('nsCertificatesInterest', 0),
                emp_emailid=employee,
            )
            return JsonResponse({"message": "Other income data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)
        

# TodoList
# @csrf_exempt
# def task_list(request):
#     if request.method == 'GET':
#         try:
#             # Fetch all tasks from the database
#             tasks = Todotasks.objects.all().values('id', 'title', 'completed')  # Adjust fields as necessary
#             # Convert the queryset to a list of dictionaries
#             task_list = list(tasks)
#             return JsonResponse(task_list, safe=False)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

#     elif request.method == 'POST':
#         try:
#             # Parse the incoming data from the request
#             data = json.loads(request.body)
#             title = data.get('text', '')
#             completed = data.get('isCompleted', False)

#             # Create a new task
#             new_task = Todotasks1(title=title, completed=completed)
#             new_task.save()

#             return JsonResponse({'message': 'Task created successfully'}, status=201)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)

# @csrf_exempt
# def task_delete(request, task_id):
#     if request.method == 'DELETE':
#         try:
#             # Fetch the task by its ID
#             task = Todotasks1.objects.get(id=task_id)
#             task.delete()
#             return JsonResponse({'message': 'Task deleted successfully'}, status=200)
#         except Todotasks1.DoesNotExist:
#             return JsonResponse({'error': 'Task not found'}, status=404)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)
#     else:
#         return JsonResponse({'error': 'Invalid request method'}, status=405)
    

# @csrf_exempt
# def task_update(request, task_id):
#     if request.method == 'PUT':
#         try:
#             # Fetch the task by its ID
#             task = Todotasks1.objects.get(id=task_id)
            
#             # Parse the request body
#             data = json.loads(request.body)
            
#             # Update the task attributes
#             task.title = data.get('title', task.title)
#             task.completed = data.get('completed', task.completed)
#             task.save()
            
#             return JsonResponse({
#                 'message': 'Task updated successfully',
#                 'task': {
#                     'id': task.id,
#                     'title': task.title,
#                     'completed': task.completed
#                 }
#             }, status=200)
#         except Todotasks1.DoesNotExist:
#             return JsonResponse({'error': 'Task not found'}, status=404)
#         except json.JSONDecodeError:
#             return JsonResponse({'error': 'Invalid JSON'}, status=400)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)
#     else:
#         return JsonResponse({'error': 'Invalid request method'}, status=405)
    


def attrition_year_data(request):
    data = list(AttritionYearData.objects.values())  # Get all objects and convert to list of dicts
    return JsonResponse(data, safe=False)

def attrition_department_data(request):
    data = list(AttritionDepartmentData.objects.values())  # Get all objects and convert to list of dicts
    return JsonResponse(data, safe=False)

def attrition_reason_data(request):
    data = list(AttritionReasonData.objects.values())  # Get all objects and convert to list of dicts
    return JsonResponse(data, safe=False)

def attrition_gender_data(request):
    data = list(AttritionGenderData.objects.values())  # Get all objects and convert to list of dicts
    return JsonResponse(data, safe=False)

def attrition_choice_data(request):
    data = list(AttritionChoiceData.objects.values())  # Get all objects and convert to list of dicts
    return JsonResponse(data, safe=False)

@csrf_exempt
def poi_list_create(request,emp_emailid):
    if request.method == 'GET':
        try:
            poi_record = Poifiles_new1.objects.filter(Emp_id__emp_emailid=emp_emailid).values(
                'Investment_1', 'actualAmount_80C', 'Status1',
                'Investment_2', 'actualAmount_80D', 'Status2',
                'Investment_3', 'OIE_actualAmount', 'Status3',
                'Investment_4', 'OSI_actualAmount', 'Status4'
            ).first()
            if poi_record:
                return JsonResponse({'poi_record': poi_record}, safe=False)
            return JsonResponse({'error': 'No records found for the employee'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    # elif request.method == 'POST':
    #     category = request.POST.get('category')
    #     declared_amount = request.POST.get('declaredAmount')
    #     actual_amount = request.POST.get('actualAmount')
    #     file = request.FILES.get('file')

    #     poi = POI(category=category, declared_amount=declared_amount, actual_amount=actual_amount, file=file)
    #     poi.save()

    #     return JsonResponse({
    #         'id': poi.id,
    #         'category': poi.category,
    #         'declaredAmount': poi.declared_amount,
    #         'actualAmount': poi.actual_amount,
    #         'fileUrl': poi.file.url if poi.file else None,
    #     }, status=201)
        
@csrf_exempt
def toggle_investment_status(request, emp_id, investment_field):
    if request.method == 'POST':
        try:
            poi_record = get_object_or_404(Poifiles_new1, Emp_id=emp_id)
            # Map the investment field to the status fields in the model
            status_field_map = {
                'Investment_1': 'Status1',
                'Investment_2': 'Status2',
                'Investment_3': 'Status3',
                'Investment_4': 'Status4'
            }
            status_field = status_field_map.get(investment_field)
            if status_field:
                current_status = getattr(poi_record, status_field)
                new_status = 'Verified' if current_status == 'Pending' else 'Pending'
                setattr(poi_record, status_field, new_status)
                poi_record.save()
                return JsonResponse({'status': 'success', 'new_status': new_status})
            return JsonResponse({'error': 'Invalid investment field'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def poi_delete(request, poi_id):
    if request.method == 'DELETE':
        poi = get_object_or_404(POI, pk=poi_id)
        if poi.file:
            default_storage.delete(poi.file.path)
        poi.delete()
        return JsonResponse({'message': 'POI deleted successfully.'}, status=200)

@csrf_exempt
def Settings(request):
    company_id = request.session.get('c_id')
    # print(company_id)
    user_id = request.session.get('emp_emailid')
    # print(user_id)
    if not company_id:
        return JsonResponse({'error': 'ID not found in session'}, status=401)
    if request.method == 'GET':
        employee_details = Employee.objects.filter(emp_emailid=user_id)
        employee_list = []
        for emp in employee_details:
            employee_list.append({
                'emp_name': emp.emp_name,
                'emp_emailid': emp.emp_emailid,
                'emp_phone': emp.emp_phone,
                'emp_profile': emp.emp_profile.url,
                'emp_skills': emp.emp_skills,
            })
        return JsonResponse(employee_list, safe=False)
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            emp_emailid = data.get('emp_emailid')
            if not emp_emailid:
                return JsonResponse({'error': 'Employee email ID is required'}, status=400)
            try:
                changes_made = False
                employee = Employee.objects.get(emp_emailid=emp_emailid)
                # Update fields only if new data is provided
                emp_emailid = data.get('emp_emailid')
                if emp_emailid and emp_emailid != employee.emp_emailid:
                    employee.emp_emailid = emp_emailid
                    changes_made = True
                emp_name = data.get('emp_name')
                if emp_name and emp_name != employee.emp_name:
                    employee.emp_name = emp_name
                    changes_made = True
                emp_phone = data.get('emp_phone')
                if emp_phone and emp_phone != employee.emp_phone:
                    employee.emp_phone = emp_phone
                    changes_made = True
                emp_profile = data.get('emp_profile')
                if emp_profile and emp_profile != str(employee.emp_profile):
                    employee.emp_profile = emp_profile
                    changes_made = True
                emp_skills = data.get('emp_skills')
                if emp_skills and emp_skills != employee.emp_skills:
                    employee.emp_skills = emp_skills
                    changes_made = True
                # Save only if changes were made
                if changes_made:
                    employee.save()
                return JsonResponse({'status': 'Employee details updated successfully'}, status=200)
            except Employee.DoesNotExist:
                return JsonResponse({'error': 'Employee not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=405)
            

    






# addded dipayan
@csrf_exempt
@role_required(['Manager', 'Super Admin','Employee'])
def JDForm(request):
    company_id = request.session.get('c_id')
    user_name = request.session.get('emp_name')

    if not company_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    if request.method == 'GET':
        try:
            # Fetch all employee email IDs
            employees = Employee.objects.values_list('emp_emailid', flat=True)
            return JsonResponse({'employee_email_ids': list(employees)}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            # Load the JSON data
            data = json.loads(request.body)

            # Extract the required fields
            jd_name = data.get('jd_name')
            responsibilities = data.get('responsiblities', [])  # Make sure to use the correct field name
            sdate = data.get('sdate')
            email_ids = data.get('email_ids', [])

            # Validate the input data
            if not jd_name or not responsibilities or not sdate or not email_ids:
                return JsonResponse({'error': 'All fields are required.'}, status=400)

            if not isinstance(responsibilities, list) or not isinstance(email_ids, list):
                return JsonResponse({'error': 'Responsibilities and email_ids should be arrays.'}, status=400)

            # Get or create the jid based on jd_name
            job_desc_entry = Job_desc.objects.filter(jd_name=jd_name).first()
            if job_desc_entry:
                jid = job_desc_entry.jid
            else:
                # Generate a new jid if jd_name is new
                max_jid = Job_desc.objects.aggregate(max_jid=Max('jid'))['max_jid']
                jid = (max_jid + 1) if max_jid is not None else 1

            # Create Job Description and Task entries for each employee
            for email_id in email_ids:
                try:
                    # Check if the employee exists
                    employee = Employee.objects.get(emp_emailid=email_id)
                    
                    # Create the Job Description entry
                    job_desc = Job_desc.objects.create(
                        jd_name=jd_name,
                        responsiblities=", ".join(responsibilities),  # Make sure the model field name matches
                        sdate=sdate,
                        email_id=email_id,
                        jid=jid
                    )
                    

                except Employee.DoesNotExist:
                    return JsonResponse({'error': f'Employee with email ID {email_id} does not exist'}, status=404)

            return JsonResponse({'status': 'JD assigned successfully.'}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ObjectDoesNotExist
from .models import Kra, Employee, Kra_table
import json
from datetime import datetime

@csrf_exempt
# # @role_required(['Manager', 'Super Admin', 'Employee'])
def KRAForm(request):
    if request.method == 'GET':
        try:
            employees = Employee.objects.values_list('emp_emailid', flat=True)
            return JsonResponse({'employee_email_ids': list(employees)}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            kras = data.get('kras', [])
            email_ids = data.get('email_ids', [])
            submission_date = data.get('submission_date')

            print(f"Received data: {data}")

            if not kras or not email_ids or not submission_date:
                return JsonResponse({'error': 'Incomplete data provided'}, status=400)

            submission_date = datetime.strptime(submission_date, "%Y-%m-%dT%H:%M:%S")

            for kra_data in kras:
                kra = kra_data.get('kra')
                weightage = int(kra_data.get('weightage', 0))
                kpi = kra_data.get('kpi')
                target = kra_data.get('target')
                ratings_scale = kra_data.get('ratings_scale')
                ratings = kra_data.get('ratings', 1)
                selfratings = kra_data.get('selfratings', 0)
                remarks = kra_data.get('remarks', '')
                status = kra_data.get('status', 0)
                measurement = kra_data.get('measurement', '0')
                kra_id = kra_data.get('kra_id')  # Optional field

                if not kra or not weightage or not kpi or not target or not ratings_scale:
                    return JsonResponse({'error': 'Incomplete KRA data'}, status=400)

                for email_id in email_ids:
                    if not Employee.objects.filter(emp_emailid=email_id).exists():
                        return JsonResponse({'error': f'Employee with email ID {email_id} does not exist'}, status=404)

                    # Handle kra_id (optional)
                    kra_table_entry = None
                    if kra_id:
                        try:
                            kra_table_entry = Kra_table.objects.get(pk=kra_id)
                        except ObjectDoesNotExist:
                            return JsonResponse({'error': f'kra_id {kra_id} does not exist'}, status=404)

                    # Create an entry in the Kra model
                    Kra.objects.create(
                        kra=kra,
                        weightage=weightage,
                        kpi=kpi,
                        target=target,
                        ratings_scale=ratings_scale,
                        ratings=ratings,
                        selfratings=selfratings,
                        remarks=remarks,
                        status=status,
                        email_id=email_id,
                        kra_id=kra_table_entry,  # Can be null
                        measurement=measurement,
                        submission_date=submission_date
                    )

            return JsonResponse({'message': 'KRA form submitted successfully!'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Kra, Kra_table

@csrf_exempt
def DeleteKRA(request, kra_no):
    if request.method == 'DELETE':
        try:
            # Fetch the KRA entry by kra_no
            kra = Kra.objects.get(kra_no=kra_no)
            # Delete the related entry in kra_table first
            Kra_table.objects.filter(kra_id=kra.kra_id_id).delete()
            # Delete the entry
            kra.delete()
            return JsonResponse({'message': 'KRA deleted successfully!'}, status=200)
        except Kra.DoesNotExist:
            return JsonResponse({'error': 'KRA not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=405)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import Kra

@csrf_exempt  # Only for testing; remove in production
def UpdateKRA(request, kra_no):
    if request.method == "PATCH":
        try:
            data = json.loads(request.body)
            kra = Kra.objects.get(kra_no=kra_no)

            new_status = data.get("status")  # Get new status from request
            if new_status not in [0, 1, 2, 3]:  # Validate status values
                return JsonResponse({"error": "Invalid status"}, status=400)

            kra.status = new_status
            kra.save()
            return JsonResponse({"message": "Status updated", "status": kra.status})

        except Kra.DoesNotExist:
            return JsonResponse({"error": "KRA not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    
    return JsonResponse({"error": "Invalid request"}, status=400)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def SOPForm(request):
    company_id = request.session.get('c_id')
    user_name = request.session.get('emp_name')
    if not company_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)
    
    if request.method == 'GET':
        try:
            # Get all departments from the database
            departments = Department.objects.all().values('d_id', 'd_name')
            department_list = list(departments)  # Convert the QuerySet to a list for JSON serialization

            return JsonResponse({'departments': department_list}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    elif request.method == 'POST':
        try:
            try:
                latest_sop = Sop.objects.latest('sop_id')
                sop_id = latest_sop.sop_id + 1
            except ObjectDoesNotExist:
                sop_id = 1  # Start at 1 if no records exist
            # Automatically generate sop_id
            latest_sop = Sop.objects.latest('sop_id')
            sop_id = latest_sop.sop_id + 1 if latest_sop else 1  # Increment sop_id or start at 1 if no records exist

            sop_type = request.POST.get('type')
            s_name = request.POST.get('s_name')
            sdate = request.POST.get('sdate')
            sop_file = request.FILES.get('sop_file')
            d_id = request.POST.get('d_id')

            # Debugging print statements
            # print(f"sop_id: {sop_id}, type: {sop_type}, s_name: {s_name}, sdate: {sdate}, sop_file: {sop_file}, d_id: {d_id}")
            
            # Save the SOP entry
            sop_entry = Sop(
                sop_id=sop_id,
                type=sop_type,
                s_name=s_name,
                sdate=sdate,
                sop_file=sop_file,
                d_id_id=d_id,
            )
            sop_entry.save()

            return JsonResponse({'status': 'SOP submitted successfully'}, status=201)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


#With excel file
# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def bulkEmployeeregistration(request):
#     if request.method == 'POST':
#         file = request.FILES.get('employeeFile')
#         if not file:
#             return JsonResponse({'error': 'No file uploaded'}, status=400)

#         # Save the uploaded file temporarily
#         file_name = default_storage.save(file.name, file)
#         file_path = default_storage.path(file_name)

#         try:
#             # Load the workbook and select the active worksheet
#             workbook = openpyxl.load_workbook(file_path)
#             worksheet = workbook.active

#             # Assuming the first row contains headers
#             company_name = worksheet['A2'].value
#             company_address = worksheet['B2'].value
#             company_phone = worksheet['C2'].value

#             # Create the company entry
#             new_company = Company.objects.create(
#                 c_name=company_name,
#                 c_addr=company_address,
#                 c_phone=company_phone
#             )

#             # Create departments and store their IDs
#             department_names = set()
#             for row in worksheet.iter_rows(min_row=2, values_only=True):
#                 dept_name = row[3]
#                 department_names.add(dept_name)

#             department_map = {}
#             for dept_name in department_names:
#                 new_dept = Department.objects.create(
#                     d_name=dept_name,
#                     c_id=new_company
#                 )
#                 department_map[dept_name] = new_dept.pk  # Map department names to their IDs

#             # Iterate through the worksheet and create employees
#             for row in worksheet.iter_rows(min_row=2, values_only=True):
#                 emp_name = row[4]
#                 emp_email = row[5]
#                 emp_phone = row[6]
#                 emp_skills = row[7]
#                 emp_role = row[8]
#                 dept_name = row[3]  # The department name in the current row

#                 # Validate employee data
#                 if not emp_name or not emp_email or not emp_phone or not emp_skills or not dept_name:
#                     return JsonResponse({'error': 'Employee data is incomplete in the uploaded file'}, status=400)

#                 # Create employee record
#                 Employee.objects.create(
#                     emp_name=emp_name,
#                     emp_emailid=emp_email,
#                     emp_skills=emp_skills,
#                     emp_role=emp_role,
#                     emp_phone=emp_phone,
#                     d_id_id=department_map[dept_name]  # Use the mapped department ID
#                 )

#             return JsonResponse({'message': 'Employees uploaded successfully'}, status=201)
#         except Exception as e:
#             return JsonResponse({'error': f'Failed to process the file: {str(e)}'}, status=500)
#         finally:
#             # Clean up: delete the file after processing
#             default_storage.delete(file_name)

#     return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
 

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def bulkUploadEmployeeDetailsUpload(request):
    company_id = request.session.get('c_id')
    user_name = request.session.get('emp_name')
    if not company_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)
    
    if request.method == 'POST':
        try:
            # Function to get or create an Employee object by email
            def get_or_create_employee(email):
                employee, created = Employee.objects.get_or_create(emp_emailid=email)
                return employee

            # A list to store errors
            errors = []

            # Handling personal details Excel file
            if 'personal_details' in request.FILES:
                personal_details_file = request.FILES['personal_details']
                df_personal = pd.read_excel(personal_details_file)

                for _, row in df_personal.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        personal_details, created = Personal_details.objects.update_or_create(
                            mail=employee,  # Use the correct ForeignKey field, which references Employee instance
                            defaults={
                                'first_name': row.get('first_name'),
                                'last_name': row.get('last_name'),
                                'Contact': row.get('Contact'),
                                'emergency_name': row.get('emergency_name'),
                                'emergency_contact': row.get('emergency_contact'),
                                'gender': row.get('gender'),
                                'birth_date': row.get('birth_date'),
                                'address': row.get('address'),
                                'city': row.get('city'),
                                'district': row.get('district'),
                                'post_code': row.get('post_code'),
                                'state': row.get('state'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            # Handling bank details Excel file
            if 'bank_details' in request.FILES:
                bank_details_file = request.FILES['bank_details']
                df_bank = pd.read_excel(bank_details_file)

                for _, row in df_bank.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        # Either update or create the bank details
                        bank_details, created = Bank_details.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'holder_name': row.get('holder_name'),
                                'bank_name': row.get('bank_name'),
                                'acc_no': row.get('acc_no'),
                                'branch': row.get('branch'),
                                'acc_type': row.get('acc_type'),
                                'ifsc': row.get('ifsc'),
                                'Pan_no': row.get('Pan_no'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            # Handling dependent details Excel file
            if 'dependent_details' in request.FILES:
                dependent_file = request.FILES['dependent_details']
                df_dependent = pd.read_excel(dependent_file)

                for _, row in df_dependent.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        # Either update or create the dependent details
                        dependent, created = Dependent.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'D_name': row.get('D_name'),
                                'D_gender': row.get('D_gender'),
                                'D_dob': row.get('D_dob'),
                                'D_relation': row.get('D_relation'),
                                'D_desc': row.get('D_desc'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            # Handling family details Excel file
            if 'family_details' in request.FILES:
                family_file = request.FILES['family_details']
                df_family = pd.read_excel(family_file)

                for _, row in df_family.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        # Either update or create the family details
                        family_details, created = Family_details.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'F_name': row.get('F_name'),
                                'F_gender': row.get('F_gender'),
                                'F_dob': row.get('F_dob'),
                                'F_contact': row.get('F_contact'),
                                'F_mail': row.get('F_mail'),
                                'F_relation': row.get('F_relation'),
                                'F_comment': row.get('F_comment'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            # Handling job info Excel file
            if 'job_info' in request.FILES:
                job_info_file = request.FILES['job_info']
                df_job = pd.read_excel(job_info_file)

                for _, row in df_job.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        job_info, created = Job_info.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'job_title': row.get('job_title'),
                                'department': row.get('department'),
                                'working_type': row.get('working_type'),
                                'start_date': row.get('start_date'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")
            else:
                return JsonResponse({'error': 'Job info file not uploaded'}, status=400)

            # Handling qualification details Excel file
            if 'qualification_details' in request.FILES:
                qualification_file = request.FILES['qualification_details']
                df_qualification = pd.read_excel(qualification_file)

                for _, row in df_qualification.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        # Either update or create the qualification details
                        qualification, created = Qualification.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'q_type': row.get('q_type'),
                                'q_degree': row.get('q_degree'),
                                'q_clg': row.get('q_clg'),
                                'q_uni': row.get('q_uni'),
                                'q_duration': row.get('q_duration'),
                                'q_yop': row.get('q_yop'),
                                'q_comment': row.get('q_comment'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            # Handling work experience Excel file
            if 'work_exp' in request.FILES:
                work_exp_file = request.FILES['work_exp']
                df_work_exp = pd.read_excel(work_exp_file)

                for _, row in df_work_exp.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        # Either update or create the work experience details
                        work_exp, created = Work_exp.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'start_date': row.get('start_date'),
                                'end_date': row.get('end_date'),
                                'comp_name': row.get('comp_name'),
                                'comp_location': row.get('comp_location'),
                                'designation': row.get('designation'),
                                'gross_salary': row.get('gross_salary'),
                                'leave_reason': row.get('leave_reason'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            if errors:
                return JsonResponse({'status': 'Some records were not saved due to errors', 'errors': errors}, status=400)

            return JsonResponse({'status': 'Files uploaded and data saved successfully'}, status=201)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)



@csrf_exempt
def SopList(request, sop_id=None):
    # Get the logged-in user's email and role
    company_id = request.session.get('c_id')
    user_email = request.session.get('emp_emailid')
    if not user_email or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    # Get the employee object to determine their role
    try:
        employee = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)

    user_role = employee.emp_role.lower()

    if request.method == 'GET':
        if sop_id:
            # Fetch the specific SOP using the sop_id
            try:
                sop = Sop.objects.get(sop_id=sop_id)
            except Sop.DoesNotExist:
                return JsonResponse({'error': 'SOP not found'}, status=404)

            # Prepare the data for the specific SOP
            sop_data = {
                'sop_id': sop.sop_id,
                'type': sop.type,
                's_name': sop.s_name,
                'sdate': sop.sdate,
                'sop_file': sop.sop_file.url if sop.sop_file else None,
                'ratings': sop.ratings,
                'selfratings': sop.selfratings,
                'remarks': sop.remarks,
                'd_id': sop.d_id_id,
            }
            return JsonResponse(sop_data, status=200)

        else:
            # Fetch all SOPs from the database
            sops = Sop.objects.all()
            sop_data_list = []
            for sop in sops:
                sop_data = {
                    'sop_id': sop.sop_id,
                    'type': sop.type,
                    's_name': sop.s_name,
                    'sdate': sop.sdate,
                    'sop_file': sop.sop_file.url if sop.sop_file else None,
                    'ratings': sop.ratings,
                    'selfratings': sop.selfratings,
                    'remarks': sop.remarks,
                    'd_id': sop.d_id_id,
                }
                sop_data_list.append(sop_data)
            return JsonResponse(sop_data_list, safe=False, status=200)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Accept sop_id from the request body if not provided in the URL
            sop_id = sop_id or data.get('sop_id')
            
            if not sop_id:
                return JsonResponse({'error': 'SOP ID is required for updating'}, status=400)

            # Fetch the specific SOP using the sop_id
            try:
                sop = Sop.objects.get(sop_id=sop_id)
            except Sop.DoesNotExist:
                return JsonResponse({'error': 'SOP not found'}, status=404)

            # Check if the user is a manager
            if user_role == 'manager' or user_role == 'super admin':
                sop.ratings = data.get('ratings', sop.ratings)
                sop.remarks = data.get('remarks', sop.remarks)
            else:
                # Check if the user is attempting to update ratings or remarks without permission
                if 'ratings' in data or 'remarks' in data:
                    return JsonResponse({'error': 'You do not have permission to update ratings or remarks.'}, status=403)

            # Allow all employees (including managers) to update selfratings
            sop.selfratings = data.get('selfratings', sop.selfratings)

            # Save the updated SOP
            sop.save()

            return JsonResponse({'status': 'SOP details updated successfully.'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def JdList(request):
    company_id = request.session.get('c_id')
    user_email = request.session.get('emp_emailid')

    if not user_email or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    if request.method == 'GET':
        # Ensure the user is logged in
        if not request.session.get('user_id'):
            return JsonResponse({'error': 'User not logged in'}, status=401)
        
        # Get the logged-in employee's name
        emp_name = request.session.get('emp_name')
        
        try:
            # Retrieve the employee instance using the name
            employee = Employee.objects.get(emp_name=emp_name)
        except Employee.DoesNotExist:
            return JsonResponse({"error": "Employee not found"}, status=404)

        # Retrieve all job descriptions assigned to this employee
        job_descs = Job_desc.objects.filter(email_id=employee.emp_emailid)

        # Prepare the response data
        job_desc_list = []
        for jd in job_descs:
            job_desc_list.append({
                'jd_id':jd.job_desc_id,
                'jd_name': jd.jd_name,
                'responsibilities': jd.responsiblities,
                'sdate': jd.sdate,
                'ratings': jd.ratings,
                'selfratings': jd.selfratings,
                'remarks': jd.remarks,
                'status': jd.status,
                'email_id': jd.email_id,
            })

        return JsonResponse({"job_descriptions": job_desc_list}, status=200)
    

    if request.method == 'POST':
        # Ensure the user is logged in
        if not request.session.get('user_id'):
            return JsonResponse({'error': 'User not logged in'}, status=401)

        # Retrieve all job descriptions assigned to this employee
        data = json.loads(request.body)
        sdate = data.get('sdate')
        # Prepare the response data
        # Fetch only job descriptions for the logged-in user and the given date
        job_details_queryset = Job_desc.objects.filter(sdate=sdate, email_id=user_email)

        # If no job descriptions are found, return an empty response
        if not job_details_queryset.exists():
            return JsonResponse({"job_details": []}, status=200)
        
        job_details_list = []
        for job in job_details_queryset:
            job_details_list.append({
                'jd_id': job.job_desc_id,
                'jd_name': job.jd_name,
                'responsibilities': job.responsiblities,
                'sdate': job.sdate,
                'ratings': job.ratings,
                'selfratings': job.selfratings,
                'remarks': job.remarks,
                'status': job.status,
                'email_id': job.email_id,
            })

        # Return the list of job descriptions
        return JsonResponse({"job_details": job_details_list}, status=200)
    return JsonResponse({"error": "Invalid request method"}, status=400)


@csrf_exempt
def JdDetails(request):
    company_id = request.session.get('c_id')
    user_email = request.session.get('emp_emailid')
    if not user_email or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    try:
        employee = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)

    user_role = employee.emp_role.lower()

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get the jd_id from the request body
            jd_id = data.get('jd_id')
            
            if not jd_id:
                return JsonResponse({'error': 'Job Description ID is required for updating'}, status=400)

            # Fetch the specific Job_desc using the jd_id
            try:
                job_desc = Job_desc.objects.get(job_desc_id=jd_id)
            except Job_desc.DoesNotExist:
                return JsonResponse({'error': 'Job Description not found'}, status=404)

            # Check if the user is a manager
            if user_role in ['manager', 'super admin']:
                job_desc.ratings = data.get('ratings', job_desc.ratings)
                job_desc.remarks = data.get('remarks', job_desc.remarks)
            else:
                # Check if the user is attempting to update ratings or remarks without permission
                if 'ratings' in data or 'remarks' in data:
                    return JsonResponse({'error': 'You do not have permission to update ratings or remarks.'}, status=403)

            # Allow all employees (including managers) to update selfratings
            job_desc.selfratings = data.get('selfratings', job_desc.selfratings)

            # Save the updated Job_desc
            job_desc.save()

            return JsonResponse({'status': 'Job Description details updated successfully.'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CreateQuiz(request):
    if request.method == 'POST':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:

            user = Employee.objects.get(emp_emailid=user_email)

            if user.emp_role not in ['HR', 'Manager', 'Super Admin']:
                return JsonResponse({'error': 'You do not have permission to create a quiz'}, status=403)

            company = user.d_id.c_id

            data = json.loads(request.body)
            quiz_title = data.get('quizTitle')
            course_id = data.get('course')
            total_questions = data.get('totalQuestions')
            marks_for_correct_answer = data.get('marksOnRightAnswer')
            marks_for_wrong_answer = data.get('minusMarksOnWrongAnswer')
            total_marks = int(data.get('total_marks', 0))
            passing_marks = data.get('passingMarks')
            time_limit = data.get('timeLimit')

            # Extract questions and options from the request
            questions = data.get('questions')

            # Check if the course exists and belongs to the user's company
            try:
                course = Courses.objects.get(course_id=course_id, c_id=company)
            except Courses.DoesNotExist:
                return JsonResponse({'error': 'Course does not exist in your company'}, status=404)

            # Create the quiz record
            quiz = Quiz(
                eid=user,
                title=quiz_title,
                course_title=course.course_title,
                total=total_questions,
                correct=marks_for_correct_answer,
                wrong=marks_for_wrong_answer,
                passing=passing_marks,
                total_marks=total_marks,
                time=time_limit,
                date=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                status='active'
            )
            quiz.save()

            # Create question and options records
            for question_data in questions:
                question_text = question_data['text']
                correct_answer = question_data['correct_answer']
                options = question_data['options']

                question = Question(quiz=quiz, text=question_text, correct_answer=correct_answer)
                question.save()

                for option_data in options:
                    option = Option(
                        question=question,
                        text=option_data['text'],
                        is_correct=option_data['is_correct']
                    )
                    option.save()

            return JsonResponse({'message': 'Quiz created successfully'}, status=201)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User does not exist'}, status=404)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])  # Custom role decorator for authorization
def GetAllQuizzes(request):
    if request.method == 'GET':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            user = Employee.objects.get(emp_emailid=user_email)

            if user.emp_role not in ['HR', 'Manager', 'Super Admin','Employee']:
                return JsonResponse({'error': 'You do not have permission to view quizzes'}, status=403)

            # Get all quizzes belonging to the user's company
            quizzes = Quiz.objects.filter(eid_d_id_c_id=user.d_id.c_id).values(
                'id', 'title', 'course_title', 'correct', 'wrong', 'total_marks', 
                'passing', 'total', 'time', 'date', 'status'
            )

            # Check if quizzes exist
            if not quizzes.exists():
                return JsonResponse({'error': 'No quizzes found for your company'}, status=404)

            # Convert to list and format date
            quiz_list = list(quizzes)

            # Return all quizzes as JSON
            return JsonResponse(quiz_list, safe=False)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User does not exist'}, status=404)

        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ToggleQuizStatus(request):
    if request.method == 'POST':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            user = Employee.objects.get(emp_emailid=user_email)

            if user.emp_role not in ['HR', 'Manager', 'Super Admin']:
                return JsonResponse({'error': 'You do not have permission to toggle quiz status'}, status=403)

            data = json.loads(request.body)
            quiz_id = data.get('quiz_id')
            new_status = data.get('status')  # Either 'active' or 'inactive'

            if new_status not in ['active', 'inactive']:
                return JsonResponse({'error': 'Invalid status value'}, status=400)

            # Find the quiz and ensure it belongs to the user's company
            try:
                quiz = Quiz.objects.get(id=quiz_id, eid_d_id_c_id=user.d_id.c_id)
            except Quiz.DoesNotExist:
                return JsonResponse({'error': 'Quiz not found or you do not have access'}, status=404)

            # Update quiz status
            quiz.status = new_status
            quiz.save()

            return JsonResponse({'message': f'Quiz status updated to {new_status}'}, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User does not exist'}, status=404)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
def AttemptQuiz(request, quiz_id=None):
    user_email = request.session.get('user_id')
    if not user_email:
        return JsonResponse({'error': 'User not authenticated'}, status=403)

    # Fetch the employee and department
    user = get_object_or_404(Employee, emp_emailid=user_email)
    user_department = user.d_id  # Get the user's department
    company_id = user_department.c_id  # Get the company ID from the department

    # Handle GET request
    if request.method == 'GET':
        if quiz_id is None:
            # Get all quizzes that belong to the same company as the user
            quizzes = Quiz.objects.filter(eid_d_id_c_id=company_id,status="active").values(
                'id', 'title', 'course_title', 'correct', 'wrong', 'total_marks', 
                'passing', 'total', 'time', 'date', 'status'
            )

            # Return full details of each quiz
            return JsonResponse(list(quizzes), safe=False)
        else:
            # Get a specific quiz if quiz_id is provided and matches the company
            quiz = get_object_or_404(Quiz, id=quiz_id, eid_d_id_c_id=company_id)

            # Fetch the related questions and options for the quiz
            questions = []
            for question in Question.objects.filter(quiz=quiz):
                options = Option.objects.filter(question=question).values('id', 'text')
                questions.append({
                    'id': question.id,
                    'text': question.text,
                    'options': list(options)
                })

            # Store the start time in the session
            request.session['quiz_start_time'] = timezone.now().isoformat()

            response_data = {
                'quiz_id': quiz.id,
                'quiz_title': quiz.title,
                'time_limit': quiz.time,
                'questions': questions
            }
            return JsonResponse(response_data, safe=False)

        

    # Handle POST request for quiz submission
    elif request.method == 'POST':
        data = json.loads(request.body)
        chosen_options = data.get('chosen_options')  # {question_id: chosen_option_id}

        # Retrieve the start time from the session
        start_time_str = request.session.get('quiz_start_time')
        if not start_time_str:
            return JsonResponse({'error': 'Start time not found'}, status=400)

        start_time = timezone.datetime.fromisoformat(start_time_str)

        # Fetch the quiz details
        quiz = get_object_or_404(Quiz, id=quiz_id)

        # Check if the user has already attempted this quiz
        attempt = QuizAttempt.objects.filter(quiz=quiz, employee=user).first()
        if attempt:
            # Update existing attempt
            attempt.chosen_options = chosen_options
            attempt.score = 0  # Reset score to recalculate
            attempt.total_correct = 0  # Reset correct count
            attempt.total_wrong = 0  # Reset wrong count
        else:
            # Create a new attempt if none exists
            attempt = QuizAttempt(
                quiz=quiz,
                employee=user,
                chosen_options=chosen_options,
                score=0,
                total_correct=0,
                total_wrong=0
            )

        # Initialize score calculation
        score = 0
        total_correct = 0
        total_wrong = 0

        for question_id, option_id in chosen_options.items():
            try:
                question = Question.objects.get(id=question_id, quiz=quiz)
                chosen_option = Option.objects.get(id=option_id, question=question)

                if chosen_option.is_correct:
                    score += quiz.correct  # Add correct marks
                    total_correct += 1
                else:
                    score -= quiz.wrong  # Subtract wrong marks
                    total_wrong += 1
            except Question.DoesNotExist:
                return JsonResponse({'error': f'Invalid question ID: {question_id}'}, status=400)
            except Option.DoesNotExist:
                return JsonResponse({'error': f'Invalid option ID: {option_id}'}, status=400)

        # Calculate elapsed time
        end_time = timezone.now()
        time_taken = (end_time - start_time)  # Time taken as a timedelta

        # Save time taken in total seconds
        total_seconds = int(time_taken.total_seconds())
        tolerance_of_time = 60 # 1 Minute Time tolerance so that request response time cost not paid by user attempting quiz
        if total_seconds > (quiz.time * 60 + tolerance_of_time):  # Check if time limit exceeded (convert to seconds)
            return JsonResponse({'error': 'Time limit exceeded'}, status=400)

        # Update the attempt details with the new score, time, and question stats
        attempt.score = score
        attempt.time_taken = total_seconds  # Save time in seconds
        attempt.is_passed = score >= quiz.passing
        attempt.total_correct = total_correct  # Save total correct answers
        attempt.total_wrong = total_wrong  # Save total wrong answers
        attempt.total_unattempted = quiz.total - (total_correct + total_wrong)  # Save total wrong answers
        attempt.save()

        # Update quiz total marks (if needed) and related fields
        # quiz.total_marks = score  # Update the total marks for the quiz
        quiz.save()

        # Format the time taken for response
        formatted_time = f"{total_seconds // 3600:02}:{(total_seconds % 3600) // 60:02}:{total_seconds % 60:02}"

        # Return the result of the quiz
        response_data = {
            'message': 'Quiz submitted successfully',
            'score': float(score),
            'total_correct': total_correct,
            'total_wrong': total_wrong,
            'is_passed': attempt.is_passed,
            'time_taken': formatted_time,  # Return formatted time
            'max_time': quiz.time,
        }

        return JsonResponse(response_data, status=201)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)

@csrf_exempt
def QuizResults(request, quiz_id=None):
    # Check if the request method is GET
    if request.method == 'GET':
        # Get the logged-in user's email ID from the session
        user_email = request.session.get('user_id')

        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            # Fetch the employee object based on the logged-in user's email ID
            user = Employee.objects.get(emp_emailid=user_email)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

        # If a specific quiz_id is provided
        if quiz_id:
            # Fetch quiz attempts for the given quiz by the logged-in user
            attempts = QuizAttempt.objects.filter(quiz_id=quiz_id, employee=user)
            if not attempts.exists():
                return JsonResponse({'error': 'No attempts found for this quiz'}, status=404)

            # Get the first attempt as an example
            attempt = attempts.first()

            # Fetch the quiz details
            quiz = attempt.quiz

            # Use pre-calculated total_correct, total_wrong, and total_unattempted from the database
            total_correct = attempt.total_correct
            total_wrong = attempt.total_wrong
            total_unattempted = attempt.total_unattempted

            # Get the negative mark deduction from the quiz
            negative_mark_deduction = quiz.wrong

            # Calculate total negative marks
            negative_marks = total_wrong * negative_mark_deduction

            # Calculate final score after applying negative marks
            final_score = attempt.score

            # Convert time_taken (in seconds) to hh:mm:ss format
            time_taken_seconds = attempt.time_taken
            time_taken_str = str(timedelta(seconds=time_taken_seconds))

            # Prepare the response data
            response_data = {
                'quiz_title': quiz.title,
                'full_marks': quiz.total_marks,
                'obtained_marks': float(final_score),
                'negative_marks': float(negative_marks),
                'time_taken': time_taken_str,
                'is_passed': final_score >= quiz.passing,
                'total_correct': total_correct,
                'total_wrong': total_wrong,
                'total_unattempted': total_unattempted
            }

            return JsonResponse(response_data, status=200)

        # If no quiz_id is provided, return all quiz attempts made by the logged-in user
        else:
            # Fetch all quiz attempts by the logged-in user
            attempts = QuizAttempt.objects.filter(employee=user)
            if not attempts.exists():
                return JsonResponse({'error': 'No quiz attempts found for the user'}, status=404)

            # List to store all the quiz results
            all_attempts_data = []

            for attempt in attempts:
                quiz = attempt.quiz

                # Use pre-calculated total_correct, total_wrong, and total_unattempted from the database
                total_correct = attempt.total_correct
                total_wrong = attempt.total_wrong
                total_unattempted = attempt.total_unattempted

                # Get the negative mark deduction from the quiz
                negative_mark_deduction = quiz.wrong

                # Calculate total negative marks
                negative_marks = total_wrong * negative_mark_deduction

                # Calculate final score after applying negative marks
                final_score = attempt.score

                # Convert time_taken (in seconds) to hh:mm:ss format
                time_taken_seconds = attempt.time_taken
                time_taken_str = str(timedelta(seconds=time_taken_seconds))

                # Prepare individual attempt data
                attempt_data = {
                    'quiz_title': quiz.title,
                    'full_marks': quiz.total_marks,
                    'obtained_marks': float(final_score),
                    'negative_marks': float(negative_marks),
                    'time_taken': time_taken_str,
                    'is_passed': final_score >= quiz.passing,
                    'total_correct': total_correct,
                    'total_wrong': total_wrong,
                    'total_unattempted': total_unattempted
                }

                # Add to the list of all attempts
                all_attempts_data.append(attempt_data)

            # Return all quiz attempts made by the logged-in user
            return JsonResponse({'quiz_attempts': all_attempts_data}, status=200)

    return JsonResponse({'error': 'Invalid request method'}, status=405)





@csrf_exempt
def DetailedDescription(request, quiz_id):
    # Get the logged-in user's email ID from the session
    user_email = request.session.get('user_id')
    
    if not user_email:
        return JsonResponse({'error': 'User not authenticated'}, status=403)

    try:
        # Fetch the employee object based on the logged-in user's email ID
        user = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)
    
    # Ensure it's a GET request
    if request.method == 'GET':
        try:
            # Fetch the quiz details
            quiz = Quiz.objects.get(id=quiz_id)

            # Get the quiz attempt for the logged-in employee
            attempt = QuizAttempt.objects.get(quiz=quiz, employee=user)

            # Fetch the questions and the chosen options for this quiz
            questions = []
            for question in quiz.questions.all():
                # Get the correct option for this question
                correct_option = question.options.filter(is_correct=True).first()

                # Get the selected option by the user for this question
                chosen_option_id = attempt.chosen_options.get(str(question.id))
                chosen_option = question.options.filter(id=chosen_option_id).first()

                # Append the question data along with the correct and chosen option
                questions.append({
                    'question_id': question.id,
                    'question_text': question.text,
                    'correct_option': correct_option.text if correct_option else None,  # Correct answer text
                    'chosen_option': chosen_option.text if chosen_option else None  # Chosen answer text by the user
                })

            # Convert time_taken (in seconds) to hh:mm:ss format
            time_taken_str = str(timedelta(seconds=attempt.time_taken))  # Use timedelta from the datetime module

            # Prepare the response data including marks and time
            response_data = {
                'quiz_id': quiz.id,
                'quiz_title': quiz.title,
                'questions': questions,
                'marks_obtained': float(attempt.score),  # The marks the employee got
                'total_marks': quiz.total_marks,  # Full marks for the quiz
                'time_taken': time_taken_str  # Time taken to complete the quiz in hh:mm:ss format
            }

            return JsonResponse(response_data, status=200)

        except Quiz.DoesNotExist:
            return JsonResponse({'error': 'Quiz not found'}, status=404)
        except QuizAttempt.DoesNotExist:
            return JsonResponse({'error': 'Quiz attempt not found for the employee'}, status=404)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)




@csrf_exempt
def AttemptedQuizzes(request):
    # Get the logged-in user's email ID from the session
    user_email = request.session.get('user_id')

    if not user_email:
        return JsonResponse({'error': 'User not authenticated'}, status=403)

    try:
        # Fetch the employee object based on the logged-in user's email ID
        user = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)

    # Ensure it's a GET request
    if request.method == 'GET':
        # Fetch all quiz attempts for the logged-in employee
        quiz_attempts = QuizAttempt.objects.filter(employee=user)

        if not quiz_attempts.exists():
            return JsonResponse({'error': 'No quizzes attempted by the user'}, status=404)

        # Prepare the list of attempted quizzes and their details
        quizzes_data = []
        for attempt in quiz_attempts:
            quiz = attempt.quiz

            # Convert time_taken (in seconds) to hh:mm:ss format
            time_taken_str = str(timedelta(seconds=attempt.time_taken))  # Use timedelta from the datetime module

            # Add the quiz and attempt details to the response

            quizzes_data.append({
                'quiz_id': quiz.id,
                'quiz_title': quiz.title,
                'total_questions': quiz.total,
                'passing_marks': quiz.passing,
                'marks_obtained': float(attempt.score),  # The marks the employee got
                'total_marks': quiz.total_marks,  # Full marks for the quiz
                'time_taken': time_taken_str,  # Time taken to complete the quiz in hh:mm:ss format
                'is_passed': attempt.is_passed,  # Whether the user passed the quiz
                'total_correct': attempt.total_correct,  # Number of correct answers
                'total_wrong': attempt.total_wrong,  # Number of wrong answers
                'attempted': (quiz.total - attempt.total_unattempted)  # Number of unattempted questions
            })

        return JsonResponse({'quizzes': quizzes_data}, status=200)

    return JsonResponse({'error': 'Invalid request method'}, status=405)
    






# Utility function to get the employee object
def get_employee(email):
    try:
        return Employee.objects.get(emp_emailid=email)
    except Employee.DoesNotExist:
        return None


        
@csrf_exempt
def investment_80C(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)

        if employee:
            # Save to ITDeclaration80c
            ITDeclaration80c.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Life Insurance Premium'),
                Investment1_Amount=data.get('amount', 0),
            )
            
            # Save or update Poifiles_new
            poifile, created = Poifiles_new1.objects.get_or_create(
                Emp_id=employee,
                defaults={
                    'Year': data.get('year', '2024'),
                    'Investment_1': '80C',
                    'actualAmount_80C': data.get('amount', 0),
                    'Doc_80C': data.get('doc', ''),
                    'Status1': 'Pending'
                }
            )
            if not created:
                poifile.actualAmount_80C = data.get('amount', 0)
                poifile.Doc_80C = data.get('doc', '')
                poifile.save()

            return JsonResponse({"message": "Investment 80C data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)




@csrf_exempt
def investment_80D(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)

        if employee:
            # Save to Itdeclaration80d_new1
            Itdeclaration80d_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Medi claim Policy for Self,Spouse,Children-80D'),
                Investment1_Amount=data.get('amount', 0),
                emp_emailid=employee,
            )
            
            # Save or update Poifiles_new
            poifile, created = Poifiles_new1.objects.get_or_create(
                Emp_id=employee,
                defaults={
                    'Year': data.get('year', '2024'),
                    'Investment_2': '80D',
                    'actualAmount_80D': data.get('amount', 0),
                    'Doc_80D': data.get('doc', ''),
                    'Status2': 'Pending'
                }
            )
            if not created:
                poifile.actualAmount_80D = data.get('amount', 0)
                poifile.Doc_80D = data.get('doc', '')
                poifile.save()

            return JsonResponse({"message": "Investment 80D data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)



        
@csrf_exempt
def other_investments(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)

        if employee:
            # Save to Itdeclaration_oie_new1
            Itdeclaration_oie_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Additional Exemptions on voluntary NPS'),
                Investment1_Amount=data.get('amount', 0),
                emp_emailid=employee,
            )
            
            # Save or update Poifiles_new
            poifile, created = Poifiles_new1.objects.get_or_create(
                Emp_id=employee,
                defaults={
                    'Year': data.get('year', '2024'),
                    'Investment_3': 'OIE',
                    'OIE_actualAmount': data.get('amount', 0),
                    'OIE_Doc': data.get('doc', ''),
                    'Status3': 'Pending'
                }
            )
            if not created:
                poifile.OIE_actualAmount = data.get('amount', 0)
                poifile.OIE_Doc = data.get('doc', '')
                poifile.save()

            return JsonResponse({"message": "Other investments data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)
       


@csrf_exempt
def other_income(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)

        if employee:
            # Save to Itdeclaration_osi_new1
            Itdeclaration_osi_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('income', 'Income from other sources'),
                Investment1_Amount=data.get('amount', 0),
                Investment2_Amount=data.get('savingsInterest', 0),
                Investment3_Amount=data.get('fixedDepositInterest', 0),
                Investment4_Amount=data.get('nsCertificatesInterest', 0),
                emp_emailid=employee,
            )
            
            # Save or update Poifiles_new
            poifile, created = Poifiles_new1.objects.get_or_create(
                Emp_id=employee,
                defaults={
                    'Year': data.get('year', '2024'),
                    'Investment_4': 'OSI',
                    'OSI_actualAmount': data.get('amount', 0),
                    'OSI_Doc': data.get('doc', ''),
                    'Status4': 'Pending'
                }
            )
            if not created:
                poifile.OSI_actualAmount = data.get('amount', 0)
                poifile.OSI_Doc = data.get('doc', '')
                poifile.save()

            return JsonResponse({"message": "Other income data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)




@csrf_exempt
def get_departments_by_company(request):
    c_id = request.session.get('c_id')  # Get the company ID from the session
    if c_id:
        departments = Department.objects.filter(c_id=c_id).values('d_id', 'd_name')
        return JsonResponse(list(departments), safe=False)
    else:
        return JsonResponse({'error': 'No company ID found in session'}, status=400)

@csrf_exempt
def get_employee_names_by_company(request):
    c_id = request.session.get('c_id')
    if c_id:
        employees = Employee.objects.filter(d_id__c_id=c_id).values('emp_name', 'emp_emailid')
        return JsonResponse({'employees': list(employees)}, safe=False)
    return JsonResponse({'employees': []}, safe=False)




@csrf_exempt
def AddBankDetails(request):
    emp_emailid = request.session.get('emp_emailid')  # Get email from session (logged-in user)
    
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Extract fields from the request data
            holder_name = data.get('holder_name')
            bank_name = data.get('bank_name')
            acc_no = data.get('acc_no')
            branch = data.get('branch')
            acc_type = data.get('acc_type')
            ifsc = data.get('ifsc')
            Pan_no = data.get('Pan_no')

            # Check if all required fields are present
            if not all([holder_name, bank_name, acc_no, branch, acc_type, ifsc, Pan_no]):
                return JsonResponse({'status': 'error', 'message': 'Missing required fields'}, status=400)
            
            # Retrieve the employee instance from the session email ID
            try:
                employee = Employee.objects.get(emp_emailid=emp_emailid)
            except Employee.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)

            # Create new Bank_details entry
            bank_details = Bank_details.objects.create(
                holder_name=holder_name,
                bank_name=bank_name,
                acc_no=acc_no,
                branch=branch,
                acc_type=acc_type,
                ifsc=ifsc,
                Pan_no=Pan_no,
                emp_emailid=employee  # Link to the employee using ForeignKey
            )

            bank_details.save()  # Save the instance

            return JsonResponse({'status': 'success', 'message': 'Bank details added successfully'})

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)




@csrf_exempt
def AddPersonalDetails(request):
    emp_emailid = request.session.get('emp_emailid')  # Get email from session (logged-in user)
    
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Extract fields from the request data
            first_name = data.get('first_name')
            last_name = data.get('last_name')
            Contact = data.get('Contact')
            emergency_name = data.get('emergency_name')
            emergency_contact = data.get('emergency_contact')
            gender = data.get('gender')
            birth_date = data.get('birth_date')  # Ensure this is sent in correct date format (YYYY-MM-DD)
            address = data.get('address')
            city = data.get('city')
            district = data.get('district')
            post_code = data.get('post_code')
            state = data.get('state')

            # Check if at least one required field is present (you can modify this based on your needs)
            if not first_name:
                return JsonResponse({'status': 'error', 'message': 'First name is required'}, status=400)

            # Retrieve the employee instance from the session email ID
            try:
                employee = Employee.objects.get(emp_emailid=emp_emailid)
            except Employee.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)

            # Create new Personal_details entry
            personal_details = Personal_details.objects.create(
                first_name=first_name,
                last_name=last_name,
                Contact=Contact,
                emergency_name=emergency_name,
                emergency_contact=emergency_contact,
                gender=gender,
                birth_date=birth_date,
                address=address,
                city=city,
                district=district,
                post_code=post_code,
                state=state,
                mail=employee  # Link to the employee using the ForeignKey
            )

            personal_details.save()  # Save the instance

            return JsonResponse({'status': 'success', 'message': 'Personal details added successfully'})

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
def Letters1(request):
    if 'emp_emailid' in request.session and 'c_id' in request.session:
        emp_emailid = request.session['emp_emailid']
        c_id = request.session['c_id']

        try:
            # Retrieve the company based on the logged-in user's c_id
            company = get_object_or_404(Company, pk=c_id)

            # Filter and fetch all custom letters associated with this company
            custom_letters = Custom_letters.objects.filter(c_id=company)

            # Prepare the list of letters with their names and content for the response
            allocated_letters = [
                {
                    'letter_name': letter.letter_name.replace('_', ' ').title(),
                    'letter_content': letter.letter_content
                }
                for letter in custom_letters
            ]

            return JsonResponse({'allocated_letters': allocated_letters})

        except Company.DoesNotExist:
            return JsonResponse({'error': 'Company not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

# @csrf_exempt
# def task_list1(request):
#     emp_emailid = request.session.get('emp_emailid')  # Get the logged-in emp_emailid from the session

#     if not emp_emailid:
#         return JsonResponse({'error': 'User not logged in'}, status=401)

#     if request.method == 'GET':
#         try:
#             # Fetch tasks only for the logged-in user's emp_emailid
#             tasks = Todotasks1.objects.filter(emp_emailid=emp_emailid).values('id', 'title', 'completed')
#             task_list = list(tasks)
#             return JsonResponse(task_list, safe=False)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

#     elif request.method == 'POST':
#         try:
#             # Parse incoming data
#             data = json.loads(request.body)
#             title = data.get('text', '')
#             completed = data.get('isCompleted', False)

#             # Create a new task with the logged-in user's emp_emailid
#             new_task = Todotasks1(title=title, completed=completed, emp_emailid=emp_emailid)
#             new_task.save()

#             return JsonResponse({'message': 'Task created successfully'}, status=201)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)

#     else:
#         return JsonResponse({'error': 'Invalid HTTP method'}, status=400)
@csrf_exempt
def task_list1(request):
    emp_emailid = request.session.get('emp_emailid')  # Get the logged-in user's email

    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        try:
            # Fetch tasks only for the logged-in user
            tasks = Todotasks1.objects.filter(emp_emailid=emp_emailid).values(
                'id', 'title', 'description', 'start_date', 'end_date', 'priority', 'completed'
            )
            return JsonResponse(list(tasks), safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            # Parse incoming JSON data
            data = json.loads(request.body)
            
            # Extract task details
            title = data.get('title', '')
            description = data.get('description', '')
            start_date = data.get('start_date', None)
            end_date = data.get('end_date', None)
            priority = data.get('priority', False)
            completed = data.get('completed', False)

            # Create and save the new task
            new_task = Todotasks1.objects.create(
                title=title,
                description=description,
                start_date=start_date,
                end_date=end_date,
                priority=priority,
                completed=completed,
                emp_emailid=emp_emailid
            )

            return JsonResponse({'message': 'Task created successfully', 'task_id': new_task.id}, status=201)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def task_update(request, task_id):
    if request.method == 'PUT':
        try:
            # Fetch the task by its ID
            task = Todotasks1.objects.get(id=task_id)
            
            # Parse the request body
            data = json.loads(request.body)
            
            # Update the task attributes
            task.title = data.get('title', task.title)
            task.completed = data.get('completed', task.completed)
            task.save()
            
            return JsonResponse({
                'message': 'Task updated successfully',
                'task': {
                    'id': task.id,
                    'title': task.title,
                    'completed': task.completed
                }
            }, status=200)
        except Todotasks1.DoesNotExist:
            return JsonResponse({'error': 'Task not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
@csrf_exempt
def task_delete(request, task_id):
    if request.method == 'DELETE':
        try:
            # Fetch the task by its ID
            task = Todotasks1.objects.get(id=task_id)
            task.delete()
            return JsonResponse({'message': 'Task deleted successfully'}, status=200)
        except Todotasks1.DoesNotExist:
            return JsonResponse({'error': 'Task not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
=======
import json
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseNotAllowed , JsonResponse, HttpResponseBadRequest
from .models import *
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.db import transaction
from django.core import serializers
from django.db import connection, IntegrityError
from django.utils import timezone
from datetime import datetime
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.contrib.auth.decorators import login_required
from .decorators import role_required
from django.db.models import Q
from django.db.models import Sum
from django.core.files.uploadedfile import UploadedFile
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.core.exceptions import ObjectDoesNotExist

import openpyxl
from django.core.files.storage import default_storage
import pandas as pd
from django.db.models import Max
from datetime import timedelta
import traceback



import logging
#added
from datetime import timedelta
from functools import wraps
from django.utils.decorators import method_decorator
import random, string
from django.contrib.auth.hashers import make_password
from django.utils.dateparse import parse_datetime
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import os

# decorators
# def role_required(allowed_roles):
#     def decorator(view_func):
#         @wraps(view_func)
#         def _wrapped_view(request, *args, **kwargs):
#             if not request.user.is_authenticated:
#                 return JsonResponse({'error': 'Authentication required'}, status=401)
            
#             user_role = getattr(request.user, 'emp_role', None)
#             if user_role not in allowed_roles:
#                 return JsonResponse({'error': 'Permission denied'}, status=403)
            
#             return view_func(request, *args, **kwargs)
#         return _wrapped_view
#     return decorator

from django.core.files.storage import default_storage
from datetime import timedelta
from functools import wraps
from django.utils.decorators import method_decorator
import random, string
from django.contrib.auth.hashers import make_password
import urllib.parse
import logging
from django.core.serializers import serialize
from rest_framework.response import Response
from rest_framework.decorators import api_view





# Set up logging
logger = logging.getLogger(__name__)



@csrf_exempt
def Home(request):
    return JsonResponse({"data":"Welcome to happy backsideeee hehe"})


@csrf_exempt
def MeetTheTeam(request):
    return JsonResponse({"data":"Meet the team here!"})

@csrf_exempt
def Contact(request):
    return JsonResponse({"data":"Contact Us Page reached without any issues"})

@csrf_exempt
def About(request):
    return JsonResponse({"data":"About Us Page reached without any issues"})

@csrf_exempt
def TermsAndConditions(request):
    return JsonResponse({"data":"Terms and Conditions Page reached without any issues"})

# Working login without checking department
# @csrf_exempt
# def Login(request):
#     if request.method == 'POST':
#         if request.session.get('user_id'):
#             return JsonResponse({'error': 'User already logged in'}, status=403)  # Changed to 403

#         data = json.loads(request.body)
#         emp_emailid = data.get('email')
#         emp_pwd = data.get('password')

#         # Validate input
#         if not emp_emailid or not emp_pwd:
#             return JsonResponse({'error': 'Email and password are required'}, status=400)

#         try:
#             # Check if the employee exists with the provided email and password
#             user = Employee.objects.get(emp_emailid=emp_emailid, emp_pwd=emp_pwd)

#             # Set session data
#             department = user.d_id
#             company = department.c_id

#             request.session['user_id'] = user.emp_emailid
#             request.session['emp_name'] = user.emp_name
#             request.session['emp_emailid'] = user.emp_emailid
#             request.session['emp_role'] = user.emp_role
#             request.session['d_id'] = department.d_id
#             request.session['c_id'] = company.c_id
#             request.session['c_name'] = company.c_name

#             request.session.save()

#             # Prepare response data
#             profile_url = settings.MEDIA_URL + str(user.emp_profile) if user.emp_profile else None

#             response_data = {
#                 'message': 'Login successful',
#                 'user_id': user.emp_emailid,
#                 'emp_name': user.emp_name,
#                 'emp_emailid': user.emp_emailid,
#                 'emp_role': user.emp_role,
#                 'd_id': department.d_id,
#                 'c_id': company.c_id,
#                 'c_name': company.c_name,
#                 'profile_url': profile_url,
#             }

#             return JsonResponse(response_data, status=200)

#         except Employee.DoesNotExist:
#             return JsonResponse({'error': 'Invalid email or password'}, status=401)
#     else:
#         return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
@csrf_exempt
def Login(request):
    if request.method == 'POST':
        if request.session.get('user_id'):
            return JsonResponse({'message': 'User already logged in'}, status=200)

        data = json.loads(request.body)
        emp_emailid = data.get('email')
        emp_pwd = data.get('password')
        emp_role = data.get('department')  # Get department (emp_role) from the request

        # Validate input
        if not emp_emailid or not emp_pwd or not emp_role:
            return JsonResponse({'error': 'Email, password, and department are required'}, status=400)

        try:
            # Check if the employee exists with the provided email, password, and emp_role
            user = Employee.objects.get(emp_emailid=emp_emailid, emp_pwd=emp_pwd, emp_role=emp_role)

            # Set session data
            department = user.d_id
            company = department.c_id

            request.session['user_id'] = user.emp_emailid
            request.session['emp_name'] = user.emp_name
            request.session['emp_emailid'] = user.emp_emailid
            request.session['emp_role'] = user.emp_role
            request.session['d_id'] = department.d_id
            request.session['c_id'] = company.c_id
            request.session['c_name'] = company.c_name

            request.session.save()

            # Prepare response data
            profile_url = settings.MEDIA_URL + str(user.emp_profile) if user.emp_profile else None

            response_data = {
                'message': 'Login successful',
                'user_id': user.emp_emailid,
                'emp_name': user.emp_name,
                'emp_emailid': user.emp_emailid,
                'emp_role': user.emp_role,
                'd_id': department.d_id,
                'c_id': company.c_id,
                'c_name': company.c_name,
                'profile_url': profile_url,
            }

            return JsonResponse(response_data, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Invalid email, password, or department'}, status=401)
    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)

from django.http import JsonResponse
from .models import Employee

@csrf_exempt
def GetDepartments(request):
    if request.method == 'GET':
        # Fetch unique emp_role values from the Employee table
        departments = Employee.objects.values_list('emp_role', flat=True).distinct()
        departments_data = [{"d_id": idx, "d_name": dept} for idx, dept in enumerate(departments)]
        return JsonResponse({"departments": departments_data}, status=200)
    else:
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)


def Users(request):
    company_id = request.session.get('c_id')
    user_id = request.session.get('user_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    employees = Employee.objects.filter(d_id__c_id=company_id, emp_emailid=user_id).values(
        'emp_name', 'emp_emailid', 'emp_phone', 'emp_role', 'd_id', 'emp_profile'
    )
    data = {
        'employees': list(employees)
    }

    return JsonResponse(data)


@csrf_exempt
def Logout(request):
    if request.method == 'POST':
        if 'user_id' in request.session:
            request.session.flush()
            return JsonResponse({'message': 'Logout successful'}, status=200)
        else:
            return JsonResponse({'error': 'No user is logged in'}, status=400)
    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


# NOTE
# ***Department name should contain Super Manager at last ***

    

    
# @csrf_exempt
# def Register(request):
#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)

#             # Extract company details
#             name = data.get('companyName')
#             addr = data.get('companyAddress')
#             phone = data.get('companyPhone')
#             dept_names_str = data.get('deptName')
#             dept_names = [name.strip() for name in dept_names_str.split(',')]

#             # Extract employee details
#             emp_name = data.get('empName')
#             emp_email = data.get('empMail')
#             emp_phone = data.get('empNum')
#             emp_pwd = data.get('emp_pwd', 'changeme')  # Default to 'changeme' if not provided
#             emp_role = data.get('empRole', 'Super Manager')  # Default to 'Super Manager' if not provided
#             emp_skills = data.get('empSkills')

#             # Check if the company with the exact details exists
#             try:
#                 company = Company.objects.get(c_name=name, c_addr=addr, c_phone=phone)
#                 created = False
#             except Company.DoesNotExist:
#                 company = Company.objects.create(c_name=name, c_addr=addr, c_phone=phone)
#                 created = True

#             # Only create departments if the company is newly created
#             if created:
#                 # Create departments and get the first department's ID
#                 first_dept_id = None
#                 for dept_name in dept_names:
#                     # Create department with the Company instance
#                     new_dept = Department.objects.create(d_name=dept_name, c_id=company)
#                     if first_dept_id is None:
#                         first_dept_id = new_dept
#             else:
#                 # If the company already exists, use the existing first department ID
#                 first_dept_id = Department.objects.filter(c_id=company).first()
                
#                 if not first_dept_id:
#                     return JsonResponse({'error': 'No departments found for existing company'}, status=404)

#             # Create the employee
#             Employee.objects.create(
#                 emp_name=emp_name,
#                 emp_emailid=emp_email,
#                 emp_skills=emp_skills,
#                 emp_role=emp_role,
#                 emp_phone=emp_phone,
#                 emp_pwd=emp_pwd,
#                 d_id=first_dept_id
#             )

#             return JsonResponse({'message': 'Company registration successful'}, status=201)

#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

#     else:
#         return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
    
@csrf_exempt
def Register(request):
    if request.method == 'POST':
        
        try:
            data = json.loads(request.body)

            # Extract company details
            name = data.get('companyName')
            addr = data.get('companyAddress')
            phone = data.get('companyPhone')
            dept_names_str = data.get('deptName')
            dept_names = [name.strip() for name in dept_names_str.split(',')]

            # Extract employee details
            emp_name = data.get('empName')
            emp_email = data.get('empMail')
            emp_phone = data.get('empNum')
            emp_pwd = data.get('empPwd')
            emp_role = data.get('deptName')
            emp_skills = data.get('empSkills')

            # Capture the client's IP address
            #office_ip = request.META.get('REMOTE_ADDR')

            # Check if the company with the exact details exists
            try:
                company = Company.objects.get(c_name=name, c_addr=addr, c_phone=phone)
                created = False
            except Company.DoesNotExist:
                # Create a new company with the fetched IP address
                company = Company.objects.create(c_name=name, c_addr=addr, c_phone=phone)
                created = True
            # Only create departments if the company is newly created
            if created:
                first_dept_id = None
                for dept_name in dept_names:
                    new_dept = Department.objects.create(d_name=dept_name, c_id=company)
                    if first_dept_id is None:
                        first_dept_id = new_dept
            else:
                first_dept_id = Department.objects.filter(c_id=company).first()
                
                if not first_dept_id:
                    return JsonResponse({'error': 'No departments found for existing company'}, status=404)

            # Create the employee
            Employee.objects.create(
                emp_name=emp_name,
                emp_emailid=emp_email,
                emp_skills=emp_skills,
                emp_role=emp_role,
                emp_phone=emp_phone,
                emp_pwd=emp_pwd,
                d_id=first_dept_id
            )
            return JsonResponse({'message': 'Company registration successful'}, status=201)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

            return JsonResponse({'message': 'Company registration successful'}, status=201)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)






# incomplete , need to complete and add session
@csrf_exempt
def Profile(request, id):
    try:
        personal_details =  personal_details.objects.filter(mail=id).values()
        if not personal_details.exists():
            return JsonResponse({'error': 'Employee Details not found'}, status=404)
        return JsonResponse(list(personal_details), safe=False)
    except Exception as e:
        print("Error fetching employee data:", e)
        return JsonResponse({'error': 'Internal Server Error'}, status=500)


@csrf_exempt
def SopAndPolicies(request):
    # sop and policies display page. to be visible to everyone
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if user_id and company_id:
        tasks_data = Tasks.objects.filter(emp_emailid=user_id).values('ratings', 'remark', 'sop_id', 'selfratings')

        sop_ids = [task['sop_id'] for task in tasks_data]

        sop_data = Sop.objects.filter(sop_id__in=sop_ids).values('sop_id', 'type', 's_name', 'sdate')
        sop_data_dict = {sop['sop_id']: sop for sop in sop_data}

        files_data = Files.objects.filter(sop_id__in=sop_ids).values('file_name', 'sop_id')
        sop_files_dict = {}
        for file in files_data:
            sop_id = file['sop_id']
            if sop_id not in sop_files_dict:
                sop_files_dict[sop_id] = []
            sop_files_dict[sop_id].append(file['file_name'])

        combined_data = []
        for task in tasks_data:
            sop_id = task['sop_id']
            sop_info = sop_data_dict.get(sop_id, {})
            files_info = sop_files_dict.get(sop_id, [])

            combined_task = {
                'ratings': task['ratings'],
                'remark': task['remark'],
                'sop_id': sop_id,
                'selfratings': task['selfratings'],
                'sop_info': sop_info,
                'files': files_info
            }
            combined_data.append(combined_task)

        return JsonResponse({'data': 
            combined_data}, safe=False)
    else:
        return JsonResponse({'error': 'User not logged in'}, status=401)


@csrf_exempt
@require_http_methods(["PUT"])
def UpdateSelfratings(request, sop_id):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    try:
        task = Tasks.objects.get(sop_id=sop_id, emp_emailid=user_id)

        data = json.loads(request.body)
        self_rating = int(data.get('selfratings'))
        task.selfratings = self_rating
        task.save()

        return JsonResponse({'message': 'Rating updated successfully'}, status=200)
    except Tasks.DoesNotExist:
        return JsonResponse({'error': 'Task does not exist'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def KraList(request):
    if request.method == 'GET':
        kras = Kra.objects.all().values('kra_no', 'kra', 'weightage', 'kpi', 'target', 'ratings_scale', 'ratings', 'selfratings', 'remarks', 'status', 'email_id', 'kra_id')
        kra_list = list(kras)
        return JsonResponse(kra_list, safe=False)
    else:
        return JsonResponse({'error': 'Invalid HTTP method'}, status=400)


@csrf_exempt
def Forms(request):
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)
    user_name = request.session.get('emp_name')

    if not c_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    if request.method == 'GET':
        try:
            employee = get_object_or_404(Employee, emp_name=user_name, d_id__c_id=company_id)

            allocated_forms = []

            forms = Custom_forms.objects.filter(c_id=company_id)

            for form in forms:
                allocated_users = [user.strip() for user in form.alloc.split(',') if user.strip()]
                if allocated_users and user_name in allocated_users:
                    formatted_form_name = form.form_name.replace(' ', '_').lower()
                    response_table_name = f"resp_{company_id}{formatted_form_name}"

                    with connection.cursor() as cursor:
                        cursor.execute(f"SELECT COUNT(*) FROM {response_table_name} WHERE emp_name = %s", [user_name])
                        responded = cursor.fetchone()[0] > 0

                    allocated_forms.append({
                        'form_name': form.form_name,
                        'responded': responded
                    })

            return JsonResponse({'forms': allocated_forms}, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def FormReviewRespose(request):
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)
    user_name = request.session.get('emp_name')

    if not c_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    if request.method == 'GET':
        try:
            form_name = request.GET.get('form_name')

            if not form_name:
                return JsonResponse({'error': 'Form name is required'}, status=400)

            form = Custom_forms.objects.get(c_id=company_id, form_name = form_name)

            formatted_form_name = form.form_name.replace(' ', '_').lower()
            response_table_name = f"resp_{company_id}{formatted_form_name}"
            response_data = []

            with connection.cursor() as cursor:
                cursor.execute(f"SELECT * FROM {response_table_name} WHERE emp_name = %s", [user_name])
                response_row = cursor.fetchone()
                if not response_row:
                    return JsonResponse({'error': 'Response not found'}, status=404)

                columns = [col[0] for col in cursor.description]
                for col, val in zip(columns, response_row):
                    if col != 'emp_name':
                        cursor.execute(f"SELECT label FROM Custom_forms_questions WHERE ID = %s AND form_name = %s AND c_id = %s", [col, form_name, c_id])
                        question = cursor.fetchone()
                        if question:
                            response_data.append({'question': question[0], 'answer': val})

            return JsonResponse({'response': response_data}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def FormsSubmitResponse(request):
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)
    user_name = request.session.get('emp_name')

    if not c_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            form_name = data.get('form_name')

            if not form_name:
                return JsonResponse({'error': 'Form name is required'}, status=400)

            form = Custom_forms.objects.get(c_id=company_id, form_name = form_name)

            formatted_form_name = form.form_name.replace(' ', '_').lower()
            response_table_name = f"resp_{company_id}{formatted_form_name}"

            columns = ', '.join(data.keys())
            values = ', '.join(['%s'] * len(data))
            query = f"INSERT INTO {response_table_name} ({columns}) VALUES ({values})"
            params = list(data.values())

            with connection.cursor() as cursor:
                cursor.execute(query, params)

            return JsonResponse({'message': 'Response submitted successfully'}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def Letters(request):
    if 'emp_emailid' in request.session and 'c_id' in request.session:
        emp_emailid = request.session['emp_emailid']
        c_id = request.session['c_id']

        try:
            user_name = Employee.objects.get(emp_emailid=emp_emailid).emp_name
            company = get_object_or_404(Company, pk=c_id)

            allocated_letters = []

            custom_letters = Custom_letters.objects.filter(c_id=company)

            for letter in custom_letters:
                allocated_emails = [email.strip() for email in letter.alloc.split(',') if email.strip()]
                if allocated_emails and emp_emailid in allocated_emails:
                    allocated_letters.append({
                        'letter_name': letter.letter_name.replace('_', ' ').title()
                    })

            return JsonResponse({'allocated_letters': allocated_letters})

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Required session data not found'}, status=401)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin', 'Employee'])
def FAQsView(request):
    c_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')

    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            faqs = Faqs.objects.filter(c_id=c_id).values('faq_id', 'question', 'answer', 'emp_emailid', 'imp')
            data = list(faqs)
            return JsonResponse({'faqs': data})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            question = data.get('question')
            emp_emailid = get_object_or_404(Employee, emp_emailid=emp_emailid)
            c_id = get_object_or_404(Company, c_id=c_id)

            with transaction.atomic():
                new_faq = Faqs.objects.create(
                    c_id=c_id,
                    emp_emailid=emp_emailid,
                    question=question,
                    answer='',
                    imp=False
                )
                return JsonResponse({'success': 'FAQ added successfully', 'faq_id': new_faq.faq_id}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

@csrf_exempt
def ApplyLeave(request):
    if request.method == 'POST':
        try:
            # Parse the request body to get the data
            data = json.loads(request.body)
            leavetype = data.get('leaveType')
            fromdate = data.get('fromDate')
            todate = data.get('toDate')
            description = data.get('leaveDescription')

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        # Check if user is authenticated through session
        emp_email = request.session.get('emp_emailid')
        if not emp_email:
            return JsonResponse({'error': 'User not authenticated'}, status=401)

        try:
            # Retrieve employee record based on email
            employee = Employee.objects.get(emp_emailid=emp_email)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

        try:
            # Parse and validate the date formats
            date1 = datetime.strptime(fromdate, "%Y-%m-%d")
            date2 = datetime.strptime(todate, "%Y-%m-%d")

            # Validate that fromdate is not later than todate
            if date1 > date2:
                return JsonResponse({'error': 'From date cannot be later than To date'}, status=400)

            days = (date2 - date1).days + 1  # Include both start and end dates
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)

        try:
            # Access the company through the employee's department
            leave_type = Leavetype.objects.get(LeaveType__iexact=leavetype, company_id=employee.d_id.c_id)


        except Leavetype.DoesNotExist:
            return JsonResponse({'error': 'Leave type does not exist'}, status=400)

        # Check if the Leavecounttemp record exists
        if Leavecounttemp.objects.filter(emp_emailid=emp_email).exists():
            leave_count = Leavecounttemp.objects.get(emp_emailid=emp_email)
        else:
            # Create a new Leavecounttemp record with default values
            leave_count = Leavecounttemp.objects.create(
                emp_emailid=emp_email,
                casualleave=15,
                medicalleave=15,
                lopleave=365,
                earnedleave=20  # Set default values as needed
            )

        # Mapping the leave type to the correct field
        leave_type_mapping = {
            'Casual Leave': 'casualleave',
            'Medical Leave': 'medicalleave',
            'LOP Leave': 'lopleave',
            'Earned Leave': 'earnedleave'
        }

        leave_limit_field = leave_type_mapping.get(leavetype, '').lower()

        # Validate if the field exists in the Leavecounttemp model
        if not leave_limit_field:
            return JsonResponse({'error': 'Invalid leave type provided'}, status=400)

        leave_limit = getattr(leave_count, leave_limit_field, 0)

        final = leave_limit - days
        if final < 0:
            return JsonResponse({'error': 'Exceeding leave limits'}, status=400)

        # Create the leave record
        leave = Tblleaves.objects.create(
            LeaveType=leave_type,
            FromDate=fromdate,
            ToDate=todate,
            Days=days,
            Description=description,
            Status=0,
            IsRead=0,
            emp_emailid=employee
        )

        return JsonResponse({'message': 'Leave submitted successfully'}, status=200)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)





@csrf_exempt
def LeaveHistory(request):
    emp_emailid = request.session.get('emp_emailid')

    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    leaves = Tblleaves.objects.filter(emp_emailid=emp_emailid).values('id', 'LeaveType__LeaveType', 'FromDate', 'ToDate', 'PostingDate', 'AdminRemark', 'Status')
    leaves_data = list(leaves)
    print(leaves_data)

    return JsonResponse({'Leaves': leaves_data}, status=200)


@csrf_exempt
def AddLoan(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'Employee email not found in session'}, status=400)

    if request.method == 'GET':
        loans = Loan.objects.filter(emp_emailid=emp_emailid).values('name', 'department', 'lamt', 'mamt', 'startdate', 'reason', 'status')

        data = []
        for index, loan in enumerate(loans, start=1):
            status = "Pending" if loan['status'] == 0 else "Approved"
            data.append({
                'name': loan['name'],
                'department': loan['department'],
                'loan_amount': loan['lamt'],
                'monthly_amount': loan['mamt'],
                'start_date': loan['startdate'],
                'reason': loan['reason'],
                'status': status
            })

        return JsonResponse({'loans': data})

    elif request.method == 'POST':
        try:
            loan_data = json.loads(request.body)
            name = loan_data.get('name')
            department = loan_data.get('department')
            lamt = loan_data.get('lamt')
            mamt = loan_data.get('mamt')
            startdate = loan_data.get('startdate')
            reason = loan_data.get('reason')
            status = loan_data.get('status')

            Loan.objects.create(name=name, emp_emailid=emp_emailid, department=department, lamt=lamt, mamt=mamt, startdate=startdate, reason=reason, status=status)

            return JsonResponse({'message': 'Loan applied successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddExpenses(request):
    emp_emailid = request.session.get('emp_emailid')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email not found in session'}, status=400)

    if request.method == 'POST':
        expense_data = json.loads(request.body)

        expense = expense_data.get('expense')
        expensedate = expense_data.get('expensedate')
        expenseitm = expense_data.get('expenseitm')

        Expenses.objects.create(emp_emailid=emp_emailid, expense=expense, expensedate=expensedate, expenseitm=expenseitm)

        return JsonResponse({'message': 'Expense added successfully'})

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin', 'Employee'])
def ManageExpenses(request):
    emp_emailid = request.session.get('emp_emailid')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email not found in session'}, status=400)

    if request.method == 'GET':
        expenses = Expenses.objects.filter(emp_emailid=emp_emailid).values('expensedate', 'expenseitm', 'expense_id').annotate(total_expense=Sum('expense')).order_by('expensedate')

        data = []
        for expense in expenses:
            data.append({
                'date': expense['expensedate'].strftime('%Y-%m-%d'),
                'amount': expense['total_expense'],
                'expense_item': expense['expenseitm'],
                'expense_id': expense['expense_id']
            })

        return JsonResponse({'expenses': data})

    elif request.method == 'POST':
        expense_data = json.loads(request.body)
        expense_id = request.GET.get('expense_id')

        if expense_id is not None:
            expense = get_object_or_404(Expenses, expense_id=expense_id, emp_emailid=emp_emailid)
            expense.expensedate = expense_data.get('date', expense.expensedate)
            expense.expense = expense_data.get('amount', expense.expense)
            expense.expenseitm = expense_data.get('expense_item', expense.expenseitm)
            expense.save()
            return JsonResponse({'message': 'Expense updated successfully'})
        else:
            return JsonResponse({'error': 'Expense ID not provided'}, status=400)


    elif request.method == 'DELETE':
        expense_id = request.GET.get('expense_id')
        expense = get_object_or_404(Expenses, expense_id=expense_id, emp_emailid=emp_emailid)
        expense.delete()
        return JsonResponse({'message': 'Expense deleted successfully'})

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ExpenseReport(request):
    emp_emailid = request.session.get('emp_emailid')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email not found in session'}, status=400)

    try:
        expenses = Expenses.objects.filter(emp_emailid=emp_emailid).values('expensedate', 'expenseitm', 'expense_id').annotate(total_expense=Sum('expense')).order_by('expensedate')

        data = []
        for expense in expenses:
            data.append({
                'date': expense['expensedate'].strftime('%Y-%m-%d'),
                'amount': expense['total_expense'],
                'expense_item': expense['expenseitm'],
                'expense_id': expense['expense_id']
            })

        return JsonResponse({'expenses': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def Resign(request):
    c_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')

    if not c_id or not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    email = get_object_or_404(Resignation, emp_emailid=emp_emailid)

    if request.method == 'GET':
        resignations = Resignation.objects.all()
        print(resignations)
        resignation_data = []
        for resignation in resignations:
            resignation_data.append({
                'emp_emailid': email,
                'submit_date': resignation.submit_date,
                'exp_leave': resignation.exp_leave,
                'leave_reason': resignation.leave_reason,
                'leave_reason_2': resignation.leave_reason_2,
                'leave_reason_3': resignation.leave_reason_3,
                'leave_date': resignation.leave_date,
                'notice_serve': resignation.notice_serve,
                'settle_date': resignation.settle_date,
                'shortfall_date': resignation.shortfall_date,
                'exit_interview': resignation.exit_interview,
                'last_working': resignation.last_working,
                'status': resignation.status,
                'approved_by': resignation.approved_by,
                'notice_per': resignation.notice_per,
            })
        return JsonResponse(resignation_data, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            emp_emailid = data.get('emp_emailid')
            exp_leave = data.get('exp_leave')
            leave_reason = data.get('leave_reason')
            leave_reason_2 = data.get('leave_reason_2')
            leave_reason_3 = data.get('leave_reason_3')
            leave_date = data.get('leave_date')
            settle_date = data.get('settle_date')
            exit_interview = data.get('exit_interview')
            last_working = data.get('last_working')
            status = data.get('status')
            approved_by = data.get('approved_by')

            if not all([emp_emailid, exp_leave, leave_reason, exit_interview, last_working]):
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            submit_date = timezone.now().date()
            date1 = submit_date
            date2 = timezone.now().date()
            diff = (date2 - date1).days
            notice_serve = abs(diff)
            notice_per = 30
            shortfall_date = max(0, 30 - notice_serve)

            resignation, created = Resignation.objects.update_or_create(
                emp_emailid=emp_emailid,
                defaults={
                    'submit_date': submit_date,
                    'exp_leave': exp_leave,
                    'leave_reason': leave_reason,
                    'leave_reason_2': leave_reason_2,
                    'leave_reason_3': leave_reason_3,
                    'leave_date': leave_date,
                    'notice_serve': notice_serve,
                    'settle_date': settle_date,
                    'shortfall_date': shortfall_date,
                    'exit_interview': exit_interview,
                    'last_working': last_working,
                    'status': status,
                    'approved_by': approved_by,
                    'notice_per': notice_per,
                }
            )

            response = {'message': 'Resignation created successfully'} if created else {'message': 'Resignation updated successfully'}
            return JsonResponse({'status': 1, 'message': response})

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['Super Admin', 'Manager', 'HR','Employee'])
def AddCourses(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    # Check if user is logged in and has a valid company ID
    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        try:
            course_title = request.POST.get('course_title')
            description = request.POST.get('description')
            thumbnail = request.FILES.get('thumbnail')  # Optional field

            # Check for required fields
            if not all([course_title, description]):
                return HttpResponseBadRequest("Missing required fields")

            # Fetch the company based on company_id from session
            company = get_object_or_404(Company, pk=company_id)

            # Ensure that the course is being added to the correct company
            if company.c_id != company_id:
                return JsonResponse({'error': 'Invalid company ID'}, status=403)

            # Prepare course data
            course_data = {
                'course_title': course_title,
                'description': description,
                'c_id': company,  # Ensure that the company is the same as user's company
                'c_name': company.c_name
            }

            # Include thumbnail only if provided
            if thumbnail:
                course_data['thumbnail'] = thumbnail

            # Create the course within the company
            course = Courses.objects.create(**course_data)

            response_data = {
                'message': 'Course added successfully',
                'course_id': course.course_id
            }

            return JsonResponse(response_data, status=201)

        except Exception as e:
            # Handle exceptions gracefully and return an error message
            return JsonResponse({'error': str(e)}, status=400)

    else:
        # Only allow POST requests
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
    


@csrf_exempt
@role_required(['Super Admin', 'Manager', 'HR','Employee'])
def GetCourses(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    # Check if user is logged in and has a valid company ID
    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in or no company ID found'}, status=401)

    if request.method == 'GET':
        try:
            # Fetch the company based on company_id from session
            company = get_object_or_404(Company, pk=company_id)

            # Fetch all courses for the logged-in user's company
            courses = Courses.objects.filter(c_id=company)

            # Prepare the response data
            course_list = []
            for course in courses:
                course_list.append({
                    'id': course.course_id,         # Ensure 'id' is used as expected in frontend
                    'title': course.course_title,   # Ensure 'title' is used as expected in frontend
                })

            return JsonResponse({'courses': course_list}, status=200)

        except Exception as e:
            # Handle exceptions gracefully and return an error message
            return JsonResponse({'error': str(e)}, status=400)

    else:
        # Only allow GET requests
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)

@csrf_exempt
@role_required(['Super Admin', 'Manager', 'HR','Employee'])
def GetCourses(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    # Check if user is logged in and has a valid company ID
    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in or no company ID found'}, status=401)

    if request.method == 'GET':
        try:
            # Fetch the company based on company_id from session
            company = get_object_or_404(Company, pk=company_id)

            # Fetch all courses for the logged-in user's company
            courses = Courses.objects.filter(c_id=company)

            # If no courses found, return a message
            if not courses.exists():
                return JsonResponse({'message': 'No courses found for this company'}, status=200)

            # Prepare the response data
            course_list = []
            for course in courses:
                course_list.append({
                    'course_id': course.course_id,
                    'course_title': course.course_title,
                    'description': course.description,
                    'thumbnail': course.thumbnail.url if course.thumbnail else None,
                    'company_name': course.c_name,
                })

            return JsonResponse({'courses': course_list}, status=200)

        except Exception as e:
            # Handle exceptions gracefully and return an error message
            return JsonResponse({'error': str(e)}, status=400)

    else:
        # Only allow GET requests
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UploadMedia(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            # video_name = data.get('video_name')
            location = data.get('location')
            descr = data.get('descr')
            course_id = data.get('course_id')

            try:
                course = Courses.objects.get(course_id=course_id)
            except Courses.DoesNotExist:
                return JsonResponse({'error': 'Course not found'}, status=404)

            existing_videos_count = Video.objects.filter(course_id=course_id).count()
            if existing_videos_count > 1:
                return JsonResponse({'message': 'Video with the same details already exists.'}, status=400)

            video = Video.objects.create(
                location=location,
                descr=descr,
                course_id=course,
                # video_name=video_name
            )

            return JsonResponse({'message': 'Video added successfully'}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UploadPdf(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'POST':
        try:
            pdf_file = request.FILES.get('pdf_file')
            pdf_description = request.POST.get('pdf_description')
            course_id = request.POST.get('course_id')

            if not pdf_file or not pdf_description or not course_id:
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            try:
                # Fetch the course using the provided course_id
                course = Courses.objects.get(course_id=course_id)
            except Courses.DoesNotExist:
                return JsonResponse({'error': 'Course not found'}, status=404)

            # Ensure the 'pdfs/' directory exists
            pdf_dir = os.path.join(settings.MEDIA_ROOT, 'pdfs')
            if not os.path.exists(pdf_dir):
                os.makedirs(pdf_dir)

            # Save the PDF file
            pdf_path = os.path.join('pdfs', pdf_file.name)
            with open(os.path.join(settings.MEDIA_ROOT, pdf_path), 'wb+') as destination:
                for chunk in pdf_file.chunks():
                    destination.write(chunk)

            # Save PDF details in the database
            pdf = Pdf.objects.create(
                pdf_name=pdf_file.name,
                location=pdf_path,
                descr=pdf_description,
                course=course
            )

            return JsonResponse({'message': 'PDF uploaded successfully'}, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
    



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddMediaContent(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        c_id = request.session.get('c_id')

        if not c_id:
            return JsonResponse({'error': 'Company ID not found in session'}, status=400)

        courses_data = Courses.objects.filter(c_id=c_id).values('course_title', 'course_id')
        data = list(courses_data)

        return JsonResponse(data, safe=False)
    else:
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UpdateDeleteMedia(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        courses_data = Courses.objects.filter(c_id=company_id).values('course_title', 'description', 'course_id')
        data = list(courses_data)
        return JsonResponse(data, safe=False)
    else:
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UpdateMedia(request, course_id):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    try:
        course = Courses.objects.get(course_id=course_id, c_id=company_id)
    except Courses.DoesNotExist:
        return JsonResponse({'error': 'Course not found or you do not have permission to access it'}, status=404)

    if request.method == 'GET':
        try:
            videos = Video.objects.filter(course_id=course_id).values('video_id', 'location', 'descr')

            course_data = {
                'course_id': course.course_id,
                'course_title': course.course_title,
                'videos': list(videos)
            }
            return JsonResponse(course_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            course = Courses.objects.get(course_id=course_id)

            if 'course_title' in data:
                course.course_title = data['course_title']
                course.save()

            if 'videos' in data:
                for video_data in data['videos']:
                    video_id = video_data.get('video_id')
                    if video_id:
                        video = Video.objects.get(video_id=video_id, course_id=course_id)
                        video.location = video_data.get('location', video.location)
                        video.descr = video_data.get('descr', video.descr)
                        video.save()

            return JsonResponse({'message': 'Course updated successfully'}, status=200)
        except (Courses.DoesNotExist, Video.DoesNotExist):
            return JsonResponse({'error': 'Course or video not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'DELETE':
        try:
            media_content = Courses.objects.get(course_id=course_id)
            media_content.delete()
            return JsonResponse({'message': 'Media content deleted successfully'})
        except Courses.DoesNotExist:
            return JsonResponse({'error': 'Media content not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ReportingStructureForm(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        employees = Employee.objects.filter(d_id__c_id=company_id, Status='Active').values('emp_emailid', 'd_id__c_id')
        employees_list = list(employees)
        return JsonResponse(employees_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            selected_employees = data.get('eemail', [])
            reporting_employee = data.get('remail')
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)

        if not selected_employees or not reporting_employee:
            return JsonResponse({'error': 'Select employees and a reporting employee first!'}, status=400)

        with transaction.atomic():
            try:
                for emp_email in selected_employees:
                    Reporting.objects.update_or_create(
                        c_id=company_id,
                        Reporting_from=emp_email,
                        defaults={'Reporting_to': reporting_employee}
                    )
                return JsonResponse({'success': 'Reporting structure updated successfully'}, status=200)
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CeoHrAnnouncements(request):
    company_id = request.session.get('c_id')
    print(f"Company ID from session: {company_id}")  # Debugging

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'POST':
        sender = request.POST.get('email')  # Correct syntax for calling 'get'
        cc = request.POST.get('cc', '')
        departments = request.POST.getlist('dept')
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        files = request.FILES.getlist('files')
        send_to_all = 'check' in request.POST and request.POST['check'] == 'yes'

        # Debugging statements to verify POST data
        print(f"Sender Email: {sender}")
        print(f"CC Email: {cc}")
        print(f"Departments: {departments}")
        print(f"Subject: {subject}")
        print(f"Message: {message}")
        print(f"Files: {files}")
        print(f"Send to All: {send_to_all}")


        emails_sent = 0
        separator = getattr(settings, 'EMAIL_SEPARATOR', ',')  # Use default if not set
        print(f"Email Separator: {separator}")  # Debugging
        try:
            if send_to_all:
                employees = Employee.objects.filter(d_id__c_id=company_id, Status='Active')
                print(f"Employees (send_to_all=True): {employees.count()} found")  # Debugging
            else:
                employees = Employee.objects.filter(d_id__in=departments, Status='Active')
                print(f"Employees (send_to_all=False): {employees.count()} found")  # Debugging
            for employee in employees:
                print(f"Employee Email: {employee.emp_emailid}")  # Debugging

            for employee in employees:
                email = EmailMessage(
                    subject=subject,
                    body=message,
                    from_email=sender,
                    to=[employee.emp_emailid],
                    cc=[cc],  # Use a list directly
                )

                for file in files:
                    email.attach(file.name, file.read(), file.content_type)

                email.send()
                print(f"Email sent to: {employee.emp_emailid}")  # Debugging
                emails_sent += 1

            return JsonResponse({'success': f'Total {emails_sent} emails sent successfully'}, status=200)

        except Exception as e:
            # Improved exception handling and logging
            print(f"Error sending emails: {str(e)}")  # Use logging in production
            return JsonResponse({'error': 'An error occurred while sending emails. Check server logs for details.'}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)



# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def CeoHrAnnouncements(request):
#     company_id = request.session.get('c_id')
#     if not company_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'GET':
#         # Fetch departments for the logged-in company
#         departments = Department.objects.filter(c_id=company_id).values('d_id', 'd_name')
#         return JsonResponse(list(departments), safe=False)

#     if request.method == 'POST':
#         sender = request.POST.get('email')
#         cc = request.POST.get('cc', '')
#         departments = request.POST.getlist('dept')  # Expecting d_id values
#         subject = request.POST.get('subject')
#         message = request.POST.get('message')
#         files = request.FILES.getlist('files')
#         send_to_all = request.POST.get('check') == 'yes'

#         try:
#             if send_to_all:
#                 employees = Employee.objects.filter(d_id__c_id=company_id, Status='Active')
#             else:
#                 employees = Employee.objects.filter(d_id__in=departments, Status='Active')

#             emails_sent = 0
#             for employee in employees:
#                 email = EmailMessage(
#                     subject=subject,
#                     body=message,
#                     from_email=sender,
#                     to=[employee.emp_emailid],
#                     cc=[cc] if cc else None,
#                 )
#                 for file in files:
#                     email.attach(file.name, file.read(), file.content_type)
#                 email.send()
#                 emails_sent += 1

#             return JsonResponse({'success': f'Total {emails_sent} emails sent successfully'}, status=200)
        
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

#     return JsonResponse({'error': 'Invalid request method'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddNewEmployee(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        emp_count = Employee.objects.filter(d_id__c_id=company_id).count()
        company_data = Company.objects.filter(c_id=company_id).values('payment_type', 'emp_limit').first()
        departments = Department.objects.filter(c_id=company_id).values('d_id', 'd_name')


        response_data = {
            'emp_count': emp_count,
            'company': company_data,
            'departments': list(departments)
        }

        return JsonResponse(response_data, status=200)

    elif request.method == 'POST':
        data = json.loads(request.body)
        name = data.get('fullName')
        email = data.get('email')
        phone = data.get('phone')
        department_id = data.get('department')
        skills = data.get('skills')
        emp_role = data.get('emp_role')

        valid_roles = ['HR', 'Super Admin', 'Manager', 'Employee']
        if emp_role not in valid_roles:
            return JsonResponse({'error': 'Invalid role selected'}, status=400)


        company = Company.objects.filter(c_id=company_id).first()
        department = Department.objects.filter(d_id=department_id).first()

        # if company.emp_limit <= Employee.objects.filter(d_id__c_id=company_id).count():
        #     return JsonResponse({'error': 'Employee limit reached'}, status=403)

        Employee.objects.create(
            emp_name=name,
            emp_emailid=email,
            emp_phone=phone,
            emp_skills=skills,
            emp_role=emp_role,  # Assign the validated role
            d_id=department,
        )

        return JsonResponse({'success': 'Employee created successfully'}, status=201)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UpdateDeleteEmployee(request):
    company_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            employees = Employee.objects.filter(d_id__c_id=company_id)
            email_ids = [employee.emp_emailid for employee in employees]
            return JsonResponse({'email_ids': email_ids}, status=200)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employees not found or you do not have permission to access them'}, status=404)

    elif request.method == 'DELETE':
        try:
            email_id = request.GET.get('emailId')
            employee = Employee.objects.get(emp_emailid=email_id, d_id__c_id=company_id)
            if Quiz.objects.filter(eid=employee).exists():
                return JsonResponse({'error': 'Cannot delete employee with existing quizzes'}, status=400)

            employee.delete()
            return JsonResponse({'success': 'Employee deleted successfully'}, status=200)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found or you do not have permission to delete it'}, status=404)


# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def UpdateEmployeeDetails(request):
#     company_id = request.session.get('c_id')
#     emp_emailid = request.session.get('emp_emailid')

#     if not company_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'GET':
#         try:
#             employee = Employee.objects.get(emp_emailid=emp_emailid, d_id__c_id=company_id)
#             employee_data = {
#                 'emp_name': employee.emp_name,
#                 'emp_emailid': employee.emp_emailid,
#                 'emp_phone': employee.emp_phone,
#                 'd_id': employee.d_id.d_id,
#                 'emp_skills': employee.emp_skills,
#             }

#             departments = Department.objects.filter(c_id=company_id).values('d_id', 'd_name')

#             response_data = {
#                 'employee_data': employee_data,
#                 'departments': list(departments)
#             }

#             return JsonResponse(response_data, status=200)
#         except Employee.DoesNotExist:
#             return JsonResponse({'error': 'Employee not found or you do not have permission to access it'}, status=404)

#     elif request.method == 'POST':
#         data = json.loads(request.body)
#         emp_name = data.get('emp_name')
#         emp_emailid = data.get('emp_emailid')
#         emp_phone = data.get('emp_phone')
#         d_id = data.get('d_id')
#         emp_skills = data.get('emp_skills')

#         try:
#             employee = Employee.objects.get(emp_emailid=emp_emailid, d_id__c_id=company_id)
#             employee.emp_name = emp_name
#             employee.emp_phone = emp_phone
#             employee.d_id = Department.objects.get(d_id=d_id, c_id=company_id)
#             employee.emp_skills = emp_skills
#             employee.save()
#             return JsonResponse({'success': 'Employee details updated successfully'}, status=200)
#         except Employee.DoesNotExist:
#             return JsonResponse({'error': 'Employee not found or you do not have permission to update it'}, status=404)
@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def UpdateEmployeeDetails(request):
    company_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')
    to_be_updated_employee_email = request.GET.get('emp_emailid')


    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            employee = Employee.objects.get(emp_emailid=to_be_updated_employee_email or emp_emailid, d_id__c_id=company_id)
            
            employee_data = {
                'emp_name': employee.emp_name,
                'emp_emailid': employee.emp_emailid,
                'emp_phone': employee.emp_phone,
                'd_id': employee.d_id.d_id,
                'emp_skills': employee.emp_skills,
                'emp_role': employee.emp_role,
            }

            departments = Department.objects.filter(c_id=company_id).values('d_id', 'd_name')

            response_data = {
                'employee_data': employee_data,
                'departments': list(departments)
            }

            return JsonResponse(response_data, status=200)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found or you do not have permission to access it'}, status=404)

    elif request.method == 'POST':
        data = json.loads(request.body)
        emp_name = data.get('emp_name')
        emp_emailid = data.get('emp_emailid')
        emp_phone = data.get('emp_phone')
        emp_role = data.get('emp_role')
        d_id = data.get('d_id')
        emp_skills = data.get('emp_skills')

        try:
            employee = Employee.objects.get(emp_emailid=emp_emailid, d_id__c_id=company_id)
            employee.emp_name = emp_name
            employee.emp_phone = emp_phone
            employee.emp_role = emp_role
            employee.d_id = Department.objects.get(d_id=d_id, c_id=company_id)
            employee.emp_skills = emp_skills
            employee.save()
            return JsonResponse({'success': 'Employee details updated successfully'}, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found or you do not have permission to update it'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def JDForms(request):
    if request.method == 'POST':
        res = request.POST.getlist('res[]')
        role = request.POST.get('role')
        sdate = request.POST.get('date')
        emp_emails = request.POST.getlist('eemail[]')

        if not (res and role and sdate and emp_emails):
            return JsonResponse({'error': 'All fields are required!'}, status=400)

        try:
            sdate = datetime.datetime.strptime(sdate, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Invalid date format!'}, status=400)

        for emp_email in emp_emails:
            if not Employee.objects.filter(emp_emailid=emp_email).exists():
                return JsonResponse({'error': f'Employee with email {emp_email} does not exist!'}, status=400)

        try:
            jd_instance = Job_desc.objects.create(jd_name=role, sdate=sdate)

            for emp_email in emp_emails:
                for responsibility in res:
                    Tasks.objects.create(job_desc_id=jd_instance, responsiblities=responsibility, emp_emailid=emp_email, sdate=sdate)

            return JsonResponse({'message': 'Form submitted successfully!'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)



@csrf_exempt
def LeaveDashboard(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        # Fetch department, employee, and leave type counts
        departments = Department.objects.filter(c_id=company_id)
        dptcount = departments.count()
        employees = Employee.objects.filter(d_id__in=departments).values('emp_emailid')
        empcount = employees.count()
        leavtypcount = Leavetype.objects.count()

        emp_emails = [e['emp_emailid'] for e in employees]
        leaves_fetched = Tblleaves.objects.filter(emp_emailid__in=emp_emails).order_by('-id')

        # Define status mapping from numbers to meaningful strings
        status_mapping = {
            0: 'Pending',
            1: 'Approved',
            2: 'Rejected',
            # Add more status mappings as needed
        }

        leaves_data = []
        for leave in leaves_fetched:
            leave_dict = {
                'lid': leave.id,
                'emp_name': leave.emp_emailid.emp_name,
                'emp_emailid': leave.emp_emailid.emp_emailid,
                'LeaveType': leave.LeaveType.LeaveType,
                'PostingDate': leave.PostingDate.strftime('%Y-%m-%d %H:%M:%S'),
                'Status': status_mapping.get(leave.Status, 'Unknown'),  # Map status number to string
            }
            leaves_data.append(leave_dict)

        response_data = {
            'empcount': empcount,
            'dptcount': dptcount,
            'leavtypcount': leavtypcount,
            'leaves_fetched': leaves_data,
        }
        return JsonResponse(response_data, safe=False)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


#leaveaction

@csrf_exempt
def update_leave_status(request, leave_id):
    # Check for session authentication
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == "POST":
        try:
            # Parse JSON body instead of using request.POST
            data = json.loads(request.body)
            
            status = data.get('status')  # 'approved' or 'not-approved'
            admin_remark = data.get('description')

            # Fetch the leave record based on the ID passed from the frontend
            leave = Tblleaves.objects.get(id=leave_id)

            # Map the frontend status to backend integer values
            if status == 'approved':
                leave.Status = 1  # Approved
            elif status == 'not-approved':
                leave.Status = 2  # Rejected
            else:
                return JsonResponse({"error": "Invalid status"}, status=400)

            # Update the remark and the date of action
            leave.AdminRemark = admin_remark
            leave.AdminRemarkDate = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

            # Save changes to the database
            leave.save()

            return JsonResponse({"message": "Leave status updated successfully"}, status=200)

        except Tblleaves.DoesNotExist:
            return JsonResponse({"error": "Leave record not found"}, status=404)

    return JsonResponse({"error": "Invalid request method"}, status=405)

# Configure logging
logging.basicConfig(level=logging.DEBUG)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def LeaveDetails(request, id):  # Accept the 'id' parameter from the URL
    logging.debug(f"Received request method: {request.method}")

    if request.method == 'GET':
        try:
            logging.debug(f"Leave ID: {id}")
            if not id:
                return JsonResponse({'error': 'Leave ID is required'}, status=400)

            company_id = request.session.get('c_id')
            logging.debug(f"Company ID from session: {company_id}")

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            leave_dict = Tblleaves.objects.filter(id=id).values(
                'id', 'LeaveType__LeaveType', 'ToDate', 'FromDate', 'Description', 'PostingDate', 'Status',
                'AdminRemark', 'AdminRemarkDate', 'emp_emailid__emp_name', 'emp_emailid__emp_phone',
                'emp_emailid__d_id__c_id', 'emp_emailid__emp_role'
            ).first()

            logging.debug(f"Leave Details: {leave_dict}")

            if not leave_dict:
                return JsonResponse({'error': 'Leave not found'}, status=404)

            user_company_id = request.session.get('c_id')
            if leave_dict['emp_emailid__d_id__c_id'] != user_company_id:
                return JsonResponse({'error': 'You are not authorized to view this leave details'}, status=403)

            # Mapping statuses: 0 = Pending, 1 = Approved, 2 = Rejected
            status_mapping = {0: "Pending", 1: "Approved", 2: "Rejected"}
            leave_dict['Status'] = status_mapping.get(leave_dict['Status'], "Unknown")

            leave_dict['emp_emailid__emp_name'] = leave_dict['emp_emailid__emp_name'].strip()
            leave_dict['PostingDate'] = leave_dict['PostingDate'].isoformat()
            leave_dict['AdminRemarkDate'] = leave_dict['AdminRemarkDate'].isoformat() if leave_dict['AdminRemarkDate'] else None

            return JsonResponse(leave_dict, safe=False)
        except Exception as e:
            logging.error(f"Exception occurred: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    else:
        logging.error(f"Invalid request method: {request.method}")
        return JsonResponse({'error': 'Invalid request method'}, status=405)





@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ManageLeaveType(request):
    if request.method == 'GET':
        try:
            company_id = request.session.get('c_id')

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            leavetypes = Leavetype.objects.filter(company_id=company_id)
            leavetypes_list = []

            for leavetype in leavetypes:
                leavetype_dict = {
                    'id': leavetype.id,
                    'LeaveType': leavetype.LeaveType,
                    'Description': leavetype.Description,
                    'Limit': leavetype.Limit,
                    'CreationDate': leavetype.CreationDate.strftime('%Y-%m-%d %H:%M:%S')
                }
                leavetypes_list.append(leavetype_dict)

            return JsonResponse(leavetypes_list, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            leave_id = data.get('id')
            
            if not leave_id:
                return JsonResponse({'error': 'Leave ID is required for updating'}, status=400)
            
            leavetype = Leavetype.objects.get(id=leave_id, company_id=request.session.get('c_id'))
            leavetype.LeaveType = data.get('LeaveType', leavetype.LeaveType)
            leavetype.Description = data.get('Description', leavetype.Description)
            leavetype.Limit = data.get('Limit', leavetype.Limit)
            leavetype.save()

            return JsonResponse({'message': 'Leave Type updated successfully'}, status=200)
        except Leavetype.DoesNotExist:
            return JsonResponse({'error': 'Leave Type not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'DELETE':
        try:
            leave_id = request.GET.get('id')
            if leave_id:
                leavetype = Leavetype.objects.get(id=leave_id, company_id=request.session.get('c_id'))
                leavetype.delete()
                return JsonResponse({'message': 'Leave Type deleted successfully'})
            else:
                return JsonResponse({'error': 'Leave ID is required for deletion'}, status=400)
        except Leavetype.DoesNotExist:
            return JsonResponse({'error': 'Leave Type not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EditLeaveType(request, id):
    if request.method == 'GET':
        try:
            leavetype = Leavetype.objects.get(id=id)

            company_id = request.session.get('c_id')
            if not company_id or leavetype.company_id != company_id:
                return JsonResponse({'error': 'Leave Type not found'}, status=404)

            leavetype_data = {
                'LeaveType': leavetype.LeaveType,
                'Description': leavetype.Description,
                'Limit': leavetype.Limit,
            }

            return JsonResponse(leavetype_data, safe=False)
        except Leavetype.DoesNotExist:
            return JsonResponse({'error': 'Leave Type not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            leavetype = Leavetype.objects.get(id=id)

            company_id = request.session.get('c_id')
            if not company_id or leavetype.company_id != company_id:
                return JsonResponse({'error': 'Leave Type not found'}, status=404)

            if 'LeaveType' in data:
                leavetype.LeaveType = data['LeaveType']
            if 'Description' in data:
                leavetype.Description = data['Description']
            if 'Limit' in data:
                leavetype.Limit = data['Limit']

            leavetype.save()

            return JsonResponse({'message': 'Leave Type updated successfully'}, status=200)
        except Leavetype.DoesNotExist:
            return JsonResponse({'error': 'Leave Type not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def Leaves(request):
    if request.method == 'GET':
        try:
            company_id = request.session.get('c_id')

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            departments = Department.objects.filter(c_id=company_id)
            dptcount = departments.count()
            employees = Employee.objects.filter(d_id__in=departments).values('emp_emailid')
            empcount = employees.count()

            leavtypcount = Leavetype.objects.count()

            emp_emails = [e['emp_emailid'] for e in employees]
            leaves_fetched = Tblleaves.objects.filter(emp_emailid__in=emp_emails).order_by('-id')

            leaves_data = []
            for leave in leaves_fetched:
                leave_dict = {
                    'lid': leave.id,
                    'emp_name': leave.emp_emailid.emp_name,
                    'emp_emailid': leave.emp_emailid.emp_emailid,
                    'LeaveType': leave.LeaveType_id,
                    'PostingDate': leave.PostingDate.strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': leave.Status,
                }
                leaves_data.append(leave_dict)

            response_data = {
                'empcount': empcount,
                'dptcount': dptcount,
                'leavtypcount': leavtypcount,
                'leaves_fetched': leaves_data,
            }
            return JsonResponse(response_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def PendingLeaves(request):
    if request.method == 'GET':
        try:
            company_id = request.session.get('c_id')

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            departments = Department.objects.filter(c_id=company_id)
            employees = Employee.objects.filter(d_id__in=departments).values('emp_emailid')

            emp_emails = [e['emp_emailid'] for e in employees]
            leaves_fetched = Tblleaves.objects.filter(emp_emailid__in=emp_emails, Status=0).order_by('-id')

            leaves_data = []
            for leave in leaves_fetched:
                leave_dict = {
                    'lid': leave.id,
                    'emp_name': leave.emp_emailid.emp_name,
                    'emp_emailid': leave.emp_emailid.emp_emailid,
                    'LeaveType': leave.LeaveType_id,
                    'PostingDate': leave.PostingDate.strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': leave.Status,
                }
                leaves_data.append(leave_dict)

            response_data = {
                'leaves_fetched': leaves_data
            }
            return JsonResponse(response_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ApprovedLeaves(request):
    if request.method == 'GET':
        try:
            company_id = request.session.get('c_id')

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            departments = Department.objects.filter(c_id=company_id)
            employees = Employee.objects.filter(d_id__in=departments).values('emp_emailid')

            emp_emails = [e['emp_emailid'] for e in employees]
            leaves_fetched = Tblleaves.objects.filter(emp_emailid__in=emp_emails, Status=1).order_by('-id')

            leaves_data = []
            for leave in leaves_fetched:
                leave_dict = {
                    'lid': leave.id,
                    'emp_name': leave.emp_emailid.emp_name,
                    'emp_emailid': leave.emp_emailid.emp_emailid,
                    'LeaveType': leave.LeaveType_id,
                    'PostingDate': leave.PostingDate.strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': leave.Status,
                }
                leaves_data.append(leave_dict)

            response_data = {
                'leaves_fetched': leaves_data
            }
            return JsonResponse(response_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def RejectedLeaves(request):
    if request.method == 'GET':
        try:
            company_id = request.session.get('c_id')

            if not company_id:
                return JsonResponse({'error': 'Company ID not found in session'}, status=401)

            departments = Department.objects.filter(c_id=company_id)
            employees = Employee.objects.filter(d_id__in=departments).values('emp_emailid')

            emp_emails = [e['emp_emailid'] for e in employees]
            leaves_fetched = Tblleaves.objects.filter(emp_emailid__in=emp_emails, Status=2).order_by('-id')

            leaves_data = []
            for leave in leaves_fetched:
                leave_dict = {
                    'lid': leave.id,
                    'emp_name': leave.emp_emailid.emp_name,
                    'emp_emailid': leave.emp_emailid.emp_emailid,
                    'LeaveType': leave.LeaveType_id,
                    'PostingDate': leave.PostingDate.strftime('%Y-%m-%d %H:%M:%S'),
                    'Status': leave.Status,
                }
                leaves_data.append(leave_dict)

            response_data = {
                'leaves_fetched': leaves_data
            }
            return JsonResponse(response_data, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


# Resignation Management
@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AllResignation(request):
    emp_emailid = request.session.get('emp_emailid')
    c_id = request.session.get('c_id')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            resignations = Resignation.objects.filter(emp_emailid__d_id__c_id=c_id).select_related('emp_emailid', 'emp_emailid__d_id', 'job_info')

            resignations_data = []

            for resignation in resignations:
                employee = resignation.emp_emailid
                job_info = resignation.emp_emailid.job_info
                if job_info:
                    interval = (job_info.start_date - resignation.submit_date.date()).days
                else:
                    interval = None

                resignation_data = {
                    'emp_profile': employee.emp_profile.url,
                    'emp_name': employee.emp_name,
                    'emp_role': employee.emp_role,
                    'emp_emailid': employee.emp_emailid,
                    'department_name': employee.d_id.d_name,
                    'job_title': job_info.job_title if job_info else '',
                    'start_date': job_info.start_date if job_info else None,
                    'leave_date': resignation.leave_date,
                    'exp_leave': resignation.exp_leave,
                    'shortfall_date': resignation.shortfall_date,
                    'days_since_resignation': interval,
                }

                resignations_data.append(resignation_data)

            return JsonResponse({'resignations': resignations_data}, status=200)

        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AllExitClearance(request):
    emp_emailid = request.session.get('emp_emailid')
    c_id = request.session.get('c_id')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email ID not found in session'}, status=401)

    try:
        employees_in_company = Employee.objects.filter(d_id__c_id=c_id).values_list('emp_emailid', flat=True)

        resignations = Resignation.objects.filter(
            emp_emailid__in=employees_in_company,
            status='Approved',
            emp_emailid__d_id__c_id=c_id
        ).select_related('emp_emailid', 'emp_emailid__d_id', 'clearance')

        clearances_data = []

        for resignation in resignations:
            employee = resignation.emp_emailid
            clearance = resignation.clearance
            try:
                job_info = Job_info.objects.get(emp_emailid=employee.emp_emailid)
            except Job_info.DoesNotExist:
                job_info = None

            today = timezone.now().date()
            interval = (today - resignation.submit_date.date()).days

            clearance_status = clearance.status if clearance else 'Pending'

            clearance_data = {
                'emp_profile': employee.emp_profile.url,
                'emp_name': employee.emp_name,
                'emp_role': employee.emp_role,
                'emp_emailid': employee.emp_emailid,
                'department_name': employee.d_id.d_name,
                'job_title': job_info.job_title if job_info else '',
                'days_since_resignation': interval,
                'clearance_status': clearance_status,
            }

            clearances_data.append(clearance_data)

        return JsonResponse({'clearances': clearances_data}, status=200)

    except Exception as e:
        return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AllFinalSettlement(request):
    emp_emailid = request.session.get('emp_emailid')
    c_id = request.session.get('c_id')

    if not emp_emailid:
        return JsonResponse({'error': 'Employee email ID not found in session'}, status=401)

    try:
        employees_in_company = Employee.objects.filter(d_id__c_id=c_id).values_list('emp_emailid', flat=True)

        resignations = Resignation.objects.filter(
            emp_emailid__in=employees_in_company,
            clearance__status='Approved',
            emp_emailid__d_id__c_id=c_id
        ).select_related('emp_emailid', 'emp_emailid__d_id', 'clearance')

        settlements_data = []

        for resignation in resignations:
            employee = resignation.emp_emailid
            clearance = resignation.clearance
            try:
                job_info = Job_info.objects.get(emp_emailid=employee.emp_emailid)
            except Job_info.DoesNotExist:
                job_info = None

            today = timezone.now().date()
            interval = (today - resignation.submit_date.date()).days

            clearance_status = clearance.status if clearance else 'Pending'

            settlement_data = {
                'emp_profile': employee.emp_profile.url,
                'emp_name': employee.emp_name,
                'emp_role': employee.emp_role,
                'emp_emailid': employee.emp_emailid,
                'department_name': employee.d_id.d_name,
                'job_title': job_info.job_title if job_info else '',
                'days_since_resignation': interval,
                'clearance_status': clearance_status,
            }

            settlements_data.append(settlement_data)

        return JsonResponse({'settlements': settlements_data}, status=200)

    except Exception as e:
        return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EditResignation(request):
    c_id = request.session.get('c_id')

    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        departments = Department.objects.filter(c_id=c_id)
        employees = Employee.objects.filter(d_id__in=departments)

        resignations = Resignation.objects.select_related('emp_emailid').filter(emp_emailid__d_id__in=departments)

        resignation_data = []
        for resignation in resignations:
            job_info = Job_info.objects.get(emp_emailid=resignation.emp_emailid.emp_emailid)
            interval = (timezone.now().date() - job_info.start_date).days
            length_of_service = '{} years {} months'.format(interval // 365, (interval % 365) // 30)
            resignation_data.append({
                'emp_name': resignation.emp_emailid.emp_name,
                'emp_profile': resignation.emp_emailid.emp_profile,
                'emp_emailid': resignation.emp_emailid.emp_emailid,
                'emp_role': resignation.emp_emailid.emp_role,
                'status': resignation.status,
                'start_date': job_info.start_date,
                'leave_date': resignation.leave_date,
                'exp_leave': resignation.exp_leave,
                'length_of_service': length_of_service,
                'shortfall_date': resignation.shortfall_date,
            })

        return JsonResponse(resignation_data, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            emp_emailid = data.get('emp_emailid')
            leave_reason = data.get('leave_reason')
            exp_leave = data.get('exp_leave')
            notice_per = data.get('notice_per')
            status = data.get('status')

            if not all([emp_emailid, leave_reason, exp_leave, notice_per, status]):
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            resignation, created = Resignation.objects.update_or_create(
                emp_emailid_id=emp_emailid,
                defaults={
                    'leave_reason': leave_reason,
                    'exp_leave': exp_leave,
                    'notice_per': notice_per,
                    'status': status
                }
            )

            response = {'message': 'Resignation created successfully'} if created else {'message': 'Resignation updated successfully'}
            return JsonResponse(response, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EditExitClearnace(request):
    c_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')
    given_to_email = request.GET.get('given_to_email')

    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            clearance = Clearance.objects.get(given_to=given_to_email)

            clearance_data = {
                'Accounts': clearance.Accounts,
                'Hr': clearance.Hr,
                'Hr_Plant': clearance.Hr_Plant,
                'IT': clearance.IT,
                'Project': clearance.Project,
                'status': clearance.status,
                'given_by': clearance.given_by.emp_emailid if clearance.given_by else None,
                'given_to': clearance.emp_emailid.emp_emailid
            }

            return JsonResponse(clearance_data, status=200)

        except Clearance.DoesNotExist:
            return JsonResponse({'error': 'Clearance not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            clearance, created = Clearance.objects.get_or_create(emp_emailid=given_to_email)

            if 'Accounts' in data:
                clearance.Accounts = data['Accounts']
            if 'Hr' in data:
                clearance.Hr = data['Hr']
            if 'Hr_Plant' in data:
                clearance.Hr_Plant = data['Hr_Plant']
            if 'IT' in data:
                clearance.IT = data['IT']
            if 'Project' in data:
                clearance.Project = data['Project']

            clearance.given_to = given_to_email
            clearance.given_by = emp_emailid

            clearance.save()

            return JsonResponse({'success': 'Clearance updated successfully'}, status=200)

        except Clearance.DoesNotExist:
            return JsonResponse({'error': 'Clearance not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def GenerateFnf(request):
    c_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
    personal_details = get_object_or_404(Personal_details, mail=emp_emailid)
    job_info = get_object_or_404(Job_info, emp_emailid=emp_emailid)
    resignation = get_object_or_404(Resignation, emp_emailid=emp_emailid)
    bank_details = get_object_or_404(Bank_details, emp_emailid=emp_emailid)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="full_and_final_settlement_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    data = [
        ["Name:", employee.emp_name, "Contact No.:", employee.emp_phone],
        ["Email Id:", employee.emp_emailid, "Role:", employee.emp_role],
        ["Gender:", personal_details.gender, "Birth Date:", personal_details.birth_date],
        ["Department:", job_info.department, "Joining Date:", job_info.start_date],
        ["Resigned On:", resignation.submit_date, "Leaving Date:", resignation.leave_date],
        ["Bank Name:", bank_details.bank_name, "Account No.:", bank_details.acc_no],
        ["Account Type:", bank_details.acc_type, "IFSC:", bank_details.ifsc],
        ["Pan no:", bank_details.Pan_no]
    ]

    table = Table(data)
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                               ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                               ('FONTNAME', (0, 0), (-1, 0), 'Courier-Bold'),
                               ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                               ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                               ('GRID', (0, 0), (-1, -1), 1, colors.black)]))

    elements.append(table)
    doc.build(elements)

    return response


import logging
logger = logging.getLogger(__name__)

@csrf_exempt
def DisplayTraining(request):
    c_id = request.session.get('c_id')
    emp_emailid = request.session.get('emp_emailid')

    logger.info(f'c_id from session: {c_id}, emp_emailid from session: {emp_emailid}')

    if not c_id or not emp_emailid:
        return JsonResponse({'error': 'Company ID or Employee Email not found in session'}, status=401)

    if request.method == 'GET':
        courses = Courses.objects.filter(
            c_id=c_id,
            course_employee__emp_emailid=emp_emailid
        ).distinct().values('course_id', 'course_title', 'thumbnail', 'description')

        return JsonResponse(list(courses), safe=False, status=200)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


# Pages section
@csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
def CreateCase(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')
    print(f"User ID: {user_id}, Company ID: {company_id}")
    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            create_for = data.get('createFor')
            case_type = data.get('caseType')
            case_title = data.get('caseTitle')
            case_desc = data.get('detailedDescription')

            if not all([create_for, case_type, case_title, case_desc]):
                return HttpResponseBadRequest("Missing required fields")

            employee = get_object_or_404(Employee, emp_emailid=user_id)

            new_case = Case(
                create_for=create_for,
                case_type=case_type,
                case_title=case_title,
                case_desc=case_desc,
                created_by=employee,
                case_status='New'
            )

            new_case.save()

            return JsonResponse({"message": "Case created successfully"}, status=201)
        except json.JSONDecodeError:
            return HttpResponseBadRequest("Invalid JSON")
        except Employee.DoesNotExist:
            return HttpResponseBadRequest("Employee not found")
    else:
        return HttpResponseBadRequest("Only POST requests are allowed")


@csrf_exempt
def MyCases(request):
    def TimeAgo(ptime):
        current_time = timezone.now()
        estimate_time = current_time - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(Q(created_by__emp_emailid=user_id) | Q(assigned_to__emp_emailid=user_id))

    case_data = []
    for case in cases:
        timeago = TimeAgo(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
        })

    return JsonResponse({'cases': case_data})


# Need to add signals ( triggers) to the employee table
@csrf_exempt
def UpdatePersonalDetails(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    # print(f"Logged in user email: {emp_emailid}")  # Debugging

    # Fetch the Employee instance using the logged-in user's email
    try:
        employee = Employee.objects.get(emp_emailid=emp_emailid)
    except Employee.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)

    # print(f"Logged in user email: {emp_emailid}")  # Debugging

    # Fetch the Employee instance using the logged-in user's email
    try:
        employee = Employee.objects.get(emp_emailid=emp_emailid)
    except Employee.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)

    if request.method == 'GET':
        try:
            # personal_detail = Personal_details.objects.get(mail=employee)
            personal_detail = Personal_details.objects.get(mail=employee)
            data = {
                'first_name': personal_detail.first_name,
                'last_name': personal_detail.last_name,
                'address': personal_detail.address,
                'state': personal_detail.state,
                'city': personal_detail.city,
                'district': personal_detail.district,
                'post_code': personal_detail.post_code,
                'Contact': personal_detail.Contact,
                'birth_date': personal_detail.birth_date,
                'gender': personal_detail.gender,
                'emergency_name': personal_detail.emergency_name,
                'emergency_contact': personal_detail.emergency_contact,
                'emp_emailid': emp_emailid,
            }
            return JsonResponse({'status': 'success', 'data': data})
        except Personal_details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)

            # Extract data from request body
            # Extract data from request body
            first_name = data.get('first_name')
            last_name = data.get('last_name')
            address = data.get('address')
            state = data.get('state')
            city = data.get('city')
            district = data.get('district')
            post_code = data.get('post_code')
            Contact = data.get('Contact')
            birth_date = data.get('birth_date')
            gender = data.get('gender')
            emergency_name = data.get('emergency_name')
            emergency_contact = data.get('emergency_contact')

            # Use get_or_create to find or create a new Personal_details entry
            personal_detail, created = Personal_details.objects.get_or_create(mail=employee)

            # Update or set details
            personal_detail.first_name = first_name
            personal_detail.last_name = last_name
            personal_detail.address = address
            personal_detail.state = state
            personal_detail.city = city
            personal_detail.district = district
            personal_detail.post_code = post_code
            personal_detail.Contact = Contact
            personal_detail.birth_date = birth_date
            personal_detail.gender = gender
            personal_detail.emergency_name = emergency_name
            personal_detail.emergency_contact = emergency_contact
            personal_detail.save()

            # Send appropriate response based on whether the record was created or updated
            if created:
                return JsonResponse({'status': 'success', 'message': 'Personal details created successfully.'}, status=201)
            else:
                return JsonResponse({'status': 'success', 'message': 'Personal details updated successfully.'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)




@csrf_exempt
def UpdateJobDetails(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':
        try:
            # Try to retrieve the job_info for the employee's email
            job_info = Job_info.objects.get(emp_emailid=emp_emailid)

            data = {
                'job_title': job_info.job_title,
                'department': job_info.department,
                'working_type': job_info.working_type,
                'start_date': job_info.start_date,
            }

            return JsonResponse({'status': 'success', 'data': data})

        except Job_info.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Job information not found. You can add your own details.'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    elif request.method == 'PUT':
        try:
            # Parse JSON data from the request body
            data = json.loads(request.body.decode('utf-8'))

            # Collect data from the parsed JSON
            job_title = data.get('job_title')
            department = data.get('department')
            working_type = data.get('working_type')
            start_date = data.get('start_date')

            # Validate the data
            if not all([job_title, department, working_type, start_date]):
                return JsonResponse({'status': 'error', 'message': 'All fields are required'}, status=400)

            # Retrieve the employee instance using the email
            employee = Employee.objects.get(emp_emailid=emp_emailid)

            # Get or create a new Job_info object for the employee
            job_info, created = Job_info.objects.get_or_create(emp_emailid=employee)
            
            # Update the job_info fields with the new data
            job_info.job_title = job_title
            job_info.department = department
            job_info.working_type = working_type
            job_info.start_date = start_date

            # Save the updated or newly created job_info object
            job_info.save()

            # Respond with appropriate message based on whether a new job_info was created or updated
            if created:
                return JsonResponse({'status': 'success', 'message': 'Job information created successfully'})
            else:
                return JsonResponse({'status': 'success', 'message': 'Job information updated successfully'})

        except Employee.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)



@csrf_exempt
def UpdateBankDetails(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':
        try:
            bank_details = Bank_details.objects.get(emp_emailid=emp_emailid)
            data = {
                'holder_name': bank_details.holder_name,
                'bank_name': bank_details.bank_name,
                'acc_no': bank_details.acc_no,
                'branch': bank_details.branch,
                'acc_type': bank_details.acc_type,
                'ifsc': bank_details.ifsc,
                'Pan_no': bank_details.Pan_no,
            }
            return JsonResponse({'status': 'success', 'data': data})

        
        except Bank_details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Bank details not found'}, status=404)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            holder_name = data.get('holder_name')
            bank_name = data.get('bank_name')
            acc_no = data.get('acc_no')
            branch = data.get('branch')
            acc_type = data.get('acc_type')
            ifsc = data.get('ifsc')
            Pan_no = data.get('Pan_no')

            if not all([holder_name, bank_name, acc_no, branch, acc_type, ifsc, Pan_no, emp_emailid]):
                return JsonResponse({'status': 'error', 'message': 'Missing fields'}, status=400)
            # Fetch the Employee instance based on the email ID
            try:
                employee = Employee.objects.get(emp_emailid=emp_emailid)

                # Check if bank details already exist and update them
                try:
                    bank_details = Bank_details.objects.get(emp_emailid=employee)
                    bank_details.holder_name = holder_name
                    bank_details.bank_name = bank_name
                    bank_details.acc_no = acc_no
                    bank_details.branch = branch
                    bank_details.acc_type = acc_type
                    bank_details.ifsc = ifsc
                    bank_details.Pan_no = Pan_no
                    bank_details.save()

                    return JsonResponse({'status': 'success', 'message': 'Bank details updated successfully'})

                except Bank_details.DoesNotExist:
                    # Create new bank details if not found
                    Bank_details.objects.create(
                        emp_emailid=employee,
                        holder_name=holder_name,
                        bank_name=bank_name,
                        acc_no=acc_no,
                        branch=branch,
                        acc_type=acc_type,
                        ifsc=ifsc,
                        Pan_no=Pan_no
                    )
                    return JsonResponse({'status': 'success', 'message': 'Bank details created successfully'})

            except Employee.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)



@csrf_exempt
def UpdateWorkExperience(request):
    # print("View is being accessed")
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    # Fetch the Employee instance using the emp_emailid from the session
    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

    if request.method == 'GET':
        # Fetch only the logged-in user's work experience
        work_experiences = Work_exp.objects.filter(emp_emailid=employee)

        work_exp_list = []
        for work_exp in work_experiences:
            work_exp_list.append({
                'W_Id': work_exp.W_Id,
                'start_date': work_exp.start_date,
                'end_date': work_exp.end_date,
                'comp_name': work_exp.comp_name,
                'comp_location': work_exp.comp_location,
                'designation': work_exp.designation,
                'gross_salary': work_exp.gross_salary,
                'leave_reason': work_exp.leave_reason
            })

        return JsonResponse(work_exp_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Fetch the Employee instance using the emp_emailid from the session
            emp_emailid = request.session.get('emp_emailid')
            if not emp_emailid:
                return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

            employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

            # Check if work experience with the same details already exists for the logged-in employee
            existing_work_exp = Work_exp.objects.filter(
                emp_emailid=employee,
                start_date=data['start_date'],
                end_date=data['end_date'],
                comp_name=data['comp_name'],
                comp_location=data['comp_location'],
                designation=data['designation'],
                gross_salary=data['gross_salary'],
                leave_reason=data['leave_reason']
            ).exists()

            if existing_work_exp:
                return JsonResponse({'status': 'error', 'message': 'Work experience with these details already exists'}, status=400)

            # If W_Id is provided, try to update the existing work experience
            if 'W_Id' in data:
                work_exp = Work_exp.objects.filter(W_Id=data['W_Id'], emp_emailid=employee).first()
                if not work_exp:
                    return JsonResponse({'status': 'error', 'message': 'Work experience not found for the provided W_Id'}, status=404)

                # Update the existing work experience
                work_exp.start_date = data['start_date']
                work_exp.end_date = data['end_date']
                work_exp.comp_name = data['comp_name']
                work_exp.comp_location = data['comp_location']
                work_exp.designation = data['designation']
                work_exp.gross_salary = data['gross_salary']
                work_exp.leave_reason = data['leave_reason']
                work_exp.save()

                return JsonResponse({'message': 'Work experience updated successfully', 'W_Id': work_exp.W_Id})

            else:
                # If W_Id is not provided, create a new work experience entry
                new_work_exp = Work_exp(
                    emp_emailid=employee,
                    start_date=data['start_date'],
                    end_date=data['end_date'],
                    comp_name=data['comp_name'],
                    comp_location=data['comp_location'],
                    designation=data['designation'],
                    gross_salary=data['gross_salary'],
                    leave_reason=data['leave_reason']
                )
                new_work_exp.save()

                return JsonResponse({'message': 'Work experience created successfully', 'W_Id': new_work_exp.W_Id})

        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    
    elif request.method == 'PUT':
        try:
            # Parse the incoming JSON data
            data = json.loads(request.body)

            # Retrieve the email and work experience ID
            email_to_delete = data.get('emp_emailid')  # The email ID of the employee whose record is to be deleted
            work_exp_id = data.get('W_Id')  # The ID of the specific work experience to delete

            # Check if the user's role has permission to delete
            if employee.emp_role not in ['Manager', 'Super Admin', 'HR','Employee']:
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

            # Ensure emp_emailid is provided
            if not email_to_delete:
                return JsonResponse({'status': 'error', 'message': 'emp_emailid is required'}, status=400)

            # Ensure W_Id is provided
            if not work_exp_id:
                return JsonResponse({'status': 'error', 'message': 'W_Id is required'}, status=400)

            # Fetch the specific work experience for the given email and W_Id
            work_experience = get_object_or_404(Work_exp, W_Id=work_exp_id, emp_emailid=email_to_delete)

            # Delete the specific work experience
            work_experience.delete()

            return JsonResponse({'message': 'Work experience deleted successfully'})

        except KeyError as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")
        except Work_exp.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Work experience not found'}, status=404)

    else:
        return HttpResponseNotAllowed(['GET', 'POST', 'PUT'])



@csrf_exempt
def UpdateDependent(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

    if request.method == 'GET':
        dependents = Dependent.objects.filter(emp_emailid=emp_emailid)

        dependent_list = []
        for dep in dependents:
            dependent_list.append({
                'D_name': dep.D_name,
                'D_gender': dep.D_gender,
                'D_dob': dep.D_dob,
                'D_relation': dep.D_relation,
                'D_desc': dep.D_desc,
                'D_Id': dep.D_Id,
            })

        return JsonResponse(dependent_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
        
            # Fetch the Employee instance
            employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
        
        # Create and save the new dependent
            dependent = Dependent(
                emp_emailid=employee,  # Use the Employee instance, not the email string
                D_name=data['D_name'],
                D_gender=data['D_gender'],
                D_dob=data['D_dob'],
                D_relation=data['D_relation'],
                D_desc=data['D_desc']
            )
            dependent.save()

            return JsonResponse({'message': 'Dependent added successfully', 'D_Id': dependent.D_Id})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")


    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            D_Id = data.get('D_Id')  # Get D_Id from the request body instead of query params
            if not D_Id:
                return JsonResponse({'status': 'error', 'message': 'D_Id is required'}, status=400)
            

            dependent = get_object_or_404(Dependent, D_Id=D_Id, emp_emailid=emp_emailid)
            employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
            
            # Update the dependent's information
            dependent.D_name = data['D_name']
            dependent.D_gender = data['D_gender']
            dependent.D_dob = data['D_dob']
            dependent.D_relation = data['D_relation']
            dependent.D_desc = data['D_desc']
            dependent.save()

            return JsonResponse({'message': 'Dependent updated successfully'})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    elif request.method == 'DELETE':
        try:
            D_Id = request.GET.get('D_Id')
            dependent = get_object_or_404(Dependent, D_Id = D_Id, emp_emailid=emp_emailid)

            if employee.emp_role not in ['Manager', 'Super Admin','Employee', 'HR']:
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
            
            dependent.delete()

            return JsonResponse({'message': 'Dependent deleted successfully'})
        except KeyError as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    else:
        return HttpResponseNotAllowed(['GET', 'POST', 'PUT', 'DELETE'])


@csrf_exempt
def UpdateAdhaar(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':

        adhaar = Adhaar.objects.filter(emp_emailid=emp_emailid).first()
        if adhaar:
            data = {
                'adhaar_no': adhaar.adhaar_no,
                'adhaar_name': adhaar.adhaar_name,
                'enroll_no': adhaar.enroll_no,
                'adhaar_pic': adhaar.adhaar_pic.url if adhaar.adhaar_pic else ''
            }
            return JsonResponse(data)
        else:
            return JsonResponse({'error': 'Adhaar details not found!'}, status=404)

    elif request.method == 'POST':
        data = request.POST
        file = request.FILES.get('adhaar_pic')

        employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
        adhaar, created = Adhaar.objects.get_or_create(
            emp_emailid=employee,
            defaults={
                'adhaar_no': data['adhaar_no'],
                'adhaar_name': data['adhaar_name'],
                'enroll_no': data['enroll_no'],
                'adhaar_pic': file
            }
        )
        if created:
            return JsonResponse({'message': 'Adhaar details created successfully!'}, status=201)
        else:
            adhaar.adhaar_no = data['adhaar_no']
            adhaar.adhaar_name = data['adhaar_name']
            adhaar.enroll_no = data['enroll_no']
            if file:
                adhaar.adhaar_pic = file
            adhaar.save()
            return JsonResponse({'message': 'Adhaar details updated successfully!'})

    elif request.method == 'DELETE':
        adhaar = Adhaar.objects.filter(emp_emailid=emp_emailid).first()
        if adhaar:
            adhaar.delete()
            return JsonResponse({'message': 'Adhaar details deleted successfully!'})
        else:
            return JsonResponse({'error': 'Adhaar details not found!'}, status=404)

    else:
        return JsonResponse({'error': 'Method not allowed!'}, status=405)


@csrf_exempt
def UpdateLicence(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

    if request.method == 'GET':
        licence = Licence.objects.filter(emp_emailid=employee).first()
        if licence:
            data = {
                'licence_no': licence.licence_no,
                'licence_name': licence.licence_name,
                'expiry_date': licence.expiry_date,
                'licence_pic': licence.licence_pic.url if licence.licence_pic else None,
            }
            return JsonResponse(data)
        else:
            return JsonResponse({'message': 'No licence found'}, status=404)

    elif request.method == 'POST':
        data = request.POST
        file = request.FILES.get('licence_pic')
        licence_record, created = Licence.objects.update_or_create(
            emp_emailid=employee,
            defaults={
                'licence_no': data['licence_no'],
                'licence_name': data['licence_name'],
                'expiry_date': data['expiry_date'],
                'licence_pic': file
            }
        )
        if created:
            return JsonResponse({'message': 'Licence details created successfully!'}, status=201)
        else:
            # Update the licence details if already exists
            licence_record.licence_no = data['licence_no']
            licence_record.licence_name = data['licence_name']
            licence_record.expiry_date = data['expiry_date']
            if file:
                licence_record.licence_pic = file
            licence_record.save()
            return JsonResponse({'message': 'Licence details updated successfully!'})

    elif request.method == 'DELETE':
        licence = get_object_or_404(Licence, emp_emailid=employee)
        if licence.licence_pic:
            licence.licence_pic.delete(save=False)
        licence.delete()
        return JsonResponse({'message': 'Licence details deleted successfully!'})

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def UpdatePassport(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':
        passport = Passport.objects.filter(emp_emailid=emp_emailid).first()
        if passport:
            data = {
                'passport_no': passport.passport_no,
                'passport_name': passport.passport_name,
                'passport_validity': passport.passport_validity,
                'passport_pic': passport.passport_pic.url if passport.passport_pic else ''
            }
            return JsonResponse(data)
        else:
            return JsonResponse({'error': 'Passport details not found!'}, status=404)

    elif request.method == 'POST':
        data = request.POST
        file = request.FILES.get('passport_pic')

        employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
        passport, created = Passport.objects.get_or_create(
            emp_emailid=employee,
            defaults={
                'passport_no': data['passport_no'],
                'passport_name': data['passport_name'],
                'passport_validity': data['passport_validity'],
                'passport_pic': file
            }
        )
        if created:
            return JsonResponse({'message': 'Passport details created successfully!'}, status=201)
        else:
            passport.passport_no = data['passport_no']
            passport.passport_name = data['passport_name']
            passport.passport_validity = data['passport_validity']
            if file:
                passport.passport_pic = file
            passport.save()
            return JsonResponse({'message': 'Passport details updated successfully!'})

    elif request.method == 'DELETE':
        passport = Passport.objects.filter(emp_emailid=emp_emailid).first()
        if passport:
            passport.delete()
            return JsonResponse({'message': 'Passport details deleted successfully!'})
        else:
            return JsonResponse({'error': 'Passport details not found!'}, status=404)

    else:
        return JsonResponse({'error': 'Method not allowed!'}, status=405)


@csrf_exempt
def UpdatePan(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':
        pan = Pan.objects.filter(emp_emailid=emp_emailid).first()
        if pan:
            data = {
                'pan_no': pan.pan_no,
                'pan_name': pan.pan_name,
                'pan_pic': pan.pan_pic.url if pan.pan_pic else ''
            }
            return JsonResponse(data)
        else:
            return JsonResponse({'error': 'PAN details not found!'}, status=404)

    elif request.method == 'POST':
        data = request.POST
        file = request.FILES.get('pan_pic')

        employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
        pan, created = Pan.objects.get_or_create(
            emp_emailid=employee,
            defaults={
                'pan_no': data['pan_no'],
                'pan_name': data['pan_name'],
                'pan_pic': file
            }
        )
        if created:
            return JsonResponse({'message': 'PAN details created successfully!'}, status=201)
        else:
            pan.pan_no = data['pan_no']
            pan.pan_name = data['pan_name']
            if file:
                pan.pan_pic = file
            pan.save()
            return JsonResponse({'message': 'PAN details updated successfully!'})

    elif request.method == 'DELETE':
        pan = Pan.objects.filter(emp_emailid=emp_emailid).first()
        if pan:
            pan.delete()
            return JsonResponse({'message': 'PAN details deleted successfully!'})
        else:
            return JsonResponse({'error': 'PAN details not found!'}, status=404)

    else:
        return JsonResponse({'error': 'Method not allowed!'}, status=405)


@csrf_exempt
def UpdateQualification(request):
    emp_emailid = request.session.get('emp_emailid')
    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)

    if request.method == 'GET':
        qualifications = Qualification.objects.filter(emp_emailid=emp_emailid)

        qualification_list = []
        for qual in qualifications:
            qualification_list.append({
                'Q_Id': qual.Q_Id,
                'q_type': qual.q_type,
                'q_degree': qual.q_degree,
                'q_clg': qual.q_clg,
                'q_uni': qual.q_uni,
                'q_duration': qual.q_duration,
                'q_yop': qual.q_yop,
                'q_comment': qual.q_comment,
            })

        return JsonResponse(qualification_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

            qualification = Qualification(
                emp_emailid = employee,
                q_type=data['q_type'],
                q_degree=data['q_degree'],
                q_clg=data['q_clg'],
                q_uni=data['q_uni'],
                q_duration=data['q_duration'],
                q_yop=data['q_yop'],
                q_comment=data['q_comment']
            )
            qualification.save()

            return JsonResponse({'message': 'Qualification added successfully', 'Q_Id': qualification.Q_Id})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            Q_Id = request.GET.get('Q_Id')
            qualification = Qualification.objects.get(Q_Id=Q_Id, emp_emailid=emp_emailid)
            emp_emailid=employee,
            qualification.q_type = data['q_type']
            qualification.q_degree = data['q_degree']
            qualification.q_clg = data['q_clg']
            qualification.q_uni = data['q_uni']
            qualification.q_duration = data['q_duration']
            qualification.q_yop = data['q_yop']
            qualification.q_comment = data['q_comment']
            qualification.save()

            return JsonResponse({'message': 'Qualification updated successfully'})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")
        except Qualification.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Qualification not found'}, status=404)

    elif request.method == 'DELETE':
        try:
            Q_Id = request.GET.get('Q_Id')
            qualification = Qualification.objects.get(Q_Id=Q_Id, emp_emailid=emp_emailid)

            employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

            if employee.emp_role not in ['Manager', 'Super Admin','Employee', 'HR']:
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
            
            qualification.delete()

            return JsonResponse({'message': 'Qualification deleted successfully'})
        except (KeyError, Qualification.DoesNotExist) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    else:
        return HttpResponseNotAllowed(['GET', 'POST', 'PUT', 'DELETE'])


@csrf_exempt
def UpdateFamilyDetails(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)
    

     # Fetch the Employee instance using emp_emailid
    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

    if request.method == 'GET':
        family_details = Family_details.objects.filter(emp_emailid=emp_emailid)

        family_list = []
        for fam in family_details:
            family_list.append({
                'F_Id': fam.F_Id,
                'F_name': fam.F_name,
                'F_gender': fam.F_gender,
                'F_dob': fam.F_dob,
                'F_contact': fam.F_contact,
                'F_mail': fam.F_mail,
                'F_relation': fam.F_relation,
                'F_comment': fam.F_comment,
            })

        return JsonResponse(family_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)

            family_detail = Family_details(
                emp_emailid=employee,
                F_name=data['F_name'],
                F_gender=data['F_gender'],
                F_dob=data['F_dob'],
                F_contact=data['F_contact'],
                F_mail=data['F_mail'],
                F_relation=data['F_relation'],
                F_comment=data['F_comment']
            )
            family_detail.save()

            return JsonResponse({'message': 'Family detail added successfully', 'F_Id': family_detail.F_Id})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            F_Id = request.GET.get('F_Id')
            family_detail = Family_details.objects.get(F_Id=F_Id, emp_emailid=emp_emailid)

            family_detail.F_name = data['F_name']
            family_detail.F_gender = data['F_gender']
            family_detail.F_dob = data['F_dob']
            family_detail.F_contact = data['F_contact']
            family_detail.F_mail = data['F_mail']
            family_detail.F_relation = data['F_relation']
            family_detail.F_comment = data['F_comment']
            family_detail.save()

            return JsonResponse({'message': 'Family detail updated successfully'})
        except (KeyError, ValueError) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")
        except Family_details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Family detail not found'}, status=404)

    elif request.method == 'DELETE':
        try:
            F_Id = request.GET.get('F_Id')
            family_detail = Family_details.objects.get(F_Id=F_Id, emp_emailid=emp_emailid)

            if employee.emp_role not in ['Manager', 'Super Admin','Employee', 'HR']:
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)


            family_detail.delete()

            return JsonResponse({'message': 'Family detail deleted successfully'})
        except (KeyError, Family_details.DoesNotExist) as e:
            return HttpResponseBadRequest(f"Invalid data: {e}")

    else:
        return HttpResponseNotAllowed(['GET', 'POST', 'PUT', 'DELETE'])


# Case Management section
@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AllCases(request):
    def TimeAgo(ptime):
        current_time = timezone.now()
        estimate_time = current_time - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(created_by__d_id__c_id=company_id)

    case_data = []
    for case in cases:
        timeago = TimeAgo(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by_profile': created_by.emp_profile.url
        })
    return JsonResponse({'cases': case_data})


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CaseInfo(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        case_id = request.GET.get('case_id')
        if not case_id:
            return JsonResponse({"error": "Case ID not provided"}, status=400)

        case = get_object_or_404(Case, case_id=case_id)
        # if case.created_by.d_id.c_id != company_id:
        #     print(case.created_by.d_id.c_id)
        #     print(company_id)
        #     return JsonResponse({"error": "Unauthorized access"}, status=403)


        case_data = {
            'create_for': case.create_for,
            'case_type': case.case_type,
            'case_title': case.case_title,
            'case_desc': case.case_desc,
            'case_date': case.case_date.strftime('%Y-%m-%d %H:%M:%S'),
            'case_status': case.case_status,
            'case_id': case.case_id,
            'created_by': {
                'emp_name': case.created_by.emp_name,
                'emp_emailid': case.created_by.emp_emailid,
                'emp_phone': case.created_by.emp_phone,
                'emp_profile': case.created_by.emp_profile.url,
            }
        }

        return JsonResponse({'case': case_data})

    elif request.method == 'POST':
        user_email = request.session.get('emp_emailid')
        case_id = request.GET.get('case_id')

        if not case_id or not user_email:
            return JsonResponse({"error": "Case ID or User Email not provided"}, status=400)

        case = get_object_or_404(Case, case_id=case_id)
        employee = get_object_or_404(Employee, emp_emailid=user_email)

        case.assigned_to = employee
        case.save()
        return JsonResponse({"message": "Case assigned successfully"}, status=200)

    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def BenefitsCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='Benefits'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url
        })

    context = {
        'cases': case_data
    }

    return JsonResponse({'cases': case_data})


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def TravelExpenseCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='Travel and Expense'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url
        })

    context = {
        'cases': case_data
    }

    return JsonResponse({'cases': case_data})


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CompensationPayrollCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='Compensation and payroll'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url
        })

    context = {
        'cases': case_data
    }

    return JsonResponse({'cases': case_data})


# incomplete pending uncomment , need to complete this (Employee Jobs)
@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ManagerRating(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        # Tasks
        tasks_data = Tasks.objects.filter(emp_emailid__d_id__c_id=c_id).values('emp_emailid', 'selfratings', 'status', 'tid__description' )
        print(tasks_data)
        tasks = list(tasks_data)
        return JsonResponse(tasks_data, safe=False)


@csrf_exempt
def FAQManagement(request):
    if request.method == 'GET':
        c_id = request.session.get('c_id')  # Retrieve company ID from session
        if not c_id:
            return JsonResponse({'error': 'User not authenticated or company ID not found'}, status=401)

        faqs = Faqs.objects.filter(c_id=c_id)  # Filter FAQs by the logged-in user's company ID

        if not faqs.exists():
            return JsonResponse({'message': 'No FAQs pending'})

        faqs_list = []
        for faq in faqs:
            faqs_list.append({
                'faq_id': faq.faq_id,
                'question': faq.question,
                'answer': faq.answer if faq.answer else None,
                'emp_emailid': faq.emp_emailid.emp_emailid,
                'imp': faq.imp,
                'c_id': faq.c_id.c_id,
            })

        return JsonResponse({'faqs': faqs_list})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            faq_id = data.get('faq_id')  # Get the FAQ ID from the POST data
            answer = data.get('answer')    # Get the answer from the POST data
            
            if not faq_id or answer is None:
                return JsonResponse({'error': 'FAQ ID and answer are required'}, status=400)

            # Update or create the FAQ with the provided FAQ ID and answer
            faq, created = Faqs.objects.update_or_create(
                faq_id=faq_id,
                defaults={'answer': answer}
            )
            return JsonResponse({'message': 'FAQ updated successfully' if not created else 'FAQ created successfully'})

        except Exception as e:
            return HttpResponseBadRequest({'error': str(e)})

    elif request.method == 'DELETE':
        faq_id = request.GET.get('faq_id')  # Get FAQ ID from the query parameters
        if not faq_id:
            return JsonResponse({'error': 'FAQ ID is required'}, status=400)

        try:
            faq = Faqs.objects.filter(faq_id=faq_id, c_id=request.session.get('c_id'))  # Ensure the user can only delete FAQs from their company
            if faq.exists():
                faq.delete()  # Delete the FAQ
                return JsonResponse({'message': 'FAQ deleted successfully.'}, status=204)  # No Content response
            return JsonResponse({'error': 'FAQ not found.'}, status=404)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)




@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EnrollEmployee(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        courses = Courses.objects.filter(c_id=c_id).values('course_id', 'course_title').distinct()

        employees = Employee.objects.filter(d_id__c_id=c_id).values('emp_emailid')

        return JsonResponse({'courses': list(courses), 'employees': list(employees)})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            course_id = data.get('course_id')
            course_title = data.get('course_title')
            emp_emailid = data.get('emp_emailid')

            if not course_title:
                return JsonResponse({'error': 'Course title cannot be empty'}, status=400)

            if not Employee.objects.filter(emp_emailid=emp_emailid, d_id__c_id=c_id).exists():
                return JsonResponse({'error': 'Employee does not belong to the same company'}, status=400)

            course_company_id = Courses.objects.get(course_id=course_id).c_id_id
            if course_company_id != c_id:
                return JsonResponse({'error': 'This course does not belong to your company'}, status=400)

            if Course_employee.objects.filter(course_id=course_id, emp_emailid=emp_emailid).exists():
                return JsonResponse({'error': 'Employee is already enrolled in this course'}, status=400)

            Course_employee.objects.create(
                course_id_id=course_id,
                emp_emailid_id=emp_emailid,
                status=0,
                course_title=course_title
            )

            return JsonResponse({'message': 'Employee enrolled successfully'}, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ViewAllEnrollments(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        enrollments = Course_employee.objects.filter(course_id__c_id=c_id).values('id', 'course_title', 'emp_emailid')

        return JsonResponse({'enrollments': list(enrollments)})

    elif request.method == 'DELETE':
        try:
            data = json.loads(request.body)
            id = data.get('id')

            enrollment = Course_employee.objects.get(id=id)
            enrollment.delete()
            return JsonResponse({'message': 'Enrollment deleted successfully'}, status=200)
        except Course_employee.DoesNotExist:
            return JsonResponse({'error': 'Enrollment not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AdhocPayments(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        allowances = Adhoc.objects.all().values('name', 'dept', 'type', 'amt', 'mon', 'year')
        allowances_list = list(allowances)
        return JsonResponse(allowances_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            adhoc_payment = Adhoc(
                name=data.get('name'),
                dept=data.get('dept'),
                type=data.get('type'),
                amt=data.get('amt'),
                mon=data.get('mon'),
                year=data.get('year')
            )
            adhoc_payment.save()

            return JsonResponse({'message': "Allowance or Deduction added successfully"}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except IntegrityError as e:
            return JsonResponse({"error": "Database integrity error: " + str(e)}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def LoanPayments(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        all_loans = list(Loan.objects.all().values('id', 'name', 'department', 'lamt', 'mamt', 'startdate', 'reason', 'status'))
        pending_loans = list(Loan.objects.filter(status=0).values('id', 'name', 'department', 'lamt', 'mamt', 'startdate', 'reason', 'status'))
        approved_loans = list(Loan.objects.filter(status=1).values('id', 'name', 'department', 'lamt', 'mamt', 'startdate', 'reason', 'status'))

        response_data = {
            'all_loans': all_loans,
            'pending_loans': pending_loans,
            'approved_loans': approved_loans
        }

        return JsonResponse(response_data, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            loan_id = data.get('id')
            if not loan_id:
                return JsonResponse({"error": "Loan ID not provided"}, status=400)

            loan = Loan.objects.get(id=loan_id)
            loan.status = 1
            loan.save()

            return JsonResponse({'message': "Loan approved successfully"}, status=200)
        except Loan.DoesNotExist:
            return JsonResponse({"error": "Loan not found"}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def LeaveEncashment(request):
#     user_id = request.session.get('user_id')
#     company_id = request.session.get('c_id')

#     if not user_id or not company_id:
#         return JsonResponse({'error': 'User not logged in'}, status=401)

#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             employee = Employee.objects.get(emp_emailid=user_id)

#             leave_encashment = Leave_Encashment(
#                 txndt = data.get('txndt'),
#                 refn = data.get('refn'),
#                 effdt = data.get('effdt'),
#                 emp = data.get('emp'),
#                 type = data.get('type'),
#                 blnc = data.get('blnc'),
#                 pdays = data.get('pdays'),
#                 sal = data.get('sal'),
#                 emp_emailid = employee
#             )
#             leave_encashment.save()

#             return JsonResponse({"message": "Leave inserted successfully"}, status=201)
#         except Employee.DoesNotExist:
#             return JsonResponse({"error": "Employee not found"}, status=404)
#         except Exception as e:
#             return JsonResponse({"error": str(e)}, status=400)
#     else:
#         return JsonResponse({"error": "Invalid request method"}, status=405)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def LeaveEncashment(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            employee = Employee.objects.get(emp_emailid=user_id)

            leave_encashment = Leave_Encashment(
                txndt=data.get('txndate'),  # Changed to match the frontend
                refn=data.get('referencenum'),
                effdt=data.get('effectivedate'),
                emp=data.get('empname'),
                type=data.get('leaveType'),
                blnc=data.get('leavesallowed'),
                pdays=data.get('totalleaves'),
                sal=data.get('salary'),
                emp_emailid=employee
            )
            leave_encashment.save()

            return JsonResponse({"message": "Leave inserted successfully"}, status=201)
        except Employee.DoesNotExist:
            return JsonResponse({"error": "Employee not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    elif request.method == 'GET':
        # Fetch the leave encashment data for the logged-in user
        try:
            employee = Employee.objects.get(emp_emailid=user_id)
            leave_encashments = Leave_Encashment.objects.filter(emp_emailid=employee)
            data = list(leave_encashments.values())
            return JsonResponse(data, safe=False)
        except Employee.DoesNotExist:
            return JsonResponse({"error": "Employee not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    
    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@role_required(['HR', 'Manager', 'Super Admin','Employee'])
@csrf_exempt
def ViewLeaveEncashment(request):
    company_id = request.session.get('c_id')

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            leave_encashments = Leave_Encashment.objects.filter(
                emp_emailid__d_id__c_id=company_id
            ).values(
                'emp',
                'emp_emailid__emp_name',
                'blnc',
                'pdays',
                'enclve'
            ).distinct('emp')
            
            employee_list = [
                {
                    'emp': item['emp'],
                    'name': item['emp'],
                    'totalLeaves': item['blnc'],
                    'leavesTaken': item['pdays'],
                    'encashment': item['enclve']
                } for item in leave_encashments
            ]
            
            return JsonResponse({'employees': employee_list}, safe=False)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def OffCyclePayments(request):
    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        off_cycle_payments = Off_cycle.objects.all().values()
        off_cycle_list = list(off_cycle_payments)
        return JsonResponse(off_cycle_list, safe=False)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            off_cycle_payment = Off_cycle(
                name=data.get('name'),
                tname=data.get('tname'),
                amt=data.get('amt'),
                effdt=data.get('effdt'),
                note=data.get('note')
            )
            off_cycle_payment.save()

            return JsonResponse({"message": "Off-cycle payment inserted successfully"}, status=201)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    elif request.method == 'DELETE':
        try:
            # Extract the ID from the URL
            payment_id = request.GET.get('id')

            if not payment_id:
                return JsonResponse({"error": "Payment ID is required"}, status=400)

            # Fetch the off-cycle payment by ID
            off_cycle_payment = Off_cycle.objects.get(id=payment_id)
            off_cycle_payment.delete()

            return JsonResponse({"message": "Off-cycle payment deleted successfully"}, status=200)
        except Off_cycle.DoesNotExist:
            return JsonResponse({"error": "Off-cycle payment not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)  
    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def BankTransferPayout(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        payouts = Salary.objects.filter(paymentmethod='bank').order_by('-payout_month').values('sal_id', 'payout_month').distinct()
        return JsonResponse({'payouts': list(payouts)})
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def BankTransfer(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        month = request.GET.get('month')
        if not month:
            return JsonResponse({'error': 'Month parameter is required'}, status=400)

    try:
        salaries = Salary.objects.filter(paymentmethod='bank', payout_month=month)
        total_employees = salaries.count()
        total_amount = salaries.filter(holdsalary=0, paid=0).aggregate(Sum('Net_Salary'))['Net_Salary__sum'] or 0
        total_employees_with_errors = salaries.filter(holdsalary=1).count()

        salary_details = salaries.select_related('emp_emailid').values(
            'sal_id', 'emp_emailid__holder_name', 'emp_emailid__bank_name', 'emp_emailid__branch',
            'emp_emailid__ifsc', 'emp_emailid__acc_no', 'emp_emailid__acc_type', 'Net_Salary',
            'holdsalary', 'paid'
        )

        response_data = {
            'month': month,
            'total_employees': total_employees,
            'total_amount': total_amount,
            'total_employees_with_errors': total_employees_with_errors,
            'salary_details': list(salary_details)
        }

        return JsonResponse(response_data, status=200)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def BankTransferUpdate(request):
    c_id = request.session.get('c_id')
    if not c_id:
        print('Company ID not found in session')
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    month = request.GET.get('month')
    if not month:
        print('Month parameter is required')
        return JsonResponse({'error': 'Month parameter is required'}, status=400)

    if request.method == 'POST':
        try:
            narration = request.POST.get('narration')
            debit_account = request.POST.get('debitac')
            file_name_start = request.POST.get('filename')
            value_date = request.POST.get('valuedate')

            print(f"Received data: narration={narration}, debit_account={debit_account}, file_name_start={file_name_start}, value_date={value_date}, month={month}")

            batch_id = Banktransferstatement.objects.all().count() + 1

            salaries = Salary.objects.filter(holdsalary=0, paymentmethod='bank', paid=0, payout_month=month)
            for salary in salaries:
                print(f"Creating Banktransferstatement for {salary.emp_emailid.holder_name}")
                Banktransferstatement.objects.create(
                    batchid=batch_id,
                    name=salary.emp_emailid.holder_name,
                    bank=salary.emp_emailid.bank_name,
                    branch=salary.emp_emailid.branch,
                    ifsc=salary.emp_emailid.ifsc,
                    accountno=salary.emp_emailid.acc_no,
                    amount=salary.Net_Salary
                )

            Banktransferstatement.objects.filter(batchid=batch_id).update(
                debitno=debit_account,
                date=datetime.strptime(value_date, '%Y-%m-%d').date(),
                narration=narration,
                filename=file_name_start
            )

            salaries.update(paid=1)

            print("Bank Transfer Statement generated successfully.")
            return JsonResponse({'message': 'Bank Transfer Statement Generated successfully.'}, status=200)

        except Exception as e:
            print(f"Error during Bank Transfer Update: {str(e)}")
            return JsonResponse({'error': str(e)}, status=400)

    else:
        print('Method not allowed')
        return JsonResponse({'error': 'Method not allowed'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CashChequeTransferPayout(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            payouts = Salary.objects.filter(paymentmethod='cash').order_by('-payout_month').values('sal_id', 'payout_month').distinct()

            payout_list = [{'sal_id': payout['sal_id'], 'payout_month': payout['payout_month']} for payout in payouts]

            return JsonResponse({'payroll_months': payout_list}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def HoldSalaryPayout(request):
#     c_id = request.session.get('c_id')
#     if not c_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'GET':
#         try:
#             payouts = Salary1.objects.filter().values('sal_id', 'holdsalary', 'emp_emailid__holder_name', 'emp_emailid__bank_name', 'emp_emailid__branch', 'emp_emailid__ifsc', 'emp_emailid__acc_no').distinct()

#             payout_list = []
#             for payout in payouts:
#                 payout_dict = {
#                     'sal_id': payout['sal_id'],
#                     'holder_name': payout['emp_emailid__holder_name'],
#                     'bank_name': payout['emp_emailid__bank_name'],
#                     'branch': payout['emp_emailid__branch'],
#                     'ifsc': payout['emp_emailid__ifsc'],
#                     'acc_no': payout['emp_emailid__acc_no'],
#                     'hold_salary': payout['holdsalary']
#                 }
#                 payout_list.append(payout_dict)

#             return JsonResponse({'payouts': payout_list}, status=200)

#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)

#     else:
#         return JsonResponse({'error': 'Method not allowed'}, status=405)







@csrf_exempt
def HoldSalaryPayout(request):
    c_id = request.session.get('c_id')
    
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        emailid = request.GET.get('emailid')

        # Decode URL-encoded email ID for debugging purposes
        decoded_emailid = urllib.parse.unquote(emailid)
        print(f"Decoded email ID: {decoded_emailid}")
        
        if not emailid:
            return JsonResponse({'error': 'Email ID not provided'}, status=400)

        try:
            payouts = Salary1.objects.filter(
                emp_emailid=emailid
            ).values(
                'emp_emailid__holder_name',
                'emp_emailid__bank_name',
                'emp_emailid__branch',
                'emp_emailid__ifsc',
                'emp_emailid__acc_no',
                'sal_id',
                'holdsalary'
            ).distinct()

            if not payouts:
                return JsonResponse({'error': 'No salary details found for this employee'}, status=404)

            payout_list = []
            for payout in payouts:
                payout_dict = {
                    'sal_id': payout['sal_id'],
                    'holder_name': payout['emp_emailid__holder_name'],
                    'bank_name': payout['emp_emailid__bank_name'],
                    'branch': payout['emp_emailid__branch'],
                    'ifsc': payout['emp_emailid__ifsc'],
                    'acc_no': payout['emp_emailid__acc_no'],
                    'holdsalary': payout['holdsalary']
                }
                payout_list.append(payout_dict)
                print("Payout List:")
                for item in payout_list:
                    print(item)


            return JsonResponse({'payouts': payout_list}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Method not allowed'}, status=405)



# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def HoldSalary(request):
#     c_id = request.session.get('c_id')
#     emailid = request.GET.get('emailid')
#     if not c_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'GET':
#         try:
#             emp_name = Salary1.objects.filter(emp_emailid=emailid).values('emp_emailid__holder_name').first()

#             if not emp_name:
#                 return JsonResponse({'error': 'Employee not found'}, status=404)
#             return JsonResponse(emp_name, status=200)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)

#     elif request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             hold_reason = data.get('hold_reason')
#             remarks = data.get('remarks')

#             if not hold_reason or not remarks:
#                 return JsonResponse({'error': 'Hold reason and remarks are required'}, status=400)

#             salary = Salary.objects.filter(emp_emailid=emailid, emp_emailid__emp_emailid__d_id__c_id=c_id).first()
#             if not salary:
#                 return JsonResponse({'error': 'Salary or Bank details not found for the given email ID'}, status=404)

#             salary.holdsalary = 1
#             salary.notes = hold_reason
#             salary.remarks = remarks
#             salary.save()

#             return JsonResponse({'message': f'{salary.emp_emailid.holder_name} has been put on hold for salary.'}, status=200)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)
#     else:
#         return JsonResponse({'error': 'Method not allowed'}, status=405)
    






@csrf_exempt
def HoldSalary(request):
    c_id = request.session.get('c_id')

    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            employees = Employee.objects.filter(d_id__c_id=c_id).values('emp_name', 'emp_emailid')
            if not employees:
                return JsonResponse({'error': 'No employees found for this company'}, status=404)
            return JsonResponse({'employees': list(employees)}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            emp_emailid = data.get('employeeId')
            hold_reason = data.get('holdReason')
            remarks = data.get('remarks')

            if not emp_emailid or not hold_reason or not remarks:
                return JsonResponse({'error': 'Employee ID, hold reason, and remarks are required'}, status=400)

            # Retrieve the Bank_details instance for the given email
            bank_details = Bank_details.objects.filter(emp_emailid=emp_emailid).first()
            if not bank_details:
                return JsonResponse({'error': 'Bank details not found for the given employee ID'}, status=404)

            # Find the salary record for the given employee and c_id
            salary, created = Salary1.objects.update_or_create(
                emp_emailid=bank_details,  # Use the Bank_details instance
                defaults={
                    'holdsalary': 1,
                    'notes': hold_reason,
                    'remarks': remarks
                }
            )

            message = 'Salary held successfully' if not created else 'Salary record created and held successfully'
            return JsonResponse({'message': message}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)




# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def UnholdSalary(request):
#     c_id = request.session.get('c_id')
#     emailid = request.GET.get('emailid')
#     if not c_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'POST':
#         try:
#             salary = Salary1.objects.filter(emp_emailid=emailid, emp_emailid_emp_emailidd_id_c_id=c_id).first()
#             if not salary:
#                 return JsonResponse({'error': 'Salary or Bank details not found for the given email ID'}, status=404)

#             salary.holdsalary = 0
#             salary.save()

#             return JsonResponse({'message': f'{salary.emp_emailid.holder_name} is no more on hold for salary.'}, status=200)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)
#     else:
#         return JsonResponse({'error': 'Method not allowed'}, status=405)
    


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])  # Assuming this decorator checks roles
def UnholdSalary(request):
    c_id = request.session.get('c_id')  # Get the company ID from session
    emailid = request.GET.get('emailid')  # Get employee email from GET parameters
    
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'POST':
        try:
            # Query to check if employee exists within the given company through their department
            employee = Employee.objects.filter(emp_emailid=emailid, d_id__c_id=c_id).first()
            
            if not employee:
                return JsonResponse({'error': 'Employee not found in the specified company'}, status=404)

            # Now that we have the employee, we fetch their salary record
            salary = Salary1.objects.filter(emp_emailid=employee.emp_emailid).first()

            if not salary:
                return JsonResponse({'error': 'Salary details not found for the given email ID'}, status=404)

            # Unhold the salary by setting the holdsalary field to 0
            salary.holdsalary = 0
            salary.save()

            return JsonResponse({'message': f'{employee.emp_name} is no more on hold for salary.'}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Method not allowed'}, status=405)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def PayslipPayout(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        payouts = Salary.objects.order_by('-payout_month').values('sal_id', 'payout_month').distinct()
        return JsonResponse({'payouts': list(payouts)})
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def GeneratePayslip(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        month = request.GET.get('month')
        if not month:
            return JsonResponse({'error': 'Month parameter is required'}, status=400)

        payslip = Salary.objects.filter(emp_emailid__emp_emailid__d_id__c_id=c_id, payout_month=month).select_related('emp_emailid').values(
            'sal_id',
            'emp_emailid__holder_name',
            'emp_emailid__bank_name',
            'emp_emailid__branch',
            'emp_emailid__acc_type',
            'emp_emailid__acc_no'
        )

        return JsonResponse({'payouts': list(payslip)})
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def HomeSalary(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        employees = Employee.objects.filter(d_id__in=Department.objects.filter(c_id=c_id)).values('emp_name', 'emp_emailid', 'emp_role')
        return JsonResponse({'employees': list(employees)})
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddSalary(request):
    print("View reached")
    if request.method == 'POST':
        print("View reached")
        emp_emailid = request.GET.get('emp_emailid')
        print("View reached")

        try:
            employee = Employee.objects.get(emp_emailid=emp_emailid)
            personal_details = Personal_details.objects.get(emp_emailid=emp_emailid)
            poifiles_new = Poifiles_new.objects.get(emp_emailid=emp_emailid)

            age = datetime.now().year - personal_details.birth_date.year
            actual80c = poifiles_new.actualAmount_80C
            actual80d = poifiles_new.actualAmount_80D
            actualoie = poifiles_new.OIE_actualAmount
            actualosi = poifiles_new.OSI_actualAmount

            eligible_deductions = actual80c + actual80d + actualoie + actualosi

            annual_ctc = 0
            remarks = ""
            notes = ""

            # Retrieve salary details
            salary_instance = Salary.objects.filter(emp_emailid=emp_emailid).order_by('-sal_id').first()
            if salary_instance:
                annual_ctc = salary_instance.annual_ctc
                remarks = salary_instance.remarks
                notes = salary_instance.notes

            payout_month = request.POST.get('payoutMonth', '')
            effective_from = request.POST.get('effectiveFrom', '')
            revise = int(request.POST.get('revision', 0))

            if not annual_ctc:
                annual_ctc = int(request.POST.get('annualCTC', 0))

            annual_ctc += annual_ctc * (revise / 100)
            monthly_ctc = annual_ctc / 12
            basic = monthly_ctc * 0.4
            hra = basic * 0.4
            conveyance = 1600
            da = 0
            special_allowance = monthly_ctc - (hra + conveyance + da + basic)

            annual_taxable = annual_ctc - eligible_deductions
            tax_liability = tax_calculation_to_add_salary(age, annual_taxable)

            monthly_tds = tax_liability / 12
            monthly_epf = basic * 0.12
            monthly_prof_tax = 200
            net_salary = monthly_ctc - (monthly_tds + monthly_epf + monthly_prof_tax)

            # Save salary details to database
            salary_instance = Salary(
                emp_emailid=emp_emailid,
                basic=basic,
                hra=hra,
                conveyance=conveyance,
                da=da,
                special_allowance=special_allowance,
                monthly_ctc=monthly_ctc,
                annual_ctc=annual_ctc,
                Eligible_Deductions=eligible_deductions,
                Yearly_Taxable_Salary=annual_taxable,
                otal_Tax_Liability=tax_liability,
                Monthly_TDS=monthly_tds,
                Monthly_EPF=monthly_epf,
                Monthly_Professional_Tax=monthly_prof_tax,
                Net_Salary=net_salary,
                payout_month=payout_month,
                effective_from=effective_from,
                notes=notes,
                remarks=remarks,
                revision=revise
            )
            salary_instance.save()
            return JsonResponse({'message': 'Salary details saved successfully!'})

        except ObjectDoesNotExist:
            return JsonResponse({'error': 'Employee with the provided email ID does not exist'}, status=404)

    else:
        emp_emailid = request.GET.get('emp_emailid')

        # Retrieve employee details from models
        employee = Employee.objects.get(emp_emailid=emp_emailid)
        personal_details = Personal_details.objects.get(emp_emailid=emp_emailid)
        poifiles_new = Poifiles_new.objects.get(emp_emailid=emp_emailid)

        age = datetime.now().year - personal_details.birth_date.year
        actual80c = poifiles_new.actualAmount_80C
        actual80d = poifiles_new.actualAmount_80D
        actualoie = poifiles_new.OIE_actualAmount
        actualosi = poifiles_new.OSI_actualAmount

        eligible_deductions = actual80c + actual80d + actualoie + actualosi

        annual_ctc = 0
        remarks = ""
        notes = ""

        # Retrieve salary details
        salary_instance = Salary.objects.filter(emp_emailid=emp_emailid).order_by('-sal_id').first()
        if salary_instance:
            annual_ctc = salary_instance.annual_ctc
            remarks = salary_instance.remarks
            notes = salary_instance.notes

        context = {
            'employee': employee,
            'personal_details': personal_details,
            'poifiles_new': poifiles_new,
            'age': age,
            'eligible_deductions': eligible_deductions,
            'annual_ctc': annual_ctc,
            'remarks': remarks,
            'notes': notes,
        }

        return JsonResponse(context)





# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def AddSalary1(request):
#     if request.method == 'GET':
#         # Get company id from session
#         c_id = request.session.get('c_id')
#         if not c_id:
#             return JsonResponse({'error': 'Company ID not found in session'}, status=400)

#         # Retrieve all employees that belong to departments under this company
#         employees = Employee.objects.filter(d_id__c_id=c_id)

#         # Prepare employee data to return to the frontend
#         employee_data = [{'emp_emailid': emp.emp_emailid, 'name': emp.emp_name} for emp in employees]

#         return JsonResponse({'employees': employee_data})

#     elif request.method == 'POST':
#         data = json.loads(request.body)
#         emp_emailid = data.get('emp_emailid')
#         payout_month = data.get('payoutMonth')
#         revision = int(data.get('revision', 0))
#         annual_ctc = int(data.get('annualCTC', 0))
#         payment_method = data.get('paymentMethod')
#         employee_remarks = data.get('employeeRemarks')
#         notes = data.get('notes')
#         effective_from = data.get('effectiveFrom')

#         c_id = request.session.get('c_id')

#         if not emp_emailid or not c_id:
#             return JsonResponse({'error': 'Missing employee or company details'}, status=400)

#         try:
#             # Retrieve the employee's bank details from Bank_details using emp_emailid
#             bank_details = Bank_details.objects.get(emp_emailid__emp_emailid=emp_emailid)

#             # Retrieve any additional employee information if necessary
#             personal_details = Personal_details.objects.get(mail__emp_emailid=emp_emailid)
#             poifiles_new = Poifiles_new1.objects.get(Emp_id__emp_emailid=emp_emailid)

#             # # Calculate salary details
#             age = datetime.now().year - personal_details.birth_date.year
#             actual80c = poifiles_new.actualAmount_80C
#             actual80d = poifiles_new.actualAmount_80D
#             actualoie = poifiles_new.OIE_actualAmount
#             actualosi = poifiles_new.OSI_actualAmount

#             eligible_deductions = actual80c + actual80d + actualoie + actualosi

#             # If revision is provided, apply it to the annual CTC
#             annual_ctc += annual_ctc * (revision / 100)
#             monthly_ctc = annual_ctc / 12
#             basic = monthly_ctc * 0.4
#             hra = basic * 0.4
#             conveyance = 1600
#             da = 0
#             special_allowance = monthly_ctc - (hra + conveyance + da + basic)

#             # Calculate tax liability
#             annual_taxable = annual_ctc - eligible_deductions
#             tax_liability = tax_calculation_to_add_salary(age, annual_taxable)

#             monthly_tds = tax_liability / 12
#             monthly_epf = basic * 0.12
#             monthly_prof_tax = 200
           
#             net_salary = monthly_ctc - (monthly_tds + monthly_epf + monthly_prof_tax)

#             # Save salary details to the database
#             salary_instance = Salary1.objects.create(
#                 emp_emailid=bank_details,  # Assign Bank_details instance to Salary
#                 basic=basic,
#                 hra=hra,
#                 conveyance=conveyance,
#                 da=da,
#                 special_allowance=special_allowance,
#                 monthly_ctc=monthly_ctc,
#                 annual_ctc=annual_ctc,
#                 Eligible_Deductions=eligible_deductions,
#                 Yearly_Taxable_Salary=annual_taxable,
#                 Total_Tax_Liability=tax_liability,
#                 Monthly_TDS=monthly_tds,
#                 Monthly_EPF=monthly_epf,
#                 Monthly_Professional_Tax=monthly_prof_tax,
#                 Net_Salary=net_salary,
#                 payout_month=payout_month,
#                 effective_from=effective_from,
#                 notes=notes,
#                 remarks=employee_remarks,
#                 revision=revision,
#                 paymentmethod=payment_method
#             )
#             salary_instance.save()

#             return JsonResponse({'message': 'Salary details saved successfully.'})

#         except Bank_details.DoesNotExist:
#             return JsonResponse({'error': 'Employee bank details not found.'}, status=404)

#         except Employee.DoesNotExist:
#             return JsonResponse({'error': 'Employee not found or does not belong to this company.'}, status=404)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddSalary1(request):
    if request.method == 'GET':
        # Get company id from session
        c_id = request.session.get('c_id')
        if not c_id:
            return JsonResponse({'error': 'Company ID not found in session'}, status=400)

        # Retrieve all employees that belong to departments under this company
        employees = Employee.objects.filter(d_id__c_id=c_id)

        # Prepare employee data to return to the frontend
        employee_data = [{'emp_emailid': emp.emp_emailid, 'name': emp.emp_name} for emp in employees]

        return JsonResponse({'employees': employee_data})

    elif request.method == 'POST':
        data = json.loads(request.body)
        emp_emailid = data.get('emp_emailid')
        payout_month = data.get('payoutMonth')
        revision = int(data.get('revision', 0))
        annual_ctc = int(data.get('annualCTC', 0))
        payment_method = data.get('paymentMethod')
        employee_remarks = data.get('employeeRemarks')
        notes = data.get('notes')
        effective_from = data.get('effectiveFrom')

        c_id = request.session.get('c_id')

        if not emp_emailid or not c_id:
            return JsonResponse({'error': 'Missing employee or company details'}, status=400)

        try:
            # Retrieve the employee's bank details from Bank_details using emp_emailid
            bank_details = Bank_details.objects.get(emp_emailid__emp_emailid=emp_emailid)

            # Retrieve any additional employee information if necessary
            personal_details = Personal_details.objects.get(mail__emp_emailid=emp_emailid)
            poifiles_new = Poifiles_new1.objects.get(Emp_id__emp_emailid=emp_emailid)

            # Calculate salary details
            age = datetime.now().year - personal_details.birth_date.year
            actual80c = poifiles_new.actualAmount_80C
            actual80d = poifiles_new.actualAmount_80D
            actualoie = poifiles_new.OIE_actualAmount
            actualosi = poifiles_new.OSI_actualAmount

            eligible_deductions = actual80c + actual80d + actualoie + actualosi

            # If revision is provided, apply it to the annual CTC
            annual_ctc += annual_ctc * (revision / 100)
            monthly_ctc = annual_ctc / 12
            basic = monthly_ctc * 0.4
            hra = basic * 0.4
            conveyance = 1600
            da = 0
            special_allowance = monthly_ctc - (hra + conveyance + da + basic)

            # Calculate tax liability
            annual_taxable = annual_ctc - eligible_deductions
            tax_liability = tax_calculation_to_add_salary(age, annual_taxable)

            monthly_tds = tax_liability / 12
            monthly_epf = basic * 0.12
            monthly_prof_tax = 200

            net_salary = monthly_ctc - (monthly_tds + monthly_epf + monthly_prof_tax)

            # Check if salary record exists for the given employee and payout month
            salary_instance, created = Salary1.objects.update_or_create(
                emp_emailid=bank_details,  # Use Bank_details instance
                # payout_month=payout_month,  # Use the payout month for identifying the record
                defaults={
                    'basic': basic,
                    'hra': hra,
                    'conveyance': conveyance,
                    'da': da,
                    'special_allowance': special_allowance,
                    'monthly_ctc': monthly_ctc,
                    'annual_ctc': annual_ctc,
                    'Eligible_Deductions': eligible_deductions,
                    'Yearly_Taxable_Salary': annual_taxable,
                    'Total_Tax_Liability': tax_liability,
                    'Monthly_TDS': monthly_tds,
                    'Monthly_EPF': monthly_epf,
                    'Monthly_Professional_Tax': monthly_prof_tax,
                    'Net_Salary': net_salary,
                    'effective_from': effective_from,
                    'notes': notes,
                    'remarks': employee_remarks,
                    'revision': revision,
                    'paymentmethod': payment_method
                }
            )

            if created:
                message = 'Salary details created successfully.'
            else:
                message = 'Salary details updated successfully.'

            return JsonResponse({'message': message})

        except Bank_details.DoesNotExist:
            return JsonResponse({'error': 'Employee bank details not found.'}, status=404)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found or does not belong to this company.'}, status=404)



# Function to calculate tax while adding salary. Check Function above Name: AddSalary
def tax_calculation_to_add_salary(age, annual_taxable):
    tax = 0
    tcredit = 0
    surch = 0
    mtr = 0

    if age < 60:
        if annual_taxable > 1000000:
            tax = 112500 + (annual_taxable - 1000000) * 30 / 100
        elif annual_taxable > 500000 and annual_taxable <= 1000000:
            tax = 12500 + (annual_taxable - 500000) * 20 / 100
        elif annual_taxable > 200000 and annual_taxable <= 500000:
            tax = (annual_taxable - 250000) * 5 / 100
        elif annual_taxable <= 250000:
            tax = 0
    elif age >= 80:
        if annual_taxable > 1000000:
            tax = 100000 + (annual_taxable - 1000000) * 30 / 100
        elif annual_taxable > 500000 and annual_taxable <= 1000000:
            tax = (annual_taxable - 500000) * 20 / 100
        elif annual_taxable <= 500000:
            tax = 0
    elif age >= 60:
        if annual_taxable > 1000000:
            tax = 110000 + (annual_taxable - 1000000) * 30 / 100
        elif annual_taxable > 500000 and annual_taxable <= 1000000:
            tax = 10000 + (annual_taxable - 500000) * 20 / 100
        elif annual_taxable > 250000 and annual_taxable <= 500000:
            tax = (annual_taxable - 300000) * 5 / 100
        elif annual_taxable < 300000:
            tax = 0

    if age < 60 and annual_taxable <= 500000 and tax > 0:
        tcredit = tax

    if age >= 60 and age < 80 and annual_taxable <= 500000 and tax > 0:
        tcredit = tax

    tax = tax - tcredit

    if annual_taxable > 5000000 and annual_taxable <= 10000000:
        surch = tax * 10 / 100
        if surch > 0 and age < 60 and (annual_taxable * 1 - tax * 1.10) <= 3687500:
            mtr = surch * 1 - (annual_taxable - 3687500 - tax)
        if surch > 0 and age >= 60 and (annual_taxable - tax * 1.10) <= 3690000:
            mtr = surch * 1 - (annual_taxable - 3690000 * 1 - tax * 1)
        if surch > 0 and age >= 80 and (annual_taxable - tax * 1.10) <= 3700000:
            mtr = surch - (annual_taxable - 3700000 * 1 - tax * 1)
    if annual_taxable > 10000000:
        surch = tax * 15 / 100
        if surch > 0 and (age < 60) and (annual_taxable - tax * 1.15) <= 6906250:
            mtr = surch - (annual_taxable - 6906250 * 1 - tax)
        if surch > 0 and age >= 60 and (annual_taxable - tax * 1.15) <= 6909000:
            mtr = surch - (annual_taxable - 6909000 * 1 - tax)
        if surch > 0 and age >= 80 and (annual_taxable - tax * 1.15) <= 6920000:
            mtr = surch - (annual_taxable - 6920000 * 1 - tax)
    if annual_taxable <= 5000000:
        surch = 0
    surch = surch - mtr
    ttax = tax + surch
    ecess = ttax * .04
    liab = ttax + ecess

    return liab


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def SalaryRevisionHistory(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        emailid=request.GET.get('emailid')

        decoded_emailid = urllib.parse.unquote(emailid)
        print(f"Decoded email ID: {decoded_emailid}")

        if not emailid:
            return JsonResponse({'error':'Email ID not provided'}, status=400)

        try:
            # Retrieve salaries based on employee email ID
            salaries = Salary1.objects.filter(
               emp_emailid=emailid
            ).values(
                    'sal_id',         
                    'emp_emailid__emp_emailid',                     
                    'revision',                     
                    'effective_from',              
                    'basic',           
                    'hra', 
                    'conveyance',             
                    'da',                     
                    'special_allowance',                      
                    'annual_ctc',         
                    'paymentmethod',
                    'notes',
                    'remarks'                                                                                       
            ).distinct()

            if not salaries:
                return JsonResponse({'error': 'No salary details found for this employee'}, status=404)

            salary_list = []
            for salary in salaries:
                salary_dict = {
                    'sal_id': salary['sal_id'],  # Correctly access email ID from Bank_details
                    'emp_emailid': salary['emp_emailid__emp_emailid'],  # Correctly access email ID from Bank_details
                    'revision': salary['revision'],
                    'effective_from': salary['effective_from'],
                    'basic': salary['basic'],
                    'hra': salary['hra'],
                    'conveyance': salary['conveyance'],
                    'da': salary['da'],
                    'special_allowance': salary['special_allowance'],
                    'annual_ctc': salary['annual_ctc'],
                    'paymentmethod': salary['paymentmethod'],
                    'notes': salary['notes'],
                    'remarks': salary['remarks'],
                }
                salary_list.append(salary_dict)
                print("salary list: ")
                for item in salary_list:
                    print(item)

            return JsonResponse({'salaries': salary_list}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)






@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def DisplaySalaryDetails(request):
    c_id = request.session.get('c_id')
    if not c_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        emailid=request.GET.get('emailid')

        decoded_emailid = urllib.parse.unquote(emailid)
        print(f"Decoded email ID: {decoded_emailid}")

        if not emailid:
            return JsonResponse({'error':'Email ID not provided'}, status=400)
        try:
           # Retrieve salaries based on employee email ID
            salaries = Salary1.objects.filter(
               emp_emailid=emailid
            ).values(
                    'emp_emailid__emp_emailid',         # Employee email ID from Bank_details model
                    'payout_month',                     # Payout month
                    'monthly_ctc',                      # Monthly CTC
                    'Eligible_Deductions',              # Eligible Deductions
                    'Yearly_Taxable_Salary',            # Yearly Taxable Salary
                    'Total_Tax_Liability',              # Total Tax Liability
                    'Monthly_TDS',                      # Monthly TDS
                    'Monthly_EPF',                      # Monthly EPF
                    'Monthly_Professional_Tax',         # Monthly Professional Tax
                    'Net_Salary'                        # Net Salary
            ).distinct()

            if not salaries:
                return JsonResponse({'error': 'No salary details found for this employee'}, status=404)
            
            salary_list = []
            for salary in salaries:
                salary_dict = {
                    'emp_emailid': salary['emp_emailid__emp_emailid'],  # Correctly access email ID from Bank_details
                    'payout_month': salary['payout_month'],
                    'monthly_ctc': salary['monthly_ctc'],
                    'Eligible_Deductions': salary['Eligible_Deductions'],
                    'Yearly_Taxable_Salary': salary['Yearly_Taxable_Salary'],
                    'Total_Tax_Liability': salary['Total_Tax_Liability'],
                    'Monthly_TDS': salary['Monthly_TDS'],
                    'Monthly_EPF': salary['Monthly_EPF'],
                    'Monthly_Professional_Tax': salary['Monthly_Professional_Tax'],
                    'Net_Salary': salary['Net_Salary'],
                }
                salary_list.append(salary_dict)
                print("salary list: ")
                for item in salary_list:
                    print(item)

            return JsonResponse({'salaries': salary_list}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)




# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def DisplaySalaryDetails(request):
#     c_id = request.session.get('c_id')
#     if not c_id:
#         return JsonResponse({'error': 'Company ID not found in session'}, status=401)

#     if request.method == 'GET':
#         # sid = request.GET.get('id')
#         # if not sid:
#         #     return JsonResponse({'error': 'Employee ID not provided'}, status=400)

#         try:
#             salaries = Salary1.objects.filter(emp_emailid__emp_emailid=sid)

#             salary_list = []
#             for salary in salaries:
#                 salary_dict = {
#                     'emp_emailid': salary.emp_emailid.emp_emailid,  # Accessing the email id through Bank_details
#                     'payout_month': salary.payout_month,
#                     'monthly_ctc': salary.monthly_ctc,
#                     'Eligible_Deductions': salary.Eligible_Deductions,
#                     'Yearly_Taxable_Salary': salary.Yearly_Taxable_Salary,
#                     'Total_Tax_Liability': salary.Total_Tax_Liability,
#                     'Monthly_TDS': salary.Monthly_TDS,
#                     'Monthly_EPF': salary.Monthly_EPF,
#                     'Monthly_Professional_Tax': salary.Monthly_Professional_Tax,
#                     'Net_Salary': salary.Net_Salary,
#                 }
#                 salary_list.append(salary_dict)

#             return JsonResponse({'salaries': salary_list}, status=200)

#         except Salary1.DoesNotExist:
#             return JsonResponse({'error': 'Salary records not found for the provided employee ID'}, status=404)

#     else:
#         return JsonResponse({'error': 'Method not allowed'}, status=405)







@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CustomForms(request):
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)
    user_name = request.session.get('emp_name')

    if not c_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    if request.method == 'GET':
        try:
            forms = Custom_forms.objects.filter(c_id=company_id).order_by('-seq').values('form_name')
            forms_list = [{'form_name': form['form_name']} for form in forms]
            return JsonResponse({'forms': forms_list})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    # elif request.method == 'POST':
    #     data = json.loads(request.body)
    #     form_name = data.get('newFormName')
    #     if not form_name:
    #         return JsonResponse({'error': 'Form name is required'}, status=400)

    #     try:
    #         formatted_form_name = form_name.replace(' ', '_').lower()
    #         response_table_name = f"resp_{c_id}{formatted_form_name}"

    #         with connection.cursor() as cursor:
    #             cursor.execute(f"""
    #                 CREATE TABLE IF NOT EXISTS "{response_table_name}" (
    #                     "emp_name" VARCHAR(255) PRIMARY KEY
    #                 );
    #             """)

    #         if Custom_forms.objects.filter(c_id=company_id, form_name=formatted_form_name).exists():
    #             return JsonResponse({'error': 'Form name already exists'}, status=400)

    #         new_form = Custom_forms(c_id=c_id, form_name=formatted_form_name)
    #         new_form.save()
    #         return JsonResponse({'message': 'Form created successfully'})
    #     except Exception as e:
    #         return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            form_name = data.get('newFormName')
            alloc = data.get('alloc', None)

            if not form_name:
                return JsonResponse({'error': 'Form name is required'}, status=400)

            # Create the Custom_forms instance using the company_id from the session
            custom_form = Custom_forms(
                form_name=form_name,
                alloc=alloc,
                c_id=c_id  # Use the company instance obtained from the session
            )
            custom_form.save()

            return JsonResponse({'message': 'Custom form created successfully', 'form_id': custom_form.seq}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

        # return JsonResponse({'error': 'Invalid request method'}, status=405)

    elif request.method == 'DELETE':
        form_name = request.GET.get('form_name')
        question_id = request.GET.get('idOfQuestion')

        if not form_name:
            return JsonResponse({'error': 'Form name is required'}, status=400)

        try:
            formatted_form_name = form_name.replace(' ', '_').lower()
            response_table_name = f"resp_{company_id}{formatted_form_name}"

            if question_id:
                # Delete specific question of the form
                Custom_forms_questions.objects.filter(ID=question_id, form_name=formatted_form_name, c_id=company_id).delete()

                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        ALTER TABLE "{response_table_name}" DROP COLUMN "{question_id}";
                    """)
                return JsonResponse({'message': 'Question deleted successfully'})
            else:
                # Delete the entire form
                Custom_forms.objects.filter(form_name=formatted_form_name, c_id=company_id).delete()
                Custom_forms_questions.objects.filter(form_name=formatted_form_name, c_id=company_id).delete()

                with connection.cursor() as cursor:
                    cursor.execute(f"""
                        DROP TABLE IF EXISTS "{response_table_name}";
                    """)
                return JsonResponse({'message': 'Form deleted successfully'})
        except Custom_forms.DoesNotExist:
            return JsonResponse({'error': 'Form not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EditFormView(request):
    user_name = request.session.get('emp_name')
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)

    if not user_name or not c_id:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    form_name = request.GET.get('form_name')
    if not form_name:
        return JsonResponse({'error': 'Form name is required'}, status=400)

    formatted_form_name = form_name.replace(' ', '_').lower()
    try:
        questions = Custom_forms_questions.objects.filter(form_name=formatted_form_name, c_id=company_id).values()
        employees = Employee.objects.filter(d_id__c_id=c_id).values_list('emp_name', flat=True)
        return JsonResponse({'questions': list(questions), 'employees': list(employees)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AllocateFormView(request):
    user_name = request.session.get('emp_name')
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)

    if not user_name or not c_id:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    # Get form name from POST body
    data = json.loads(request.body)
    form_name = data.get('form_name')
    print("Form name from POST:", form_name)

    if not form_name:
        return JsonResponse({'error': 'Form name is required'}, status=400)

    if request.method == 'POST':
        try:
            allocated_employees = data.get('allocated_employee')
            if not allocated_employees:
                return JsonResponse({'error': 'Allocated employees are missing'}, status=400)

            allocated_employee_str = ", ".join(allocated_employees)
            print("Allocated employees as string:", allocated_employee_str)

            custom_form = Custom_forms.objects.get(form_name=form_name, c_id=company_id)
            print("Custom form found:", custom_form)

            custom_form.alloc = allocated_employee_str
            custom_form.save()
            print("Form allocation updated successfully.")

            return JsonResponse({'message': 'Form allocated successfully'}, status=200)  # Added explicit status code

        except Custom_forms.DoesNotExist:
            print("Form not found with name:", form_name, "and company_id:", company_id)
            return JsonResponse({'error': 'Form not found'}, status=404)
        except Exception as e:
            print("Unexpected error:", str(e))
            return JsonResponse({'error': str(e)}, status=500)

    else:
        print("Received non-POST request. Method:", request.method)
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddTextQuestionFormView(request):
    user_name = request.session.get('emp_name')
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)

    if not user_name or not c_id:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    form_name = request.GET.get('form_name')
    if not form_name:
        return JsonResponse({'error': 'Form name is required'}, status=400)

    if request.method == 'POST':
        data = json.loads(request.body)
        label = data.get('question')
        ID = data.get('ID')
        type = data.get('Type')
        name = ID

        if not label or not ID or not type:
            return JsonResponse({'error': 'All fields are required'}, status=400)

        if Custom_forms_questions.objects.filter(ID=ID, name=name, form_name=form_name, c_id=company_id).exists():
                return JsonResponse({'error': 'Question with the same ID already exists'}, status=400)

        formatted_form_name = form_name.replace(' ', '_').lower()
        response_table_name = f"resp_{company_id}{formatted_form_name}"

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name='{response_table_name}' AND column_name='{name}'")
                if not cursor.fetchone():
                    cursor.execute(f"ALTER TABLE {response_table_name} ADD {name} VARCHAR(255);")
                else:
                    return JsonResponse({'error': f"Column '{name}' already exists in the table"}, status=400)

            custom_form_question = Custom_forms_questions.objects.create(
                label=label,
                type=type,
                ID=ID,
                name=name,
                form_name=formatted_form_name,
                c_id=c_id
            )
            return JsonResponse({'success': 'Text question added successfully'}, status=201)
        except IntegrityError as e:
            return JsonResponse({'error': 'Database integrity error: ' + str(e)}, status=500)
        except Exception as e:
            return JsonResponse({'error': 'An unexpected error occurred: ' + str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AddRadioQuestionView(request):
    user_name = request.session.get('emp_name')
    company_id = request.session.get('c_id')
    c_id = get_object_or_404(Company, pk=company_id)

    if not user_name or not c_id:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    form_name = request.GET.get('form_name')
    if not form_name:
        return JsonResponse({'error': 'Form name is required'}, status=400)

    if request.method == 'POST':
        data = json.loads(request.body)
        label = data.get('radioLabel')
        ID = data.get('radioID')
        options_arr = data.get('radioOptionName')
        options_str = ", ".join(options_arr) if options_arr else ''
        name = ID

        if Custom_forms_questions.objects.filter(ID=ID, form_name=form_name, c_id=company_id).exists():
            return JsonResponse({'error': 'Question with the same ID already exists'}, status=400)

        formatted_form_name = form_name.replace(' ', '_').lower()
        response_table_name = f"resp_{company_id}{formatted_form_name}"

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name='{response_table_name}' AND column_name='{name}'")
                if not cursor.fetchone():
                    cursor.execute(f"ALTER TABLE {response_table_name} ADD {name} VARCHAR(255);")
                else:
                    return JsonResponse({'error': f"Column '{name}' already exists in the table"}, status=400)


            custom_form_question = Custom_forms_questions.objects.create(
                label=label,
                type='radio',
                ID=ID,
                name=name,
                optionName=options_str,
                form_name=formatted_form_name,
                c_id=c_id
            )
            return JsonResponse({'success': 'Radio question added successfully'}, status=201)
        except IntegrityError as e:
            return JsonResponse({'error': 'Database integrity error: ' + str(e)}, status=500)
        except Exception as e:
            return JsonResponse({'error': 'An unexpected error occurred: ' + str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CustomLetter(request):
    company_id = request.session.get('c_id')
    user_name = request.session.get('emp_name')

    if not company_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    c_id = get_object_or_404(Company, pk=company_id)
    employee = get_object_or_404(Employee, emp_name=user_name)

    if request.method == 'GET':
        try:
            letters = Custom_letters.objects.filter(c_id=c_id).order_by('-seq').values('letter_name')
            if not letters.exists():
                return JsonResponse({'message': 'No letters available. Please create a letter to see it here.'}, status=200)

            letters_list = [{'letter_name': letter['letter_name'], 'display_name': letter['letter_name'].replace('_', ' ').title()} for letter in letters]
            return JsonResponse({'letters': letters_list}, status=200)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_letter_name = data.get('newLetterName')

            if not new_letter_name:
                return JsonResponse({'error': 'Letter name is required'}, status=400)

            new_letter_name = new_letter_name.replace(' ', '_').lower()

            if Custom_letters.objects.filter(letter_name=new_letter_name, c_id=c_id).exists():
                return JsonResponse({'error': 'Letter name already exists!'}, status=400)

            Custom_letters.objects.create(letter_name=new_letter_name, c_id=c_id)
            return JsonResponse({ 'message': 'Letter deleted successfully'}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)
    


    elif request.method == 'DELETE':
        this_letter = request.GET.get('letter_name')

        if not this_letter:
            return JsonResponse({'error': 'Letter name is required'}, status=400)

        formatted_letter_name = this_letter.replace(' ', '_').lower()

        try:
            letter_to_delete = get_object_or_404(Custom_letters, letter_name=formatted_letter_name, c_id=c_id)
            letter_to_delete.delete()
            return JsonResponse({'message': 'Letter deleted successfully'}, status=200)
        except Custom_letters.DoesNotExist:
            return JsonResponse({'error': 'Letter not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EditLetterView(request):
    company_id = request.session.get('c_id')
    user_name = request.session.get('emp_name')
    emp_emailid = request.session.get('emp_emailid')
    department_id = request.session.get('d_id')

    if not company_id or not user_name or not department_id:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    company = get_object_or_404(Company, pk=company_id)
    employee = get_object_or_404(Employee, emp_emailid=emp_emailid)

    if request.method == 'GET':
        this_letter = request.GET.get('letter_name')
        if not this_letter:
            return JsonResponse({'error': 'Letter name is required'}, status=400)

        this_letter = this_letter.replace(' ', '_').lower()
        custom_letter = get_object_or_404(Custom_letters, letter_name=this_letter, c_id=company)
        departments = Department.objects.filter(c_id=company).values('d_id')

        employees = []
        for dept in departments:
            emp_in_dept = Employee.objects.filter(Q(d_id=dept['d_id']) & ~Q(emp_role='HR')).values('emp_name', 'emp_emailid')
            employees.extend(emp_in_dept)

        return JsonResponse({
            'letter_content': custom_letter.letter_content,
            'letter_name': custom_letter.letter_name.replace('_', ' ').title(),
            'employees': [{'name': emp['emp_name'], 'email': emp['emp_emailid']} for emp in employees]
        }, status=200)

    elif request.method == 'POST':
        data = json.loads(request.body)
        this_letter = data.get('thisLetter')
        letter_content = data.get('letterContent')

        if not this_letter or not letter_content:
            return JsonResponse({'error': 'Letter name and content are required'}, status=400)

        this_letter = this_letter.replace(' ', '_').lower()

        try:
            custom_letter = get_object_or_404(Custom_letters, letter_name=this_letter, c_id=company)
            custom_letter.letter_content = letter_content
            custom_letter.save()

            return JsonResponse({'success': 'Letter updated successfully'}, status=201)

        except Custom_letters.DoesNotExist:
            return JsonResponse({'error': f'Letter "{this_letter}" not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EmployeeDetails(request):
    company_id = request.session.get('c_id')
    logged_in_email = request.session.get('emp_emailid')  # Ensure user is logged in

    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if not logged_in_email:  # If the user is not logged in
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        try:
            # Fetch departments that belong to the logged-in user's company
            departments = Department.objects.filter(c_id=company_id)

            # Fetch employees who are in the departments of the logged-in user's company
            employees = Employee.objects.filter(d_id__in=departments)

            # Apply search filter if search term is provided
            search_term = request.GET.get('search', '')
            if search_term:
                employees = employees.filter(emp_name__icontains=search_term)

            # Sorting logic
            sort_column = request.GET.get('sort_column', 'emp_name')  # Default to 'emp_name' if not provided
            sort_direction = request.GET.get('sort_direction', 'asc')
            if sort_direction == 'desc':
                sort_column = '-' + sort_column
            employees = employees.order_by(sort_column)

            # Pagination logic
            page = request.GET.get('page', 1)
            entries_per_page = request.GET.get('entries_per_page', 10)
            paginator = Paginator(employees, entries_per_page)

            try:
                employees_page = paginator.page(page)
            except PageNotAnInteger:
                employees_page = paginator.page(1)
            except EmptyPage:
                employees_page = paginator.page(paginator.num_pages)

            # Serialize employee data
            employee_data = employees_page.object_list.values(
                'emp_name', 'emp_emailid', 'emp_phone', 'emp_role'
            )

            data = {
                'employees': list(employee_data),
                'total_pages': paginator.num_pages,
                'current_page': employees_page.number,
                'total_entries': paginator.count,
            }
            return JsonResponse(data)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'DELETE':
        emp_emailid = request.GET.get('emp_emailid')

        if not emp_emailid:
            return JsonResponse({'error': 'Employee email ID is required'}, status=400)

        try:
            # Ensure the employee to be deleted belongs to the same company as the logged-in user
            employee = Employee.objects.get(emp_emailid=emp_emailid, d_id__c_id=company_id)
            # Ensure the employee to be deleted belongs to the same company as the logged-in user
            employee = Employee.objects.get(emp_emailid=emp_emailid, d_id__c_id=company_id)
            employee.delete()
            return JsonResponse({'message': 'Employee deleted successfully'})
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found or does not belong to your company'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)



@csrf_exempt
def AttendanceDetails(request):
    user_email = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_email or not company_id:
        return JsonResponse({'error': 'User not authenticated or company ID not found'}, status=403)

    if request.method == 'GET':
        try:
            user = Employee.objects.get(emp_emailid=user_email)

            # If user is HR, fetch attendance for all employees in their company
            if user.emp_role == 'HR':
                # Fetch all departments under the user's company
                departments = Department.objects.filter(c_id=company_id).values_list('d_id', flat=True)
                
                # Fetch all employees under those departments
                employees = Employee.objects.filter(d_id__in=departments)
                
                # Fetch attendance records for those employees
                attendance_data = Attendance.objects.filter(emp_emailid__in=employees.values_list('emp_emailid', flat=True)).values('id', 'datetime_log', 'log_type', 'emp_emailid')
                
                # Create a mapping of employee emails to names
                employee_name_mapping = {employee.emp_emailid: employee.emp_name for employee in employees}
                
            else:
                # If not HR, fetch attendance only for the logged-in user
                attendance_data = Attendance.objects.filter(emp_emailid=user_email).values('id', 'datetime_log', 'log_type', 'emp_emailid')
                
                # Only need the name of the logged-in user
                employee_name_mapping = {user.emp_emailid: user.emp_name}

            data = []
            for attendance in attendance_data:
                datetime_log = attendance['datetime_log']

                # Parse the datetime_log if it's a string
                if isinstance(datetime_log, str):
                    datetime_log = parse_datetime(datetime_log)
                    if datetime_log is None:
                        raise ValueError(f"Invalid datetime format: {attendance['datetime_log']}")

                emp_email = attendance['emp_emailid']
                emp_name = employee_name_mapping.get(emp_email, "Unknown")  # Fetch the name from the mapping

                data.append({
                    'id': attendance['id'],
                    'emp_email': emp_email,
                    'emp_name': emp_name,  # Include name
                    'date': datetime_log.date(),
                    'log_type': attendance['log_type'],
                    'time': datetime_log.time()
                })

            # Return 'results' and 'total' to match the frontend expectation
            return JsonResponse({
                'results': data,
                'total': len(data)
            })

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User does not exist'}, status=404)
        except ValueError as ve:
            return JsonResponse({'error': str(ve)}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'DELETE':
        attendance_id = request.GET.get('id')
        if not attendance_id:
            return JsonResponse({'error': 'Attendance ID is required'}, status=400)

        try:
            # Fetch the attendance record and ensure it belongs to the correct company
            attendance = get_object_or_404(Attendance, id=attendance_id)
            
            # Debug: Print attendance details
            print(f"Attendance Record: {attendance.id}, Employee Email: {attendance.emp_emailid.emp_emailid}")

            # Fetch the employee associated with the attendance record
            employee = get_object_or_404(Employee, emp_emailid=attendance.emp_emailid.emp_emailid)
            
            # Debug: Print employee details
            print(f"Employee Found: {employee.emp_emailid}")

            # Check if the employee's department belongs to the user's company
            if employee.d_id.c_id.c_id != company_id:
                return JsonResponse({'error': 'Unauthorized action'}, status=403)

            # Delete the attendance record
            attendance.delete()
            return JsonResponse({'message': 'Attendance record deleted successfully'})

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee associated with attendance record not found'}, status=404)
        except Attendance.DoesNotExist:
            return JsonResponse({'error': 'Attendance record not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)




@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def EmployeeMaster(request):
    company_id = request.session.get('c_id')
    print(company_id)
    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    employees = Employee.objects.filter(d_id__c_id=company_id).values(
        'emp_name', 'emp_emailid', 'emp_phone', 'emp_role', 'd_id', 'emp_profile'
    )

    data = {
        'employees': list(employees)
    }

    return JsonResponse(data)


from psycopg2.extras import DateTimeTZRange
from datetime import datetime
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Todotasks, Employee, Tasks

@csrf_exempt
def task_list(request):
    emp_emailid = request.session.get('emp_emailid')

    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            print(f"📩 [RECEIVED] Data from frontend: {data}")

            title = data.get('title')
            description = data.get('description')
            priority = data.get('priority', False)
            completed = data.get('completed', False)
            start_date = data.get('start_date')
            end_date = data.get('end_date')

            if not title or not description or not start_date or not end_date:
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            # Convert date strings to datetime objects
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d")

            # Store in proper PostgreSQL range format
            date_range = DateTimeTZRange(start_date_obj, end_date_obj, '[)')  # Inclusive start, exclusive end

            new_task = Todotasks(
                title=title,
                description=description,
                date=date_range,  # Corrected format
                priority=priority,
                completed=completed
            )
            new_task.save()
            print(f"✅ [DB SUCCESS] Task Created: {new_task.__dict__}")

            # Associate with Employee
            employee = Employee.objects.get(emp_emailid=emp_emailid)
            task_entry = Tasks.objects.create(tid=new_task, emp_emailid=employee)
            print(f"✅ [DB SUCCESS] Task linked in Tasks table: {task_entry.__dict__}")

            return JsonResponse({'message': 'Task added successfully!'}, status=201)

        except Exception as e:
            print(f"❌ [ERROR] Task creation failed: {str(e)}")
            return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def task_detail(request, task_id):
    """
    Handles requests for a specific to-do task.

    GET: Retrieves details of a task.
    PUT: Updates a task.
    DELETE: Deletes a task.
    """
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    try:
        # Get the task, ensuring it belongs to the user
        task = Todotasks.objects.get(tid=task_id, tasks_tid__emp_emailid=emp_emailid)
    except Todotasks.DoesNotExist:
        return HttpResponse(status=404)

    if request.method == 'GET':
        return JsonResponse({'id': task.tid, 'description': task.description, 'date': task.date})

    elif request.method == 'PUT':
        data = json.loads(request.body)
        task.description = data.get('description', task.description)
        task.save()
        return JsonResponse({'id': task.tid, 'description': task.description, 'date': task.date})

    elif request.method == 'DELETE':
        task.delete()
        return HttpResponse(status=204)

@csrf_exempt
def task_search(request):
    """
    Handles search requests for to-do tasks.

    GET: Retrieves tasks matching the search term for the authenticated user.
    """
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        search_term = request.GET.get('search', '')
        tasks = Todotasks.objects.filter(
            description__icontains=search_term,
            tasks_tid__emp_emailid=emp_emailid
        ).values()
        tasks_list = list(tasks)
        return JsonResponse(tasks_list, safe=False)
    
    
@csrf_exempt
@require_http_methods(["GET"])
def get_performance_data(request, employee_email, year):
    # Authenticate user
    user_email = request.session.get('user_id')
    
    # Fetch the employee record using the email from the session
    try:
        logged_in_employee = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'User not authenticated'}, status=401)
    
    # Fetch the employee record for the requested performance data
    try:
        employee = Employee.objects.get(emp_emailid=employee_email, d_id=logged_in_employee.d_id)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)
    
    # Filter tasks based on the employee and year
    tasks = Tasks.objects.filter(emp_emailid=employee, sop__sdate__year=year)
    
    # Calculate task statistics
    total_sop_assigned = tasks.count()
    total_sop_accomplished = tasks.filter(status=True).count()
    total_kra_assigned = tasks.filter(kra_id__isnull=False).count()
    total_kra_accomplished = tasks.filter(kra_id__isnull=False, status=True).count()
    total_jd_assigned = tasks.filter(job_desc_id__isnull=False).count()
    total_jd_accomplished = tasks.filter(job_desc_id__isnull=False, status=True).count()
    total_training_assigned = tasks.filter(tid__isnull=False).count()
    total_training_accomplished = tasks.filter(tid__isnull=False, status=True).count()
    
    # Prepare the response data
    data = {
        'totalSopAssigned': total_sop_assigned,
        'totalSopAccomplished': total_sop_accomplished,
        'totalKraAssigned': total_kra_assigned,
        'totalKraAccomplished': total_kra_accomplished,
        'totalJdAssigned': total_jd_assigned,
        'totalJdAccomplished': total_jd_accomplished,
        'totalTrainingAssigned': total_training_assigned,
        'totalTrainingAccomplished': total_training_accomplished,
    }

    return JsonResponse(data, status=200)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def jd_form(request):
    """
    Handles requests related to Job Description forms.

    GET: Not supported for this endpoint.
    POST: Creates a new JD form and associated tasks.
    PUT: Not supported for this endpoint.
    DELETE: Not supported for this endpoint.
    """
    company_id = request.session.get('c_id')
    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            jd_name = data.get('jdName')
            date = data.get('date')
            responsibilities = data.get('responsibilities')

            if not all([jd_name, date, responsibilities]):
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            # Assuming you are sending employee emails in the request
            employee_emails = data.get('employeeEmails', [])  # Get employee emails from data

            with transaction.atomic():
                jd_instance = Job_desc.objects.create(
                    jd_name=jd_name,
                    responsiblities=', '.join(responsibilities),  # Store responsibilities as comma-separated string
                    sdate=date
                )

                for emp_email in employee_emails:
                    employee = Employee.objects.get(emp_emailid=emp_email)
                    if employee.d_id.c_id_id != company_id:
                        raise ValueError(f"Employee {emp_email} does not belong to this company.")

                    Tasks.objects.create(
                        job_desc_id=jd_instance,
                        responsiblities=', '.join(responsibilities),  # Store responsibilities as comma-separated string
                        emp_emailid=employee,
                        sdate=date,
                        d_id=employee.d_id
                    )

            return JsonResponse({'message': 'JD form and tasks created successfully'}, status=201)

        except (Employee.DoesNotExist, ValueError) as e:
            return JsonResponse({'error': str(e)}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def AttritionReport(request):
    """
    Generates an attrition report for the company.

    GET: Retrieves the attrition report data.
    """
    company_id = request.session.get('c_id')
    if not company_id:
        return JsonResponse({'error': 'Company ID not found in session'}, status=401)

    if request.method == 'GET':
        try:
            # Calculate attrition rate for the past 12 months
            today = datetime.now()
            start_date = today - timedelta(days=365)  # Past 12 months

            # Get all employees who joined within the past year
            joined_employees = Job_info.objects.filter(
                emp_emailid__d_id__c_id=company_id, 
                start_date__gte=start_date
            ).annotate(
                join_month=TruncMonth('start_date')
            )

            # Get employees who resigned within the past year
            resigned_employees = Resignation.objects.filter(
                emp_emailid__d_id__c_id=company_id,
                submit_date__gte=start_date
            ).annotate(
                resignation_month=TruncMonth('submit_date')
            )

            # Aggregate data by month
            attrition_data = joined_employees.values('join_month').annotate(
                joined_count=Count('emp_emailid'),
                resigned_count=Count('emp_emailid__R_Id', filter=Q(emp_emailid__R_Id__in=resigned_employees)),
            ).order_by('join_month')

            # Calculate attrition rate and format the data for response
            report_data = []
            for data in attrition_data:
                month_name = calendar.month_name[data['join_month'].month]
                attrition_rate = (data['resigned_count'] / data['joined_count']) * 100 if data['joined_count'] > 0 else 0
                report_data.append({
                    'month': month_name,
                    'joined_count': data['joined_count'],
                    'resigned_count': data['resigned_count'],
                    'attrition_rate': round(attrition_rate, 2)
                })

            return JsonResponse({'attrition_report': report_data}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
@csrf_exempt
@require_http_methods(["GET"])
def get_performance_data(request, year):
    # Fetch the user ID from the session
    user_email = request.session.get('user_id')

    if not user_email:
        return JsonResponse({'error': 'User not authenticated'}, status=401)

    # Fetch the employee record using the email from the session
    try:
        logged_in_employee = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'User not authenticated'}, status=401)
    
    # Filter tasks based on the employee and year, assuming sop_id has a field like `sdate`
    tasks = Tasks.objects.filter(emp_emailid=logged_in_employee, sop_id__sdate__year=year)
    
    # Calculate task statistics
    total_sop_assigned = tasks.count()
    total_sop_accomplished = tasks.filter(status=True).count()
    total_kra_assigned = tasks.filter(kra_id__isnull=False).count()
    total_kra_accomplished = tasks.filter(kra_id__isnull=False, status=True).count()
    total_jd_assigned = tasks.filter(job_desc_id__isnull=False).count()
    total_jd_accomplished = tasks.filter(job_desc_id__isnull=False, status=True).count()
    total_training_assigned = tasks.filter(tid__isnull=False).count()
    total_training_accomplished = tasks.filter(tid__isnull=False, status=True).count()
    
    # Prepare the response data
    data = {
        'totalSopAssigned': total_sop_assigned,
        'totalSopAccomplished': total_sop_accomplished,
        'totalKraAssigned': total_kra_assigned,
        'totalKraAccomplished': total_kra_accomplished,
        'totalJdAssigned': total_jd_assigned,
        'totalJdAccomplished': total_jd_accomplished,
        'totalTrainingAssigned': total_training_assigned,
        'totalTrainingAccomplished': total_training_accomplished,
    }

    return JsonResponse(data, status=200)

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def BenefitsCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='Benefits'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url
        })

    context = {
        'cases': case_data
    }

    return JsonResponse({'cases': case_data})


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def TravelExpenseCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='TravelCases'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url
        })

    context = {
        'cases': case_data
    }

    return JsonResponse({'cases': case_data})


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CompensationPayrollCases(request):
    def get_timeago(ptime):
        estimate_time = timezone.now() - ptime
        if estimate_time.days < 1:
            return 'Today'

        condition = {
            365: 'year',
            30: 'month',
            1: 'day'
        }

        for secs, unit in condition.items():
            delta = estimate_time.days // secs
            if delta > 0:
                return f"about {delta} {unit}{'s' if delta > 1 else ''} ago"

    user_id = request.session.get('user_id')
    company_id = request.session.get('c_id')

    if not user_id or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    cases = Case.objects.filter(
        created_by__d_id__c_id=company_id,
        case_type='CompensationPayroll'
    ).select_related('created_by', 'assigned_to')

    case_data = []
    for case in cases:
        timeago = get_timeago(case.case_date)
        assigned_to = case.assigned_to.emp_name if case.assigned_to else "Not Assigned"
        created_by = case.created_by

        case_data.append({
            'case_id': case.case_id,
            'create_for': case.create_for,
            'case_title': case.case_title,
            'case_type': case.case_type,
            'case_date': timeago,
            'assigned_to': assigned_to,
            'case_status': case.case_status,
            'created_by': created_by.emp_name,
            'created_by_profile': created_by.emp_profile.url if created_by.emp_profile else None
        })

    return JsonResponse({'cases': case_data})

from django.contrib.auth import get_user_model
from django.db import transaction
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import Company, Department, Employee

@csrf_exempt
def RegisterHR(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Extract HR details
            username = data.get('username')
            email = data.get('email')
            password = data.get('password')
            first_name = data.get('first_name')
            last_name = data.get('last_name')
            phone = data.get('phone')
            skills = data.get('skills', '')

            # Extract company details
            company_name = data.get('company_name')
            company_address = data.get('company_address')
            company_phone = data.get('company_phone')

            # Validate required fields
            if not all([username, email, password, first_name, last_name, phone, company_name, company_address, company_phone]):
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            with transaction.atomic():
                # Create or get the company
                company, created = Company.objects.get_or_create(
                    c_name=company_name,
                    defaults={
                        'c_addr': company_address,
                        'c_phone': company_phone
                    }
                )

                # Create or get the HR department
                hr_department, _ = Department.objects.get_or_create(
                    d_name='Human Resources',
                    c_id=company
                )

                # Create the HR employee
                hr_employee = Employee.objects.create(
                    emp_name=f"{first_name} {last_name}",
                    emp_emailid=email,
                    emp_phone=phone,
                    emp_pwd=password,  # Note: You should hash this password
                    emp_role='HR',
                    emp_skills=skills,
                    d_id=hr_department
                )

            return JsonResponse({
                'message': 'HR registration successful',
                'hr_id': hr_employee.emp_emailid,
                'company_id': company.c_id
            }, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except IntegrityError as e:
            return JsonResponse({'error': f'Database integrity error: {str(e)}'}, status=409)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
    
    
    
@csrf_exempt
@require_http_methods(["GET", "POST"])
def Resign(request):
    if request.method == "GET":
        try:
            resignations = Resignation.objects.all().values()
            return JsonResponse(list(resignations), safe=False)
        except Exception as e:
            print("Error fetching resignations:", str(e))  # Debugging statement
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            print("Received data:", data)  # Debugging statement

            # Validate required fields
            required_fields = ["emp_emailid", "exp_leave", "leave_reason", "leave_date"]
            for field in required_fields:
                if field not in data:
                    print(f"Missing required field: {field}")  # Debugging statement
                    return JsonResponse({"error": f"Missing required field: {field}"}, status=400)

            # Convert date fields to datetime objects
            try:
                exp_leave = datetime.strptime(data.get("exp_leave"), "%Y-%m-%d").date()
                leave_date = datetime.strptime(data.get("leave_date"), "%Y-%m-%d").date()
                settle_date = datetime.strptime(data.get("settle_date"), "%Y-%m-%d").date() if data.get("settle_date") else None
                exit_interview = datetime.strptime(data.get("exit_interview"), "%Y-%m-%d").date() if data.get("exit_interview") else None
                last_working = datetime.strptime(data.get("last_working"), "%Y-%m-%d").date() if data.get("last_working") else None
            except ValueError as e:
                print("Date conversion error:", str(e))  # Debugging statement
                return JsonResponse({"error": "Invalid date format"}, status=400)

            resignation, created = Resignation.objects.update_or_create(
                emp_emailid=data.get("emp_emailid"),
                defaults={
                    'exp_leave': exp_leave,
                    'leave_reason': data.get("leave_reason"),
                    'leave_reason_2': data.get("leave_reason_2"),
                    'leave_reason_3': data.get("leave_reason_3"),
                    'leave_date': leave_date,
                    'settle_date': settle_date,
                    'exit_interview': exit_interview,
                    'last_working': last_working,
                    'status': data.get("status"),
                    'approved_by': data.get("approved_by"),
                    'notice_per': data.get("notice_per"),
                }
            )

            response = {'message': 'Resignation created successfully'} if created else {'message': 'Resignation updated successfully'}
            return JsonResponse({'status': 1, 'message': response})

        except json.JSONDecodeError:
            print("Invalid JSON data")  # Debugging statement
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            print("Error:", str(e))  # Debugging statement
            return JsonResponse({'error': f'An unexpected error occurred: {e}'}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    

@csrf_exempt
def social_submit_feedback_get(request):
    if request.method == 'GET':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            user = Employee.objects.select_related('d_id__c_id').get(emp_emailid=user_email)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)

        feedbacks = Feedback.objects.filter(emp_emailid=user).values(
            'fid', 'reason', 'from_email', 'date'
        )

        feedback_list = []
        for fb in feedbacks:
            reactions = Reaction.objects.filter(feedback_id=fb['fid']).values('user_email', 'emoji', 'date')
            feedback_list.append({
                'feedback_id': fb['fid'],
                'content': fb['reason'],
                'sender': fb['from_email'],
                'created_at': fb['date'],
                'reactions': list(reactions)  # Include reactions for each feedback
            })

        return JsonResponse(feedback_list, safe=False)

    return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)



@csrf_exempt
def social_add_reaction(request, feedback_id):  # Add feedback_id here
    if request.method == 'POST':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            data = json.loads(request.body)
            emoji = data.get('emoji')

            if not emoji:
                return JsonResponse({'error': 'Emoji is required'}, status=400)

            # Check if feedback exists
            try:
                feedback = Feedback.objects.get(fid=feedback_id)
            except Feedback.DoesNotExist:
                return JsonResponse({'error': 'Feedback not found'}, status=404)

            # Add the reaction
            Reaction.objects.create(
                feedback=feedback,
                user_email=user_email,
                emoji=emoji
            )

            return JsonResponse({'message': 'Reaction added successfully'}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)





@csrf_exempt
def social_submit_feedback_post(request):
    if request.method == 'POST':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        # Get the authenticated user's company
        try:
            user = Employee.objects.select_related('d_id__c_id').get(emp_emailid=user_email)
            user_company = user.d_id.c_id
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)

        try:
            data = json.loads(request.body)
            emp_emailid = data.get('emp_emailid')
            skill = data.get('skill')
            from_email = user_email
            reason = data.get('reason')

            if not emp_emailid or not skill or not from_email or not reason:
                return JsonResponse({'error': 'All fields are required'}, status=400)

            # Ensure the employee exists and belongs to the same company
            try:
                employee = Employee.objects.select_related('d_id__c_id').get(emp_emailid=emp_emailid, d_id__c_id=user_company)
            except Employee.DoesNotExist:
                return JsonResponse({'error': 'Employee does not exist or does not belong to your company'}, status=404)

            # Save feedback, using the correct field for the foreign key reference
            feedback = Feedback.objects.create(
                emp_emailid_id=employee.emp_emailid,  # Use emp_emailid_id instead of emp_emailid
                skill=skill,
                from_email=from_email,
                reason=reason
            )

            # Retrieve feedbacks only for employees in the same company
            employee_emails = Employee.objects.filter(d_id__c_id=user_company).values_list('emp_emailid', flat=True)
            feedbacks = Feedback.objects.filter(emp_emailid__in=employee_emails).values(
                'emp_emailid__emp_emailid', 'skill', 'from_email', 'reason', 'date'
            )

            feedback_list = list(feedbacks)

            return JsonResponse({'message': 'Feedback submitted successfully', 'feedbacks': feedback_list}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except TypeError:
            return JsonResponse({'error': 'TypeError occurred'}, status=500)
        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)
    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)




@csrf_exempt
def employee_view(request):
    if request.method == 'GET':
        # Check if user is authenticated
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        # Get the authenticated user's company
        try:
            user = Employee.objects.select_related('d_id__c_id').get(emp_emailid=user_email)
            user_company = user.d_id.c_id
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)

        # Retrieve all employee details for employees in the same company
        employees = Employee.objects.filter(d_id__c_id=user_company).values(
            'emp_emailid', 'd_id'
        )
        employee_list = [{'id': emp['d_id'], 'name': emp['emp_emailid']} for emp in employees]
        
        return JsonResponse(employee_list, safe=False)

    else:
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)





@csrf_exempt
def allquiz(request):
    if request.method == 'GET':
        # Check if user is authenticated
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            # Retrieve the user from the Employee model
            user = Employee.objects.get(emp_emailid=user_email)

            # Get the user's company through the department
            company = user.d_id.c_id

            # Get the list of courses that belong to the user's company
            company_courses = Courses.objects.filter(c_id=company).values_list('course_title', flat=True)

            # Fetch only quizzes that are related to courses in the user's company
            quizzes = Quiz.objects.filter(course_title__in=company_courses)
            quiz_list = quizzes.values('id', 'title', 'course_title', 'total', 'total_marks', 'passing', 'time', 'status')

            return JsonResponse(list(quiz_list), safe=False)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User does not exist'}, status=404)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    
# Configure logging
logging.basicConfig(level=logging.DEBUG)

@csrf_exempt
def markattendance(request):
    if request.method == 'POST':
        try:
            # Parse request body
            data = json.loads(request.body)
            emp_email = data.get('email')
            latitude = data.get('latitude')
            longitude = data.get('longitude')
            log_type = data.get('log_type')
            employee_ip = request.META.get('REMOTE_ADDR')

            logging.debug(f"Employee IP: {employee_ip}")

            # Check for missing data
            if not emp_email or not latitude or not longitude:
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            # Fetch employee details
            try:
                employee = Employee.objects.get(emp_emailid=emp_email)
                logging.debug(f"Employee found: {employee}")
            except Employee.DoesNotExist:
                logging.error(f"Employee with email {emp_email} not found")
                return JsonResponse({'error': 'Employee not found'}, status=404)

            # Proceed with marking attendance
            Attendance.objects.create(
                emp_emailid=employee,  # Use the correct field name
                user_ip=employee_ip,  # Use the correct field name
                log_type=log_type,
                latitude=latitude,
                longitude=longitude,
                datetime_log=timezone.now().isoformat(),  # Use current datetime in ISO format
                date_updated=timezone.now()
            )
            logging.debug("Attendance marked successfully")
            return JsonResponse({'message': 'Attendance marked successfully'}, status=201)

        except Exception as e:
            logging.error(f"Error marking attendance: {str(e)}")
            logging.error(traceback.format_exc())  # Log the full traceback for debugging
            return JsonResponse({'error': 'Internal Server Error'}, status=500)

    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)





#Settings views
@csrf_exempt
@require_http_methods(["POST"])
def update_settings_account(request):
    try:
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'Authentication required'}, status=401)

        new_name = request.POST.get('name')
        new_phone = request.POST.get('phone')
        new_skills = request.POST.get('skills')
        new_photo = request.FILES.get('photo')  # Get the uploaded file

        if not any([new_name, new_phone, new_skills, new_photo]):
            return JsonResponse({'error': 'No data provided for update'}, status=400)

        try:
            employee = Employee.objects.get(emp_emailid=user_email)

            should_save = False
            if new_name:
                employee.emp_name = new_name
                should_save = True
            if new_phone:
                employee.emp_phone = new_phone
                should_save = True
            if new_skills:
                employee.emp_skills = new_skills
                should_save = True
            if new_photo:
                employee.emp_photo = new_photo  # Assuming you have a field named 'emp_photo'
                should_save = True

            if should_save:
                employee.save()
                return JsonResponse({'message': 'Account details updated successfully'}, status=200)
            else:
                return JsonResponse({'message': 'No changes made'}, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        return JsonResponse({'error': 'An internal error occurred'}, status=500)



@csrf_exempt
@require_http_methods(["POST"])
def update_settings_password(request):
    try:
        # Check if the user is logged in
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'Authentication required'}, status=401)

        # Extract form data
        old_password = request.POST.get('oldPassword')
        new_password = request.POST.get('newPassword')
        new_confirm_password = request.POST.get('confirmPassword')

        # Validate password fields
        if not old_password:
            return JsonResponse({'error': 'Old password is required to change password'}, status=400)
        if not new_password or not new_confirm_password:
            return JsonResponse({'error': 'New password and confirm password are required'}, status=400)

        try:
            # Fetch the employee object using the email from the session
            employee = Employee.objects.get(emp_emailid=user_email)

            if employee.emp_pwd == old_password:  # Check if the old password matches
                if new_password == new_confirm_password:
                    employee.emp_pwd = new_password  # Update to the new password (plain text storage)
                    employee.save()
                    return JsonResponse({'message': 'Password updated successfully'}, status=200)
                else:
                    return JsonResponse({'error': 'Passwords do not match'}, status=400)
            else:
                return JsonResponse({'error': 'Old password is incorrect'}, status=400)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        return JsonResponse({'error': 'An internal error occurred'}, status=500)
    

    

def generate_otp():
    return ''.join(random.choices(string.digits, k=6))

@csrf_exempt
def send_otp(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')

        try:
            # Verify if the email exists in the Employee model
            user = Employee.objects.get(emp_emailid=email)
            
            # Generate OTP
            otp = generate_otp()

            # Check if an OTP record already exists for the given email
            otp_record, created = OTPVerification.objects.update_or_create(
                emp_emailid=email,
                defaults={
                    'otp': otp,
                    'created_at': timezone.now()
                    }
                
            )
            
            # Send the OTP via email
            send_mail(
                'Your OTP for Password Reset',
                f'Your OTP is: {otp}',
                settings.EMAIL_HOST_USER,
                [email],
                fail_silently=False,
            )
            
            return JsonResponse({'message': 'OTP sent successfully!'}, status=200)
        
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Email not found'}, status=404)
    
    return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)

@csrf_exempt
def verify_otp(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')
        otp = data.get('otp')

        try:
            # Fetch the OTPVerification object
            otp_record = OTPVerification.objects.get(emp_emailid=email, otp=otp)
            
            # Get the current time
            now = timezone.now()
            created_at = otp_record.created_at
            
            # Set the OTP expiration time (e.g., 10 minutes from creation)
            expiration_time = created_at + timedelta(minutes=10)
            
            # Debugging information
            print(f"Current Time: {now}")
            print(f"Created At: {created_at}")
            print(f"Expiration Time: {expiration_time}")

            # Check if the OTP is still valid
            if now > expiration_time:
                return JsonResponse({'error': 'OTP expired'}, status=400)
            
            # OTP is valid; allow the user to reset their password
            return JsonResponse({'message': 'OTP verified successfully'}, status=200)
        
        except OTPVerification.DoesNotExist:
            return JsonResponse({'error': 'Invalid OTP'}, status=400)
    
    return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


@csrf_exempt
def reset_password(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')
        password = data.get('password')  # Update this line

        try:
            user = Employee.objects.get(emp_emailid=email)

            if user.emp_pwd == password:
                return JsonResponse({'error': 'New password cannot be the same as the current password'}, status=400)

            user.emp_pwd = password
            user.save()
            return JsonResponse({'message': 'Password reset successful'}, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)

    return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)

def employee_details(request, emp_emailid):
    if request.method == 'GET':
        # Retrieve the employee based on their email ID
        employee = get_object_or_404(Employee, emp_emailid=emp_emailid)
        
        # Serialize the data to send as JSON
        employee_data = {
            "emp_name": employee.emp_name,
            "emp_emailid": employee.emp_emailid,
            "emp_skills": employee.emp_skills,
            "emp_role": employee.emp_role,
            "emp_phone": employee.emp_phone,
            "emp_profile": employee.emp_profile.url if employee.emp_profile else None,
            "add_date": employee.add_date,
            "pay_sts": employee.pay_sts,
            "status": employee.Status,
            "department_id": employee.d_id.d_id,  # ForeignKey field
            "department_name": employee.d_id.d_name,  # Fetching department name
        }

        return JsonResponse({"employee": employee_data}, status=200)

# @csrf_exempt
# with hashing
# def reset_password(request):
#     if request.method == 'POST':
#         data = json.loads(request.body)
#         email = data.get('email')
#         new_password = data.get('new_password')

#         try:
#             user = Employee.objects.get(emp_emailid=email)
#             user.emp_pwd = make_password(new_password)
#             user.save()
#             return JsonResponse({'message': 'Password reset successful'}, status=200)
        
#         except Employee.DoesNotExist:
#             return JsonResponse({'error': 'User not found'}, status=404)
    
#     return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


# Calender
@csrf_exempt
def event_handler(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        # Add new event
        data = json.loads(request.body)
        event = Events.objects.create(
            evt_type=data.get('evt_type'),
            evt_start=data.get('evt_start'),
            evt_end=data.get('evt_end'),
            evt_text=data.get('evt_text'),
            evt_color=data.get('evt_color'),
            status=data.get('status', False),
            emp_emailid=emp_emailid
        )
        return JsonResponse({
            'evt_id': event.evt_id,
            'evt_type': event.evt_type,
            'evt_start': event.evt_start,
            'evt_end': event.evt_end,
            'evt_text': event.evt_text,
            'evt_color': event.evt_color,
            'status': event.status
        }, status=201)


    elif request.method == 'PUT':
        # Update existing event
        data = json.loads(request.body)
        evt_id = data.get('evt_id')
        event = get_object_or_404(Events, pk=evt_id, emp_emailid=emp_emailid)

        event.evt_type = data.get('evt_type', event.evt_type)
        event.evt_start = data.get('evt_start', event.evt_start)
        event.evt_end = data.get('evt_end', event.evt_end)
        event.evt_text = data.get('evt_text', event.evt_text)
        event.evt_color = data.get('evt_color', event.evt_color)
        event.status = data.get('status', event.status)
        event.save()

        return JsonResponse({
            'evt_id': event.evt_id,
            'evt_type': event.evt_type,
            'evt_start': event.evt_start,
            'evt_end': event.evt_end,
            'evt_text': event.evt_text,
            'evt_color': event.evt_color,
            'status': event.status
        }, status=200)

    elif request.method == 'DELETE':
        # Delete an event
        data = json.loads(request.body)
        evt_id = data.get('evt_id')
        event = get_object_or_404(Events, pk=evt_id, emp_emailid=emp_emailid)
        event.delete()
        return HttpResponse(status=204)

    elif request.method == 'GET':
        # Retrieve all events for the logged-in user
        events = Events.objects.filter(emp_emailid=emp_emailid)
        events_list = list(events.values(
            'evt_id', 'evt_type', 'evt_start', 'evt_end', 'evt_text', 'evt_color', 'status'
        ))
        return JsonResponse(events_list, safe=False, status=200)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


# IT Declaration
# Utility function to get the employee object
def get_employee(email):
    try:
        return Employee.objects.get(emp_emailid=email)
    except Employee.DoesNotExist:
        return None

@csrf_exempt
def investment_80C(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)
        
        if employee:
            ITDeclaration80c.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Life Insurance Premium'),
                Investment1_Amount=data.get('amount', 0),
            )
            return JsonResponse({"message": "Investment 80C data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)

@csrf_exempt
def investment_80D(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)
        
        if employee:
            Itdeclaration80d_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Medi claim Policy for Self,Spouse,Children-80D'),
                Investment1_Amount=data.get('amount', 0),
                emp_emailid=employee,
            )
            return JsonResponse({"message": "Investment 80D data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)

@csrf_exempt
def other_investments(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)
        
        if employee:
            Itdeclaration_oie_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Additional Exemptions on voluntary NPS'),
                Investment1_Amount=data.get('amount', 0),
                emp_emailid=employee,
            )
            return JsonResponse({"message": "Other investments data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)

@csrf_exempt
def other_income(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)
        
        if employee:
            Itdeclaration_osi_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('income', 'Income from other sources'),
                Investment1_Amount=data.get('amount', 0),
                Investment2_Amount=data.get('savingsInterest', 0),
                Investment3_Amount=data.get('fixedDepositInterest', 0),
                Investment4_Amount=data.get('nsCertificatesInterest', 0),
                emp_emailid=employee,
            )
            return JsonResponse({"message": "Other income data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)
        

# TodoList
# @csrf_exempt
# def task_list(request):
#     if request.method == 'GET':
#         try:
#             # Fetch all tasks from the database
#             tasks = Todotasks.objects.all().values('id', 'title', 'completed')  # Adjust fields as necessary
#             # Convert the queryset to a list of dictionaries
#             task_list = list(tasks)
#             return JsonResponse(task_list, safe=False)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

#     elif request.method == 'POST':
#         try:
#             # Parse the incoming data from the request
#             data = json.loads(request.body)
#             title = data.get('text', '')
#             completed = data.get('isCompleted', False)

#             # Create a new task
#             new_task = Todotasks1(title=title, completed=completed)
#             new_task.save()

#             return JsonResponse({'message': 'Task created successfully'}, status=201)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)

# @csrf_exempt
# def task_delete(request, task_id):
#     if request.method == 'DELETE':
#         try:
#             # Fetch the task by its ID
#             task = Todotasks1.objects.get(id=task_id)
#             task.delete()
#             return JsonResponse({'message': 'Task deleted successfully'}, status=200)
#         except Todotasks1.DoesNotExist:
#             return JsonResponse({'error': 'Task not found'}, status=404)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)
#     else:
#         return JsonResponse({'error': 'Invalid request method'}, status=405)
    

# @csrf_exempt
# def task_update(request, task_id):
#     if request.method == 'PUT':
#         try:
#             # Fetch the task by its ID
#             task = Todotasks1.objects.get(id=task_id)
            
#             # Parse the request body
#             data = json.loads(request.body)
            
#             # Update the task attributes
#             task.title = data.get('title', task.title)
#             task.completed = data.get('completed', task.completed)
#             task.save()
            
#             return JsonResponse({
#                 'message': 'Task updated successfully',
#                 'task': {
#                     'id': task.id,
#                     'title': task.title,
#                     'completed': task.completed
#                 }
#             }, status=200)
#         except Todotasks1.DoesNotExist:
#             return JsonResponse({'error': 'Task not found'}, status=404)
#         except json.JSONDecodeError:
#             return JsonResponse({'error': 'Invalid JSON'}, status=400)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)
#     else:
#         return JsonResponse({'error': 'Invalid request method'}, status=405)
    


def attrition_year_data(request):
    data = list(AttritionYearData.objects.values())  # Get all objects and convert to list of dicts
    return JsonResponse(data, safe=False)

def attrition_department_data(request):
    data = list(AttritionDepartmentData.objects.values())  # Get all objects and convert to list of dicts
    return JsonResponse(data, safe=False)

def attrition_reason_data(request):
    data = list(AttritionReasonData.objects.values())  # Get all objects and convert to list of dicts
    return JsonResponse(data, safe=False)

def attrition_gender_data(request):
    data = list(AttritionGenderData.objects.values())  # Get all objects and convert to list of dicts
    return JsonResponse(data, safe=False)

def attrition_choice_data(request):
    data = list(AttritionChoiceData.objects.values())  # Get all objects and convert to list of dicts
    return JsonResponse(data, safe=False)

@csrf_exempt
def poi_list_create(request,emp_emailid):
    if request.method == 'GET':
        try:
            poi_record = Poifiles_new1.objects.filter(Emp_id__emp_emailid=emp_emailid).values(
                'Investment_1', 'actualAmount_80C', 'Status1',
                'Investment_2', 'actualAmount_80D', 'Status2',
                'Investment_3', 'OIE_actualAmount', 'Status3',
                'Investment_4', 'OSI_actualAmount', 'Status4'
            ).first()
            if poi_record:
                return JsonResponse({'poi_record': poi_record}, safe=False)
            return JsonResponse({'error': 'No records found for the employee'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    # elif request.method == 'POST':
    #     category = request.POST.get('category')
    #     declared_amount = request.POST.get('declaredAmount')
    #     actual_amount = request.POST.get('actualAmount')
    #     file = request.FILES.get('file')

    #     poi = POI(category=category, declared_amount=declared_amount, actual_amount=actual_amount, file=file)
    #     poi.save()

    #     return JsonResponse({
    #         'id': poi.id,
    #         'category': poi.category,
    #         'declaredAmount': poi.declared_amount,
    #         'actualAmount': poi.actual_amount,
    #         'fileUrl': poi.file.url if poi.file else None,
    #     }, status=201)
        
@csrf_exempt
def toggle_investment_status(request, emp_id, investment_field):
    if request.method == 'POST':
        try:
            poi_record = get_object_or_404(Poifiles_new1, Emp_id=emp_id)
            # Map the investment field to the status fields in the model
            status_field_map = {
                'Investment_1': 'Status1',
                'Investment_2': 'Status2',
                'Investment_3': 'Status3',
                'Investment_4': 'Status4'
            }
            status_field = status_field_map.get(investment_field)
            if status_field:
                current_status = getattr(poi_record, status_field)
                new_status = 'Verified' if current_status == 'Pending' else 'Pending'
                setattr(poi_record, status_field, new_status)
                poi_record.save()
                return JsonResponse({'status': 'success', 'new_status': new_status})
            return JsonResponse({'error': 'Invalid investment field'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def poi_delete(request, poi_id):
    if request.method == 'DELETE':
        poi = get_object_or_404(POI, pk=poi_id)
        if poi.file:
            default_storage.delete(poi.file.path)
        poi.delete()
        return JsonResponse({'message': 'POI deleted successfully.'}, status=200)

@csrf_exempt
def Settings(request):
    company_id = request.session.get('c_id')
    # print(company_id)
    user_id = request.session.get('emp_emailid')
    # print(user_id)
    if not company_id:
        return JsonResponse({'error': 'ID not found in session'}, status=401)
    if request.method == 'GET':
        employee_details = Employee.objects.filter(emp_emailid=user_id)
        employee_list = []
        for emp in employee_details:
            employee_list.append({
                'emp_name': emp.emp_name,
                'emp_emailid': emp.emp_emailid,
                'emp_phone': emp.emp_phone,
                'emp_profile': emp.emp_profile.url,
                'emp_skills': emp.emp_skills,
            })
        return JsonResponse(employee_list, safe=False)
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            emp_emailid = data.get('emp_emailid')
            if not emp_emailid:
                return JsonResponse({'error': 'Employee email ID is required'}, status=400)
            try:
                changes_made = False
                employee = Employee.objects.get(emp_emailid=emp_emailid)
                # Update fields only if new data is provided
                emp_emailid = data.get('emp_emailid')
                if emp_emailid and emp_emailid != employee.emp_emailid:
                    employee.emp_emailid = emp_emailid
                    changes_made = True
                emp_name = data.get('emp_name')
                if emp_name and emp_name != employee.emp_name:
                    employee.emp_name = emp_name
                    changes_made = True
                emp_phone = data.get('emp_phone')
                if emp_phone and emp_phone != employee.emp_phone:
                    employee.emp_phone = emp_phone
                    changes_made = True
                emp_profile = data.get('emp_profile')
                if emp_profile and emp_profile != str(employee.emp_profile):
                    employee.emp_profile = emp_profile
                    changes_made = True
                emp_skills = data.get('emp_skills')
                if emp_skills and emp_skills != employee.emp_skills:
                    employee.emp_skills = emp_skills
                    changes_made = True
                # Save only if changes were made
                if changes_made:
                    employee.save()
                return JsonResponse({'status': 'Employee details updated successfully'}, status=200)
            except Employee.DoesNotExist:
                return JsonResponse({'error': 'Employee not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=405)
            

    






# addded dipayan
@csrf_exempt
@role_required(['Manager', 'Super Admin','Employee'])
def JDForm(request):
    company_id = request.session.get('c_id')
    user_name = request.session.get('emp_name')

    if not company_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

    if request.method == 'GET':
        try:
            # Fetch all employee email IDs
            employees = Employee.objects.values_list('emp_emailid', flat=True)
            return JsonResponse({'employee_email_ids': list(employees)}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            # Load the JSON data
            data = json.loads(request.body)

            # Extract the required fields
            jd_name = data.get('jd_name')
            responsibilities = data.get('responsiblities', [])  # Make sure to use the correct field name
            sdate = data.get('sdate')
            email_ids = data.get('email_ids', [])

            # Validate the input data
            if not jd_name or not responsibilities or not sdate or not email_ids:
                return JsonResponse({'error': 'All fields are required.'}, status=400)

            if not isinstance(responsibilities, list) or not isinstance(email_ids, list):
                return JsonResponse({'error': 'Responsibilities and email_ids should be arrays.'}, status=400)

            # Get or create the jid based on jd_name
            job_desc_entry = Job_desc.objects.filter(jd_name=jd_name).first()
            if job_desc_entry:
                jid = job_desc_entry.jid
            else:
                # Generate a new jid if jd_name is new
                max_jid = Job_desc.objects.aggregate(max_jid=Max('jid'))['max_jid']
                jid = (max_jid + 1) if max_jid is not None else 1

            # Create Job Description and Task entries for each employee
            for email_id in email_ids:
                try:
                    # Check if the employee exists
                    employee = Employee.objects.get(emp_emailid=email_id)
                    
                    # Create the Job Description entry
                    job_desc = Job_desc.objects.create(
                        jd_name=jd_name,
                        responsiblities=", ".join(responsibilities),  # Make sure the model field name matches
                        sdate=sdate,
                        email_id=email_id,
                        jid=jid
                    )
                    

                except Employee.DoesNotExist:
                    return JsonResponse({'error': f'Employee with email ID {email_id} does not exist'}, status=404)

            return JsonResponse({'status': 'JD assigned successfully.'}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ObjectDoesNotExist
from .models import Kra, Employee, Kra_table
import json
from datetime import datetime

@csrf_exempt
# # @role_required(['Manager', 'Super Admin', 'Employee'])
def KRAForm(request):
    if request.method == 'GET':
        try:
            employees = Employee.objects.values_list('emp_emailid', flat=True)
            return JsonResponse({'employee_email_ids': list(employees)}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            kras = data.get('kras', [])
            email_ids = data.get('email_ids', [])
            submission_date = data.get('submission_date')

            print(f"Received data: {data}")

            if not kras or not email_ids or not submission_date:
                return JsonResponse({'error': 'Incomplete data provided'}, status=400)

            submission_date = datetime.strptime(submission_date, "%Y-%m-%dT%H:%M:%S")

            for kra_data in kras:
                kra = kra_data.get('kra')
                weightage = int(kra_data.get('weightage', 0))
                kpi = kra_data.get('kpi')
                target = kra_data.get('target')
                ratings_scale = kra_data.get('ratings_scale')
                ratings = kra_data.get('ratings', 1)
                selfratings = kra_data.get('selfratings', 0)
                remarks = kra_data.get('remarks', '')
                status = kra_data.get('status', 0)
                measurement = kra_data.get('measurement', '0')
                kra_id = kra_data.get('kra_id')  # Optional field

                if not kra or not weightage or not kpi or not target or not ratings_scale:
                    return JsonResponse({'error': 'Incomplete KRA data'}, status=400)

                for email_id in email_ids:
                    if not Employee.objects.filter(emp_emailid=email_id).exists():
                        return JsonResponse({'error': f'Employee with email ID {email_id} does not exist'}, status=404)

                    # Handle kra_id (optional)
                    kra_table_entry = None
                    if kra_id:
                        try:
                            kra_table_entry = Kra_table.objects.get(pk=kra_id)
                        except ObjectDoesNotExist:
                            return JsonResponse({'error': f'kra_id {kra_id} does not exist'}, status=404)

                    # Create an entry in the Kra model
                    Kra.objects.create(
                        kra=kra,
                        weightage=weightage,
                        kpi=kpi,
                        target=target,
                        ratings_scale=ratings_scale,
                        ratings=ratings,
                        selfratings=selfratings,
                        remarks=remarks,
                        status=status,
                        email_id=email_id,
                        kra_id=kra_table_entry,  # Can be null
                        measurement=measurement,
                        submission_date=submission_date
                    )

            return JsonResponse({'message': 'KRA form submitted successfully!'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Kra, Kra_table

@csrf_exempt
def DeleteKRA(request, kra_no):
    if request.method == 'DELETE':
        try:
            # Fetch the KRA entry by kra_no
            kra = Kra.objects.get(kra_no=kra_no)
            # Delete the related entry in kra_table first
            Kra_table.objects.filter(kra_id=kra.kra_id_id).delete()
            # Delete the entry
            kra.delete()
            return JsonResponse({'message': 'KRA deleted successfully!'}, status=200)
        except Kra.DoesNotExist:
            return JsonResponse({'error': 'KRA not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=405)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import Kra

@csrf_exempt  # Only for testing; remove in production
def UpdateKRA(request, kra_no):
    if request.method == "PATCH":
        try:
            data = json.loads(request.body)
            kra = Kra.objects.get(kra_no=kra_no)

            new_status = data.get("status")  # Get new status from request
            if new_status not in [0, 1, 2, 3]:  # Validate status values
                return JsonResponse({"error": "Invalid status"}, status=400)

            kra.status = new_status
            kra.save()
            return JsonResponse({"message": "Status updated", "status": kra.status})

        except Kra.DoesNotExist:
            return JsonResponse({"error": "KRA not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    
    return JsonResponse({"error": "Invalid request"}, status=400)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def SOPForm(request):
    company_id = request.session.get('c_id')
    user_name = request.session.get('emp_name')
    if not company_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)
    
    if request.method == 'GET':
        try:
            # Get all departments from the database
            departments = Department.objects.all().values('d_id', 'd_name')
            department_list = list(departments)  # Convert the QuerySet to a list for JSON serialization

            return JsonResponse({'departments': department_list}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    elif request.method == 'POST':
        try:
            try:
                latest_sop = Sop.objects.latest('sop_id')
                sop_id = latest_sop.sop_id + 1
            except ObjectDoesNotExist:
                sop_id = 1  # Start at 1 if no records exist
            # Automatically generate sop_id
            latest_sop = Sop.objects.latest('sop_id')
            sop_id = latest_sop.sop_id + 1 if latest_sop else 1  # Increment sop_id or start at 1 if no records exist

            sop_type = request.POST.get('type')
            s_name = request.POST.get('s_name')
            sdate = request.POST.get('sdate')
            sop_file = request.FILES.get('sop_file')
            d_id = request.POST.get('d_id')

            # Debugging print statements
            # print(f"sop_id: {sop_id}, type: {sop_type}, s_name: {s_name}, sdate: {sdate}, sop_file: {sop_file}, d_id: {d_id}")
            
            # Save the SOP entry
            sop_entry = Sop(
                sop_id=sop_id,
                type=sop_type,
                s_name=s_name,
                sdate=sdate,
                sop_file=sop_file,
                d_id_id=d_id,
            )
            sop_entry.save()

            return JsonResponse({'status': 'SOP submitted successfully'}, status=201)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


#With excel file
# @csrf_exempt
# @role_required(['HR', 'Manager', 'Super Manager'])
# def bulkEmployeeregistration(request):
#     if request.method == 'POST':
#         file = request.FILES.get('employeeFile')
#         if not file:
#             return JsonResponse({'error': 'No file uploaded'}, status=400)

#         # Save the uploaded file temporarily
#         file_name = default_storage.save(file.name, file)
#         file_path = default_storage.path(file_name)

#         try:
#             # Load the workbook and select the active worksheet
#             workbook = openpyxl.load_workbook(file_path)
#             worksheet = workbook.active

#             # Assuming the first row contains headers
#             company_name = worksheet['A2'].value
#             company_address = worksheet['B2'].value
#             company_phone = worksheet['C2'].value

#             # Create the company entry
#             new_company = Company.objects.create(
#                 c_name=company_name,
#                 c_addr=company_address,
#                 c_phone=company_phone
#             )

#             # Create departments and store their IDs
#             department_names = set()
#             for row in worksheet.iter_rows(min_row=2, values_only=True):
#                 dept_name = row[3]
#                 department_names.add(dept_name)

#             department_map = {}
#             for dept_name in department_names:
#                 new_dept = Department.objects.create(
#                     d_name=dept_name,
#                     c_id=new_company
#                 )
#                 department_map[dept_name] = new_dept.pk  # Map department names to their IDs

#             # Iterate through the worksheet and create employees
#             for row in worksheet.iter_rows(min_row=2, values_only=True):
#                 emp_name = row[4]
#                 emp_email = row[5]
#                 emp_phone = row[6]
#                 emp_skills = row[7]
#                 emp_role = row[8]
#                 dept_name = row[3]  # The department name in the current row

#                 # Validate employee data
#                 if not emp_name or not emp_email or not emp_phone or not emp_skills or not dept_name:
#                     return JsonResponse({'error': 'Employee data is incomplete in the uploaded file'}, status=400)

#                 # Create employee record
#                 Employee.objects.create(
#                     emp_name=emp_name,
#                     emp_emailid=emp_email,
#                     emp_skills=emp_skills,
#                     emp_role=emp_role,
#                     emp_phone=emp_phone,
#                     d_id_id=department_map[dept_name]  # Use the mapped department ID
#                 )

#             return JsonResponse({'message': 'Employees uploaded successfully'}, status=201)
#         except Exception as e:
#             return JsonResponse({'error': f'Failed to process the file: {str(e)}'}, status=500)
#         finally:
#             # Clean up: delete the file after processing
#             default_storage.delete(file_name)

#     return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)
 

@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def bulkUploadEmployeeDetailsUpload(request):
    company_id = request.session.get('c_id')
    user_name = request.session.get('emp_name')
    if not company_id or not user_name:
        return JsonResponse({'error': 'Required session data not found'}, status=401)
    
    if request.method == 'POST':
        try:
            # Function to get or create an Employee object by email
            def get_or_create_employee(email):
                employee, created = Employee.objects.get_or_create(emp_emailid=email)
                return employee

            # A list to store errors
            errors = []

            # Handling personal details Excel file
            if 'personal_details' in request.FILES:
                personal_details_file = request.FILES['personal_details']
                df_personal = pd.read_excel(personal_details_file)

                for _, row in df_personal.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        personal_details, created = Personal_details.objects.update_or_create(
                            mail=employee,  # Use the correct ForeignKey field, which references Employee instance
                            defaults={
                                'first_name': row.get('first_name'),
                                'last_name': row.get('last_name'),
                                'Contact': row.get('Contact'),
                                'emergency_name': row.get('emergency_name'),
                                'emergency_contact': row.get('emergency_contact'),
                                'gender': row.get('gender'),
                                'birth_date': row.get('birth_date'),
                                'address': row.get('address'),
                                'city': row.get('city'),
                                'district': row.get('district'),
                                'post_code': row.get('post_code'),
                                'state': row.get('state'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            # Handling bank details Excel file
            if 'bank_details' in request.FILES:
                bank_details_file = request.FILES['bank_details']
                df_bank = pd.read_excel(bank_details_file)

                for _, row in df_bank.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        # Either update or create the bank details
                        bank_details, created = Bank_details.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'holder_name': row.get('holder_name'),
                                'bank_name': row.get('bank_name'),
                                'acc_no': row.get('acc_no'),
                                'branch': row.get('branch'),
                                'acc_type': row.get('acc_type'),
                                'ifsc': row.get('ifsc'),
                                'Pan_no': row.get('Pan_no'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            # Handling dependent details Excel file
            if 'dependent_details' in request.FILES:
                dependent_file = request.FILES['dependent_details']
                df_dependent = pd.read_excel(dependent_file)

                for _, row in df_dependent.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        # Either update or create the dependent details
                        dependent, created = Dependent.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'D_name': row.get('D_name'),
                                'D_gender': row.get('D_gender'),
                                'D_dob': row.get('D_dob'),
                                'D_relation': row.get('D_relation'),
                                'D_desc': row.get('D_desc'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            # Handling family details Excel file
            if 'family_details' in request.FILES:
                family_file = request.FILES['family_details']
                df_family = pd.read_excel(family_file)

                for _, row in df_family.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        # Either update or create the family details
                        family_details, created = Family_details.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'F_name': row.get('F_name'),
                                'F_gender': row.get('F_gender'),
                                'F_dob': row.get('F_dob'),
                                'F_contact': row.get('F_contact'),
                                'F_mail': row.get('F_mail'),
                                'F_relation': row.get('F_relation'),
                                'F_comment': row.get('F_comment'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            # Handling job info Excel file
            if 'job_info' in request.FILES:
                job_info_file = request.FILES['job_info']
                df_job = pd.read_excel(job_info_file)

                for _, row in df_job.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        job_info, created = Job_info.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'job_title': row.get('job_title'),
                                'department': row.get('department'),
                                'working_type': row.get('working_type'),
                                'start_date': row.get('start_date'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")
            else:
                return JsonResponse({'error': 'Job info file not uploaded'}, status=400)

            # Handling qualification details Excel file
            if 'qualification_details' in request.FILES:
                qualification_file = request.FILES['qualification_details']
                df_qualification = pd.read_excel(qualification_file)

                for _, row in df_qualification.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        # Either update or create the qualification details
                        qualification, created = Qualification.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'q_type': row.get('q_type'),
                                'q_degree': row.get('q_degree'),
                                'q_clg': row.get('q_clg'),
                                'q_uni': row.get('q_uni'),
                                'q_duration': row.get('q_duration'),
                                'q_yop': row.get('q_yop'),
                                'q_comment': row.get('q_comment'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            # Handling work experience Excel file
            if 'work_exp' in request.FILES:
                work_exp_file = request.FILES['work_exp']
                df_work_exp = pd.read_excel(work_exp_file)

                for _, row in df_work_exp.iterrows():
                    email = row.get('emp_emailid').strip().lower()
                    employee = get_or_create_employee(email)
                    try:
                        # Either update or create the work experience details
                        work_exp, created = Work_exp.objects.update_or_create(
                            emp_emailid=employee,  # Use the Employee instance
                            defaults={
                                'start_date': row.get('start_date'),
                                'end_date': row.get('end_date'),
                                'comp_name': row.get('comp_name'),
                                'comp_location': row.get('comp_location'),
                                'designation': row.get('designation'),
                                'gross_salary': row.get('gross_salary'),
                                'leave_reason': row.get('leave_reason'),
                            }
                        )
                    except Exception as e:
                        errors.append(f"{email}: {str(e)}")

            if errors:
                return JsonResponse({'status': 'Some records were not saved due to errors', 'errors': errors}, status=400)

            return JsonResponse({'status': 'Files uploaded and data saved successfully'}, status=201)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)



@csrf_exempt
def SopList(request, sop_id=None):
    # Get the logged-in user's email and role
    company_id = request.session.get('c_id')
    user_email = request.session.get('emp_emailid')
    if not user_email or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    # Get the employee object to determine their role
    try:
        employee = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)

    user_role = employee.emp_role.lower()

    if request.method == 'GET':
        if sop_id:
            # Fetch the specific SOP using the sop_id
            try:
                sop = Sop.objects.get(sop_id=sop_id)
            except Sop.DoesNotExist:
                return JsonResponse({'error': 'SOP not found'}, status=404)

            # Prepare the data for the specific SOP
            sop_data = {
                'sop_id': sop.sop_id,
                'type': sop.type,
                's_name': sop.s_name,
                'sdate': sop.sdate,
                'sop_file': sop.sop_file.url if sop.sop_file else None,
                'ratings': sop.ratings,
                'selfratings': sop.selfratings,
                'remarks': sop.remarks,
                'd_id': sop.d_id_id,
            }
            return JsonResponse(sop_data, status=200)

        else:
            # Fetch all SOPs from the database
            sops = Sop.objects.all()
            sop_data_list = []
            for sop in sops:
                sop_data = {
                    'sop_id': sop.sop_id,
                    'type': sop.type,
                    's_name': sop.s_name,
                    'sdate': sop.sdate,
                    'sop_file': sop.sop_file.url if sop.sop_file else None,
                    'ratings': sop.ratings,
                    'selfratings': sop.selfratings,
                    'remarks': sop.remarks,
                    'd_id': sop.d_id_id,
                }
                sop_data_list.append(sop_data)
            return JsonResponse(sop_data_list, safe=False, status=200)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Accept sop_id from the request body if not provided in the URL
            sop_id = sop_id or data.get('sop_id')
            
            if not sop_id:
                return JsonResponse({'error': 'SOP ID is required for updating'}, status=400)

            # Fetch the specific SOP using the sop_id
            try:
                sop = Sop.objects.get(sop_id=sop_id)
            except Sop.DoesNotExist:
                return JsonResponse({'error': 'SOP not found'}, status=404)

            # Check if the user is a manager
            if user_role == 'manager' or user_role == 'super admin':
                sop.ratings = data.get('ratings', sop.ratings)
                sop.remarks = data.get('remarks', sop.remarks)
            else:
                # Check if the user is attempting to update ratings or remarks without permission
                if 'ratings' in data or 'remarks' in data:
                    return JsonResponse({'error': 'You do not have permission to update ratings or remarks.'}, status=403)

            # Allow all employees (including managers) to update selfratings
            sop.selfratings = data.get('selfratings', sop.selfratings)

            # Save the updated SOP
            sop.save()

            return JsonResponse({'status': 'SOP details updated successfully.'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def JdList(request):
    company_id = request.session.get('c_id')
    user_email = request.session.get('emp_emailid')

    if not user_email or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    if request.method == 'GET':
        # Ensure the user is logged in
        if not request.session.get('user_id'):
            return JsonResponse({'error': 'User not logged in'}, status=401)
        
        # Get the logged-in employee's name
        emp_name = request.session.get('emp_name')
        
        try:
            # Retrieve the employee instance using the name
            employee = Employee.objects.get(emp_name=emp_name)
        except Employee.DoesNotExist:
            return JsonResponse({"error": "Employee not found"}, status=404)

        # Retrieve all job descriptions assigned to this employee
        job_descs = Job_desc.objects.filter(email_id=employee.emp_emailid)

        # Prepare the response data
        job_desc_list = []
        for jd in job_descs:
            job_desc_list.append({
                'jd_id':jd.job_desc_id,
                'jd_name': jd.jd_name,
                'responsibilities': jd.responsiblities,
                'sdate': jd.sdate,
                'ratings': jd.ratings,
                'selfratings': jd.selfratings,
                'remarks': jd.remarks,
                'status': jd.status,
                'email_id': jd.email_id,
            })

        return JsonResponse({"job_descriptions": job_desc_list}, status=200)
    

    if request.method == 'POST':
        # Ensure the user is logged in
        if not request.session.get('user_id'):
            return JsonResponse({'error': 'User not logged in'}, status=401)

        # Retrieve all job descriptions assigned to this employee
        data = json.loads(request.body)
        sdate = data.get('sdate')
        # Prepare the response data
        # Fetch only job descriptions for the logged-in user and the given date
        job_details_queryset = Job_desc.objects.filter(sdate=sdate, email_id=user_email)

        # If no job descriptions are found, return an empty response
        if not job_details_queryset.exists():
            return JsonResponse({"job_details": []}, status=200)
        
        job_details_list = []
        for job in job_details_queryset:
            job_details_list.append({
                'jd_id': job.job_desc_id,
                'jd_name': job.jd_name,
                'responsibilities': job.responsiblities,
                'sdate': job.sdate,
                'ratings': job.ratings,
                'selfratings': job.selfratings,
                'remarks': job.remarks,
                'status': job.status,
                'email_id': job.email_id,
            })

        # Return the list of job descriptions
        return JsonResponse({"job_details": job_details_list}, status=200)
    return JsonResponse({"error": "Invalid request method"}, status=400)


@csrf_exempt
def JdDetails(request):
    company_id = request.session.get('c_id')
    user_email = request.session.get('emp_emailid')
    if not user_email or not company_id:
        return JsonResponse({'error': 'User not logged in'}, status=401)
    
    try:
        employee = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)

    user_role = employee.emp_role.lower()

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get the jd_id from the request body
            jd_id = data.get('jd_id')
            
            if not jd_id:
                return JsonResponse({'error': 'Job Description ID is required for updating'}, status=400)

            # Fetch the specific Job_desc using the jd_id
            try:
                job_desc = Job_desc.objects.get(job_desc_id=jd_id)
            except Job_desc.DoesNotExist:
                return JsonResponse({'error': 'Job Description not found'}, status=404)

            # Check if the user is a manager
            if user_role in ['manager', 'super admin']:
                job_desc.ratings = data.get('ratings', job_desc.ratings)
                job_desc.remarks = data.get('remarks', job_desc.remarks)
            else:
                # Check if the user is attempting to update ratings or remarks without permission
                if 'ratings' in data or 'remarks' in data:
                    return JsonResponse({'error': 'You do not have permission to update ratings or remarks.'}, status=403)

            # Allow all employees (including managers) to update selfratings
            job_desc.selfratings = data.get('selfratings', job_desc.selfratings)

            # Save the updated Job_desc
            job_desc.save()

            return JsonResponse({'status': 'Job Description details updated successfully.'}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)



@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def CreateQuiz(request):
    if request.method == 'POST':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:

            user = Employee.objects.get(emp_emailid=user_email)

            if user.emp_role not in ['HR', 'Manager', 'Super Admin']:
                return JsonResponse({'error': 'You do not have permission to create a quiz'}, status=403)

            company = user.d_id.c_id

            data = json.loads(request.body)
            quiz_title = data.get('quizTitle')
            course_id = data.get('course')
            total_questions = data.get('totalQuestions')
            marks_for_correct_answer = data.get('marksOnRightAnswer')
            marks_for_wrong_answer = data.get('minusMarksOnWrongAnswer')
            total_marks = int(data.get('total_marks', 0))
            passing_marks = data.get('passingMarks')
            time_limit = data.get('timeLimit')

            # Extract questions and options from the request
            questions = data.get('questions')

            # Check if the course exists and belongs to the user's company
            try:
                course = Courses.objects.get(course_id=course_id, c_id=company)
            except Courses.DoesNotExist:
                return JsonResponse({'error': 'Course does not exist in your company'}, status=404)

            # Create the quiz record
            quiz = Quiz(
                eid=user,
                title=quiz_title,
                course_title=course.course_title,
                total=total_questions,
                correct=marks_for_correct_answer,
                wrong=marks_for_wrong_answer,
                passing=passing_marks,
                total_marks=total_marks,
                time=time_limit,
                date=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                status='active'
            )
            quiz.save()

            # Create question and options records
            for question_data in questions:
                question_text = question_data['text']
                correct_answer = question_data['correct_answer']
                options = question_data['options']

                question = Question(quiz=quiz, text=question_text, correct_answer=correct_answer)
                question.save()

                for option_data in options:
                    option = Option(
                        question=question,
                        text=option_data['text'],
                        is_correct=option_data['is_correct']
                    )
                    option.save()

            return JsonResponse({'message': 'Quiz created successfully'}, status=201)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User does not exist'}, status=404)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])  # Custom role decorator for authorization
def GetAllQuizzes(request):
    if request.method == 'GET':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            user = Employee.objects.get(emp_emailid=user_email)

            if user.emp_role not in ['HR', 'Manager', 'Super Admin','Employee']:
                return JsonResponse({'error': 'You do not have permission to view quizzes'}, status=403)

            # Get all quizzes belonging to the user's company
            quizzes = Quiz.objects.filter(eid_d_id_c_id=user.d_id.c_id).values(
                'id', 'title', 'course_title', 'correct', 'wrong', 'total_marks', 
                'passing', 'total', 'time', 'date', 'status'
            )

            # Check if quizzes exist
            if not quizzes.exists():
                return JsonResponse({'error': 'No quizzes found for your company'}, status=404)

            # Convert to list and format date
            quiz_list = list(quizzes)

            # Return all quizzes as JSON
            return JsonResponse(quiz_list, safe=False)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User does not exist'}, status=404)

        except Exception as e:
            return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
@role_required(['HR', 'Manager', 'Super Admin','Employee'])
def ToggleQuizStatus(request):
    if request.method == 'POST':
        user_email = request.session.get('user_id')
        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            user = Employee.objects.get(emp_emailid=user_email)

            if user.emp_role not in ['HR', 'Manager', 'Super Admin']:
                return JsonResponse({'error': 'You do not have permission to toggle quiz status'}, status=403)

            data = json.loads(request.body)
            quiz_id = data.get('quiz_id')
            new_status = data.get('status')  # Either 'active' or 'inactive'

            if new_status not in ['active', 'inactive']:
                return JsonResponse({'error': 'Invalid status value'}, status=400)

            # Find the quiz and ensure it belongs to the user's company
            try:
                quiz = Quiz.objects.get(id=quiz_id, eid_d_id_c_id=user.d_id.c_id)
            except Quiz.DoesNotExist:
                return JsonResponse({'error': 'Quiz not found or you do not have access'}, status=404)

            # Update quiz status
            quiz.status = new_status
            quiz.save()

            return JsonResponse({'message': f'Quiz status updated to {new_status}'}, status=200)

        except Employee.DoesNotExist:
            return JsonResponse({'error': 'User does not exist'}, status=404)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
def AttemptQuiz(request, quiz_id=None):
    user_email = request.session.get('user_id')
    if not user_email:
        return JsonResponse({'error': 'User not authenticated'}, status=403)

    # Fetch the employee and department
    user = get_object_or_404(Employee, emp_emailid=user_email)
    user_department = user.d_id  # Get the user's department
    company_id = user_department.c_id  # Get the company ID from the department

    # Handle GET request
    if request.method == 'GET':
        if quiz_id is None:
            # Get all quizzes that belong to the same company as the user
            quizzes = Quiz.objects.filter(eid_d_id_c_id=company_id,status="active").values(
                'id', 'title', 'course_title', 'correct', 'wrong', 'total_marks', 
                'passing', 'total', 'time', 'date', 'status'
            )

            # Return full details of each quiz
            return JsonResponse(list(quizzes), safe=False)
        else:
            # Get a specific quiz if quiz_id is provided and matches the company
            quiz = get_object_or_404(Quiz, id=quiz_id, eid_d_id_c_id=company_id)

            # Fetch the related questions and options for the quiz
            questions = []
            for question in Question.objects.filter(quiz=quiz):
                options = Option.objects.filter(question=question).values('id', 'text')
                questions.append({
                    'id': question.id,
                    'text': question.text,
                    'options': list(options)
                })

            # Store the start time in the session
            request.session['quiz_start_time'] = timezone.now().isoformat()

            response_data = {
                'quiz_id': quiz.id,
                'quiz_title': quiz.title,
                'time_limit': quiz.time,
                'questions': questions
            }
            return JsonResponse(response_data, safe=False)

        

    # Handle POST request for quiz submission
    elif request.method == 'POST':
        data = json.loads(request.body)
        chosen_options = data.get('chosen_options')  # {question_id: chosen_option_id}

        # Retrieve the start time from the session
        start_time_str = request.session.get('quiz_start_time')
        if not start_time_str:
            return JsonResponse({'error': 'Start time not found'}, status=400)

        start_time = timezone.datetime.fromisoformat(start_time_str)

        # Fetch the quiz details
        quiz = get_object_or_404(Quiz, id=quiz_id)

        # Check if the user has already attempted this quiz
        attempt = QuizAttempt.objects.filter(quiz=quiz, employee=user).first()
        if attempt:
            # Update existing attempt
            attempt.chosen_options = chosen_options
            attempt.score = 0  # Reset score to recalculate
            attempt.total_correct = 0  # Reset correct count
            attempt.total_wrong = 0  # Reset wrong count
        else:
            # Create a new attempt if none exists
            attempt = QuizAttempt(
                quiz=quiz,
                employee=user,
                chosen_options=chosen_options,
                score=0,
                total_correct=0,
                total_wrong=0
            )

        # Initialize score calculation
        score = 0
        total_correct = 0
        total_wrong = 0

        for question_id, option_id in chosen_options.items():
            try:
                question = Question.objects.get(id=question_id, quiz=quiz)
                chosen_option = Option.objects.get(id=option_id, question=question)

                if chosen_option.is_correct:
                    score += quiz.correct  # Add correct marks
                    total_correct += 1
                else:
                    score -= quiz.wrong  # Subtract wrong marks
                    total_wrong += 1
            except Question.DoesNotExist:
                return JsonResponse({'error': f'Invalid question ID: {question_id}'}, status=400)
            except Option.DoesNotExist:
                return JsonResponse({'error': f'Invalid option ID: {option_id}'}, status=400)

        # Calculate elapsed time
        end_time = timezone.now()
        time_taken = (end_time - start_time)  # Time taken as a timedelta

        # Save time taken in total seconds
        total_seconds = int(time_taken.total_seconds())
        tolerance_of_time = 60 # 1 Minute Time tolerance so that request response time cost not paid by user attempting quiz
        if total_seconds > (quiz.time * 60 + tolerance_of_time):  # Check if time limit exceeded (convert to seconds)
            return JsonResponse({'error': 'Time limit exceeded'}, status=400)

        # Update the attempt details with the new score, time, and question stats
        attempt.score = score
        attempt.time_taken = total_seconds  # Save time in seconds
        attempt.is_passed = score >= quiz.passing
        attempt.total_correct = total_correct  # Save total correct answers
        attempt.total_wrong = total_wrong  # Save total wrong answers
        attempt.total_unattempted = quiz.total - (total_correct + total_wrong)  # Save total wrong answers
        attempt.save()

        # Update quiz total marks (if needed) and related fields
        # quiz.total_marks = score  # Update the total marks for the quiz
        quiz.save()

        # Format the time taken for response
        formatted_time = f"{total_seconds // 3600:02}:{(total_seconds % 3600) // 60:02}:{total_seconds % 60:02}"

        # Return the result of the quiz
        response_data = {
            'message': 'Quiz submitted successfully',
            'score': float(score),
            'total_correct': total_correct,
            'total_wrong': total_wrong,
            'is_passed': attempt.is_passed,
            'time_taken': formatted_time,  # Return formatted time
            'max_time': quiz.time,
        }

        return JsonResponse(response_data, status=201)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)

@csrf_exempt
def QuizResults(request, quiz_id=None):
    # Check if the request method is GET
    if request.method == 'GET':
        # Get the logged-in user's email ID from the session
        user_email = request.session.get('user_id')

        if not user_email:
            return JsonResponse({'error': 'User not authenticated'}, status=403)

        try:
            # Fetch the employee object based on the logged-in user's email ID
            user = Employee.objects.get(emp_emailid=user_email)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)

        # If a specific quiz_id is provided
        if quiz_id:
            # Fetch quiz attempts for the given quiz by the logged-in user
            attempts = QuizAttempt.objects.filter(quiz_id=quiz_id, employee=user)
            if not attempts.exists():
                return JsonResponse({'error': 'No attempts found for this quiz'}, status=404)

            # Get the first attempt as an example
            attempt = attempts.first()

            # Fetch the quiz details
            quiz = attempt.quiz

            # Use pre-calculated total_correct, total_wrong, and total_unattempted from the database
            total_correct = attempt.total_correct
            total_wrong = attempt.total_wrong
            total_unattempted = attempt.total_unattempted

            # Get the negative mark deduction from the quiz
            negative_mark_deduction = quiz.wrong

            # Calculate total negative marks
            negative_marks = total_wrong * negative_mark_deduction

            # Calculate final score after applying negative marks
            final_score = attempt.score

            # Convert time_taken (in seconds) to hh:mm:ss format
            time_taken_seconds = attempt.time_taken
            time_taken_str = str(timedelta(seconds=time_taken_seconds))

            # Prepare the response data
            response_data = {
                'quiz_title': quiz.title,
                'full_marks': quiz.total_marks,
                'obtained_marks': float(final_score),
                'negative_marks': float(negative_marks),
                'time_taken': time_taken_str,
                'is_passed': final_score >= quiz.passing,
                'total_correct': total_correct,
                'total_wrong': total_wrong,
                'total_unattempted': total_unattempted
            }

            return JsonResponse(response_data, status=200)

        # If no quiz_id is provided, return all quiz attempts made by the logged-in user
        else:
            # Fetch all quiz attempts by the logged-in user
            attempts = QuizAttempt.objects.filter(employee=user)
            if not attempts.exists():
                return JsonResponse({'error': 'No quiz attempts found for the user'}, status=404)

            # List to store all the quiz results
            all_attempts_data = []

            for attempt in attempts:
                quiz = attempt.quiz

                # Use pre-calculated total_correct, total_wrong, and total_unattempted from the database
                total_correct = attempt.total_correct
                total_wrong = attempt.total_wrong
                total_unattempted = attempt.total_unattempted

                # Get the negative mark deduction from the quiz
                negative_mark_deduction = quiz.wrong

                # Calculate total negative marks
                negative_marks = total_wrong * negative_mark_deduction

                # Calculate final score after applying negative marks
                final_score = attempt.score

                # Convert time_taken (in seconds) to hh:mm:ss format
                time_taken_seconds = attempt.time_taken
                time_taken_str = str(timedelta(seconds=time_taken_seconds))

                # Prepare individual attempt data
                attempt_data = {
                    'quiz_title': quiz.title,
                    'full_marks': quiz.total_marks,
                    'obtained_marks': float(final_score),
                    'negative_marks': float(negative_marks),
                    'time_taken': time_taken_str,
                    'is_passed': final_score >= quiz.passing,
                    'total_correct': total_correct,
                    'total_wrong': total_wrong,
                    'total_unattempted': total_unattempted
                }

                # Add to the list of all attempts
                all_attempts_data.append(attempt_data)

            # Return all quiz attempts made by the logged-in user
            return JsonResponse({'quiz_attempts': all_attempts_data}, status=200)

    return JsonResponse({'error': 'Invalid request method'}, status=405)





@csrf_exempt
def DetailedDescription(request, quiz_id):
    # Get the logged-in user's email ID from the session
    user_email = request.session.get('user_id')
    
    if not user_email:
        return JsonResponse({'error': 'User not authenticated'}, status=403)

    try:
        # Fetch the employee object based on the logged-in user's email ID
        user = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)
    
    # Ensure it's a GET request
    if request.method == 'GET':
        try:
            # Fetch the quiz details
            quiz = Quiz.objects.get(id=quiz_id)

            # Get the quiz attempt for the logged-in employee
            attempt = QuizAttempt.objects.get(quiz=quiz, employee=user)

            # Fetch the questions and the chosen options for this quiz
            questions = []
            for question in quiz.questions.all():
                # Get the correct option for this question
                correct_option = question.options.filter(is_correct=True).first()

                # Get the selected option by the user for this question
                chosen_option_id = attempt.chosen_options.get(str(question.id))
                chosen_option = question.options.filter(id=chosen_option_id).first()

                # Append the question data along with the correct and chosen option
                questions.append({
                    'question_id': question.id,
                    'question_text': question.text,
                    'correct_option': correct_option.text if correct_option else None,  # Correct answer text
                    'chosen_option': chosen_option.text if chosen_option else None  # Chosen answer text by the user
                })

            # Convert time_taken (in seconds) to hh:mm:ss format
            time_taken_str = str(timedelta(seconds=attempt.time_taken))  # Use timedelta from the datetime module

            # Prepare the response data including marks and time
            response_data = {
                'quiz_id': quiz.id,
                'quiz_title': quiz.title,
                'questions': questions,
                'marks_obtained': float(attempt.score),  # The marks the employee got
                'total_marks': quiz.total_marks,  # Full marks for the quiz
                'time_taken': time_taken_str  # Time taken to complete the quiz in hh:mm:ss format
            }

            return JsonResponse(response_data, status=200)

        except Quiz.DoesNotExist:
            return JsonResponse({'error': 'Quiz not found'}, status=404)
        except QuizAttempt.DoesNotExist:
            return JsonResponse({'error': 'Quiz attempt not found for the employee'}, status=404)

    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)




@csrf_exempt
def AttemptedQuizzes(request):
    # Get the logged-in user's email ID from the session
    user_email = request.session.get('user_id')

    if not user_email:
        return JsonResponse({'error': 'User not authenticated'}, status=403)

    try:
        # Fetch the employee object based on the logged-in user's email ID
        user = Employee.objects.get(emp_emailid=user_email)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)

    # Ensure it's a GET request
    if request.method == 'GET':
        # Fetch all quiz attempts for the logged-in employee
        quiz_attempts = QuizAttempt.objects.filter(employee=user)

        if not quiz_attempts.exists():
            return JsonResponse({'error': 'No quizzes attempted by the user'}, status=404)

        # Prepare the list of attempted quizzes and their details
        quizzes_data = []
        for attempt in quiz_attempts:
            quiz = attempt.quiz

            # Convert time_taken (in seconds) to hh:mm:ss format
            time_taken_str = str(timedelta(seconds=attempt.time_taken))  # Use timedelta from the datetime module

            # Add the quiz and attempt details to the response

            quizzes_data.append({
                'quiz_id': quiz.id,
                'quiz_title': quiz.title,
                'total_questions': quiz.total,
                'passing_marks': quiz.passing,
                'marks_obtained': float(attempt.score),  # The marks the employee got
                'total_marks': quiz.total_marks,  # Full marks for the quiz
                'time_taken': time_taken_str,  # Time taken to complete the quiz in hh:mm:ss format
                'is_passed': attempt.is_passed,  # Whether the user passed the quiz
                'total_correct': attempt.total_correct,  # Number of correct answers
                'total_wrong': attempt.total_wrong,  # Number of wrong answers
                'attempted': (quiz.total - attempt.total_unattempted)  # Number of unattempted questions
            })

        return JsonResponse({'quizzes': quizzes_data}, status=200)

    return JsonResponse({'error': 'Invalid request method'}, status=405)
    






# Utility function to get the employee object
def get_employee(email):
    try:
        return Employee.objects.get(emp_emailid=email)
    except Employee.DoesNotExist:
        return None


        
@csrf_exempt
def investment_80C(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)

        if employee:
            # Save to ITDeclaration80c
            ITDeclaration80c.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Life Insurance Premium'),
                Investment1_Amount=data.get('amount', 0),
            )
            
            # Save or update Poifiles_new
            poifile, created = Poifiles_new1.objects.get_or_create(
                Emp_id=employee,
                defaults={
                    'Year': data.get('year', '2024'),
                    'Investment_1': '80C',
                    'actualAmount_80C': data.get('amount', 0),
                    'Doc_80C': data.get('doc', ''),
                    'Status1': 'Pending'
                }
            )
            if not created:
                poifile.actualAmount_80C = data.get('amount', 0)
                poifile.Doc_80C = data.get('doc', '')
                poifile.save()

            return JsonResponse({"message": "Investment 80C data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)




@csrf_exempt
def investment_80D(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)

        if employee:
            # Save to Itdeclaration80d_new1
            Itdeclaration80d_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Medi claim Policy for Self,Spouse,Children-80D'),
                Investment1_Amount=data.get('amount', 0),
                emp_emailid=employee,
            )
            
            # Save or update Poifiles_new
            poifile, created = Poifiles_new1.objects.get_or_create(
                Emp_id=employee,
                defaults={
                    'Year': data.get('year', '2024'),
                    'Investment_2': '80D',
                    'actualAmount_80D': data.get('amount', 0),
                    'Doc_80D': data.get('doc', ''),
                    'Status2': 'Pending'
                }
            )
            if not created:
                poifile.actualAmount_80D = data.get('amount', 0)
                poifile.Doc_80D = data.get('doc', '')
                poifile.save()

            return JsonResponse({"message": "Investment 80D data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)



        
@csrf_exempt
def other_investments(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)

        if employee:
            # Save to Itdeclaration_oie_new1
            Itdeclaration_oie_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('type', 'Additional Exemptions on voluntary NPS'),
                Investment1_Amount=data.get('amount', 0),
                emp_emailid=employee,
            )
            
            # Save or update Poifiles_new
            poifile, created = Poifiles_new1.objects.get_or_create(
                Emp_id=employee,
                defaults={
                    'Year': data.get('year', '2024'),
                    'Investment_3': 'OIE',
                    'OIE_actualAmount': data.get('amount', 0),
                    'OIE_Doc': data.get('doc', ''),
                    'Status3': 'Pending'
                }
            )
            if not created:
                poifile.OIE_actualAmount = data.get('amount', 0)
                poifile.OIE_Doc = data.get('doc', '')
                poifile.save()

            return JsonResponse({"message": "Other investments data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)
       


@csrf_exempt
def other_income(request):
    emp_emailid = request.session.get('emp_emailid')
    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'POST':
        data = json.loads(request.body)
        employee = get_employee(emp_emailid)

        if employee:
            # Save to Itdeclaration_osi_new1
            Itdeclaration_osi_new1.objects.create(
                Emp_id=employee.emp_emailid,
                Investment1=data.get('income', 'Income from other sources'),
                Investment1_Amount=data.get('amount', 0),
                Investment2_Amount=data.get('savingsInterest', 0),
                Investment3_Amount=data.get('fixedDepositInterest', 0),
                Investment4_Amount=data.get('nsCertificatesInterest', 0),
                emp_emailid=employee,
            )
            
            # Save or update Poifiles_new
            poifile, created = Poifiles_new1.objects.get_or_create(
                Emp_id=employee,
                defaults={
                    'Year': data.get('year', '2024'),
                    'Investment_4': 'OSI',
                    'OSI_actualAmount': data.get('amount', 0),
                    'OSI_Doc': data.get('doc', ''),
                    'Status4': 'Pending'
                }
            )
            if not created:
                poifile.OSI_actualAmount = data.get('amount', 0)
                poifile.OSI_Doc = data.get('doc', '')
                poifile.save()

            return JsonResponse({"message": "Other income data saved successfully"})
        else:
            return JsonResponse({"error": "Employee not found"}, status=400)




@csrf_exempt
def get_departments_by_company(request):
    c_id = request.session.get('c_id')  # Get the company ID from the session
    if c_id:
        departments = Department.objects.filter(c_id=c_id).values('d_id', 'd_name')
        return JsonResponse(list(departments), safe=False)
    else:
        return JsonResponse({'error': 'No company ID found in session'}, status=400)

@csrf_exempt
def get_employee_names_by_company(request):
    c_id = request.session.get('c_id')
    if c_id:
        employees = Employee.objects.filter(d_id__c_id=c_id).values('emp_name', 'emp_emailid')
        return JsonResponse({'employees': list(employees)}, safe=False)
    return JsonResponse({'employees': []}, safe=False)




@csrf_exempt
def AddBankDetails(request):
    emp_emailid = request.session.get('emp_emailid')  # Get email from session (logged-in user)
    
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Extract fields from the request data
            holder_name = data.get('holder_name')
            bank_name = data.get('bank_name')
            acc_no = data.get('acc_no')
            branch = data.get('branch')
            acc_type = data.get('acc_type')
            ifsc = data.get('ifsc')
            Pan_no = data.get('Pan_no')

            # Check if all required fields are present
            if not all([holder_name, bank_name, acc_no, branch, acc_type, ifsc, Pan_no]):
                return JsonResponse({'status': 'error', 'message': 'Missing required fields'}, status=400)
            
            # Retrieve the employee instance from the session email ID
            try:
                employee = Employee.objects.get(emp_emailid=emp_emailid)
            except Employee.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)

            # Create new Bank_details entry
            bank_details = Bank_details.objects.create(
                holder_name=holder_name,
                bank_name=bank_name,
                acc_no=acc_no,
                branch=branch,
                acc_type=acc_type,
                ifsc=ifsc,
                Pan_no=Pan_no,
                emp_emailid=employee  # Link to the employee using ForeignKey
            )

            bank_details.save()  # Save the instance

            return JsonResponse({'status': 'success', 'message': 'Bank details added successfully'})

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)




@csrf_exempt
def AddPersonalDetails(request):
    emp_emailid = request.session.get('emp_emailid')  # Get email from session (logged-in user)
    
    if not emp_emailid:
        return JsonResponse({'status': 'error', 'message': 'User not logged in'}, status=401)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Extract fields from the request data
            first_name = data.get('first_name')
            last_name = data.get('last_name')
            Contact = data.get('Contact')
            emergency_name = data.get('emergency_name')
            emergency_contact = data.get('emergency_contact')
            gender = data.get('gender')
            birth_date = data.get('birth_date')  # Ensure this is sent in correct date format (YYYY-MM-DD)
            address = data.get('address')
            city = data.get('city')
            district = data.get('district')
            post_code = data.get('post_code')
            state = data.get('state')

            # Check if at least one required field is present (you can modify this based on your needs)
            if not first_name:
                return JsonResponse({'status': 'error', 'message': 'First name is required'}, status=400)

            # Retrieve the employee instance from the session email ID
            try:
                employee = Employee.objects.get(emp_emailid=emp_emailid)
            except Employee.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Employee not found'}, status=404)

            # Create new Personal_details entry
            personal_details = Personal_details.objects.create(
                first_name=first_name,
                last_name=last_name,
                Contact=Contact,
                emergency_name=emergency_name,
                emergency_contact=emergency_contact,
                gender=gender,
                birth_date=birth_date,
                address=address,
                city=city,
                district=district,
                post_code=post_code,
                state=state,
                mail=employee  # Link to the employee using the ForeignKey
            )

            personal_details.save()  # Save the instance

            return JsonResponse({'status': 'success', 'message': 'Personal details added successfully'})

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@csrf_exempt
def Letters1(request):
    if 'emp_emailid' in request.session and 'c_id' in request.session:
        emp_emailid = request.session['emp_emailid']
        c_id = request.session['c_id']

        try:
            # Retrieve the company based on the logged-in user's c_id
            company = get_object_or_404(Company, pk=c_id)

            # Filter and fetch all custom letters associated with this company
            custom_letters = Custom_letters.objects.filter(c_id=company)

            # Prepare the list of letters with their names and content for the response
            allocated_letters = [
                {
                    'letter_name': letter.letter_name.replace('_', ' ').title(),
                    'letter_content': letter.letter_content
                }
                for letter in custom_letters
            ]

            return JsonResponse({'allocated_letters': allocated_letters})

        except Company.DoesNotExist:
            return JsonResponse({'error': 'Company not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    else:
        return JsonResponse({'error': 'Required session data not found'}, status=401)

# @csrf_exempt
# def task_list1(request):
#     emp_emailid = request.session.get('emp_emailid')  # Get the logged-in emp_emailid from the session

#     if not emp_emailid:
#         return JsonResponse({'error': 'User not logged in'}, status=401)

#     if request.method == 'GET':
#         try:
#             # Fetch tasks only for the logged-in user's emp_emailid
#             tasks = Todotasks1.objects.filter(emp_emailid=emp_emailid).values('id', 'title', 'completed')
#             task_list = list(tasks)
#             return JsonResponse(task_list, safe=False)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

#     elif request.method == 'POST':
#         try:
#             # Parse incoming data
#             data = json.loads(request.body)
#             title = data.get('text', '')
#             completed = data.get('isCompleted', False)

#             # Create a new task with the logged-in user's emp_emailid
#             new_task = Todotasks1(title=title, completed=completed, emp_emailid=emp_emailid)
#             new_task.save()

#             return JsonResponse({'message': 'Task created successfully'}, status=201)
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=400)

#     else:
#         return JsonResponse({'error': 'Invalid HTTP method'}, status=400)
@csrf_exempt
def task_list1(request):
    emp_emailid = request.session.get('emp_emailid')  # Get the logged-in user's email

    if not emp_emailid:
        return JsonResponse({'error': 'User not logged in'}, status=401)

    if request.method == 'GET':
        try:
            # Fetch tasks only for the logged-in user
            tasks = Todotasks1.objects.filter(emp_emailid=emp_emailid).values(
                'id', 'title', 'description', 'start_date', 'end_date', 'priority', 'completed'
            )
            return JsonResponse(list(tasks), safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            # Parse incoming JSON data
            data = json.loads(request.body)
            
            # Extract task details
            title = data.get('title', '')
            description = data.get('description', '')
            start_date = data.get('start_date', None)
            end_date = data.get('end_date', None)
            priority = data.get('priority', False)
            completed = data.get('completed', False)

            # Create and save the new task
            new_task = Todotasks1.objects.create(
                title=title,
                description=description,
                start_date=start_date,
                end_date=end_date,
                priority=priority,
                completed=completed,
                emp_emailid=emp_emailid
            )

            return JsonResponse({'message': 'Task created successfully', 'task_id': new_task.id}, status=201)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def task_update(request, task_id):
    if request.method == 'PUT':
        try:
            # Fetch the task by its ID
            task = Todotasks1.objects.get(id=task_id)
            
            # Parse the request body
            data = json.loads(request.body)
            
            # Update the task attributes
            task.title = data.get('title', task.title)
            task.completed = data.get('completed', task.completed)
            task.save()
            
            return JsonResponse({
                'message': 'Task updated successfully',
                'task': {
                    'id': task.id,
                    'title': task.title,
                    'completed': task.completed
                }
            }, status=200)
        except Todotasks1.DoesNotExist:
            return JsonResponse({'error': 'Task not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
@csrf_exempt
def task_delete(request, task_id):
    if request.method == 'DELETE':
        try:
            # Fetch the task by its ID
            task = Todotasks1.objects.get(id=task_id)
            task.delete()
            return JsonResponse({'message': 'Task deleted successfully'}, status=200)
        except Todotasks1.DoesNotExist:
            return JsonResponse({'error': 'Task not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
>>>>>>> 174a167 (backend1)
        return JsonResponse({'error': 'Invalid request method'}, status=405)